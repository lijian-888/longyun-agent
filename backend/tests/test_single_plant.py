import io
import unittest

from openpyxl import Workbook

from app.single_plant import (
    SinglePlantError,
    SinglePlantObservationRequest,
    _survey_unit_code,
    _task_code,
    parse_single_plant_package,
    parse_single_plant_workbook,
)


class SinglePlantWorkbookTests(unittest.TestCase):
    def _workbook(self, rows: list[list[object]]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "单株主表"
        for row in rows:
            sheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def test_parses_chinese_headers_after_instruction_rows(self) -> None:
        content = self._workbook([
            ["填写说明：编号必须来自主数据"],
            [],
            ["育种项目编号", "材料编号", "单株编号", "世代", "株号", "试验编号", "小区号"],
            ["PROGRAM-01", "MAT-01", "PLANT-001", "F6", 1, "TRIAL-01", "A01"],
        ])

        records = parse_single_plant_workbook(content)

        self.assertEqual(1, len(records))
        self.assertEqual("PROGRAM-01", records[0]["program_code"])
        self.assertEqual("MAT-01", records[0]["material_code"])
        self.assertEqual("PLANT-001", records[0]["sample_code"])
        self.assertEqual(4, records[0]["_source_row"])

    def test_rejects_sheet_without_required_headers(self) -> None:
        content = self._workbook([["材料名称", "备注"], ["水稻材料", "无"]])

        with self.assertRaisesRegex(SinglePlantError, "至少需要"):
            parse_single_plant_workbook(content)

    def test_parses_complete_package_with_phenotype_sheet(self) -> None:
        workbook = Workbook()
        master = workbook.active
        master.title = "单株主表"
        master.append(["育种项目编号", "材料编号", "单株编号"])
        master.append(["PROGRAM-01", "MAT-01", "PLANT-001"])
        phenotype = workbook.create_sheet("表型观测")
        phenotype.append([
            "育种项目编号", "单株编号", "观测时期", "指标编号", "指标名称", "数值", "单位"
        ])
        phenotype.append(["PROGRAM-01", "PLANT-001", "成熟期", "yield", "单株产量", 38.5, "g/株"])
        output = io.BytesIO()
        workbook.save(output)

        package = parse_single_plant_package(output.getvalue())

        self.assertEqual(1, len(package["single_plants"]))
        self.assertEqual(1, len(package["phenotypes"]))
        self.assertEqual("yield", package["phenotypes"][0]["trait_code"])


class SinglePlantContractTests(unittest.TestCase):
    def test_task_and_survey_unit_codes_are_stable(self) -> None:
        self.assertEqual(_task_code("trial-1", "分蘖期"), _task_code("trial-1", "分蘖期"))
        self.assertNotEqual(_task_code("trial-1", "分蘖期"), _task_code("trial-1", "成熟期"))
        code = _survey_unit_code("单株 / 001", "12345678-0000-0000-0000-000000000000")
        self.assertTrue(code.startswith("SP-"))
        self.assertLessEqual(len(code), 160)

    def test_observation_contract_accepts_numeric_or_text_values(self) -> None:
        numeric = SinglePlantObservationRequest(
            observation_stage="成熟期",
            trait_code="plant_height",
            trait_name="株高",
            value_numeric=102.5,
            unit="cm",
        )
        text_value = SinglePlantObservationRequest(
            observation_stage="苗期",
            trait_code="leaf_color",
            trait_name="叶色",
            value_text="深绿",
        )

        self.assertEqual(102.5, numeric.value_numeric)
        self.assertEqual("深绿", text_value.value_text)


if __name__ == "__main__":
    unittest.main()
