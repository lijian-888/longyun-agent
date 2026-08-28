import gzip
import json
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook

from app.institution_data import (
    GENOTYPE_ARCHIVE_MAX_BYTES,
    REGULAR_MAX_BYTES,
    InstitutionDataError,
    complete_suffix,
    default_access_policy,
    extract_literature_text,
    extract_plink_samples,
    extract_vcf_samples,
    normalize_records,
    parse_field_mapping,
    parse_tabular_file,
    safe_identifier,
    safe_object_name,
    stage_upload,
    upload_limit,
    validate_file_contract,
)


class InstitutionFileContractTests(unittest.TestCase):
    def test_all_required_formats_are_accepted(self):
        cases = {
            "germplasm": ["demo.csv", "demo.xlsx", "demo.json"],
            "pedigree": ["demo.csv", "demo.xlsx", "demo.json"],
            "phenotype": ["demo.csv", "demo.xlsx", "demo.json"],
            "environment": ["demo.csv", "demo.xlsx", "demo.json"],
            "genotype": ["demo.vcf", "demo.vcf.gz", "demo.zip"],
            "literature": ["demo.pdf", "demo.docx", "demo.txt"],
        }
        for dataset_type, names in cases.items():
            for name in names:
                with self.subTest(dataset_type=dataset_type, name=name):
                    self.assertTrue(validate_file_contract(dataset_type, name, 1))

    def test_regular_and_compressed_genotype_limits(self):
        self.assertEqual(upload_limit("germplasm", "demo.csv"), REGULAR_MAX_BYTES)
        self.assertEqual(upload_limit("genotype", "demo.vcf"), REGULAR_MAX_BYTES)
        self.assertEqual(upload_limit("genotype", "demo.vcf.gz"), GENOTYPE_ARCHIVE_MAX_BYTES)
        self.assertEqual(upload_limit("genotype", "demo.zip"), GENOTYPE_ARCHIVE_MAX_BYTES)
        with self.assertRaisesRegex(InstitutionDataError, "200MB"):
            validate_file_contract("literature", "paper.pdf", REGULAR_MAX_BYTES + 1)
        with self.assertRaisesRegex(InstitutionDataError, "2GB"):
            validate_file_contract("genotype", "markers.vcf.gz", GENOTYPE_ARCHIVE_MAX_BYTES + 1)

    def test_unsupported_cross_dataset_format_is_rejected(self):
        with self.assertRaisesRegex(InstitutionDataError, "不支持"):
            validate_file_contract("germplasm", "markers.vcf", 100)

    def test_complete_vcfgz_suffix(self):
        self.assertEqual(complete_suffix("DATA.VCF.GZ"), ".vcf.gz")

    def test_chinese_file_name_keeps_complete_extension(self):
        safe_name = safe_object_name("连续性状表型测试数据.xlsx")
        self.assertEqual(safe_name, "连续性状表型测试数据.xlsx")
        self.assertEqual(complete_suffix(safe_name), ".xlsx")
        self.assertTrue(validate_file_contract("germplasm", safe_name, 1024))
        self.assertEqual(complete_suffix(safe_object_name("南繁基因型.vcf.gz")), ".vcf.gz")

    def test_path_components_are_removed_without_losing_extension(self):
        self.assertEqual(safe_object_name("../../目录/材料数据.xlsx"), "材料数据.xlsx")


class InstitutionUploadStagingTests(unittest.IsolatedAsyncioTestCase):
    async def test_chinese_xlsx_upload_is_staged_with_extension(self):
        class Upload:
            filename = "连续性状表型测试数据.xlsx"

            def __init__(self, payload):
                self.payload = payload

            async def read(self, _size):
                payload, self.payload = self.payload, b""
                return payload

        with tempfile.TemporaryDirectory() as temp_dir:
            upload = await stage_upload(Upload(b"xlsx-test"), "germplasm", Path(temp_dir))
            try:
                self.assertEqual(upload.file_name, "连续性状表型测试数据.xlsx")
                self.assertEqual(upload.suffix, ".xlsx")
                self.assertEqual(upload.size_bytes, 9)
            finally:
                upload.path.unlink(missing_ok=True)


class InstitutionTabularParserTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_csv_xlsx_json_and_source_to_standard_mapping(self):
        csv_path = self.root / "germplasm.csv"
        csv_path.write_text("原材料号,名称\nG-1,材料一\n", encoding="utf-8")
        csv_rows = parse_tabular_file(csv_path, ".csv")
        normalized = normalize_records(
            "germplasm",
            csv_rows,
            {"原材料号": "germplasm_id", "名称": "name"},
        )
        self.assertEqual(normalized[0]["germplasm_id"], "G-1")
        self.assertEqual(normalized[0]["name"], "材料一")

        xlsx_path = self.root / "environment.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["环境编号", "地点", "年份"])
        sheet.append(["ENV-1", "三亚", 2026])
        workbook.save(xlsx_path)
        xlsx_rows = normalize_records("environment", parse_tabular_file(xlsx_path, ".xlsx"))
        self.assertEqual(xlsx_rows[0]["environment_id"], "ENV-1")
        self.assertEqual(xlsx_rows[0]["year"], 2026)

        phenotype_template = self.root / "连续性状表型测试数据.xlsx"
        workbook = Workbook()
        instructions = workbook.active
        instructions.title = "填写说明"
        instructions.append(["项目", "要求"])
        instructions.append(["样本标识", "不要修改"])
        data_sheet = workbook.create_sheet("连续性状表型")
        data_sheet.append(["FID", "IID", "material_code", "material_name", "analysis_environment", "trait_value"])
        data_sheet.append(["SEQ-1", "SEQ-1", "G-1", "材料一", "ENV-1", 92.4])
        workbook.save(phenotype_template)
        template_rows = normalize_records(
            "germplasm",
            parse_tabular_file(phenotype_template, ".xlsx", "germplasm"),
        )
        self.assertEqual(template_rows[0]["germplasm_id"], "G-1")
        self.assertEqual(template_rows[0]["name"], "材料一")

        json_path = self.root / "phenotype.json"
        json_path.write_text(json.dumps({"records": [{"材料编号": "G-1", "性状": "plant_height", "观测值": 100}]}), encoding="utf-8")
        json_rows = normalize_records("phenotype", parse_tabular_file(json_path, ".json"))
        self.assertEqual(json_rows[0]["germplasm_id"], "G-1")
        self.assertEqual(json_rows[0]["trait_code"], "plant_height")

    def test_mapping_contract_rejects_non_string_values(self):
        self.assertEqual(parse_field_mapping('{"原材料号":"germplasm_id"}'), {"原材料号": "germplasm_id"})
        with self.assertRaisesRegex(InstitutionDataError, "字符串对象"):
            parse_field_mapping('{"原材料号":1}')

    def test_mapping_rejects_unknown_standard_target(self):
        with self.assertRaisesRegex(InstitutionDataError, "不支持的标准字段"):
            normalize_records("germplasm", [{"原材料号": "G-1"}], {"原材料号": "unknown_field"})


class InstitutionGenotypeAndLiteratureTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_vcf_vcfgz_and_plink_samples(self):
        content = "##fileformat=VCFv4.2\n#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tG-1\tG-2\n1\t1\t.\tA\tG\t.\tPASS\t.\tGT\t0/1\t0/0\n"
        vcf_path = self.root / "demo.vcf"
        vcf_path.write_text(content, encoding="utf-8")
        self.assertEqual(extract_vcf_samples(vcf_path, False), ["G-1", "G-2"])
        gz_path = self.root / "demo.vcf.gz"
        with gzip.open(gz_path, "wt", encoding="utf-8") as handle:
            handle.write(content)
        self.assertEqual(extract_vcf_samples(gz_path, True), ["G-1", "G-2"])

        zip_path = self.root / "demo.zip"
        with zipfile.ZipFile(zip_path, "w") as archive:
            archive.writestr("demo.bed", b"\x6c\x1b\x01")
            archive.writestr("demo.bim", "1 rs1 0 1 A G\n")
            archive.writestr("demo.fam", "F1 G-1 0 0 0 -9\nF2 G-2 0 0 0 -9\n")
        self.assertEqual(extract_plink_samples(zip_path), ["G-1", "G-2"])

    def test_txt_and_docx_literature(self):
        txt_path = self.root / "paper.txt"
        txt_path.write_text("公开水稻育种文献", encoding="utf-8")
        value, warnings = extract_literature_text(txt_path, ".txt")
        self.assertIn("水稻育种", value)
        self.assertEqual(warnings, [])

        docx_path = self.root / "paper.docx"
        with zipfile.ZipFile(docx_path, "w") as archive:
            archive.writestr("word/document.xml", "<w:document><w:p><w:t>南繁公开文献</w:t></w:p></w:document>")
        value, warnings = extract_literature_text(docx_path, ".docx")
        self.assertIn("南繁公开文献", value)
        self.assertEqual(warnings, [])

    def test_policy_and_identifiers_are_deterministic(self):
        self.assertEqual(safe_identifier("HNNF", "longyun"), "longyun_hnnf")
        policy = default_access_policy("hainan-nanfan", "longyun-hnnf")
        self.assertEqual(policy["effect"], "private")
        self.assertEqual(policy["principal"], "institution:hainan-nanfan")


if __name__ == "__main__":
    unittest.main()
