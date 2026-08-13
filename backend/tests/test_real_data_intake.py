from __future__ import annotations

import json
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook

from app.real_data_intake import IntakeError, parse_structured_file


class RealInstitutionDataParserTests(unittest.TestCase):
    def test_csv_preserves_institution_headers_and_deduplicates_repeated_columns(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "phenotype.csv"
            path.write_text(
                "材料编号,株高,株高\nYN-001,101.2,复测101.0\n",
                encoding="utf-8-sig",
            )

            tables = parse_structured_file(path, path.name)

        self.assertEqual(len(tables), 1)
        self.assertEqual(tables[0].columns, ("材料编号", "株高", "株高_2"))
        self.assertEqual(tables[0].rows[0][0], 2)
        self.assertEqual(tables[0].rows[0][1]["材料编号"], "YN-001")
        self.assertEqual(tables[0].rows[0][1]["株高_2"], "复测101.0")

    def test_json_unions_non_uniform_institution_fields_without_dropping_values(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "trial.json"
            path.write_text(
                json.dumps(
                    [
                        {"试验编号": "TRIAL-2026-01", "地点": "昆明"},
                        {"试验编号": "TRIAL-2026-01", "海拔": 1890},
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            tables = parse_structured_file(path, path.name)

        self.assertEqual(tables[0].columns, ("试验编号", "地点", "海拔"))
        self.assertEqual(tables[0].rows[1][1]["地点"], None)
        self.assertEqual(tables[0].rows[1][1]["海拔"], 1890)

    def test_xlsx_reads_each_non_empty_sheet_as_an_independent_table(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "institution-data.xlsx"
            workbook = Workbook()
            phenotype = workbook.active
            phenotype.title = "表型"
            phenotype.append(["材料编号", "产量"])
            phenotype.append(["YN-001", 612.5])
            environment = workbook.create_sheet("环境")
            environment.append(["试验编号", "年份", "地点"])
            environment.append(["TRIAL-2026-01", 2026, "昆明"])
            workbook.create_sheet("空表")
            workbook.save(path)
            workbook.close()

            tables = parse_structured_file(path, path.name)

        self.assertEqual([table.sheet_name for table in tables], ["表型", "环境"])
        self.assertEqual(tables[0].rows[0][1]["产量"], 612.5)
        self.assertEqual(tables[1].rows[0][1]["年份"], 2026)

    def test_unsupported_file_is_left_for_non_structured_ingestion(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "paper.pdf"
            path.write_bytes(b"%PDF-placeholder")
            self.assertEqual(parse_structured_file(path, path.name), [])

    def test_invalid_json_returns_a_safe_intake_error(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "bad.json"
            path.write_text("{broken", encoding="utf-8")
            with self.assertRaises(IntakeError):
                parse_structured_file(path, path.name)


if __name__ == "__main__":
    unittest.main()
