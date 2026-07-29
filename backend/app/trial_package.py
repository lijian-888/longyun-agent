"""Regional-trial package governance and controlled research evidence.

The first version deliberately accepts a ZIP package rather than pretending
that scattered research files are already one clean spreadsheet.  It keeps the
original archive, records each source row, creates a reviewable staging record,
and only writes trial-level standard records after a data processor confirms
publication.

This module does not make causal claims.  Its environmental and management
outputs are descriptive associations based on the published trial records.
"""

from __future__ import annotations

import hashlib
import io
import json
import math
import re
import statistics
import uuid
import zipfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from .trial_demo import ensure_trial_demo_schema
from .trial_statistics import TrialStatisticsError, run_controlled_trial_analysis


MAX_ARCHIVE_BYTES = 50 * 1024 * 1024
MAX_ARCHIVE_FILES = 80
TRIAL_TRAITS: dict[str, tuple[str, str, str]] = {
    "yield_per_mu": ("亩产", "产量表现", "kg/亩"),
    "plant_height": ("株高", "农艺性状", "cm"),
    "thousand_grain_weight": ("千粒重", "产量构成", "g"),
    "seed_setting_rate": ("结实率", "产量构成", "%"),
    "head_rice_rate": ("整精米率", "加工品质", "%"),
    "chalkiness_degree": ("垩白度", "外观品质", "%"),
    "panicle_blast_score": ("穗瘟等级", "抗病性", "级"),
    "lodging_score": ("倒伏等级", "抗倒伏性", "级"),
}
ENVIRONMENT_METRICS: dict[str, tuple[str, str]] = {
    "soil_ph": ("土壤 pH", ""),
    "available_phosphorus": ("有效磷", "mg/kg"),
    "organic_matter": ("有机质", "g/kg"),
    "rainfall": ("生育期降雨量", "mm"),
    "mean_temperature": ("平均温度", "℃"),
    "disease_pressure": ("病害压力", "级"),
}


def _new_id() -> str:
    return str(uuid.uuid4())


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _as_json(value: Any, fallback: Any) -> Any:
    if value is None:
        return fallback
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return fallback


def _clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _key(value: Any) -> str:
    return re.sub(r"[\s()（）\[\]【】_\-—:/：,.，]+", "", _clean_text(value)).lower()


def _number(value: Any) -> float | None:
    if value is None or _clean_text(value) == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", _clean_text(value))
    return float(match.group(0)) if match else None


def _year_from_name(file_name: str) -> int | None:
    match = re.search(r"(20\d{2})", file_name)
    return int(match.group(1)) if match else None


def _safe_file_name(name: str) -> str:
    safe = Path(name).name
    return re.sub(r"[^\w.\-\u4e00-\u9fff]+", "_", safe) or "trial-package.zip"


def _unit_number(value: Any, unit: str) -> tuple[float | None, str]:
    number = _number(value)
    if number is None:
        return None, ""
    raw = _clean_text(value).lower()
    if unit == "kg/亩" and ("kg/ha" in raw or "公斤/公顷" in raw or "千克/公顷" in raw):
        return round(number / 15, 3), "kg/ha 已按 1 亩 = 1/15 公顷换算为 kg/亩"
    if unit == "%" and 0 < number <= 1 and ("." in raw or isinstance(value, float)):
        return round(number * 100, 3), "小数比例已换算为百分比"
    return number, ""


def _treatment_code(value: Any) -> tuple[str, str]:
    raw = _clean_text(value)
    normalized = _key(raw)
    if normalized in {"m1", "常规n", "标准施氮", "标准n", "常规施氮"} or any(token in raw for token in ("标准", "常规")):
        return "M1", "标准施氮"
    if normalized in {"m2", "高n", "较高施氮", "高施氮", "高氮处理"} or any(token in raw for token in ("较高", "高氮", "高施")):
        return "M2", "较高施氮"
    return "M-UNKNOWN", raw or "未识别处理"


SITE_ALIASES = {
    _key("南昌试验点"): ("NC", "南昌试验点"),
    _key("赣州试验点"): ("GZ", "赣州试验点"),
    _key("九江试验点"): ("JJ", "九江试验点"),
    _key("抚州试验点"): ("FZ", "抚州试验点"),
}

MATERIAL_ALIASES: dict[str, tuple[str, str, bool, list[str]]] = {}
for index in range(1, 9):
    code = f"ME-A{index:02d}"
    name = f"候选A-{index:02d}"
    aliases = [name, f"候选A{index:02d}", f"HZ-{index:02d}", f"HZ{index:02d}", code]
    for alias in aliases:
        MATERIAL_ALIASES[_key(alias)] = (code, name, False, aliases)

# The sample deliberately uses different names for the same material across
# years.  These aliases exercise the material-mapping step that real regional
# trial packages need before multi-year comparisons are trustworthy.
for code, extra_aliases in {
    "ME-A02": ["高产2号"],
    "ME-A03": ["稳产3号"],
    "ME-A04": ["优质4号"],
    "ME-A05": ["耐酸5号"],
}.items():
    existing = MATERIAL_ALIASES[_key(code)]
    merged_aliases = list(dict.fromkeys([*existing[3], *extra_aliases]))
    for alias in merged_aliases:
        MATERIAL_ALIASES[_key(alias)] = (existing[0], existing[1], existing[2], merged_aliases)
for code, name, aliases in (
    ("CK-01", "对照CK01", ["对照CK01", "对照一号", "CK1", "CK-01"]),
    ("CK-02", "对照CK02", ["对照CK02", "对照二号", "CK2", "CK-02"]),
):
    for alias in aliases:
        MATERIAL_ALIASES[_key(alias)] = (code, name, True, list(aliases))


def _material_identity(raw_name: Any) -> tuple[str, str, bool, list[str], bool]:
    raw = _clean_text(raw_name)
    known = MATERIAL_ALIASES.get(_key(raw))
    if known:
        return (*known, False)
    token = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff]+", "", raw).upper()[:44] or "UNNAMED"
    return f"MAT-{token}", raw or "未命名材料", False, [raw] if raw else [], True


def _site_identity(raw_site: Any) -> tuple[str, str]:
    name = _clean_text(raw_site)
    known = SITE_ALIASES.get(_key(name))
    if known:
        return known
    token = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff]+", "", name).upper()[:32] or "UNKNOWN"
    return f"SITE-{token}", name or "未命名试验点"


def _first_row_value(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        if _key(alias) in row and row[_key(alias)] not in (None, ""):
            return row[_key(alias)]
    return None


def _sheet_rows(file_name: str, content: bytes) -> tuple[str, list[dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        values = sheet.iter_rows(values_only=True)
        headers = next(values, None)
        if not headers:
            continue
        header_keys = [_key(item) for item in headers]
        if not any(header_keys):
            continue
        rows: list[dict[str, Any]] = []
        for row_number, values_row in enumerate(values, start=2):
            if not any(value not in (None, "") for value in values_row):
                continue
            data = {header_keys[index]: values_row[index] if index < len(values_row) else None for index in range(len(header_keys))}
            data["__row__"] = row_number
            rows.append(data)
        return sheet.title, rows
    raise ValueError(f"{file_name} 没有可解析的数据工作表")


def _file_role(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return "unknown"
    headers = set(rows[0])
    if any(item in headers for item in (_key("小区号"), _key("区组"))) and any(item in headers for item in (_key("重复"), _key("重复号"), _key("rep"))):
        return "layout"
    if any(item in headers for item in (_key("pH"), _key("速效磷(mg/kg)"), _key("生态区"))):
        return "environment"
    # A phenotype table can also contain a nitrogen-treatment column.  Check
    # yield/plant-height columns before treating it as a management table.
    if any(item in headers for item in (_key("产量(kg/亩)"), _key("产量kg/ha"), _key("亩产"), _key("株高cm"))):
        return "phenotype"
    if any(item in headers for item in (_key("施氮量"), _key("氮肥用量(kg/ha)"), _key("氮处理"))):
        return "management"
    return "unknown"


def _parse_archive(content: bytes) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = []
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        names = [name for name in archive.namelist() if not name.endswith("/")]
        xlsx_names = [name for name in names if Path(name).suffix.lower() in {".xlsx", ".xls"}]
        if len(names) > MAX_ARCHIVE_FILES:
            raise ValueError(f"资料包内文件超过 {MAX_ARCHIVE_FILES} 个，请先拆分后再导入")
        if not xlsx_names:
            raise ValueError("资料包中未找到 Excel 文件。第一版区域试验资料包需要包含 Excel 原始表")

        source_files: list[dict[str, Any]] = []
        layouts: dict[tuple[int, str, str, str, int], dict[str, Any]] = {}
        environments: dict[tuple[int, str], dict[str, Any]] = {}
        management: dict[tuple[int, str, str], dict[str, Any]] = {}
        phenotypes: list[dict[str, Any]] = []
        materials: dict[str, dict[str, Any]] = {}
        sites: dict[str, dict[str, Any]] = {}

        for member in xlsx_names:
            file_bytes = archive.read(member)
            if len(file_bytes) > 25 * 1024 * 1024:
                raise ValueError(f"{Path(member).name} 超过 25 MB，请拆分后再导入")
            year = _year_from_name(Path(member).name)
            if not year:
                warnings.append(f"{Path(member).name} 未从文件名识别出试验年份，已跳过")
                continue
            sheet_name, rows = _sheet_rows(member, file_bytes)
            role = _file_role(rows)
            source_files.append({
                "file_name": Path(member).name,
                "source_role": role,
                "source_format": Path(member).suffix.lstrip(".").lower(),
                "relative_path": member,
                "sheet_name": sheet_name,
                "checksum": hashlib.sha256(file_bytes).hexdigest(),
                "row_count": len(rows),
            })
            if role == "unknown":
                warnings.append(f"{Path(member).name} 的表头暂未识别为布局、环境、管理或表型表，未写入标准数据")
                continue

            for row in rows:
                locator = f"{Path(member).name}#{sheet_name}!{row['__row__']}"
                if role == "layout":
                    raw_site = _first_row_value(row, ("试验地点", "试验点", "地点名称", "测试地点"))
                    raw_material = _first_row_value(row, ("材料名称/材料代号", "材料名称", "材料代号", "供试材料"))
                    treatment_raw = _first_row_value(row, ("处理", "N处理", "氮处理", "处理名称"))
                    rep_raw = _first_row_value(row, ("重复", "重复号", "Rep"))
                    block_raw = _first_row_value(row, ("区组", "区组号", "Block"))
                    plot_raw = _first_row_value(row, ("小区号", "小区", "Plot"))
                    parsed_year = int(_number(_first_row_value(row, ("试验年份", "年份"))) or year)
                    if not raw_site or not raw_material or not treatment_raw or not rep_raw:
                        warnings.append(f"{locator} 缺少地点、材料、处理或重复，未作为参试记录导入")
                        continue
                    site_code, site_name = _site_identity(raw_site)
                    material_code, material_name, is_check, aliases, needs_mapping = _material_identity(raw_material)
                    if needs_mapping:
                        warnings.append(f"{locator} 的材料“{raw_material}”没有命中材料别名表，已保留为待确认材料")
                    treatment_code, treatment_name = _treatment_code(treatment_raw)
                    rep = int(_number(rep_raw) or 0)
                    block = int(_number(block_raw) or rep or 1)
                    if rep < 1:
                        warnings.append(f"{locator} 的重复号无法识别，未作为参试记录导入")
                        continue
                    materials.setdefault(material_code, {"material_code": material_code, "material_name": material_name, "is_check": is_check, "aliases": aliases})
                    sites.setdefault(site_code, {"site_code": site_code, "site_name": site_name})
                    layouts[(parsed_year, site_code, material_code, treatment_code, rep)] = {
                        "trial_key": f"{parsed_year}-{site_code}",
                        "site_code": site_code,
                        "site_name": site_name,
                        "year": parsed_year,
                        "material_code": material_code,
                        "treatment_code": treatment_code,
                        "treatment_name": treatment_name,
                        "replicate_no": rep,
                        "block_no": block,
                        "plot_no": _clean_text(plot_raw) or f"{site_code}-{parsed_year}-{material_code}-{treatment_code}-{rep}",
                        "raw_material_name": _clean_text(raw_material),
                        "source_locator": locator,
                        "traits": [],
                    }
                elif role == "environment":
                    raw_site = _first_row_value(row, ("测试地点", "试验地点", "试验点", "地点名称"))
                    if not raw_site:
                        warnings.append(f"{locator} 缺少试验地点，未作为环境记录导入")
                        continue
                    site_code, site_name = _site_identity(raw_site)
                    sites.setdefault(site_code, {"site_code": site_code, "site_name": site_name})
                    metric_values = {
                        "soil_ph": _first_row_value(row, ("pH", "土壤pH")),
                        "available_phosphorus": _first_row_value(row, ("速效磷(mg/kg)", "速效磷", "有效磷")),
                        "organic_matter": _first_row_value(row, ("有机质", "有机质(g/kg)")),
                        "rainfall": _first_row_value(row, ("生育期降雨", "降雨量", "生育期降雨量")),
                        "mean_temperature": _first_row_value(row, ("平均温度", "生育期平均温度")),
                        "disease_pressure": _first_row_value(row, ("病害压力", "病害等级")),
                    }
                    numeric_metrics = []
                    for code, raw in metric_values.items():
                        value = _number(raw)
                        if value is not None:
                            numeric_metrics.append({"metric_code": code, "value_numeric": value, "original_value": _clean_text(raw), "source_locator": locator})
                    environments[(year, site_code)] = {
                        "trial_key": f"{year}-{site_code}",
                        "site_code": site_code,
                        "site_name": site_name,
                        "county": _clean_text(_first_row_value(row, ("县区", "县市区"))),
                        "ecological_zone": _clean_text(_first_row_value(row, ("生态区", "生态区域"))) or "未填写生态区",
                        "soil_type": _clean_text(_first_row_value(row, ("土类", "土壤类型"))),
                        "metrics": numeric_metrics,
                    }
                elif role == "management":
                    raw_site = _first_row_value(row, ("试验点", "测试地点", "试验地点", "地点名称"))
                    treatment_raw = _first_row_value(row, ("处理名称", "氮处理", "处理", "N处理"))
                    raw_rate = _first_row_value(row, ("施氮量", "氮肥用量(kg/ha)", "氮肥用量", "施氮"))
                    if not raw_site or not treatment_raw:
                        warnings.append(f"{locator} 缺少地点或处理，未作为管理记录导入")
                        continue
                    site_code, site_name = _site_identity(raw_site)
                    treatment_code, treatment_name = _treatment_code(treatment_raw)
                    raw_rate_text = _clean_text(raw_rate)
                    rate = _number(raw_rate)
                    if "ha" in _key(" ".join(row.keys())) or "kg/ha" in raw_rate_text.lower():
                        if rate is not None:
                            rate = round(rate / 15, 3)
                            warnings.append(f"{locator} 的施氮量已由 kg/ha 换算为 kg/亩")
                    management[(year, site_code, treatment_code)] = {
                        "trial_key": f"{year}-{site_code}",
                        "site_code": site_code,
                        "site_name": site_name,
                        "treatment_code": treatment_code,
                        "treatment_name": treatment_name,
                        "nitrogen_rate": rate,
                        "raw_nitrogen_rate": raw_rate_text,
                        "fertilization_stage": _clean_text(_first_row_value(row, ("施肥时期", "施肥方式"))),
                        "water_management": _clean_text(_first_row_value(row, ("水分管理", "灌溉方式"))),
                        "planting_density": _clean_text(_first_row_value(row, ("种植密度", "栽培密度"))),
                        "source_locator": locator,
                    }
                elif role == "phenotype":
                    raw_site = _first_row_value(row, ("地点名称", "地点", "试验点", "试验地点", "测试地点"))
                    raw_material = _first_row_value(row, ("供试材料", "材料", "材料代号", "材料名称", "材料名称/材料代号"))
                    treatment_raw = _first_row_value(row, ("N处理", "氮处理", "处理", "处理名称", "氮处理"))
                    rep_raw = _first_row_value(row, ("重复号", "重复", "Rep"))
                    if not raw_site or not raw_material or not treatment_raw or not rep_raw:
                        warnings.append(f"{locator} 缺少地点、材料、处理或重复，未作为表型记录导入")
                        continue
                    site_code, site_name = _site_identity(raw_site)
                    material_code, material_name, is_check, aliases, needs_mapping = _material_identity(raw_material)
                    if needs_mapping:
                        warnings.append(f"{locator} 的表型材料“{raw_material}”没有命中材料别名表")
                    materials.setdefault(material_code, {"material_code": material_code, "material_name": material_name, "is_check": is_check, "aliases": aliases})
                    sites.setdefault(site_code, {"site_code": site_code, "site_name": site_name})
                    treatment_code, _ = _treatment_code(treatment_raw)
                    rep = int(_number(rep_raw) or 0)
                    metric_columns = {
                        "yield_per_mu": ("产量(kg/亩)", "产量kg/ha", "亩产", "产量"),
                        "plant_height": ("株高cm", "株高", "株高(cm)"),
                        "thousand_grain_weight": ("千粒重", "粒重", "千粒质量", "千粒质量克"),
                        "seed_setting_rate": ("结实率", "结实率(%)"),
                        "head_rice_rate": ("整精米率%", "整精米率", "整精米"),
                        "chalkiness_degree": ("垩白度", "垩白度%"),
                        "panicle_blast_score": ("穗瘟", "穗瘟等级"),
                        "lodging_score": ("倒伏等级", "倒伏级", "倒伏"),
                    }
                    traits = []
                    for trait_code, aliases_for_trait in metric_columns.items():
                        raw = _first_row_value(row, aliases_for_trait)
                        trait_name, category, unit = TRIAL_TRAITS[trait_code]
                        value, conversion_note = _unit_number(raw, unit)
                        if trait_code == "yield_per_mu" and raw is not None and (_key("产量kg/ha") in row or "kg/ha" in _clean_text(raw).lower()):
                            raw_value = _number(raw)
                            value = round(raw_value / 15, 3) if raw_value is not None else None
                            conversion_note = "kg/ha 已按 1 亩 = 1/15 公顷换算为 kg/亩"
                        if value is None:
                            continue
                        traits.append({"trait_code": trait_code, "trait_name": trait_name, "trait_category": category, "value_numeric": value, "unit": unit, "original_value": _clean_text(raw), "source_locator": locator, "conversion_note": conversion_note})
                    phenotypes.append({
                        "year": year,
                        "site_code": site_code,
                        "site_name": site_name,
                        "material_code": material_code,
                        "treatment_code": treatment_code,
                        "replicate_no": rep,
                        "traits": traits,
                        "source_locator": locator,
                    })

    if not layouts:
        raise ValueError("未解析到材料参试与小区布局数据。资料包需要至少包含地点、材料、处理、重复和小区号")
    if not phenotypes:
        raise ValueError("未解析到农艺/品质表型数据。资料包需要至少包含产量、株高等观测列")

    entries: list[dict[str, Any]] = []
    unmatched = 0
    for phenotype in phenotypes:
        match_key = (phenotype["year"], phenotype["site_code"], phenotype["material_code"], phenotype["treatment_code"], phenotype["replicate_no"])
        layout = layouts.get(match_key)
        if not layout:
            unmatched += 1
            warnings.append(f"{phenotype['source_locator']} 未找到同年、同点、同材料、同处理、同重复的小区布局，未入库")
            continue
        layout["traits"].extend(phenotype["traits"])
    for layout in layouts.values():
        if layout["traits"]:
            entries.append(layout)

    if unmatched:
        warnings.append(f"共有 {unmatched} 行表型记录未能匹配到材料参试布局")

    trials = {}
    for layout in entries:
        trial_key = layout["trial_key"]
        trials.setdefault(trial_key, {
            "trial_key": trial_key,
            "trial_year": layout["year"],
            "site_code": layout["site_code"],
            "site_name": layout["site_name"],
            "entries": 0,
        })["entries"] += 1

    payload = {
        "materials": list(materials.values()),
        "sites": list(sites.values()),
        "trials": list(trials.values()),
        "environment": list(environments.values()),
        "management": list(management.values()),
        "entries": entries,
        "source_files": source_files,
    }
    return payload, warnings


def _validate_rcbd_payload(payload: dict[str, Any]) -> dict[str, Any]:
    """Check whether the staged records can support the first RCBD analyses.

    This deliberately validates the smallest analysis unit: one material under
    one treatment in one block of one field trial.  It does not rewrite the
    uploaded rows.  A data processor can see exactly why a package is blocked
    before the package is published into the formal trial tables.
    """
    by_trial: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in payload.get("entries", []):
        by_trial[entry["trial_key"]].append(entry)

    trial_results: list[dict[str, Any]] = []
    blocking_issues: list[str] = []
    warnings: list[str] = []
    environment_keys = {item.get("trial_key") for item in payload.get("environment", [])}
    management_keys = {item.get("trial_key") for item in payload.get("management", [])}

    for trial in payload.get("trials", []):
        trial_key = trial["trial_key"]
        entries = by_trial.get(trial_key, [])
        blocks = sorted({int(item.get("block_no") or 0) for item in entries if int(item.get("block_no") or 0) > 0})
        materials = sorted({str(item.get("material_code") or "") for item in entries if item.get("material_code")})
        treatments = sorted({str(item.get("treatment_code") or "") for item in entries if item.get("treatment_code")})
        expected = len(blocks) * len(materials) * len(treatments)
        combinations: dict[tuple[str, str, int], list[dict[str, Any]]] = defaultdict(list)
        plot_numbers: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for entry in entries:
            key = (str(entry.get("material_code") or ""), str(entry.get("treatment_code") or ""), int(entry.get("block_no") or 0))
            combinations[key].append(entry)
            plot_numbers[str(entry.get("plot_no") or "")].append(entry)
        expected_combinations = {(material, treatment, block) for material in materials for treatment in treatments for block in blocks}
        actual_combinations = set(combinations)
        missing = sorted(expected_combinations - actual_combinations)
        duplicate = sorted(key for key, values in combinations.items() if len(values) > 1)
        duplicate_plots = sorted(plot for plot, values in plot_numbers.items() if plot and len(values) > 1)
        issues: list[str] = []
        if len(blocks) < 2:
            issues.append("有效区组少于 2 个，无法进行随机区组误差估计")
        if len(materials) < 2:
            issues.append("参试材料少于 2 个，无法进行材料比较")
        if not treatments:
            issues.append("未识别处理，无法建立材料 × 管理分析单元")
        if missing:
            issues.append(f"缺少 {len(missing)} 个材料 × 处理 × 区组组合")
        if duplicate:
            issues.append(f"存在 {len(duplicate)} 个重复材料 × 处理 × 区组组合")
        if duplicate_plots:
            issues.append(f"存在 {len(duplicate_plots)} 个重复小区号")
        status = "passed" if not issues else "blocked"
        trial_result = {
            "trial_key": trial_key,
            "trial_year": trial.get("trial_year"),
            "site_name": trial.get("site_name"),
            "design_type": "随机区组设计（材料 × 施氮）",
            "block_count": len(blocks),
            "material_count": len(materials),
            "treatment_count": len(treatments),
            "expected_entry_count": expected,
            "observed_entry_count": len(entries),
            "missing_combination_count": len(missing),
            "duplicate_combination_count": len(duplicate),
            "duplicate_plot_count": len(duplicate_plots),
            "missing_examples": [" / ".join(map(str, item)) for item in missing[:5]],
            "duplicate_examples": [" / ".join(map(str, item)) for item in duplicate[:5]],
            "status": status,
            "issues": issues,
        }
        trial_results.append(trial_result)
        blocking_issues.extend(f"{trial_key}：{item}" for item in issues)
        if trial_key not in environment_keys:
            warnings.append(f"{trial_key} 缺少环境与土壤记录：可发布基础试验数据，但不能完成环境影响分析")
        if trial_key not in management_keys:
            warnings.append(f"{trial_key} 缺少管理记录：可发布基础试验数据，但不能完成施氮或管理措施影响分析")
        if len(treatments) < 2:
            warnings.append(f"{trial_key} 仅识别 {len(treatments)} 个处理：可进行材料比较，不能进行材料 × 施氮交互分析")

    if not trial_results:
        blocking_issues.append("未形成可核验的试验记录")
    return {
        "template": "水稻区域试验：随机区组设计（材料 × 施氮）v1.0",
        "overall_status": "passed" if not blocking_issues else "blocked",
        "blocking_issues": blocking_issues,
        "warnings": warnings,
        "trial_results": trial_results,
    }


def ensure_trial_package_schema(session: Session) -> None:
    """Create the staging layer while preserving the trial-level standard tables."""
    ensure_trial_demo_schema(session)
    session.execute(text("""
        CREATE TABLE IF NOT EXISTS trial_import_batch (
            id VARCHAR(36) PRIMARY KEY,
            display_name VARCHAR(300) NOT NULL,
            archive_name VARCHAR(500) NOT NULL,
            archive_path TEXT NOT NULL,
            archive_sha256 VARCHAR(128) NOT NULL,
            uploaded_by VARCHAR(200) NOT NULL,
            parse_status VARCHAR(40) NOT NULL DEFAULT 'uploaded',
            parse_summary JSONB NOT NULL DEFAULT '{}'::jsonb,
            validation_report JSONB NOT NULL DEFAULT '{}'::jsonb,
            parsed_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            warnings JSONB NOT NULL DEFAULT '[]'::jsonb,
            error_message TEXT,
            published_package_id VARCHAR(36) REFERENCES trial_data_package(id),
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            parsed_at TIMESTAMPTZ,
            published_at TIMESTAMPTZ
        )
    """))
    session.execute(text("ALTER TABLE trial_import_batch ADD COLUMN IF NOT EXISTS validation_report JSONB NOT NULL DEFAULT '{}'::jsonb"))
    session.execute(text("CREATE INDEX IF NOT EXISTS ix_trial_import_batch_status ON trial_import_batch(parse_status, created_at DESC)"))
    session.commit()


def retire_legacy_seeded_trial_demo(session: Session) -> None:
    """Hide the former auto-seeded parallel demo from the real import flow."""
    session.execute(text("""
        UPDATE trial_data_package
        SET governance_status = 'legacy_seed_hidden'
        WHERE package_code = 'RICE-MET-2023-2025-DEMO'
          AND is_simulated = TRUE
          AND governance_status <> 'legacy_seed_hidden'
    """))
    session.commit()


def _serialize_batch(row: Any, include_payload: bool = False) -> dict[str, Any]:
    data = dict(row)
    payload = _as_json(data.get("parsed_payload"), {})
    summary = _as_json(data.get("parse_summary"), {})
    result = {
        "id": str(data["id"]),
        "display_name": data["display_name"],
        "archive_name": data["archive_name"],
        "uploaded_by": data["uploaded_by"],
        "parse_status": data["parse_status"],
        "parse_summary": summary,
        "validation_report": _as_json(data.get("validation_report"), {}),
        "warnings": _as_json(data.get("warnings"), []),
        "error_message": data.get("error_message"),
        "published_package_id": str(data["published_package_id"]) if data.get("published_package_id") else None,
        "created_at": data["created_at"].isoformat() if data.get("created_at") else None,
        "parsed_at": data["parsed_at"].isoformat() if data.get("parsed_at") else None,
        "published_at": data["published_at"].isoformat() if data.get("published_at") else None,
    }
    if include_payload:
        result["preview"] = {
            "source_files": payload.get("source_files", []),
            "trials": payload.get("trials", []),
            "materials": payload.get("materials", []),
            "environment_count": len(payload.get("environment", [])),
            "management_count": len(payload.get("management", [])),
            "entry_preview": payload.get("entries", [])[:12],
        }
    return result


def list_trial_import_batches(session: Session) -> list[dict[str, Any]]:
    rows = session.execute(text("SELECT * FROM trial_import_batch ORDER BY created_at DESC")).mappings().all()
    return [_serialize_batch(row) for row in rows]


def get_trial_import_batch(session: Session, batch_id: str) -> dict[str, Any]:
    row = session.execute(text("SELECT * FROM trial_import_batch WHERE id = :id"), {"id": batch_id}).mappings().first()
    if not row:
        raise ValueError("未找到该区域试验资料包")
    return _serialize_batch(row, include_payload=True)


def upload_trial_package(session: Session, file_name: str, content: bytes, actor: str, raw_storage_dir: Path) -> dict[str, Any]:
    if not file_name.lower().endswith(".zip"):
        raise ValueError("区域试验资料包第一版仅支持 ZIP。请将材料布局、环境土壤、管理和表型 Excel 打包后上传")
    if len(content) > MAX_ARCHIVE_BYTES:
        raise ValueError("资料包超过 50 MB，请按年份或试验组拆分后再上传")
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            if any(Path(name).is_absolute() or ".." in Path(name).parts for name in archive.namelist()):
                raise ValueError("资料包包含不安全的内部路径")
    except zipfile.BadZipFile as exc:
        raise ValueError("上传文件不是有效的 ZIP 资料包") from exc

    batch_id = _new_id()
    safe_name = _safe_file_name(file_name)
    target_dir = raw_storage_dir / "trial-packages" / batch_id
    target_dir.mkdir(parents=True, exist_ok=True)
    archive_path = target_dir / safe_name
    archive_path.write_bytes(content)
    digest = hashlib.sha256(content).hexdigest()
    session.execute(text("""
        INSERT INTO trial_import_batch (id, display_name, archive_name, archive_path, archive_sha256, uploaded_by, parse_status)
        VALUES (:id, :display_name, :archive_name, :archive_path, :archive_sha256, :uploaded_by, 'parsing')
    """), {
        "id": batch_id,
        "display_name": Path(file_name).stem,
        "archive_name": safe_name,
        "archive_path": str(archive_path),
        "archive_sha256": digest,
        "uploaded_by": actor,
    })
    session.commit()

    try:
        payload, warnings = _parse_archive(content)
        validation_report = _validate_rcbd_payload(payload)
        warnings.extend(validation_report["warnings"])
        warnings.extend(f"随机区组核验未通过：{item}" for item in validation_report["blocking_issues"])
        summary = {
            "source_file_count": len(payload["source_files"]),
            "material_count": len(payload["materials"]),
            "trial_count": len(payload["trials"]),
            "environment_count": len(payload["environment"]),
            "management_count": len(payload["management"]),
            "entry_count": len(payload["entries"]),
            "observation_count": sum(len(item.get("traits", [])) for item in payload["entries"]),
            "unresolved_count": sum("没有命中" in item for item in warnings),
            "design_validation_status": validation_report["overall_status"],
            "design_blocking_count": len(validation_report["blocking_issues"]),
            "design_warning_count": len(validation_report["warnings"]),
        }
        session.execute(text("""
            UPDATE trial_import_batch
            SET parse_status = 'ready_for_review', parse_summary = CAST(:summary AS jsonb),
                validation_report = CAST(:validation_report AS jsonb), parsed_payload = CAST(:payload AS jsonb),
                warnings = CAST(:warnings AS jsonb), parsed_at = now()
            WHERE id = :id
        """), {
            "id": batch_id,
            "summary": _json(summary),
            "validation_report": _json(validation_report),
            "payload": _json(payload),
            "warnings": _json(warnings),
        })
        session.commit()
    except Exception as exc:
        session.execute(text("UPDATE trial_import_batch SET parse_status = 'failed', error_message = :message, parsed_at = now() WHERE id = :id"), {"id": batch_id, "message": str(exc)})
        session.commit()
        raise
    return get_trial_import_batch(session, batch_id)


def _trial_id(package_id: str, trial_key: str) -> str:
    return str(uuid.uuid5(uuid.UUID(package_id), f"trial:{trial_key}"))


def publish_trial_package(session: Session, batch_id: str, actor: str) -> dict[str, Any]:
    row = session.execute(text("SELECT * FROM trial_import_batch WHERE id = :id FOR UPDATE"), {"id": batch_id}).mappings().first()
    if not row:
        raise ValueError("未找到该区域试验资料包")
    if row["parse_status"] == "published":
        return get_trial_import_batch(session, batch_id)
    if row["parse_status"] != "ready_for_review":
        raise ValueError("资料包尚未完成解析，不能发布入库")
    validation_report = _as_json(row.get("validation_report"), {})
    if validation_report.get("overall_status") != "passed":
        issues = validation_report.get("blocking_issues") or ["随机区组设计核验未通过"]
        raise ValueError("资料包暂不能发布：" + "；".join(issues[:3]))
    payload = _as_json(row["parsed_payload"], {})
    if not payload.get("entries"):
        raise ValueError("资料包没有可发布的参试与表型记录")

    package_id = _new_id()
    package_code = f"RTP-{datetime.now().strftime('%Y%m%d')}-{batch_id[:8].upper()}"
    is_simulated = "模拟" in row["display_name"] or "sample" in row["display_name"].lower()
    session.execute(text("""
        INSERT INTO trial_data_package (id, package_code, package_name, dataset_type, governance_status, description, is_simulated)
        VALUES (:id, :code, :name, '多环境区域试验', 'published', :description, :simulated)
    """), {
        "id": package_id,
        "code": package_code,
        "name": row["display_name"],
        "description": f"由数据处理员 {actor} 核验发布；原始资料包 {row['archive_name']} 保留于本地原始数据区。",
        "simulated": is_simulated,
    })

    material_ids: dict[str, str] = {}
    for material in payload.get("materials", []):
        existing = session.execute(text("SELECT id FROM breeding_material WHERE material_code = :code"), {"code": material["material_code"]}).scalar()
        material_id = str(existing) if existing else _new_id()
        material_ids[material["material_code"]] = material_id
        if existing:
            session.execute(text("""
                UPDATE breeding_material
                SET material_name = :name, is_check = :is_check,
                    aliases = CAST(:aliases AS jsonb)
                WHERE id = :id
            """), {"id": material_id, "name": material["material_name"], "is_check": material["is_check"], "aliases": _json(material.get("aliases", []))})
        else:
            session.execute(text("""
                INSERT INTO breeding_material (id, material_code, material_name, material_type, is_check, aliases, pedigree_summary)
                VALUES (:id, :code, :name, '水稻育种材料', :is_check, CAST(:aliases AS jsonb), '资料包导入；系谱待后续补充')
            """), {"id": material_id, "code": material["material_code"], "name": material["material_name"], "is_check": material["is_check"], "aliases": _json(material.get("aliases", []))})

    site_payload = {item["site_code"]: item for item in payload.get("sites", [])}
    for environment in payload.get("environment", []):
        site_payload.setdefault(environment["site_code"], {"site_code": environment["site_code"], "site_name": environment["site_name"]})
        site_payload[environment["site_code"]].update({
            "county": environment.get("county") or site_payload[environment["site_code"]].get("county"),
            "ecological_zone": environment.get("ecological_zone") or site_payload[environment["site_code"]].get("ecological_zone"),
            "soil_type": environment.get("soil_type") or site_payload[environment["site_code"]].get("soil_type"),
        })
    site_ids: dict[str, str] = {}
    for site in site_payload.values():
        existing = session.execute(text("SELECT id FROM trial_site WHERE site_code = :code"), {"code": site["site_code"]}).scalar()
        site_id = str(existing) if existing else _new_id()
        site_ids[site["site_code"]] = site_id
        if existing:
            session.execute(text("""
                UPDATE trial_site SET site_name = :name, county = COALESCE(:county, county),
                    ecological_zone = COALESCE(NULLIF(:zone, ''), ecological_zone), soil_type = COALESCE(:soil, soil_type)
                WHERE id = :id
            """), {"id": site_id, "name": site["site_name"], "county": site.get("county"), "zone": site.get("ecological_zone"), "soil": site.get("soil_type")})
        else:
            session.execute(text("""
                INSERT INTO trial_site (id, site_code, site_name, province, county, ecological_zone, soil_type)
                VALUES (:id, :code, :name, '江西省', :county, :zone, :soil)
            """), {"id": site_id, "code": site["site_code"], "name": site["site_name"], "county": site.get("county"), "zone": site.get("ecological_zone") or "未填写生态区", "soil": site.get("soil_type")})

    environments = {item["trial_key"]: item for item in payload.get("environment", [])}
    management_rows = defaultdict(list)
    for item in payload.get("management", []):
        management_rows[item["trial_key"]].append(item)
    entries_by_trial = defaultdict(list)
    for item in payload.get("entries", []):
        entries_by_trial[item["trial_key"]].append(item)

    published_counts = {"trial_count": 0, "entry_count": 0, "observation_count": 0}
    for trial in payload.get("trials", []):
        trial_key = trial["trial_key"]
        trial_validation = next(
            (item for item in validation_report.get("trial_results", []) if item.get("trial_key") == trial_key),
            {},
        )
        trial_id = _trial_id(package_id, trial_key)
        trial_code = f"{package_code}-{trial_key}"
        session.execute(text("""
            INSERT INTO field_trial (
                id, trial_code, package_id, site_id, trial_year, trial_name, design_type, replicate_count,
                design_metadata, design_validation_status, data_status, source_note
            ) VALUES (
                :id, :code, :package_id, :site_id, :year, :name, :design_type, :replicate_count,
                CAST(:design_metadata AS jsonb), 'passed', 'published', :note
            )
        """), {
            "id": trial_id,
            "code": trial_code,
            "package_id": package_id,
            "site_id": site_ids[trial["site_code"]],
            "year": trial["trial_year"],
            "name": f"{trial['trial_year']} 年 {trial['site_name']} 水稻区域试验",
            "design_type": trial_validation.get("design_type", "随机区组设计（材料 × 施氮）"),
            "replicate_count": trial_validation.get("block_count", 0),
            "design_metadata": _json(trial_validation),
            "note": f"导入批次 {batch_id}；已通过随机区组完整性核验",
        })
        published_counts["trial_count"] += 1
        environment = environments.get(trial_key)
        if environment:
            for metric in environment.get("metrics", []):
                metric_name, unit = ENVIRONMENT_METRICS.get(metric["metric_code"], (metric["metric_code"], ""))
                session.execute(text("""
                    INSERT INTO trial_environment_metric (id, trial_id, metric_code, metric_name, value_numeric, unit, original_value, collection_method, source_locator)
                    VALUES (:id, :trial_id, :code, :name, :value, :unit, :original, '资料包环境与土壤检测表', :locator)
                """), {"id": _new_id(), "trial_id": trial_id, "code": metric["metric_code"], "name": metric_name, "value": metric["value_numeric"], "unit": unit, "original": metric["original_value"], "locator": metric["source_locator"]})

        treatment_ids: dict[str, str] = {}
        trial_management = management_rows.get(trial_key, [])
        treatment_groups: dict[str, dict[str, Any]] = {}
        for entry in entries_by_trial[trial_key]:
            treatment_groups.setdefault(entry["treatment_code"], {"treatment_name": entry["treatment_name"]})
        for management in trial_management:
            treatment_groups.setdefault(management["treatment_code"], {"treatment_name": management["treatment_name"]})
        for treatment_code, treatment in treatment_groups.items():
            treatment_id = _new_id()
            treatment_ids[treatment_code] = treatment_id
            session.execute(text("""
                INSERT INTO trial_treatment (id, trial_id, treatment_code, treatment_name, treatment_description)
                VALUES (:id, :trial_id, :code, :name, :description)
            """), {"id": treatment_id, "trial_id": trial_id, "code": treatment_code, "name": treatment["treatment_name"], "description": "由区域试验资料包解析"})
            for management in [item for item in trial_management if item["treatment_code"] == treatment_code]:
                events = [
                    ("施氮", "氮肥用量", management.get("nitrogen_rate"), "kg/亩", management.get("fertilization_stage")),
                    ("灌溉", "水分管理", None, None, management.get("water_management")),
                    ("栽培", "种植密度", None, None, management.get("planting_density")),
                ]
                for event_type, input_name, rate, unit, notes in events:
                    if rate is None and not notes:
                        continue
                    session.execute(text("""
                        INSERT INTO trial_management_event (id, treatment_id, event_type, input_name, rate_per_mu, unit, event_stage, notes)
                        VALUES (:id, :treatment_id, :event_type, :input_name, :rate, :unit, '区域试验管理', :notes)
                    """), {"id": _new_id(), "treatment_id": treatment_id, "event_type": event_type, "input_name": input_name, "rate": rate, "unit": unit, "notes": notes})

        for entry in entries_by_trial[trial_key]:
            entry_id = _new_id()
            session.execute(text("""
                INSERT INTO trial_entry (id, trial_id, treatment_id, material_id, replicate_no, block_no, plot_no, raw_material_name, source_locator)
                VALUES (:id, :trial_id, :treatment_id, :material_id, :replicate_no, :block_no, :plot_no, :raw_material_name, :locator)
            """), {"id": entry_id, "trial_id": trial_id, "treatment_id": treatment_ids[entry["treatment_code"]], "material_id": material_ids[entry["material_code"]], "replicate_no": entry["replicate_no"], "block_no": entry["block_no"], "plot_no": entry["plot_no"], "raw_material_name": entry["raw_material_name"], "locator": entry["source_locator"]})
            published_counts["entry_count"] += 1
            for trait in entry.get("traits", []):
                session.execute(text("""
                    INSERT INTO trial_phenotype_observation (id, entry_id, trait_code, trait_name, trait_category, value_numeric, unit, original_value, observation_stage, evaluation_method, source_locator, quality_status, publish_status)
                    VALUES (:id, :entry_id, :trait_code, :trait_name, :trait_category, :value_numeric, :unit, :original_value, '成熟期', '区域试验观测', :locator, 'passed', 'published')
                """), {"id": _new_id(), "entry_id": entry_id, "trait_code": trait["trait_code"], "trait_name": trait["trait_name"], "trait_category": trait["trait_category"], "value_numeric": trait["value_numeric"], "unit": trait["unit"], "original_value": trait["original_value"], "locator": trait["source_locator"]})
                published_counts["observation_count"] += 1

    for source_file in payload.get("source_files", []):
        session.execute(text("""
            INSERT INTO trial_source_file (id, package_id, file_name, source_role, source_format, relative_path, raw_schema_note, processing_status, checksum)
            VALUES (:id, :package_id, :file_name, :source_role, :source_format, :relative_path, :note, 'published', :checksum)
        """), {"id": _new_id(), "package_id": package_id, "file_name": source_file["file_name"], "source_role": source_file["source_role"], "source_format": source_file["source_format"], "relative_path": source_file["relative_path"], "note": f"工作表 {source_file['sheet_name']}；共 {source_file['row_count']} 行", "checksum": source_file["checksum"]})

    session.execute(text("""
        UPDATE trial_import_batch
        SET parse_status = 'published', published_package_id = :package_id, published_at = now()
        WHERE id = :id
    """), {"id": batch_id, "package_id": package_id})
    session.commit()
    result = get_trial_import_batch(session, batch_id)
    result["published_counts"] = published_counts
    return result


def _published_package(session: Session) -> dict[str, Any] | None:
    row = session.execute(text("""
        SELECT id, package_code, package_name, created_at
        FROM trial_data_package
        WHERE governance_status = 'published'
        ORDER BY created_at DESC
        LIMIT 1
    """)).mappings().first()
    return dict(row) if row else None


def _trial_summaries(session: Session, package_id: str) -> list[dict[str, Any]]:
    rows = session.execute(text("""
        SELECT summary.*, trial.package_id
        FROM v_trial_material_summary summary
        JOIN field_trial trial ON trial.id = summary.trial_id
        WHERE trial.package_id = :package_id AND trial.data_status = 'published'
        ORDER BY summary.trial_year, summary.site_code, summary.treatment_code, summary.material_code
    """), {"package_id": package_id}).mappings().all()
    return [dict(row) for row in rows]


def _environment_by_trial(session: Session, package_id: str) -> dict[str, dict[str, Any]]:
    rows = session.execute(text("""
        SELECT trial.id AS trial_id, trial.trial_year, site.site_name, site.ecological_zone,
               MAX(metric.value_numeric) FILTER (WHERE metric.metric_code = 'soil_ph') AS soil_ph,
               MAX(metric.value_numeric) FILTER (WHERE metric.metric_code = 'available_phosphorus') AS available_phosphorus,
               MAX(metric.value_numeric) FILTER (WHERE metric.metric_code = 'rainfall') AS rainfall,
               MAX(metric.value_numeric) FILTER (WHERE metric.metric_code = 'mean_temperature') AS mean_temperature,
               MAX(metric.value_numeric) FILTER (WHERE metric.metric_code = 'disease_pressure') AS disease_pressure
        FROM field_trial trial
        JOIN trial_site site ON site.id = trial.site_id
        LEFT JOIN trial_environment_metric metric ON metric.trial_id = trial.id
        WHERE trial.package_id = :package_id AND trial.data_status = 'published'
        GROUP BY trial.id, trial.trial_year, site.site_name, site.ecological_zone
    """), {"package_id": package_id}).mappings().all()
    return {str(row["trial_id"]): dict(row) for row in rows}


def _mean(values: list[float | int | None]) -> float | None:
    numeric = [float(value) for value in values if value is not None]
    return round(statistics.mean(numeric), 3) if numeric else None


def _pearson(left: list[float], right: list[float]) -> float | None:
    if len(left) < 3 or len(left) != len(right):
        return None
    left_mean = statistics.mean(left)
    right_mean = statistics.mean(right)
    denominator = math.sqrt(sum((item - left_mean) ** 2 for item in left) * sum((item - right_mean) ** 2 for item in right))
    if denominator == 0:
        return None
    return round(sum((a - left_mean) * (b - right_mean) for a, b in zip(left, right)) / denominator, 3)


def _question_year(question: str) -> int | None:
    match = re.search(r"(20\d{2})", question)
    return int(match.group(1)) if match else None


def _question_site(question: str, rows: list[dict[str, Any]]) -> str | None:
    seen: dict[str, str] = {}
    for row in rows:
        name = row["site_name"]
        seen[_key(name)] = name
        seen[_key(name.replace("试验点", ""))] = name
    question_key = _key(question)
    for key, name in seen.items():
        if key and key in question_key:
            return name
    return None


def _question_treatment(question: str) -> str:
    if any(token in question for token in ("较高施氮", "高施氮", "高氮", "M2")):
        return "M2"
    return "M1"


def _source_cards(package: dict[str, Any], rows: list[dict[str, Any]], title: str, detail: str) -> list[dict[str, Any]]:
    years = sorted({row["trial_year"] for row in rows})
    sites = sorted({row["site_name"] for row in rows})
    return [{
        "priority": 1,
        "type": "published_regional_trial_data",
        "title": title,
        "detail": f"资料包：{package['package_name']}；覆盖 {', '.join(map(str, years))} 年、{', '.join(sites)}。{detail}",
        "query_template": "controlled_trial_analysis",
        "query_parameters": {"package_code": package["package_code"], "published_only": True},
        "query_planner": "区域试验受控分析器",
    }]


def _same_trial_analysis(rows: list[dict[str, Any]], question: str) -> dict[str, Any]:
    year = _question_year(question) or max(row["trial_year"] for row in rows)
    site = _question_site(question, rows)
    treatment = _question_treatment(question)
    selected = [row for row in rows if row["trial_year"] == year and row["treatment_code"] == treatment and (site is None or row["site_name"] == site)]
    if not selected:
        return {"analysis_type": "same_trial_comparison", "message": "没有命中指定的同试验条件，请补充年份、地点或处理。", "filters": {"year": year, "site": site, "treatment": treatment}}
    yield_values = [float(row["yield_per_mu"]) for row in selected if row["yield_per_mu"] is not None]
    height_values = [float(row["plant_height"]) for row in selected if row["plant_height"] is not None]
    yield_threshold = statistics.median(yield_values) if yield_values else 0
    height_threshold = statistics.median(height_values) if height_values else float("inf")
    ranking = []
    for row in selected:
        score = (float(row["yield_per_mu"] or 0) - yield_threshold) - 0.35 * (float(row["plant_height"] or height_threshold) - height_threshold)
        ranking.append({
            "material": row["material_name"], "material_code": row["material_code"], "is_check": bool(row["is_check"]),
            "yield_kg_per_mu": row["yield_per_mu"], "plant_height_cm": row["plant_height"],
            "lodging_score": row["lodging_score"], "panicle_blast_score": row["panicle_blast_score"],
            "high_yield_low_height_candidate": bool((row["yield_per_mu"] or 0) >= yield_threshold and (row["plant_height"] or height_threshold) <= height_threshold),
            "source_trial": row["trial_code"], "composite_rank_score": round(score, 3),
        })
    ranking.sort(key=lambda item: item["composite_rank_score"], reverse=True)
    return {"analysis_type": "same_trial_comparison", "filters": {"year": year, "site": site or "全部试验点", "treatment": treatment}, "sample_size": len(selected), "ranking": ranking, "method_note": "同年、同试验点、同处理下按小区重复均值比较；高产低株高候选按该条件内产量不低于中位数且株高不高于中位数标记。"}


def _stability_analysis(rows: list[dict[str, Any]], question: str) -> dict[str, Any]:
    treatment = _question_treatment(question)
    selected = [row for row in rows if row["treatment_code"] == treatment and row["yield_per_mu"] is not None]
    by_material: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_environment_check: dict[tuple[str, str], list[float]] = defaultdict(list)
    for row in selected:
        key = (row["trial_code"], row["treatment_code"])
        if row["is_check"]:
            by_environment_check[key].append(float(row["yield_per_mu"]))
    for row in selected:
        by_material[row["material_code"]].append(row)
    output = []
    for material_code, records in by_material.items():
        yields = [float(item["yield_per_mu"]) for item in records]
        relative = []
        for item in records:
            checks = by_environment_check.get((item["trial_code"], item["treatment_code"]), [])
            if checks:
                relative.append((float(item["yield_per_mu"]) / statistics.mean(checks) - 1) * 100)
        mean_yield = statistics.mean(yields)
        stdev = statistics.stdev(yields) if len(yields) > 1 else 0
        output.append({
            "material": records[0]["material_name"], "material_code": material_code, "is_check": bool(records[0]["is_check"]),
            "environment_count": len({item["trial_code"] for item in records}), "average_yield_kg_per_mu": round(mean_yield, 2),
            "relative_yield_vs_checks_percent": round(statistics.mean(relative), 2) if relative else None,
            "yield_standard_deviation": round(stdev, 2), "yield_cv_percent": round(stdev / mean_yield * 100, 2) if mean_yield else None,
            "valid_years": sorted({item["trial_year"] for item in records}),
        })
    output.sort(key=lambda item: (-item["average_yield_kg_per_mu"], item["yield_cv_percent"] or 999))
    return {"analysis_type": "multi_environment_stability", "treatment": treatment, "records": output, "method_note": "按同一材料在各年各点、同一施氮处理下的试验均值计算平均产量、相对对照增产、标准差和变异系数；有效环境数是实际有观测值的试验环境数。"}


def _environment_analysis(session: Session, package_id: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    environment = _environment_by_trial(session, package_id)
    by_trial: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_trial[str(row["trial_id"])].append(row)
    observation_rows = []
    for trial_id, trial_rows in by_trial.items():
        env = environment.get(trial_id)
        if not env:
            continue
        observation_rows.append({
            "trial_year": env["trial_year"], "site_name": env["site_name"], "ecological_zone": env["ecological_zone"],
            "soil_ph": env["soil_ph"], "available_phosphorus": env["available_phosphorus"], "rainfall": env["rainfall"],
            "mean_yield_kg_per_mu": _mean([row["yield_per_mu"] for row in trial_rows]),
            "mean_seed_setting_rate": _mean([row["seed_setting_rate"] for row in trial_rows]),
            "mean_thousand_grain_weight": _mean([row["thousand_grain_weight"] for row in trial_rows]),
        })
    relations = []
    for env_code, env_name in (("soil_ph", "土壤 pH"), ("available_phosphorus", "有效磷"), ("rainfall", "降雨量")):
        for outcome_code, outcome_name in (("mean_seed_setting_rate", "结实率"), ("mean_thousand_grain_weight", "千粒重"), ("mean_yield_kg_per_mu", "产量")):
            pairs = [(float(item[env_code]), float(item[outcome_code])) for item in observation_rows if item.get(env_code) is not None and item.get(outcome_code) is not None]
            relations.append({"environment_factor": env_name, "outcome": outcome_name, "sample_size": len(pairs), "pearson_r": _pearson([pair[0] for pair in pairs], [pair[1] for pair in pairs])})
    return {"analysis_type": "environment_association", "trial_environment_records": observation_rows, "relations": relations, "limitation_note": "相关系数仅描述本资料包覆盖环境中的线性关联，样本量有限且环境、管理和材料效应可能共同变化，不能据此判定因果。"}


def _management_analysis(rows: list[dict[str, Any]]) -> dict[str, Any]:
    grouped: dict[tuple[str, str], dict[str, dict[str, Any]]] = defaultdict(dict)
    for row in rows:
        grouped[(row["trial_code"], row["material_code"])][row["treatment_code"]] = row
    output = []
    for (_, _), pair in grouped.items():
        standard, high = pair.get("M1"), pair.get("M2")
        if not standard or not high:
            continue
        output.append({
            "trial": standard["trial_code"], "year": standard["trial_year"], "site": standard["site_name"],
            "material": standard["material_name"], "material_code": standard["material_code"],
            "yield_delta_kg_per_mu": round(float(high["yield_per_mu"] or 0) - float(standard["yield_per_mu"] or 0), 2),
            "height_delta_cm": round(float(high["plant_height"] or 0) - float(standard["plant_height"] or 0), 2),
            "lodging_delta": round(float(high["lodging_score"] or 0) - float(standard["lodging_score"] or 0), 2),
            "standard_yield": standard["yield_per_mu"], "high_n_yield": high["yield_per_mu"],
        })
    output.sort(key=lambda item: item["yield_delta_kg_per_mu"], reverse=True)
    return {"analysis_type": "management_effect", "pairs": output, "method_note": "同年同点同材料的标准施氮（M1）与较高施氮（M2）成对比较；增产和倒伏变化均为观测差值，不代表施氮的唯一因果作用。"}


def _tradeoff_analysis(rows: list[dict[str, Any]]) -> dict[str, Any]:
    selected = [row for row in rows if row["treatment_code"] == "M1"]
    by_material: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in selected:
        by_material[row["material_code"]].append(row)
    records = []
    for material_code, values in by_material.items():
        records.append({
            "material": values[0]["material_name"], "material_code": material_code, "is_check": bool(values[0]["is_check"]),
            "average_yield_kg_per_mu": _mean([item["yield_per_mu"] for item in values]),
            "average_head_rice_rate": _mean([item["head_rice_rate"] for item in values]),
            "average_chalkiness_degree": _mean([item["chalkiness_degree"] for item in values]),
            "average_lodging_score": _mean([item["lodging_score"] for item in values]),
        })
    records.sort(key=lambda item: item["average_yield_kg_per_mu"] or 0, reverse=True)
    valid_yield = [item for item in records if item["average_yield_kg_per_mu"] is not None]
    head = [item for item in valid_yield if item["average_head_rice_rate"] is not None]
    lodging = [item for item in valid_yield if item["average_lodging_score"] is not None]
    return {
        "analysis_type": "trait_tradeoff",
        "records": records,
        "yield_head_rice_correlation": _pearson([item["average_yield_kg_per_mu"] for item in head], [item["average_head_rice_rate"] for item in head]),
        "yield_lodging_correlation": _pearson([item["average_yield_kg_per_mu"] for item in lodging], [item["average_lodging_score"] for item in lodging]),
        "limitation_note": "该结果是材料在标准施氮处理下的跨环境汇总描述；米质、产量和倒伏之间的相关性不能替代多因素试验或遗传机制验证。",
    }


def _decline_analysis(session: Session, package_id: str, rows: list[dict[str, Any]], question: str) -> dict[str, Any]:
    material_rows = [row for row in rows if not row["is_check"]]
    requested = None
    question_key = _key(question)
    for row in material_rows:
        if _key(row["material_name"]) in question_key or _key(row["material_code"]) in question_key:
            requested = row["material_code"]
            break
    if requested is None:
        # Select the material with the largest 2025 vs history decline so the sample question has a concrete subject.
        by_material = defaultdict(list)
        for row in material_rows:
            by_material[row["material_code"]].append(row)
        deltas = []
        for code, values in by_material.items():
            current = _mean([item["yield_per_mu"] for item in values if item["trial_year"] == 2025])
            previous = _mean([item["yield_per_mu"] for item in values if item["trial_year"] < 2025])
            if current is not None and previous is not None:
                deltas.append((current - previous, code))
        requested = min(deltas)[1] if deltas else material_rows[0]["material_code"]
    selected = [row for row in rows if row["material_code"] == requested and row["treatment_code"] == "M1"]
    env = _environment_by_trial(session, package_id)
    details = []
    for row in selected:
        environment = env.get(str(row["trial_id"]), {})
        details.append({
            "year": row["trial_year"], "site": row["site_name"], "yield_kg_per_mu": row["yield_per_mu"],
            "seed_setting_rate": row["seed_setting_rate"], "thousand_grain_weight": row["thousand_grain_weight"],
            "panicle_blast_score": row["panicle_blast_score"], "soil_ph": environment.get("soil_ph"),
            "available_phosphorus": environment.get("available_phosphorus"), "rainfall": environment.get("rainfall"),
            "disease_pressure": environment.get("disease_pressure"), "source_trial": row["trial_code"],
        })
    material_name = selected[0]["material_name"] if selected else requested
    return {"analysis_type": "decline_evidence", "material": material_name, "material_code": requested, "records": details, "method_note": "将该材料同一标准施氮处理下的年点表现、土壤、天气和病害压力并列。只能拆解同时出现的证据变化，不能仅凭本资料判定某一因素导致表现下降。"}


def build_published_trial_evidence(session: Session, question: str, requested_by: str = "系统") -> tuple[str, list[dict[str, Any]]]:
    """Return only actually published trial-package evidence for the assistant."""
    package = _published_package(session)
    if not package:
        return "", []
    rows = _trial_summaries(session, str(package["id"]))
    if not rows:
        return "", []
    normalized = _key(question)
    formal_analysis: dict[str, Any] | None = None
    formal_unavailable: str | None = None
    try:
        formal_analysis = run_controlled_trial_analysis(session, package, question, requested_by)
    except TrialStatisticsError as exc:
        # Keep a traceable descriptive fallback for a question that matches the
        # published package but lacks the design completeness for formal stats.
        formal_unavailable = str(exc)
    if any(token in normalized for token in ("同一试验", "哪些材料", "株高更低", "南昌点")):
        analysis = _same_trial_analysis(rows, question)
        title = "已发布区域试验：同试验材料比较"
    elif any(token in normalized for token in ("稳定性", "多年多点", "相对增产", "有效环境")):
        analysis = _stability_analysis(rows, question)
        title = "已发布区域试验：多年多点稳定性"
    elif any(token in normalized for token in ("土壤ph", "有效磷", "降雨", "环境影响", "环境关联")):
        analysis = _environment_analysis(session, str(package["id"]), rows)
        title = "已发布区域试验：环境与土壤关联"
    elif any(token in normalized for token in ("施氮", "管理措施", "高氮", "标准氮", "倒伏风险")):
        analysis = _management_analysis(rows)
        title = "已发布区域试验：管理措施影响"
    elif any(token in normalized for token in ("权衡", "取舍", "米质", "高产材料")):
        analysis = _tradeoff_analysis(rows)
        title = "已发布区域试验：性状权衡"
    elif any(token in normalized for token in ("下降", "变差", "异常", "证据拆解")):
        analysis = _decline_analysis(session, str(package["id"]), rows, question)
        title = "已发布区域试验：表现变差证据拆解"
    elif any(token in normalized for token in ("候选", "高产", "稳产", "适应生态区")):
        analysis = _stability_analysis(rows, question)
        title = "已发布区域试验：高产稳产候选筛选"
    else:
        return "", []
    if formal_analysis:
        analysis = formal_analysis
        title = f"已发布区域试验：{formal_analysis.get('title', title)}"
    elif formal_unavailable:
        analysis = {
            **analysis,
            "formal_statistics_status": "unavailable",
            "formal_statistics_reason": formal_unavailable,
            "formal_statistics_note": "当前回答仅保留可追溯的描述性汇总；请先核验试验设计、缺失组合和环境/管理资料，再用于正式统计推断。",
        }
    cards = _source_cards(package, rows, title, "仅引用本次问题实际调用的已发布区域试验标准记录。")
    if formal_analysis:
        cards.append({
            "kind": "analysis_run",
            "title": "本次正式统计分析记录",
            "detail": f"运行编号 {formal_analysis['analysis_run_id']}；{formal_analysis['engine']}；模型：{formal_analysis.get('model_formula') or '受控描述性过程'}。",
            "source": package["package_name"],
        })
    context = {
        "evidence_scope": "平台已发布区域试验标准数据",
        "package": {"code": package["package_code"], "name": package["package_name"]},
        "analysis": analysis,
        "response_rules": [
            "仅根据这个 JSON 中的记录回答，不补造数值或资料来源。",
            "将环境和管理结果表述为关联、差异或证据，不表述为已证实的因果。",
            "如包含 analysis_run_id、模型公式和局限，应在回答中说明统计口径，不把它包装成超出数据范围的结论。",
            "回答中说明分析条件、可追溯范围和限制。",
        ],
    }
    return "区域试验已发布数据（本轮可追溯证据 JSON）：\n" + _json(context), cards
