"""Private genotype assets and a narrow, reproducible rice QC pipeline.

This module deliberately keeps raw genotype files out of the chat path.  A
researcher uploads a VCF/VCF.GZ or a PLINK ZIP once, the local worker converts
and quality-controls it, then a manually confirmed immutable version can be
selected by controlled downstream workflows such as continuous-trait GWAS.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import shutil
import subprocess
import time
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import text
from sqlalchemy.orm import Session


MAX_UPLOAD_BYTES = 5 * 1024 * 1024 * 1024
UPLOAD_CHUNK_BYTES = 10 * 1024 * 1024
RICE_QC_TEMPLATE_CODE = "rice_standard_breeding_qc"
RICE_QC_TEMPLATE_VERSION = "v1.0"
DEFAULT_PROJECT_ID = "00000000-0000-4000-8000-000000000001"
SUPPORTED_POPULATIONS = {
    "stable_breeding": "稳定育种材料",
    "segregating": "分离群体",
    "natural_germplasm": "自然种质群体",
    "unknown": "待确认",
}
SUPPORTED_FORMATS = {"vcf", "plink_zip"}


class GenotypeAssetError(ValueError):
    """A clear, researcher-facing error for genotype asset intake."""


class CreateGenotypeAssetRequest(BaseModel):
    title: str = Field(min_length=1, max_length=300)
    source_format: str = Field(pattern="^(vcf|plink_zip)$")
    reference_assembly: str = Field(default="IRGSP-1.0", min_length=1, max_length=120)
    population_type: str = Field(default="stable_breeding")


class UploadInitRequest(BaseModel):
    file_name: str = Field(min_length=1, max_length=500)
    total_bytes: int = Field(gt=0, le=MAX_UPLOAD_BYTES)
    total_chunks: int = Field(gt=0, le=10000)


class MappingUpdateRequest(BaseModel):
    material_id: str | None = Field(default=None, max_length=36)
    note: str = Field(default="", max_length=1000)


class GovernanceRequestCreate(BaseModel):
    request_type: str = Field(pattern="^(material_master|mapping_conflict|reference_review|phenotype_governance)$")
    description: str = Field(min_length=1, max_length=3000)
    affected_samples: list[str] = Field(default_factory=list, max_length=1000)


class GovernanceRequestResolution(BaseModel):
    """A data processor can resolve the metadata request without receiving raw genotypes."""

    status: str = Field(pattern="^(accepted|needs_info|resolved)$")
    resolution_note: str = Field(min_length=1, max_length=3000)


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)


def _safe_name(value: str, fallback: str) -> str:
    name = Path(value or fallback).name
    clean = re.sub(r"[^A-Za-z0-9._()\-]+", "_", name).strip("._")
    return clean or fallback


def _artifact_display_stem(value: str, fallback: str = "未命名基因型数据") -> str:
    """Keep researcher-provided Chinese asset labels while making download names safe."""
    name = Path(str(value or fallback)).name.strip()
    clean = re.sub(r'[\\/:*?"<>|\x00-\x1f]+', "_", name)
    clean = re.sub(r"\s+", " ", clean).strip(" ._")
    return (clean or fallback)[:80]


def _asset_dir(storage_dir: Path, owner_id: str, asset_id: str) -> Path:
    path = storage_dir / "genotype-assets" / owner_id / asset_id
    path.mkdir(parents=True, exist_ok=True)
    return path


def _row_dict(row: Any) -> dict[str, Any]:
    return dict(row) if row is not None else {}


def _canonical_chromosome(value: str) -> tuple[str, bool]:
    raw = str(value or "").strip()
    number = re.fullmatch(r"(?:chr)?0*(\d{1,2})", raw, flags=re.IGNORECASE)
    if number and 1 <= int(number.group(1)) <= 12:
        return f"Chr{int(number.group(1))}", True
    return raw, False


def ensure_genotype_asset_schema(session: Session) -> None:
    """Create private assets, immutable QC versions and persistent jobs."""
    statements = (
        """
        CREATE TABLE IF NOT EXISTS genotype_qc_template (
          id VARCHAR(36) PRIMARY KEY,
          template_code VARCHAR(100) NOT NULL,
          version VARCHAR(30) NOT NULL,
          template_name VARCHAR(300) NOT NULL,
          population_type VARCHAR(60) NOT NULL,
          parameters JSONB NOT NULL DEFAULT '{}'::jsonb,
          status VARCHAR(30) NOT NULL DEFAULT 'published',
          change_summary TEXT NOT NULL,
          created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          UNIQUE(template_code, version)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS genotype_asset (
          id VARCHAR(36) PRIMARY KEY,
          owner_id VARCHAR(120) NOT NULL,
          title VARCHAR(300) NOT NULL,
          source_format VARCHAR(30) NOT NULL,
          reference_assembly VARCHAR(120) NOT NULL,
          population_type VARCHAR(60) NOT NULL,
          status VARCHAR(40) NOT NULL DEFAULT 'draft',
          raw_file_name VARCHAR(500),
          raw_storage_path TEXT,
          raw_size_bytes BIGINT,
          raw_sha256 VARCHAR(64),
          current_version_id VARCHAR(36),
          archived_at TIMESTAMPTZ,
          created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS genotype_asset_version (
          id VARCHAR(36) PRIMARY KEY,
          asset_id VARCHAR(36) NOT NULL REFERENCES genotype_asset(id) ON DELETE CASCADE,
          owner_id VARCHAR(120) NOT NULL,
          version_number INTEGER NOT NULL,
          status VARCHAR(40) NOT NULL DEFAULT 'processing',
          parent_version_id VARCHAR(36),
          qc_template_code VARCHAR(100) NOT NULL,
          qc_template_version VARCHAR(30) NOT NULL,
          reference_assembly VARCHAR(120) NOT NULL,
          source_snapshot JSONB NOT NULL DEFAULT '{}'::jsonb,
          qc_summary JSONB NOT NULL DEFAULT '{}'::jsonb,
          plink_prefix VARCHAR(200),
          plink_directory TEXT,
          report_path TEXT,
          package_path TEXT,
          published_at TIMESTAMPTZ,
          published_by VARCHAR(120),
          created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          UNIQUE(asset_id, version_number)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS genotype_processing_job (
          id VARCHAR(36) PRIMARY KEY,
          asset_id VARCHAR(36) NOT NULL REFERENCES genotype_asset(id) ON DELETE CASCADE,
          version_id VARCHAR(36) NOT NULL REFERENCES genotype_asset_version(id) ON DELETE CASCADE,
          owner_id VARCHAR(120) NOT NULL,
          job_type VARCHAR(60) NOT NULL DEFAULT 'convert_and_qc',
          status VARCHAR(30) NOT NULL DEFAULT 'queued',
          progress_label VARCHAR(300) NOT NULL DEFAULT '等待本地处理服务',
          error_detail TEXT,
          created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          started_at TIMESTAMPTZ,
          completed_at TIMESTAMPTZ
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS genotype_upload_session (
          id VARCHAR(36) PRIMARY KEY,
          asset_id VARCHAR(36) NOT NULL REFERENCES genotype_asset(id) ON DELETE CASCADE,
          owner_id VARCHAR(120) NOT NULL,
          file_name VARCHAR(500) NOT NULL,
          total_bytes BIGINT NOT NULL,
          total_chunks INTEGER NOT NULL,
          received_chunks JSONB NOT NULL DEFAULT '[]'::jsonb,
          created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          completed_at TIMESTAMPTZ
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS genotype_sample_mapping (
          id VARCHAR(36) PRIMARY KEY,
          version_id VARCHAR(36) NOT NULL REFERENCES genotype_asset_version(id) ON DELETE CASCADE,
          owner_id VARCHAR(120) NOT NULL,
          fid VARCHAR(160) NOT NULL,
          iid VARCHAR(160) NOT NULL,
          material_id VARCHAR(36) REFERENCES breeding_material(id),
          status VARCHAR(30) NOT NULL DEFAULT 'unmapped',
          note TEXT,
          updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          UNIQUE(version_id, fid, iid)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS genotype_governance_request (
          id VARCHAR(36) PRIMARY KEY,
          owner_id VARCHAR(120) NOT NULL,
          asset_id VARCHAR(36) REFERENCES genotype_asset(id) ON DELETE SET NULL,
          version_id VARCHAR(36) REFERENCES genotype_asset_version(id) ON DELETE SET NULL,
          asset_title_snapshot VARCHAR(300),
          version_number_snapshot INTEGER,
          request_type VARCHAR(60) NOT NULL,
          status VARCHAR(30) NOT NULL DEFAULT 'submitted',
          description TEXT NOT NULL,
          affected_samples JSONB NOT NULL DEFAULT '[]'::jsonb,
          created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          resolved_at TIMESTAMPTZ,
          resolved_by VARCHAR(120),
          resolution_note TEXT
        )
        """,
        "CREATE INDEX IF NOT EXISTS ix_genotype_asset_owner_updated ON genotype_asset(owner_id, updated_at DESC)",
        "CREATE INDEX IF NOT EXISTS ix_genotype_version_owner_status ON genotype_asset_version(owner_id, status, created_at DESC)",
        "CREATE INDEX IF NOT EXISTS ix_genotype_job_status_created ON genotype_processing_job(status, created_at)",
        "CREATE INDEX IF NOT EXISTS ix_genotype_mapping_version ON genotype_sample_mapping(version_id)",
    )
    for statement in statements:
        session.execute(text(statement))
    # Governance requests intentionally carry only the minimal asset label that
    # a data processor needs. This keeps the processor queue independent from
    # raw-file metadata and private genotype storage.
    session.execute(text("ALTER TABLE genotype_governance_request ADD COLUMN IF NOT EXISTS asset_title_snapshot VARCHAR(300)"))
    session.execute(text("ALTER TABLE genotype_governance_request ADD COLUMN IF NOT EXISTS version_number_snapshot INTEGER"))
    for table_name in (
        "genotype_asset", "genotype_asset_version", "genotype_processing_job",
        "genotype_upload_session", "genotype_sample_mapping", "genotype_governance_request",
    ):
        session.execute(text(
            f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS project_id VARCHAR(36) "
            f"NOT NULL DEFAULT '{DEFAULT_PROJECT_ID}'"
        ))
        session.execute(text(f"CREATE INDEX IF NOT EXISTS ix_{table_name}_project_id ON {table_name}(project_id)"))
    session.execute(text("""
        CREATE OR REPLACE FUNCTION inherit_genotype_project_from_asset() RETURNS trigger AS $$
        BEGIN
          NEW.project_id := COALESCE((SELECT project_id FROM genotype_asset WHERE id = NEW.asset_id), NEW.project_id);
          RETURN NEW;
        END;
        $$ LANGUAGE plpgsql
    """))
    session.execute(text("""
        CREATE OR REPLACE FUNCTION inherit_genotype_project_from_version() RETURNS trigger AS $$
        BEGIN
          NEW.project_id := COALESCE((SELECT project_id FROM genotype_asset_version WHERE id = NEW.version_id), NEW.project_id);
          RETURN NEW;
        END;
        $$ LANGUAGE plpgsql
    """))
    for table_name in ("genotype_asset_version", "genotype_processing_job", "genotype_upload_session", "genotype_governance_request"):
        session.execute(text(f"DROP TRIGGER IF EXISTS trg_{table_name}_project ON {table_name}"))
        session.execute(text(
            f"CREATE TRIGGER trg_{table_name}_project BEFORE INSERT OR UPDATE OF asset_id ON {table_name} "
            "FOR EACH ROW EXECUTE FUNCTION inherit_genotype_project_from_asset()"
        ))
    session.execute(text("DROP TRIGGER IF EXISTS trg_genotype_sample_mapping_project ON genotype_sample_mapping"))
    session.execute(text(
        "CREATE TRIGGER trg_genotype_sample_mapping_project BEFORE INSERT OR UPDATE OF version_id ON genotype_sample_mapping "
        "FOR EACH ROW EXECUTE FUNCTION inherit_genotype_project_from_version()"
    ))
    template_exists = session.execute(text("""
        SELECT 1 FROM genotype_qc_template
        WHERE template_code = :code AND version = :version
    """), {"code": RICE_QC_TEMPLATE_CODE, "version": RICE_QC_TEMPLATE_VERSION}).scalar()
    if not template_exists:
        session.execute(text("""
            INSERT INTO genotype_qc_template
            (id, template_code, version, template_name, population_type, parameters, status, change_summary)
            VALUES (:id, :code, :version, :name, 'stable_breeding', CAST(:parameters AS jsonb), 'published', :summary)
        """), {
            "id": str(uuid.uuid4()),
            "code": RICE_QC_TEMPLATE_CODE,
            "version": RICE_QC_TEMPLATE_VERSION,
            "name": "水稻常规育种材料 QC v1.0",
            "parameters": _json({
                "maf_min": 0.05,
                "snp_missing_max": 0.05,
                "sample_missing_max": 0.05,
                "hwe": "diagnostic_only",
                "heterozygosity": "flag_zscore_gt_3",
                "duplicate_material_mapping": "block_publish",
                "recognized_chromosomes": [f"Chr{i}" for i in range(1, 13)],
            }),
            "summary": "首版水稻常规育种材料质控：MAF、样本/SNP 缺失率为硬过滤；杂合率异常仅标记；HWE 仅诊断；材料重复映射阻断发布。",
        })

    # Private genotype records are database-isolated by the verified user id.
    for table_name in (
        "genotype_asset", "genotype_asset_version", "genotype_processing_job",
        "genotype_upload_session", "genotype_sample_mapping", "genotype_governance_request",
    ):
        session.execute(text(f"ALTER TABLE {table_name} ENABLE ROW LEVEL SECURITY"))
        session.execute(text(f"ALTER TABLE {table_name} FORCE ROW LEVEL SECURITY"))
        session.execute(text(f"DROP POLICY IF EXISTS genotype_owner_only ON {table_name}"))
        session.execute(text(
            f"CREATE POLICY genotype_owner_only ON {table_name} FOR ALL "
            "USING ((owner_id = current_setting('app.research_user_id', true) "
            "AND project_id = current_setting('app.project_id', true)) OR current_user = 'rice') "
            "WITH CHECK ((owner_id = current_setting('app.research_user_id', true) "
            "AND project_id = current_setting('app.project_id', true)) OR current_user = 'rice')"
        ))
    # A data processor can only read and update governance request summaries.
    # It does not receive a policy on genotype assets, uploads or mappings.
    session.execute(text("DROP POLICY IF EXISTS genotype_governance_processor_read ON genotype_governance_request"))
    session.execute(text("DROP POLICY IF EXISTS genotype_governance_processor_update ON genotype_governance_request"))
    session.execute(text(
        "CREATE POLICY genotype_governance_processor_read ON genotype_governance_request FOR SELECT "
        "USING ((current_setting('app.genotype_governance_processor', true) = 'true' "
        "AND project_id = current_setting('app.project_id', true)) OR current_user = 'rice')"
    ))
    session.execute(text(
        "CREATE POLICY genotype_governance_processor_update ON genotype_governance_request FOR UPDATE "
        "USING ((current_setting('app.genotype_governance_processor', true) = 'true' "
        "AND project_id = current_setting('app.project_id', true)) OR current_user = 'rice') "
        "WITH CHECK ((current_setting('app.genotype_governance_processor', true) = 'true' "
        "AND project_id = current_setting('app.project_id', true)) OR current_user = 'rice')"
    ))
    session.execute(text("ALTER TABLE genotype_qc_template ENABLE ROW LEVEL SECURITY"))
    session.execute(text("ALTER TABLE genotype_qc_template FORCE ROW LEVEL SECURITY"))
    session.execute(text("DROP POLICY IF EXISTS genotype_qc_template_read ON genotype_qc_template"))
    session.execute(text("CREATE POLICY genotype_qc_template_read ON genotype_qc_template FOR SELECT USING (status = 'published')"))
    session.commit()


def _version_summary(session: Session, version_id: str) -> dict[str, Any]:
    version = session.execute(text("""
        SELECT version.*, asset.title AS asset_title, asset.source_format, asset.population_type,
               asset.raw_file_name, asset.raw_sha256
        FROM genotype_asset_version version
        JOIN genotype_asset asset ON asset.id = version.asset_id
        WHERE version.id = :version_id
    """), {"version_id": version_id}).mappings().first()
    if not version:
        raise GenotypeAssetError("未找到当前账号的基因型版本。")
    mappings = session.execute(text("""
        SELECT mapping.*, material.material_code, material.material_name, material.aliases
        FROM genotype_sample_mapping mapping
        LEFT JOIN breeding_material material ON material.id = mapping.material_id
        WHERE mapping.version_id = :version_id
        ORDER BY mapping.iid, mapping.fid
    """), {"version_id": version_id}).mappings().all()
    job = session.execute(text("""
        SELECT id, status, progress_label, error_detail, created_at, started_at, completed_at
        FROM genotype_processing_job WHERE version_id = :version_id
        ORDER BY created_at DESC LIMIT 1
    """), {"version_id": version_id}).mappings().first()
    mapped = sum(1 for item in mappings if item["status"] == "mapped")
    material_ids = [str(item["material_id"]) for item in mappings if item["status"] == "mapped" and item["material_id"]]
    duplicate_material_count = len(material_ids) - len(set(material_ids))
    return {
        "id": version["id"], "asset_id": version["asset_id"], "title": version["asset_title"],
        "project_id": version["project_id"],
        "version_number": version["version_number"], "status": version["status"],
        "source_format": version["source_format"], "reference_assembly": version["reference_assembly"],
        "population_type": version["population_type"], "population_type_label": SUPPORTED_POPULATIONS.get(version["population_type"], version["population_type"]),
        "qc_template": {"code": version["qc_template_code"], "version": version["qc_template_version"]},
        "source_snapshot": version["source_snapshot"] or {}, "qc_summary": version["qc_summary"] or {},
        "raw_file_name": version["raw_file_name"], "raw_sha256": version["raw_sha256"],
        "mapping_summary": {
            "total": len(mappings), "mapped": mapped, "unmapped": len(mappings) - mapped,
            "duplicate_material_count": duplicate_material_count,
        },
        "error_message": (job["error_detail"] if job and job["status"] == "failed" else None),
        "published_at": version["published_at"].isoformat() if version["published_at"] else None,
        "created_at": version["created_at"].isoformat(), "updated_at": version["updated_at"].isoformat(),
        "job": ({**dict(job), "created_at": job["created_at"].isoformat(), "started_at": job["started_at"].isoformat() if job["started_at"] else None, "completed_at": job["completed_at"].isoformat() if job["completed_at"] else None} if job else None),
        "mappings": [{
            "id": item["id"], "fid": item["fid"], "iid": item["iid"], "material_id": item["material_id"],
            "material_code": item["material_code"], "material_name": item["material_name"],
            "aliases": item["aliases"] or [], "status": item["status"], "note": item["note"] or "",
        } for item in mappings],
        # Sample/SNP QC can be inspected before mapping, but a downloadable
        # result package represents a formally traceable analysis version.
        # Keep that boundary explicit in both the API and the UI.
        "report_available": bool(version["report_path"]),
        "package_available": bool(version["package_path"]) and version["status"] == "analysis_ready",
    }


def list_assets(session: Session) -> list[dict[str, Any]]:
    rows = session.execute(text("""
        SELECT id, current_version_id FROM genotype_asset
        WHERE archived_at IS NULL ORDER BY updated_at DESC
    """)).mappings().all()
    items: list[dict[str, Any]] = []
    for row in rows:
        if row["current_version_id"]:
            items.append(_version_summary(session, row["current_version_id"]))
        else:
            asset = session.execute(text("SELECT * FROM genotype_asset WHERE id = :id"), {"id": row["id"]}).mappings().first()
            if asset:
                items.append({
                    "id": None, "asset_id": asset["id"], "title": asset["title"], "version_number": 0,
                    "project_id": asset["project_id"],
                    "status": asset["status"], "source_format": asset["source_format"], "reference_assembly": asset["reference_assembly"],
                    "population_type": asset["population_type"], "population_type_label": SUPPORTED_POPULATIONS.get(asset["population_type"], asset["population_type"]),
                    "qc_summary": {}, "mapping_summary": {"total": 0, "mapped": 0, "unmapped": 0}, "job": None,
                    "created_at": asset["created_at"].isoformat(), "updated_at": asset["updated_at"].isoformat(),
                })
    return items


def get_asset_version(session: Session, asset_id: str, version_id: str | None = None) -> dict[str, Any]:
    if version_id:
        row = session.execute(text("SELECT id FROM genotype_asset_version WHERE id = :id AND asset_id = :asset_id"), {"id": version_id, "asset_id": asset_id}).scalar()
    else:
        row = session.execute(text("SELECT current_version_id FROM genotype_asset WHERE id = :asset_id"), {"asset_id": asset_id}).scalar()
    if not row:
        raise GenotypeAssetError("该基因型资产尚未生成可查看的版本。")
    return _version_summary(session, str(row))


def create_asset(session: Session, owner_id: str, payload: CreateGenotypeAssetRequest, project_id: str = DEFAULT_PROJECT_ID) -> dict[str, Any]:
    if payload.population_type not in SUPPORTED_POPULATIONS:
        raise GenotypeAssetError("请选择稳定育种材料、分离群体、自然种质群体或待确认。")
    asset_id = str(uuid.uuid4())
    session.execute(text("""
        INSERT INTO genotype_asset
        (id, owner_id, project_id, title, source_format, reference_assembly, population_type, status, created_at, updated_at)
        VALUES (:id, :owner_id, :project_id, :title, :source_format, :assembly, :population_type, 'awaiting_upload', :now, :now)
    """), {"id": asset_id, "owner_id": owner_id, "project_id": project_id, "title": payload.title.strip(), "source_format": payload.source_format,
             "assembly": payload.reference_assembly.strip(), "population_type": payload.population_type, "now": _now()})
    return {"asset_id": asset_id, "project_id": project_id, "status": "awaiting_upload", "upload_chunk_bytes": UPLOAD_CHUNK_BYTES}


def create_upload_session(session: Session, owner_id: str, asset_id: str, payload: UploadInitRequest, storage_dir: Path) -> dict[str, Any]:
    asset = session.execute(text("SELECT * FROM genotype_asset WHERE id = :id"), {"id": asset_id}).mappings().first()
    if not asset:
        raise GenotypeAssetError("未找到当前账号的基因型资产。")
    suffix = Path(payload.file_name).suffix.lower()
    valid = (asset["source_format"] == "vcf" and (suffix == ".vcf" or payload.file_name.lower().endswith(".vcf.gz"))) or (asset["source_format"] == "plink_zip" and suffix == ".zip")
    if not valid:
        expected = "VCF / VCF.GZ" if asset["source_format"] == "vcf" else "含 .bed/.bim/.fam 的 ZIP"
        raise GenotypeAssetError(f"当前资产类型需要上传 {expected}。")
    upload_id = str(uuid.uuid4())
    session.execute(text("""
        INSERT INTO genotype_upload_session(id, asset_id, owner_id, file_name, total_bytes, total_chunks, received_chunks)
        VALUES (:id, :asset_id, :owner_id, :file_name, :total_bytes, :total_chunks, '[]'::jsonb)
    """), {"id": upload_id, "asset_id": asset_id, "owner_id": owner_id, "file_name": _safe_name(payload.file_name, "genotype.bin"), "total_bytes": payload.total_bytes, "total_chunks": payload.total_chunks})
    (_asset_dir(storage_dir, owner_id, asset_id) / "uploads" / upload_id).mkdir(parents=True, exist_ok=True)
    # Keep ``id`` as the canonical resource identifier returned by the upload
    # API. ``upload_id`` remains for backwards compatibility with callers that
    # already consumed the initial prototype response.
    return {
        "id": upload_id,
        "upload_id": upload_id,
        "chunk_bytes": UPLOAD_CHUNK_BYTES,
        "total_chunks": payload.total_chunks,
        "received_chunks": [],
    }


def get_upload_session(session: Session, asset_id: str, upload_id: str) -> dict[str, Any]:
    row = session.execute(text("""
        SELECT id, asset_id, file_name, total_bytes, total_chunks, received_chunks, completed_at
        FROM genotype_upload_session
        WHERE id = :upload_id AND asset_id = :asset_id
    """), {"upload_id": upload_id, "asset_id": asset_id}).mappings().first()
    if not row:
        raise GenotypeAssetError("未找到可恢复的上传任务。")
    return {
        "id": row["id"],
        "upload_id": row["id"],
        "file_name": row["file_name"],
        "total_bytes": int(row["total_bytes"]),
        "total_chunks": int(row["total_chunks"]),
        "received_chunks": sorted(int(item) for item in (row["received_chunks"] or [])),
        "completed": row["completed_at"] is not None,
        "chunk_bytes": UPLOAD_CHUNK_BYTES,
    }


def upload_chunk_path(storage_dir: Path, owner_id: str, asset_id: str, upload_id: str, index: int) -> Path:
    directory = _asset_dir(storage_dir, owner_id, asset_id) / "uploads" / upload_id
    directory.mkdir(parents=True, exist_ok=True)
    return directory / f"{index:06d}.part"


def record_upload_chunk(session: Session, upload_id: str, index: int) -> None:
    row = session.execute(text("SELECT total_chunks, received_chunks FROM genotype_upload_session WHERE id = :id"), {"id": upload_id}).mappings().first()
    if not row:
        raise GenotypeAssetError("上传会话不存在或已失效。")
    if index < 0 or index >= int(row["total_chunks"]):
        raise GenotypeAssetError("上传分片编号不合法。")
    chunks = {int(value) for value in (row["received_chunks"] or [])}
    chunks.add(index)
    session.execute(text("UPDATE genotype_upload_session SET received_chunks = CAST(:chunks AS jsonb) WHERE id = :id"), {"id": upload_id, "chunks": _json(sorted(chunks))})


def complete_upload(session: Session, owner_id: str, asset_id: str, upload_id: str, storage_dir: Path) -> dict[str, Any]:
    upload = session.execute(text("SELECT * FROM genotype_upload_session WHERE id = :id AND asset_id = :asset_id"), {"id": upload_id, "asset_id": asset_id}).mappings().first()
    asset = session.execute(text("SELECT * FROM genotype_asset WHERE id = :id"), {"id": asset_id}).mappings().first()
    if not upload or not asset:
        raise GenotypeAssetError("上传会话或基因型资产不存在。")
    received = {int(value) for value in (upload["received_chunks"] or [])}
    expected = set(range(int(upload["total_chunks"])))
    if received != expected:
        raise GenotypeAssetError(f"文件尚未上传完整：已接收 {len(received)}/{len(expected)} 个分片。")
    upload_dir = _asset_dir(storage_dir, owner_id, asset_id) / "uploads" / upload_id
    raw_dir = _asset_dir(storage_dir, owner_id, asset_id) / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    target = raw_dir / _safe_name(str(upload["file_name"]), "genotype.bin")
    temporary = target.with_suffix(f"{target.suffix}.tmp")
    digest = hashlib.sha256()
    written = 0
    with temporary.open("wb") as sink:
        for index in range(int(upload["total_chunks"])):
            part = upload_dir / f"{index:06d}.part"
            if not part.is_file():
                raise GenotypeAssetError(f"缺少第 {index + 1} 个上传分片。")
            with part.open("rb") as source:
                while data := source.read(1024 * 1024):
                    sink.write(data)
                    digest.update(data)
                    written += len(data)
    if written != int(upload["total_bytes"]):
        temporary.unlink(missing_ok=True)
        raise GenotypeAssetError("合并后的文件大小与上传声明不一致，请重新上传。")
    temporary.replace(target)
    shutil.rmtree(upload_dir, ignore_errors=True)
    version_id = str(uuid.uuid4())
    version_number = int(session.execute(text("SELECT COALESCE(MAX(version_number), 0) + 1 FROM genotype_asset_version WHERE asset_id = :asset_id"), {"asset_id": asset_id}).scalar_one())
    session.execute(text("""
        INSERT INTO genotype_asset_version
        (id, asset_id, owner_id, version_number, status, qc_template_code, qc_template_version, reference_assembly, source_snapshot, created_at, updated_at)
        VALUES (:id, :asset_id, :owner_id, :version_number, 'queued', :template_code, :template_version, :assembly, CAST(:snapshot AS jsonb), :now, :now)
    """), {"id": version_id, "asset_id": asset_id, "owner_id": owner_id, "version_number": version_number,
             "template_code": RICE_QC_TEMPLATE_CODE, "template_version": RICE_QC_TEMPLATE_VERSION, "assembly": asset["reference_assembly"],
             "snapshot": _json({"raw_file_name": target.name, "raw_storage_path": str(target), "raw_size_bytes": written, "raw_sha256": digest.hexdigest(), "source_format": asset["source_format"]}), "now": _now()})
    job_id = str(uuid.uuid4())
    session.execute(text("""
        INSERT INTO genotype_processing_job(id, asset_id, version_id, owner_id, status, progress_label)
        VALUES (:id, :asset_id, :version_id, :owner_id, 'queued', '等待本地 PLINK 转换与质控')
    """), {"id": job_id, "asset_id": asset_id, "version_id": version_id, "owner_id": owner_id})
    session.execute(text("""
        UPDATE genotype_asset SET status = 'queued', raw_file_name = :name, raw_storage_path = :path,
          raw_size_bytes = :size, raw_sha256 = :sha, current_version_id = :version_id, updated_at = :now
        WHERE id = :asset_id
    """), {"name": target.name, "path": str(target), "size": written, "sha": digest.hexdigest(), "version_id": version_id, "now": _now(), "asset_id": asset_id})
    session.execute(text("UPDATE genotype_upload_session SET completed_at = :now WHERE id = :id"), {"now": _now(), "id": upload_id})
    return {"asset_id": asset_id, "version_id": version_id, "job_id": job_id, "status": "queued"}


def _read_text(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise GenotypeAssetError(f"无法读取文件编码：{path.name}")


def _fam_pairs(path: Path) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    for line_no, line in enumerate(_read_text(path).splitlines(), start=1):
        values = line.split()
        if not values:
            continue
        if len(values) < 6:
            raise GenotypeAssetError(f"{path.name} 第 {line_no} 行少于 6 列，无法识别 FID/IID。")
        rows.append((values[0], values[1]))
    if not rows:
        raise GenotypeAssetError("PLINK FAM 文件中没有可用样本。")
    return rows


def _bim_count(path: Path) -> tuple[int, list[str]]:
    count, unknown = 0, []
    for line_no, line in enumerate(_read_text(path).splitlines(), start=1):
        values = line.split()
        if not values:
            continue
        if len(values) < 6:
            raise GenotypeAssetError(f"{path.name} 第 {line_no} 行少于 6 列。")
        count += 1
        _, recognized = _canonical_chromosome(values[0])
        if not recognized and values[0] not in unknown:
            unknown.append(values[0])
    if not count:
        raise GenotypeAssetError("PLINK BIM 文件中没有可用 SNP。")
    return count, unknown


def _normalize_bim(path: Path) -> list[str]:
    output: list[str] = []
    unknown: list[str] = []
    for line in _read_text(path).splitlines():
        values = line.split()
        if not values:
            continue
        chromosome, recognized = _canonical_chromosome(values[0])
        values[0] = chromosome
        if not recognized and chromosome not in unknown:
            unknown.append(chromosome)
        output.append("\t".join(values))
    path.write_text("\n".join(output) + "\n", encoding="utf-8")
    return unknown


def _run(command: list[str], cwd: Path) -> None:
    try:
        completed = subprocess.run(command, cwd=cwd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, timeout=6 * 3600, check=False)
    except FileNotFoundError as exc:
        raise GenotypeAssetError("本地 PLINK2 未就绪。请检查基因型处理容器是否已按新版镜像构建。") from exc
    if completed.returncode != 0:
        detail = (completed.stdout or "")[-2400:]
        raise GenotypeAssetError(f"PLINK2 处理失败：{detail}")


def _safe_extract_plink(raw_path: Path, destination: Path) -> tuple[Path, Path, Path, str]:
    try:
        archive = zipfile.ZipFile(raw_path)
    except zipfile.BadZipFile as exc:
        raise GenotypeAssetError("PLINK 数据必须是包含 .bed/.bim/.fam 的 ZIP。") from exc
    with archive:
        candidates: dict[str, dict[str, zipfile.ZipInfo]] = {}
        for member in archive.infolist():
            if member.is_dir() or member.file_size > MAX_UPLOAD_BYTES:
                continue
            name = Path(member.filename).name
            suffix = Path(name).suffix.lower()
            if suffix in {".bed", ".bim", ".fam"}:
                candidates.setdefault(Path(name).stem, {})[suffix] = member
        matches = [(prefix, files) for prefix, files in candidates.items() if {".bed", ".bim", ".fam"}.issubset(files)]
        if len(matches) != 1:
            raise GenotypeAssetError("ZIP 中必须恰好有一组同前缀的 .bed/.bim/.fam 文件。")
        prefix, files = matches[0]
        destination.mkdir(parents=True, exist_ok=True)
        result: dict[str, Path] = {}
        for suffix, member in files.items():
            target = destination / f"{prefix}{suffix}"
            with archive.open(member) as source, target.open("wb") as sink:
                shutil.copyfileobj(source, sink, length=1024 * 1024)
            result[suffix] = target
    if result[".bed"].read_bytes()[:3] != b"\x6c\x1b\x01":
        raise GenotypeAssetError("BED 不是 SNP-major PLINK 二进制文件。")
    return result[".bed"], result[".bim"], result[".fam"], prefix


def _parse_whitespace_table(path: Path) -> list[dict[str, str]]:
    lines = [line.split() for line in _read_text(path).splitlines() if line.strip()]
    if not lines:
        return []
    header = [column.lstrip("#") for column in lines[0]]
    values = lines[1:]
    return [dict(zip(header, row)) for row in values if len(row) >= len(header)]


def _numeric(value: str | None) -> float | None:
    try:
        return float(value) if value not in (None, "", "NA") else None
    except (TypeError, ValueError):
        return None


def _qc_diagnostics(prefix: Path, sample_pairs: list[tuple[str, str]], variant_count: int) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    # PLINK 1.9 writes .imiss/.lmiss, while PLINK2 writes .smiss/.vmiss.
    # Accept both formats so the QC reader remains compatible with either engine.
    sample_path = prefix.with_suffix(".smiss")
    if not sample_path.is_file():
        sample_path = prefix.with_suffix(".imiss")
    snp_path = prefix.with_suffix(".vmiss")
    if not snp_path.is_file():
        snp_path = prefix.with_suffix(".lmiss")
    sample_rows = _parse_whitespace_table(sample_path)
    het_rows = {(row.get("FID"), row.get("IID")): row for row in _parse_whitespace_table(prefix.with_suffix(".het"))}
    f_values = [_numeric(row.get("F")) for row in het_rows.values()]
    finite = [value for value in f_values if value is not None]
    mean = sum(finite) / len(finite) if finite else 0.0
    std = (sum((value - mean) ** 2 for value in finite) / len(finite)) ** 0.5 if finite else 0.0
    samples: list[dict[str, Any]] = []
    for row in sample_rows:
        key = (row.get("FID", ""), row.get("IID", ""))
        heterozygosity = _numeric(het_rows.get(key, {}).get("F"))
        z_score = ((heterozygosity - mean) / std) if heterozygosity is not None and std > 0 else None
        samples.append({"FID": key[0], "IID": key[1], "missing_rate": _numeric(row.get("F_MISS")), "heterozygosity_f": heterozygosity, "heterozygosity_z": z_score, "heterozygosity_flag": bool(z_score is not None and abs(z_score) > 3)})
    snps = []
    for row in _parse_whitespace_table(snp_path):
        snps.append({"chromosome": row.get("CHR") or row.get("CHROM", ""), "snp": row.get("ID", ""), "missing_rate": _numeric(row.get("F_MISS"))})
    summary = {
        "input_sample_count": len(sample_pairs), "input_variant_count": variant_count,
        "heterozygosity_outlier_count": sum(1 for row in samples if row["heterozygosity_flag"]),
        "heterozygosity_rule": "仅标记 |Z| > 3 的异常样本，不自动删除",
    }
    return summary, samples, snps


def _qc_report_value(key: str, value: Any) -> str:
    """Format machine summary values for a readable, narrow PDF table."""
    if key == "reference_confirmed":
        return "是" if value else "否，需核验参考坐标或染色体命名"
    if key == "formal_analysis_ready":
        return "是" if value else "否，仍需完成材料映射并人工发布"
    if key == "unknown_chromosomes":
        return "无" if not value else "、".join(map(str, value))
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (list, tuple)):
        return "无" if not value else "、".join(map(str, value))
    if value in (None, ""):
        return "-"
    return str(value)


def _qc_report_rows(summary: dict[str, Any]) -> list[tuple[str, str]]:
    fields = (
        ("input_sample_count", "输入样本数"),
        ("input_variant_count", "输入 SNP 数"),
        ("qc_sample_count", "质控后样本数"),
        ("qc_variant_count", "质控后 SNP 数"),
        ("removed_sample_count", "剔除样本数"),
        ("removed_variant_count", "剔除 SNP 数"),
        ("heterozygosity_outlier_count", "杂合率异常样本数"),
        ("heterozygosity_rule", "杂合率异常规则"),
        ("reference_confirmed", "参考坐标核验"),
        ("unknown_chromosomes", "未识别染色体命名"),
        ("formal_analysis_ready", "是否可进入正式分析"),
        ("allele_frequency_source", "等位基因频率来源"),
    )
    return [(label, _qc_report_value(key, summary.get(key))) for key, label in fields if key in summary]


def _qc_parameter_text(parameters: dict[str, Any]) -> str:
    return (
        f"MAF >= {parameters.get('maf_min', '-')}; "
        f"SNP 缺失率 <= {parameters.get('snp_missing_max', '-')}; "
        f"样本缺失率 <= {parameters.get('sample_missing_max', '-')}; "
        "HWE 仅用于诊断，不自动过滤。"
    )


def _write_csv(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def _style_workbook_sheet(sheet, widths: list[int]) -> None:
    header_fill = PatternFill("solid", fgColor="E9F3EF")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="124C3B")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _build_researcher_qc_exports(
    report_dir: Path,
    artifact_stem: str,
    version_no: int,
    samples: list[dict[str, Any]],
    snps: list[dict[str, Any]],
    mappings: list[Any],
) -> tuple[Path, Path, Path]:
    """Create readable exports while the internal PLINK-compatible tables stay unchanged."""
    sample_export = report_dir / f"{artifact_stem}_样本质控明细_v{version_no}.csv"
    snp_export = report_dir / f"{artifact_stem}_SNP质控明细_v{version_no}.csv"
    mapping_export = report_dir / f"{artifact_stem}_样本材料确认映射表_v{version_no}.csv"
    _write_csv(sample_export, [
        "家系编号（FID）", "样本编号（IID）", "样本缺失率（F_MISS，比例）",
        "近交系数/杂合度指标（PLINK F）", "杂合度异常 Z 分数", "杂合度异常提示",
    ], [[
        row["FID"], row["IID"], row["missing_rate"], row["heterozygosity_f"],
        row["heterozygosity_z"], "是" if row["heterozygosity_flag"] else "否",
    ] for row in samples])
    _write_csv(snp_export, ["染色体（CHR）", "SNP 标记（ID）", "SNP 缺失率（F_MISS，比例）"], [[
        row["chromosome"], row["snp"], row["missing_rate"],
    ] for row in snps])
    _write_csv(mapping_export, ["家系编号（FID）", "样本编号（IID）", "平台材料编码（material_code）", "映射备注（note）"], [[
        item.get("fid", "") if isinstance(item, dict) else item[0],
        item.get("iid", "") if isinstance(item, dict) else item[1],
        item.get("material_code", "") if isinstance(item, dict) else "",
        (item.get("note", "") if isinstance(item, dict) else "") or "人工确认材料映射",
    ] for item in mappings])
    return sample_export, snp_export, mapping_export


def _build_qc_artifacts(
    version_dir: Path,
    title: str,
    version_no: int,
    summary: dict[str, Any],
    samples: list[dict[str, Any]],
    snps: list[dict[str, Any]],
    mappings: list[Any],
    *,
    include_package: bool,
) -> tuple[Path, Path | None, Path, Path]:
    report_dir = version_dir / "qc-report"
    report_dir.mkdir(parents=True, exist_ok=True)
    sample_csv = report_dir / "sample_qc.csv"
    snp_csv = report_dir / "snp_qc.csv"
    mapping_csv = report_dir / "material_mapping_template.csv"
    with sample_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=["FID", "IID", "missing_rate", "heterozygosity_f", "heterozygosity_z", "heterozygosity_flag"])
        writer.writeheader(); writer.writerows(samples)
    with snp_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=["chromosome", "snp", "missing_rate"])
        writer.writeheader(); writer.writerows(snps)
    with mapping_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle); writer.writerow(["FID", "IID", "material_code", "note"])
        for item in mappings:
            if isinstance(item, dict):
                writer.writerow([
                    item.get("fid", ""), item.get("iid", ""), item.get("material_code", ""),
                    item.get("note", "") or "人工确认材料映射" if item.get("material_code") else "请从平台已有材料档案选择或填写可识别材料编码",
                ])
            else:
                fid, iid = item
                writer.writerow([fid, iid, "", "请从平台已有材料档案选择或填写可识别材料编码"])
    artifact_stem = _artifact_display_stem(title)
    book = Workbook(); sheet = book.active; sheet.title = "质控摘要"
    sheet.append(["项目", "结果"])
    for label, value in _qc_report_rows(summary):
        sheet.append([label, value])
    sheet.append(["本次质控阈值", _qc_parameter_text(summary.get("parameters") or {})])
    if summary.get("small_cohort_note"):
        sheet.append(["小样本提示", summary["small_cohort_note"]])
    _style_workbook_sheet(sheet, [28, 92])
    sample_sheet = book.create_sheet("样本质控明细")
    sample_sheet.append(["家系编号（FID）", "样本编号（IID）", "样本缺失率（F_MISS，比例）", "近交系数/杂合度指标（PLINK F）", "杂合度异常 Z 分数", "杂合度异常提示"])
    for row in samples:
        sample_sheet.append([row["FID"], row["IID"], row["missing_rate"], row["heterozygosity_f"], row["heterozygosity_z"], "是" if row["heterozygosity_flag"] else "否"])
    _style_workbook_sheet(sample_sheet, [22, 22, 24, 30, 22, 20])
    snp_sheet = book.create_sheet("SNP质控明细")
    snp_sheet.append(["染色体（CHR）", "SNP 标记（ID）", "SNP 缺失率（F_MISS，比例）"])
    for row in snps:
        snp_sheet.append([row["chromosome"], row["snp"], row["missing_rate"]])
    _style_workbook_sheet(snp_sheet, [18, 36, 28])
    guide = book.create_sheet("字段说明")
    guide.append(["字段", "科研阅读说明"])
    for item in [
        ("FID", "PLINK 家系编号；无家系信息时常与 IID 相同，保留原值以支持后续 PLINK/GWAS 复现。"),
        ("IID", "个体/样本编号；对应本次测序或芯片样本，不等同于平台材料编码。"),
        ("F_MISS", "缺失率，取值为比例；0.05 表示 5%。样本和 SNP 均使用该口径。"),
        ("PLINK F", "PLINK --het 输出的近交系数/杂合度指标，用于同一数据集内部异常筛查，不能直接当作普通杂合率百分比。"),
        ("Z 分数", "基于本批样本 PLINK F 的标准化分数；绝对值大于 3 时仅提示复核，不自动剔除。"),
        ("CHR / ID", "CHR 为染色体，ID 为 SNP 标记名；保留国际通行缩写，便于衔接 PLINK、GWAS 和参考基因组。"),
    ]:
        guide.append(item)
    _style_workbook_sheet(guide, [22, 104])
    workbook_path = report_dir / f"{artifact_stem}_水稻基因型质控汇总_v{version_no}.xlsx"; book.save(workbook_path)
    sample_export, snp_export, mapping_export = _build_researcher_qc_exports(
        report_dir, artifact_stem, version_no, samples, snps, mappings,
    )
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    pdf_path = report_dir / f"{artifact_stem}_水稻基因型质控报告_v{version_no}.pdf"
    styles = getSampleStyleSheet(); normal = styles["BodyText"]; normal.fontName = "STSong-Light"; normal.fontSize = 9.5; normal.leading = 15
    heading = styles["Heading1"]; heading.fontName = "STSong-Light"; heading.fontSize = 16
    subheading = styles["Heading2"]; subheading.fontName = "STSong-Light"; subheading.fontSize = 12; subheading.leading = 18
    table_label = styles["BodyText"].clone("QcTableLabel"); table_label.fontName = "STSong-Light"; table_label.fontSize = 8.5; table_label.leading = 12
    table_value = styles["BodyText"].clone("QcTableValue"); table_value.fontName = "STSong-Light"; table_value.fontSize = 8.5; table_value.leading = 12
    table_header = styles["BodyText"].clone("QcTableHeader"); table_header.fontName = "STSong-Light"; table_header.fontSize = 9; table_header.leading = 12
    story = [Paragraph(f"基因型导入与水稻专用质控报告：{title} v{version_no}", heading), Spacer(1, 5 * mm), Paragraph("本报告由本地 PLINK2 按固定、可追溯规则生成；杂合率异常仅提示，不自动删除。材料映射必须由科研人员确认后才能发布为正式分析版本。", normal), Spacer(1, 4 * mm)]
    table_rows = [[Paragraph("项目", table_header), Paragraph("结果", table_header)]] + [
        [Paragraph(escape(label), table_label), Paragraph(escape(value), table_value)]
        for label, value in _qc_report_rows(summary)
    ]
    table = Table(table_rows, colWidths=[54 * mm, 120 * mm]); table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9f3ef")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#b8d3c8")), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])); story.append(table)
    story.extend([
        Spacer(1, 5 * mm),
        Paragraph("本次质控阈值", subheading),
        Paragraph(escape(_qc_parameter_text(summary.get("parameters") or {})), normal),
    ])
    if summary.get("small_cohort_note"):
        notice = Table([[Paragraph(f"小样本科研解释提示：{escape(str(summary['small_cohort_note']))}", table_value)]], colWidths=[174 * mm])
        notice.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff7e6")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e8b85b")),
            ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.extend([Spacer(1, 4 * mm), notice])
    SimpleDocTemplate(str(pdf_path), pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=18 * mm, bottomMargin=18 * mm).build(story)
    package_path = None
    if include_package:
        package_path = report_dir / f"{artifact_stem}_水稻基因型正式质控结果包_v{version_no}.zip"
        with zipfile.ZipFile(package_path, "w", zipfile.ZIP_DEFLATED) as archive:
            for file_path in (pdf_path, sample_export, snp_export, mapping_export, workbook_path):
                archive.write(file_path, file_path.name)
    return pdf_path, package_path, workbook_path, mapping_csv


def _archive_qc_result(session: Session, owner_id: str, project_id: str, version_id: str, title: str, package_path: Path, summary: dict[str, Any]) -> None:
    existing = session.execute(text("""
        SELECT id FROM research_result WHERE owner_id = :owner_id AND analysis_run_id = :version_id AND result_type = 'genotype_qc_package'
    """), {"owner_id": owner_id, "version_id": version_id}).scalar()
    if existing:
        session.execute(text("""
            UPDATE research_result
            SET project_id=:project_id, title=:title, file_name=:file_name, size_bytes=:size, storage_path=:path,
                summary=:summary, metadata=CAST(:metadata AS jsonb)
            WHERE id=:id
        """), {"id": existing, "project_id": project_id, "title": f"{title} · 基因型导入与质控结果包",
               "file_name": package_path.name, "size": package_path.stat().st_size, "path": str(package_path),
               "summary": "材料映射已完整确认后的正式质控结果包，包含 PDF、样本/SNP 质控表、已确认映射表和处理工作簿；PLINK 三件套仅供后续受控分析调用。",
               "metadata": _json({"generated_from": "local_genotype_qc", "formal_analysis_ready": True, "qc_summary": summary})})
        return
    session.execute(text("""
        INSERT INTO research_result
        (id, owner_id, project_id, session_id, source_message_id, analysis_run_id, result_type, title, content_type, file_name, size_bytes, storage_path, summary, metadata, created_at)
        VALUES (:id, :owner_id, :project_id, NULL, NULL, :version_id, 'genotype_qc_package', :title, 'application/zip', :file_name, :size, :path, :summary, CAST(:metadata AS jsonb), :created_at)
    """), {"id": str(uuid.uuid4()), "owner_id": owner_id, "project_id": project_id, "version_id": version_id, "title": f"{title} · 基因型导入与质控结果包",
             "file_name": package_path.name, "size": package_path.stat().st_size, "path": str(package_path),
             "summary": "材料映射已完整确认后的正式质控结果包，包含 PDF、样本/SNP 质控表、已确认映射表和处理工作簿；PLINK 三件套仅供后续受控分析调用。",
             "metadata": _json({"generated_from": "local_genotype_qc", "formal_analysis_ready": True, "qc_summary": summary}), "created_at": _now()})


def run_next_job(session: Session, storage_dir: Path) -> bool:
    """Process one queued conversion/QC job. Designed for the persistent worker."""
    job = session.execute(text("""
        SELECT * FROM genotype_processing_job WHERE status = 'queued'
        ORDER BY created_at FOR UPDATE SKIP LOCKED LIMIT 1
    """)).mappings().first()
    if not job:
        return False
    session.execute(text("UPDATE genotype_processing_job SET status='running', progress_label='正在转换为统一 PLINK 格式', started_at=:now WHERE id=:id"), {"now": _now(), "id": job["id"]})
    session.execute(text("UPDATE genotype_asset SET status='processing', updated_at=:now WHERE id=:id"), {"now": _now(), "id": job["asset_id"]})
    session.execute(text("UPDATE genotype_asset_version SET status='processing', updated_at=:now WHERE id=:id"), {"now": _now(), "id": job["version_id"]})
    session.commit()
    try:
        asset = session.execute(text("SELECT * FROM genotype_asset WHERE id=:id"), {"id": job["asset_id"]}).mappings().one()
        version = session.execute(text("SELECT * FROM genotype_asset_version WHERE id=:id"), {"id": job["version_id"]}).mappings().one()
        snapshot = version["source_snapshot"] or {}
        raw_path = Path(str(snapshot.get("raw_storage_path") or ""))
        if not raw_path.is_file(): raise GenotypeAssetError("原始基因型文件不存在，无法继续处理。")
        version_dir = _asset_dir(storage_dir, str(job["owner_id"]), str(job["asset_id"])) / f"version-{version['version_number']}"
        work_dir = version_dir / "work"; work_dir.mkdir(parents=True, exist_ok=True)
        plink_dir = version_dir / "plink"; plink_dir.mkdir(parents=True, exist_ok=True)
        if asset["source_format"] == "vcf":
            _run([os.getenv("PLINK2_BINARY", "plink2"), "--vcf", str(raw_path), "--vcf-half-call", "missing", "--double-id", "--allow-extra-chr", "--make-bed", "--out", str(work_dir / "imported")], work_dir)
            prefix = "imported"
            input_bed, input_bim, input_fam = work_dir / "imported.bed", work_dir / "imported.bim", work_dir / "imported.fam"
        else:
            input_bed, input_bim, input_fam, prefix = _safe_extract_plink(raw_path, work_dir)
        if not input_bed.is_file() or not input_bim.is_file() or not input_fam.is_file():
            raise GenotypeAssetError("PLINK 转换未生成完整的 .bed/.bim/.fam 文件。")
        unknown_chromosomes = _normalize_bim(input_bim)
        sample_pairs = _fam_pairs(input_fam); variant_count, remaining_unknown = _bim_count(input_bim)
        unknown_chromosomes = sorted(set(unknown_chromosomes + remaining_unknown))
        session.execute(text("UPDATE genotype_processing_job SET progress_label='正在生成样本与 SNP 质控诊断' WHERE id=:id"), {"id": job["id"]}); session.commit()
        diagnostics = work_dir / "diagnostics"
        allele_frequency = work_dir / "allele-frequency"
        # PLINK2 requires an explicit reference frequency file for --het/--hardy
        # when a small cohort cannot provide reliable allele-frequency estimates.
        _run([
            os.getenv("PLINK2_BINARY", "plink2"), "--bfile", str(input_bed.with_suffix("")),
            "--allow-extra-chr", "--freq", "--out", str(allele_frequency),
        ], work_dir)
        _run([
            os.getenv("PLINK2_BINARY", "plink2"), "--bfile", str(input_bed.with_suffix("")),
            "--allow-extra-chr", "--read-freq", str(allele_frequency.with_suffix(".afreq")),
            "--missing", "--het", "--hardy", "--out", str(diagnostics),
        ], work_dir)
        diagnostic_summary, sample_rows, snp_rows = _qc_diagnostics(diagnostics, sample_pairs, variant_count)
        session.execute(text("UPDATE genotype_processing_job SET progress_label='正在执行水稻常规育种材料 QC v1.0' WHERE id=:id"), {"id": job["id"]}); session.commit()
        qc_prefix = plink_dir / "qc"
        _run([os.getenv("PLINK2_BINARY", "plink2"), "--bfile", str(input_bed.with_suffix("")), "--allow-extra-chr", "--geno", "0.05", "--mind", "0.05", "--maf", "0.05", "--make-bed", "--out", str(qc_prefix)], work_dir)
        qc_samples = _fam_pairs(qc_prefix.with_suffix(".fam")); qc_variants, qc_unknown = _bim_count(qc_prefix.with_suffix(".bim"))
        unknown_chromosomes = sorted(set(unknown_chromosomes + qc_unknown))
        if not qc_samples or not qc_variants:
            raise GenotypeAssetError("按当前 QC 规则过滤后没有可用于分析的样本或 SNP，请调整数据或由管理员新增合适模板。")
        summary = {
            **diagnostic_summary, "qc_sample_count": len(qc_samples), "qc_variant_count": qc_variants,
            "removed_sample_count": len(sample_pairs) - len(qc_samples), "removed_variant_count": variant_count - qc_variants,
            "unknown_chromosomes": unknown_chromosomes, "reference_confirmed": not bool(unknown_chromosomes),
            "formal_analysis_ready": False,
            "allele_frequency_source": "当前导入队列估计值",
            "small_cohort_note": (
                "当前样本少于 50 个：频率仅用于完成本次缺失率、杂合率和 HWE 诊断；"
                "不得据此形成群体频率或 HWE 的正式科研结论，正式分析建议加载同类群体参考频率。"
                if len(sample_pairs) < 50 else ""
            ),
            "parameters": {"maf_min": 0.05, "snp_missing_max": 0.05, "sample_missing_max": 0.05, "hwe": "diagnostic_only"},
        }
        pdf_path, _, _, _ = _build_qc_artifacts(
            version_dir, str(asset["title"]), int(version["version_number"]), summary,
            sample_rows, snp_rows, qc_samples, include_package=False,
        )
        for fid, iid in qc_samples:
            session.execute(text("""
                INSERT INTO genotype_sample_mapping(id, version_id, owner_id, fid, iid, status, note, updated_at)
                VALUES (:id, :version_id, :owner_id, :fid, :iid, 'unmapped', '', :now)
                ON CONFLICT(version_id, fid, iid) DO NOTHING
            """), {"id": str(uuid.uuid4()), "version_id": version["id"], "owner_id": job["owner_id"], "fid": fid, "iid": iid, "now": _now()})
        next_status = "awaiting_mapping" if not unknown_chromosomes else "reference_review_required"
        session.execute(text("""
            UPDATE genotype_asset_version SET status=:status, qc_summary=CAST(:summary AS jsonb), plink_prefix='qc', plink_directory=:plink_dir,
              report_path=:report_path, package_path=:package_path, updated_at=:now WHERE id=:id
        """), {"status": next_status, "summary": _json(summary), "plink_dir": str(plink_dir), "report_path": str(pdf_path), "package_path": None, "now": _now(), "id": version["id"]})
        session.execute(text("UPDATE genotype_asset SET status=:status, updated_at=:now WHERE id=:id"), {"status": next_status, "now": _now(), "id": asset["id"]})
        session.execute(text("UPDATE genotype_processing_job SET status='completed', progress_label='质控已完成，等待科研人员确认材料映射', completed_at=:now WHERE id=:id"), {"now": _now(), "id": job["id"]})
        session.commit()
    except Exception as exc:
        detail = str(exc)
        session.execute(text("UPDATE genotype_processing_job SET status='failed', progress_label='处理失败，请查看错误说明', error_detail=:detail, completed_at=:now WHERE id=:id"), {"detail": detail[:8000], "now": _now(), "id": job["id"]})
        session.execute(text("UPDATE genotype_asset_version SET status='failed', updated_at=:now WHERE id=:id"), {"now": _now(), "id": job["version_id"]})
        session.execute(text("UPDATE genotype_asset SET status='failed', updated_at=:now WHERE id=:id"), {"now": _now(), "id": job["asset_id"]})
        session.commit()
    return True


def list_material_suggestions(session: Session, keyword: str, limit: int = 30) -> list[dict[str, Any]]:
    term = (keyword or "").strip()
    pattern = f"%{term}%"
    rows = session.execute(text("""
        SELECT id, material_code, material_name, aliases, material_type
        FROM breeding_material
        WHERE :term = '' OR material_code ILIKE :pattern OR material_name ILIKE :pattern OR CAST(aliases AS text) ILIKE :pattern
        ORDER BY material_code LIMIT :limit
    """), {"term": term, "pattern": pattern, "limit": limit}).mappings().all()
    return [{"id": row["id"], "material_code": row["material_code"], "material_name": row["material_name"], "aliases": row["aliases"] or [], "material_type": row["material_type"]} for row in rows]


def update_mapping(session: Session, owner_id: str, version_id: str, fid: str, iid: str, payload: MappingUpdateRequest) -> dict[str, Any]:
    version = session.execute(text("SELECT * FROM genotype_asset_version WHERE id=:id"), {"id": version_id}).mappings().first()
    if not version:
        raise GenotypeAssetError("未找到基因型版本。")
    if version["status"] == "analysis_ready":
        raise GenotypeAssetError("已发布分析版本不可修改。请创建材料映射修订版，保留原版本的可追溯性。")
    mapping = session.execute(text("SELECT id FROM genotype_sample_mapping WHERE version_id=:version_id AND fid=:fid AND iid=:iid"), {"version_id": version_id, "fid": fid, "iid": iid}).scalar()
    if not mapping:
        raise GenotypeAssetError("该样本不属于当前基因型 QC 结果。")
    status, material_id = "unmapped", None
    if payload.material_id:
        material = session.execute(text("SELECT id FROM breeding_material WHERE id=:id"), {"id": payload.material_id}).scalar()
        if not material:
            raise GenotypeAssetError("选择的材料档案不存在；请提交材料主档补充申请。")
        duplicate = session.execute(text("""
            SELECT iid FROM genotype_sample_mapping WHERE version_id=:version_id AND material_id=:material_id
              AND NOT (fid=:fid AND iid=:iid) AND status='mapped'
        """), {"version_id": version_id, "material_id": payload.material_id, "fid": fid, "iid": iid}).scalar()
        if duplicate:
            raise GenotypeAssetError(f"该材料已映射到样本 {duplicate}。首版不支持同一材料多个 DNA 样本或技术重复，不能静默合并。")
        material_id, status = payload.material_id, "mapped"
    session.execute(text("""
        UPDATE genotype_sample_mapping SET material_id=:material_id, status=:status, note=:note, updated_at=:now WHERE id=:id
    """), {"material_id": material_id, "status": status, "note": payload.note.strip(), "now": _now(), "id": mapping})
    return _version_summary(session, version_id)


def batch_update_mapping(session: Session, owner_id: str, version_id: str, content: bytes) -> dict[str, Any]:
    """Apply a small, explicit CSV mapping sheet without guessing material IDs."""
    try:
        rows = list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
    except UnicodeDecodeError:
        rows = list(csv.DictReader(io.StringIO(content.decode("gb18030"))))
    if not rows:
        raise GenotypeAssetError("映射文件没有可用数据行。请使用系统提供的 CSV 模板。")
    def read_column(row: dict[str, str], names: tuple[str, ...]) -> str:
        for name in names:
            value = row.get(name)
            if value not in (None, ""):
                return str(value).strip()
        return ""

    required_columns = (
        ("FID", "家系编号（FID）", "家系编号"),
        ("IID", "样本编号（IID）", "样本编号"),
        ("material_code", "平台材料编码（material_code）", "平台材料编码"),
    )
    headers = set(rows[0].keys() if rows else [])
    if any(not headers.intersection(names) for names in required_columns):
        raise GenotypeAssetError("映射 CSV 必须包含家系编号（FID）、样本编号（IID）、平台材料编码（material_code）三列。")
    applied, issues = 0, []
    for index, row in enumerate(rows, start=2):
        fid = read_column(row, required_columns[0])
        iid = read_column(row, required_columns[1])
        material_code = read_column(row, required_columns[2])
        if not fid or not iid:
            issues.append({"row": index, "message": "FID 或 IID 为空"})
            continue
        material_id = None
        if material_code:
            material_id = session.execute(text("SELECT id FROM breeding_material WHERE material_code=:code"), {"code": material_code}).scalar()
            if not material_id:
                issues.append({"row": index, "message": f"未找到材料编码 {material_code}"})
                continue
        try:
            note = read_column(row, ("note", "映射备注（note）", "映射备注"))
            update_mapping(session, owner_id, version_id, fid, iid, MappingUpdateRequest(material_id=str(material_id) if material_id else None, note=note))
            applied += 1
        except GenotypeAssetError as exc:
            issues.append({"row": index, "message": str(exc)})
    return {"applied": applied, "issues": issues, "version": _version_summary(session, version_id)}


def build_phenotype_template(session: Session, version_id: str) -> bytes:
    """Return a version-specific phenotype sheet with QC-retained sample identities."""
    version = session.execute(text("SELECT status FROM genotype_asset_version WHERE id=:id"), {"id": version_id}).mappings().first()
    if not version or version["status"] != "analysis_ready":
        raise GenotypeAssetError("请先完成材料映射并发布分析就绪基因型版本，再下载专用表型模板。")
    rows = session.execute(text("""
        SELECT mapping.fid, mapping.iid, material.material_code, material.material_name
        FROM genotype_sample_mapping mapping
        JOIN breeding_material material ON material.id=mapping.material_id
        WHERE mapping.version_id=:version_id AND mapping.status='mapped'
        ORDER BY material.material_code, mapping.iid
    """), {"version_id": version_id}).mappings().all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "连续性状表型"
    sheet.append(["FID", "IID", "material_code", "material_name", "analysis_environment", "trait_value"])
    for row in rows:
        sheet.append([row["fid"], row["iid"], row["material_code"], row["material_name"], "请填写明确分析环境", "请填写连续性状数值"])
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 22
    sheet.column_dimensions["E"].width = 28
    sheet.column_dimensions["F"].width = 20
    guide = workbook.create_sheet("填写说明")
    guide.append(["项目", "要求"])
    guide.append(["适用范围", "首版仅用于一个明确分析环境中的连续性状；多年多点或多重复数据请先由区域试验资料包治理，或计算 BLUP 后再导入。"])
    guide.append(["样本标识", "FID、IID、材料编码由系统预填，不要修改；只填写 analysis_environment 和 trait_value。"])
    guide.append(["性状值", "trait_value 必须是可计算的连续数值，例如株高、产量、千粒重或垩白度。"])
    guide.append(["材料新建", "若材料未建档或样本映射冲突，请提交数据治理申请，不要自行修改内部材料编码。"])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_mapping_template(session: Session, version_id: str) -> bytes:
    """Return the version-specific mapping CSV without exposing raw genotype files."""
    rows = session.execute(text("""
        SELECT mapping.fid, mapping.iid, material.material_code, mapping.note
        FROM genotype_sample_mapping mapping
        LEFT JOIN breeding_material material ON material.id = mapping.material_id
        WHERE mapping.version_id = :version_id
        ORDER BY mapping.iid, mapping.fid
    """), {"version_id": version_id}).mappings().all()
    if not rows:
        raise GenotypeAssetError("当前版本尚未生成可映射的 QC 样本。")
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["家系编号（FID）", "样本编号（IID）", "平台材料编码（material_code）", "映射备注（note）"])
    for row in rows:
        writer.writerow([row["fid"], row["iid"], row["material_code"] or "", row["note"] or ""])
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def publish_analysis_ready(session: Session, owner_id: str, version_id: str) -> dict[str, Any]:
    version = session.execute(text("SELECT * FROM genotype_asset_version WHERE id=:id"), {"id": version_id}).mappings().first()
    if not version:
        raise GenotypeAssetError("未找到当前账号的基因型版本。")
    if version["status"] == "analysis_ready": return _version_summary(session, version_id)
    if version["status"] != "awaiting_mapping":
        raise GenotypeAssetError("当前版本尚未完成可发布前置条件。请先完成质控、参考坐标确认和材料映射。")
    rows = session.execute(text("SELECT * FROM genotype_sample_mapping WHERE version_id=:version_id"), {"version_id": version_id}).mappings().all()
    if not rows or any(row["status"] != "mapped" or not row["material_id"] for row in rows):
        raise GenotypeAssetError("仍有 QC 合格样本未映射到材料档案；请逐条确认或提交数据治理申请。")
    materials = [str(row["material_id"]) for row in rows]
    if len(materials) != len(set(materials)):
        raise GenotypeAssetError("存在同一材料对应多个样本的重复映射；首版不支持技术重复，请先拆分或提交治理申请。")
    summary = dict(version["qc_summary"] or {}); summary["formal_analysis_ready"] = True
    asset = session.execute(text("SELECT title FROM genotype_asset WHERE id=:id"), {"id": version["asset_id"]}).mappings().one()
    version_dir = Path(str(version["plink_directory"] or "")).parent
    report_dir = version_dir / "qc-report"
    try:
        with (report_dir / "sample_qc.csv").open("r", encoding="utf-8-sig", newline="") as handle:
            sample_rows = list(csv.DictReader(handle))
        with (report_dir / "snp_qc.csv").open("r", encoding="utf-8-sig", newline="") as handle:
            snp_rows = list(csv.DictReader(handle))
    except OSError as exc:
        raise GenotypeAssetError("无法读取预质控明细，不能生成正式结果包；请重新执行质控。") from exc
    confirmed_mappings = session.execute(text("""
        SELECT mapping.fid, mapping.iid, mapping.note, material.material_code
        FROM genotype_sample_mapping mapping
        JOIN breeding_material material ON material.id = mapping.material_id
        WHERE mapping.version_id=:version_id
        ORDER BY mapping.iid, mapping.fid
    """), {"version_id": version_id}).mappings().all()
    pdf_path, package_path, _, _ = _build_qc_artifacts(
        version_dir, str(asset["title"]), int(version["version_number"]), summary,
        sample_rows, snp_rows, [dict(row) for row in confirmed_mappings], include_package=True,
    )
    if package_path is None:
        raise GenotypeAssetError("未能生成正式质控结果包；请稍后重试发布。")
    now = _now()
    session.execute(text("""
        UPDATE genotype_asset_version
        SET status='analysis_ready', qc_summary=CAST(:summary AS jsonb), report_path=:report_path,
            package_path=:package_path, published_at=:now, published_by=:owner_id, updated_at=:now
        WHERE id=:id
    """), {"summary": _json(summary), "report_path": str(pdf_path), "package_path": str(package_path), "now": now, "owner_id": owner_id, "id": version_id})
    session.execute(text("UPDATE genotype_asset SET status='analysis_ready', current_version_id=:version_id, updated_at=:now WHERE id=:asset_id"), {"version_id": version_id, "now": now, "asset_id": version["asset_id"]})
    _archive_qc_result(session, owner_id, str(version["project_id"]), version_id, str(asset["title"]), package_path, summary)
    return _version_summary(session, version_id)


def create_mapping_revision(session: Session, owner_id: str, asset_id: str, version_id: str) -> dict[str, Any]:
    previous = session.execute(text("SELECT * FROM genotype_asset_version WHERE id=:id AND asset_id=:asset_id"), {"id": version_id, "asset_id": asset_id}).mappings().first()
    if not previous:
        raise GenotypeAssetError("未找到可修订的基因型版本。")
    new_id = str(uuid.uuid4())
    number = int(session.execute(text("SELECT COALESCE(MAX(version_number), 0) + 1 FROM genotype_asset_version WHERE asset_id=:asset_id"), {"asset_id": asset_id}).scalar_one())
    session.execute(text("""
        INSERT INTO genotype_asset_version
        (id, asset_id, owner_id, version_number, status, parent_version_id, qc_template_code, qc_template_version, reference_assembly, source_snapshot, qc_summary, plink_prefix, plink_directory, report_path, package_path, created_at, updated_at)
        VALUES (:id, :asset_id, :owner_id, :number, 'awaiting_mapping', :parent_id, :template_code, :template_version, :assembly, CAST(:source_snapshot AS jsonb), CAST(:qc_summary AS jsonb), :plink_prefix, :plink_directory, :report_path, :package_path, :now, :now)
    """), {"id": new_id, "asset_id": asset_id, "owner_id": owner_id, "number": number, "parent_id": previous["id"], "template_code": previous["qc_template_code"], "template_version": previous["qc_template_version"], "assembly": previous["reference_assembly"], "source_snapshot": _json(previous["source_snapshot"] or {}), "qc_summary": _json(previous["qc_summary"] or {}), "plink_prefix": previous["plink_prefix"], "plink_directory": previous["plink_directory"], "report_path": previous["report_path"], "package_path": previous["package_path"], "now": _now()})
    previous_mappings = session.execute(text("SELECT * FROM genotype_sample_mapping WHERE version_id=:version_id"), {"version_id": version_id}).mappings().all()
    for item in previous_mappings:
        session.execute(text("""
            INSERT INTO genotype_sample_mapping(id, version_id, owner_id, fid, iid, material_id, status, note, updated_at)
            VALUES (:id, :version_id, :owner_id, :fid, :iid, :material_id, :status, :note, :now)
        """), {"id": str(uuid.uuid4()), "version_id": new_id, "owner_id": owner_id, "fid": item["fid"], "iid": item["iid"], "material_id": item["material_id"], "status": item["status"], "note": item["note"], "now": _now()})
    session.execute(text("UPDATE genotype_asset SET current_version_id=:version_id, status='awaiting_mapping', updated_at=:now WHERE id=:asset_id"), {"version_id": new_id, "now": _now(), "asset_id": asset_id})
    return _version_summary(session, new_id)


def create_governance_request(session: Session, owner_id: str, asset_id: str, version_id: str, payload: GovernanceRequestCreate) -> dict[str, Any]:
    request_id = str(uuid.uuid4())
    version = session.execute(text("""
        SELECT asset.title, version.version_number
        FROM genotype_asset_version version
        JOIN genotype_asset asset ON asset.id = version.asset_id
        WHERE version.id = :version_id AND asset.id = :asset_id
    """), {"asset_id": asset_id, "version_id": version_id}).mappings().first()
    if not version:
        raise GenotypeAssetError("未找到可提交治理申请的基因型版本。")
    session.execute(text("""
        INSERT INTO genotype_governance_request
        (id, owner_id, asset_id, version_id, asset_title_snapshot, version_number_snapshot, request_type, description, affected_samples)
        VALUES
        (:id, :owner_id, :asset_id, :version_id, :asset_title, :version_number, :request_type, :description, CAST(:samples AS jsonb))
    """), {
        "id": request_id,
        "owner_id": owner_id,
        "asset_id": asset_id,
        "version_id": version_id,
        "asset_title": version["title"],
        "version_number": version["version_number"],
        "request_type": payload.request_type,
        "description": payload.description.strip(),
        "samples": _json(payload.affected_samples),
    })
    return {"id": request_id, "status": "submitted", "message": "数据治理申请已提交。原始基因型文件仍保持私有，处理人员仅会看到申请范围与经授权的任务文件。"}


def list_governance_requests(session: Session) -> list[dict[str, Any]]:
    """Return metadata-only requests for the data-processing workbench.

    This deliberately omits storage paths, hashes and raw-file names. The current
    first version lets a processor maintain master-data work only; authorising raw
    genotype transfer is a separate formal workflow for the delivery version.
    """
    rows = session.execute(text("""
        SELECT request.id, request.owner_id, request.asset_id, request.version_id,
               request.request_type, request.status, request.description,
               request.affected_samples, request.created_at, request.resolved_at,
               request.resolved_by, request.resolution_note,
               request.asset_title_snapshot, request.version_number_snapshot
        FROM genotype_governance_request request
        ORDER BY CASE request.status
          WHEN 'submitted' THEN 0 WHEN 'needs_info' THEN 1 WHEN 'accepted' THEN 2 ELSE 3 END,
          request.created_at DESC
    """)).mappings().all()
    return [{
        "id": row["id"],
        "asset_id": row["asset_id"],
        "version_id": row["version_id"],
        "asset_title": row["asset_title_snapshot"] or "未命名基因型资产",
        "version_number": row["version_number_snapshot"],
        "request_type": row["request_type"],
        "status": row["status"],
        "description": row["description"],
        "affected_samples": row["affected_samples"] or [],
        "created_at": row["created_at"].isoformat() if row["created_at"] else None,
        "resolved_at": row["resolved_at"].isoformat() if row["resolved_at"] else None,
        "resolved_by": row["resolved_by"],
        "resolution_note": row["resolution_note"],
    } for row in rows]


def resolve_governance_request(
    session: Session,
    request_id: str,
    processor: str,
    payload: GovernanceRequestResolution,
) -> dict[str, Any]:
    row = session.execute(text("SELECT id FROM genotype_governance_request WHERE id=:id"), {"id": request_id}).scalar()
    if not row:
        raise GenotypeAssetError("未找到该基因型数据治理申请。")
    resolved_at = _now() if payload.status == "resolved" else None
    session.execute(text("""
        UPDATE genotype_governance_request
        SET status=:status, resolution_note=:note, resolved_by=:processor, resolved_at=:resolved_at
        WHERE id=:id
    """), {
        "id": request_id,
        "status": payload.status,
        "note": payload.resolution_note.strip(),
        "processor": processor,
        "resolved_at": resolved_at,
    })
    return {"id": request_id, "status": payload.status, "resolution_note": payload.resolution_note.strip()}


def analysis_ready_versions(session: Session) -> list[dict[str, Any]]:
    rows = session.execute(text("""
        SELECT id FROM genotype_asset_version WHERE status='analysis_ready' ORDER BY published_at DESC
    """)).scalars().all()
    return [_version_summary(session, str(version_id)) for version_id in rows]


def artifact_path(session: Session, version_id: str, kind: str) -> tuple[Path, str]:
    version = session.execute(text("SELECT status, report_path, package_path FROM genotype_asset_version WHERE id=:id"), {"id": version_id}).mappings().first()
    if not version:
        raise GenotypeAssetError("未找到基因型版本。")
    if kind == "package" and version["status"] != "analysis_ready":
        raise GenotypeAssetError("正式结果包将在材料映射全部确认并人工发布后开放下载。当前可下载预质控报告用于核验。")
    key = "report_path" if kind == "report" else "package_path"
    path = Path(str(version[key] or ""))
    if not path.is_file():
        raise GenotypeAssetError("该质控产物尚未生成或已被清理。")
    return path, "application/pdf" if kind == "report" else "application/zip"


def worker_loop(database_url: str, storage_dir: Path, poll_seconds: float = 2.0) -> None:
    """Run forever inside the compose worker container."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    engine = create_engine(database_url, pool_pre_ping=True)
    maker = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    while True:
        with maker() as session:
            try:
                processed = run_next_job(session, storage_dir)
            except Exception:
                session.rollback()
                processed = False
        if not processed:
            time.sleep(poll_seconds)
