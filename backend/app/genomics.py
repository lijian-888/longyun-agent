"""Controlled intake and confirmation gates for rice GWAS.

This module deliberately implements a narrow workflow: a continuous rice trait,
one PLINK binary genotype package, optional researcher-supplied covariates, and
a fixed association plan.  It does *not* turn chat text into shell commands.
The actual bioinformatics runner is a separately deployed, allow-listed service;
this intake layer is responsible for making a job runnable only after the
researcher has reviewed the exact inputs and defaults.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import shutil
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from pydantic import BaseModel, Field
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from .local_gwas import LocalGwasError, run_local_gwas


MAX_GENOTYPE_ARCHIVE_BYTES = 512 * 1024 * 1024
MAX_TABULAR_BYTES = 32 * 1024 * 1024
MAX_UNCOMPRESSED_ARCHIVE_BYTES = 1024 * 1024 * 1024
WORKFLOW_CODE = "rice_gwas_lmm_v1"
DEFAULT_PROJECT_ID = "00000000-0000-4000-8000-000000000001"
DEFAULT_PARAMETERS = {
    "maf": 0.05,
    "snp_missing_rate": 0.05,
    "sample_missing_rate": 0.05,
    "ld_pruning": "50 5 0.2",
    "principal_components": 3,
    "hwe_filter": "disabled_for_rice_default",
    "candidate_window_kb": 100,
    "association_model": "mixed_linear_model",
}


class GenomicsError(ValueError):
    """A researcher-facing validation error for the controlled GWAS flow."""


class CreateGwasPlanRequest(BaseModel):
    trait_name: str = Field(min_length=1, max_length=120)
    reference_assembly: str = Field(default="IRGSP-1.0", min_length=1, max_length=120)
    candidate_window_kb: int = Field(default=100, ge=1, le=2000)
    purpose: str = Field(default="寻找与连续性状相关的候选位点", min_length=1, max_length=500)


class AttachGenotypeAssetRequest(BaseModel):
    asset_id: str = Field(min_length=36, max_length=36)
    version_id: str = Field(min_length=36, max_length=36)


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)


def ensure_genomics_schema(session: Session) -> None:
    """Create the private plan table and enforce the same owner isolation as chat."""
    session.execute(text("""
        CREATE TABLE IF NOT EXISTS gwas_analysis_plan (
          id VARCHAR(36) PRIMARY KEY,
          owner_id VARCHAR(120) NOT NULL,
          project_id VARCHAR(36) NOT NULL DEFAULT '00000000-0000-4000-8000-000000000001',
          status VARCHAR(40) NOT NULL DEFAULT 'collecting',
          trait_name VARCHAR(120) NOT NULL,
          reference_assembly VARCHAR(120) NOT NULL,
          purpose TEXT NOT NULL,
          workflow_code VARCHAR(80) NOT NULL,
          parameters JSONB NOT NULL DEFAULT '{}'::jsonb,
          input_manifest JSONB NOT NULL DEFAULT '{}'::jsonb,
          preflight JSONB NOT NULL DEFAULT '{}'::jsonb,
          confirmation JSONB NOT NULL DEFAULT '{}'::jsonb,
          result_manifest JSONB NOT NULL DEFAULT '{}'::jsonb,
          created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
        )
    """))
    session.execute(text("ALTER TABLE gwas_analysis_plan ADD COLUMN IF NOT EXISTS result_manifest JSONB NOT NULL DEFAULT '{}'::jsonb"))
    session.execute(text("ALTER TABLE gwas_analysis_plan ADD COLUMN IF NOT EXISTS project_id VARCHAR(36) NOT NULL DEFAULT '00000000-0000-4000-8000-000000000001'"))
    session.execute(text("CREATE INDEX IF NOT EXISTS ix_gwas_analysis_plan_owner_updated ON gwas_analysis_plan(owner_id, updated_at DESC)"))
    session.execute(text("ALTER TABLE gwas_analysis_plan ENABLE ROW LEVEL SECURITY"))
    session.execute(text("ALTER TABLE gwas_analysis_plan FORCE ROW LEVEL SECURITY"))
    session.execute(text("DROP POLICY IF EXISTS gwas_analysis_plan_owner_only ON gwas_analysis_plan"))
    session.execute(text("""
        CREATE POLICY gwas_analysis_plan_owner_only ON gwas_analysis_plan
        FOR ALL
        USING (owner_id = current_setting('app.research_user_id', true)
               AND project_id = current_setting('app.project_id', true))
        WITH CHECK (owner_id = current_setting('app.research_user_id', true)
                    AND project_id = current_setting('app.project_id', true))
    """))


def _safe_name(name: str, fallback: str) -> str:
    cleaned = Path(str(name or fallback)).name
    cleaned = re.sub(r"[^A-Za-z0-9._()\-]+", "_", cleaned).strip("._")
    return cleaned or fallback


def _private_dir(storage_dir: Path, owner_id: str, plan_id: str) -> Path:
    path = storage_dir / "genomics" / owner_id / plan_id
    path.mkdir(parents=True, exist_ok=True)
    return path


def _write_private_file(directory: Path, filename: str, content: bytes) -> dict[str, Any]:
    final_path = directory / _safe_name(filename, "upload.bin")
    temporary_path = final_path.with_suffix(f"{final_path.suffix}.tmp")
    temporary_path.write_bytes(content)
    temporary_path.replace(final_path)
    return {
        "file_name": final_path.name,
        "storage_path": str(final_path),
        "size_bytes": len(content),
        "sha256": hashlib.sha256(content).hexdigest(),
    }


def _row_to_dict(row: Any) -> dict[str, Any]:
    row = dict(row)
    return {
        "id": row["id"],
        "project_id": row["project_id"],
        "status": row["status"],
        "trait_name": row["trait_name"],
        "reference_assembly": row["reference_assembly"],
        "purpose": row["purpose"],
        "workflow_code": row["workflow_code"],
        "parameters": row["parameters"] or {},
        "input_manifest": row["input_manifest"] or {},
        "preflight": row["preflight"] or {},
        "confirmation": row["confirmation"] or {},
        "result_manifest": row["result_manifest"] or {},
        "created_at": row["created_at"].isoformat(),
        "updated_at": row["updated_at"].isoformat(),
    }


def _plan_row(session: Session, plan_id: str) -> Any:
    row = session.execute(text("SELECT * FROM gwas_analysis_plan WHERE id = :id"), {"id": plan_id}).mappings().first()
    if not row:
        raise GenomicsError("未找到当前研究人员的 GWAS 分析计划。")
    return row


def _manifest_value(row: Any, key: str, default: Any) -> Any:
    manifest = row["input_manifest"] or {}
    return manifest.get(key, default)


def _require_collecting_plan(plan: Any) -> None:
    if plan["status"] != "collecting":
        raise GenomicsError("分析计划已确认或已提交；为保证审计一致性，不能再替换输入文件。请新建一份计划。")


def _preflight(manifest: dict[str, Any], parameters: dict[str, Any]) -> dict[str, Any]:
    genotype = manifest.get("genotype") or {}
    phenotype = manifest.get("phenotype") or {}
    missing: list[str] = []
    warnings: list[str] = []
    if not genotype:
        missing.append("PLINK 基因型 ZIP（含同前缀 bed/bim/fam）")
    if not phenotype:
        missing.append("包含 FID、IID 和连续性状值的表型文件")
    shared = int((phenotype or {}).get("matched_sample_count") or 0)
    if phenotype and shared <= 0:
        missing.append("可与 PLINK 样本匹配的表型记录")
    if phenotype and shared and shared < 80:
        warnings.append("可匹配表型样本少于 80；可以做流程演示，但统计效力可能有限。")
    genotyping_rate = genotype.get("genotyping_rate") if genotype else None
    if genotyping_rate is not None and float(genotyping_rate) < 0.9:
        warnings.append("导入前总基因型调用率低于 90%；请重点审阅缺失率质控结果。")
    if manifest.get("covariates"):
        warnings.append("自定义协变量会与基因型 PCA 一并纳入固定模型；请确认其不会与性状定义重复。")
    return {
        "status": "ready" if not missing else "incomplete",
        "missing": missing,
        "warnings": warnings,
        "matched_sample_count": shared,
        "workflow_defaults": parameters,
        "next_action": "确认分析计划" if not missing else "继续补充必填输入",
    }


def _update_plan(session: Session, plan_id: str, *, manifest: dict[str, Any], preflight: dict[str, Any], status: str | None = None, confirmation: dict[str, Any] | None = None, results: dict[str, Any] | None = None) -> Any:
    updates = {
        "manifest": _json(manifest),
        "preflight": _json(preflight),
        "updated_at": _utcnow(),
        "id": plan_id,
    }
    fragments = ["input_manifest = CAST(:manifest AS jsonb)", "preflight = CAST(:preflight AS jsonb)", "updated_at = :updated_at"]
    if status:
        fragments.append("status = :status")
        updates["status"] = status
    if confirmation is not None:
        fragments.append("confirmation = CAST(:confirmation AS jsonb)")
        updates["confirmation"] = _json(confirmation)
    if results is not None:
        fragments.append("result_manifest = CAST(:results AS jsonb)")
        updates["results"] = _json(results)
    session.execute(text(f"UPDATE gwas_analysis_plan SET {', '.join(fragments)} WHERE id = :id"), updates)
    session.flush()
    return _plan_row(session, plan_id)


def create_plan(session: Session, owner_id: str, payload: CreateGwasPlanRequest, project_id: str = DEFAULT_PROJECT_ID) -> dict[str, Any]:
    plan_id = str(uuid.uuid4())
    parameters = {**DEFAULT_PARAMETERS, "candidate_window_kb": payload.candidate_window_kb}
    session.execute(text("""
        INSERT INTO gwas_analysis_plan
        (id, owner_id, project_id, status, trait_name, reference_assembly, purpose, workflow_code, parameters, input_manifest, preflight, confirmation, created_at, updated_at)
        VALUES (:id, :owner_id, :project_id, 'collecting', :trait_name, :reference_assembly, :purpose, :workflow_code,
                CAST(:parameters AS jsonb), '{}'::jsonb, CAST(:preflight AS jsonb), '{}'::jsonb, :now, :now)
    """), {
        "id": plan_id,
        "owner_id": owner_id,
        "project_id": project_id,
        "trait_name": payload.trait_name.strip(),
        "reference_assembly": payload.reference_assembly.strip(),
        "purpose": payload.purpose.strip(),
        "workflow_code": WORKFLOW_CODE,
        "parameters": _json(parameters),
        "preflight": _json(_preflight({}, parameters)),
        "now": _utcnow(),
    })
    session.flush()
    return _row_to_dict(_plan_row(session, plan_id))


def list_plans(session: Session) -> list[dict[str, Any]]:
    rows = session.execute(text("SELECT * FROM gwas_analysis_plan ORDER BY updated_at DESC")).mappings().all()
    return [_row_to_dict(row) for row in rows]


def get_plan(session: Session, plan_id: str) -> dict[str, Any]:
    return _row_to_dict(_plan_row(session, plan_id))


def _read_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise GenomicsError("无法读取文本文件编码；请导出为 UTF-8 CSV 或 TSV。")


def _count_fam(content: bytes) -> tuple[int, set[tuple[str, str]]]:
    pairs: set[tuple[str, str]] = set()
    for number, line in enumerate(_read_text(content).splitlines(), start=1):
        if not line.strip():
            continue
        parts = line.split()
        if len(parts) < 6:
            raise GenomicsError(f".fam 第 {number} 行少于 6 列，无法识别 FID/IID。")
        pairs.add((parts[0], parts[1]))
    if not pairs:
        raise GenomicsError(".fam 中没有可用样本。")
    return len(pairs), pairs


def _count_bim(content: bytes) -> int:
    count = 0
    for number, line in enumerate(_read_text(content).splitlines(), start=1):
        if not line.strip():
            continue
        if len(line.split()) < 4:
            raise GenomicsError(f".bim 第 {number} 行少于 4 列。")
        count += 1
    if not count:
        raise GenomicsError(".bim 中没有可用 SNP 位点。")
    return count


def upload_genotype(session: Session, owner_id: str, plan_id: str, filename: str, content: bytes, storage_dir: Path) -> dict[str, Any]:
    row = _plan_row(session, plan_id)
    _require_collecting_plan(row)
    if len(content) > MAX_GENOTYPE_ARCHIVE_BYTES:
        raise GenomicsError("基因型 ZIP 超过 512 MB；请通过受控大文件导入通道提交。")
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile as exc:
        raise GenomicsError("基因型必须以 ZIP 上传，并包含同前缀的 bed、bim、fam 文件。") from exc
    with archive:
        members = [item for item in archive.infolist() if not item.is_dir()]
        if sum(item.file_size for item in members) > MAX_UNCOMPRESSED_ARCHIVE_BYTES:
            raise GenomicsError("解压后的基因型文件超过 1 GB，已拒绝以防止压缩包滥用。")
        grouped: dict[str, dict[str, zipfile.ZipInfo]] = {}
        for item in members:
            basename = Path(item.filename).name
            suffix = Path(basename).suffix.lower()
            if suffix not in {".bed", ".bim", ".fam"}:
                continue
            grouped.setdefault(Path(basename).stem, {})[suffix] = item
        matches = [(prefix, items) for prefix, items in grouped.items() if {".bed", ".bim", ".fam"}.issubset(items)]
        if len(matches) != 1:
            raise GenomicsError("ZIP 中必须恰好有一组同前缀的 .bed/.bim/.fam 文件。")
        prefix, items = matches[0]
        bed = archive.read(items[".bed"])
        bim = archive.read(items[".bim"])
        fam = archive.read(items[".fam"])
    if len(bed) < 3 or bed[:3] != b"\x6c\x1b\x01":
        raise GenomicsError(".bed 文件不是标准 SNP-major 二进制 PLINK 文件。")
    sample_count, sample_pairs = _count_fam(fam)
    variant_count = _count_bim(bim)
    directory = _private_dir(storage_dir, owner_id, plan_id)
    saved = _write_private_file(directory, "genotype.zip", content)
    metadata = {
        **saved,
        "source_archive_name": _safe_name(filename, "genotype.zip"),
        "prefix": prefix,
        "sample_count": sample_count,
        "variant_count": variant_count,
        "sample_pairs": ["\t".join(pair) for pair in sorted(sample_pairs)],
        "format": "PLINK_BED_BIM_FAM",
    }
    manifest = dict(row["input_manifest"] or {})
    manifest["genotype"] = metadata
    parameters = dict(row["parameters"] or DEFAULT_PARAMETERS)
    return _row_to_dict(_update_plan(session, plan_id, manifest=manifest, preflight=_preflight(manifest, parameters), status="collecting"))


def attach_analysis_ready_genotype(
    session: Session,
    owner_id: str,
    plan_id: str,
    asset_id: str,
    version_id: str,
    storage_dir: Path,
) -> dict[str, Any]:
    """Attach a published private QC version to a GWAS plan without exposing raw PLINK files.

    The runner still receives a plan-local ZIP because it is intentionally isolated from the
    genotype workspace.  The ZIP is an internal copy, not a browser download endpoint.
    """
    plan = _plan_row(session, plan_id)
    _require_collecting_plan(plan)
    version = session.execute(text("""
        SELECT version.*, asset.title AS asset_title, asset.reference_assembly AS asset_reference_assembly
        FROM genotype_asset_version version
        JOIN genotype_asset asset ON asset.id = version.asset_id
        WHERE version.id = :version_id AND asset.id = :asset_id
    """), {"version_id": version_id, "asset_id": asset_id}).mappings().first()
    if not version:
        raise GenomicsError("未找到可由当前科研账号使用的基因型版本。")
    if version["status"] != "analysis_ready":
        raise GenomicsError("请选择已完成样本映射并发布为“分析就绪”的基因型版本。")
    if str(plan["reference_assembly"]).strip() != str(version["asset_reference_assembly"]).strip():
        raise GenomicsError(
            f"参考版本不一致：当前 GWAS 计划为 {plan['reference_assembly']}，基因型版本为 {version['asset_reference_assembly']}。"
            "请新建使用相同参考版本的计划，避免混用坐标。"
        )
    summary = dict(version["qc_summary"] or {})
    if not summary.get("reference_confirmed", False):
        raise GenomicsError("该版本存在未确认的染色体/参考坐标，不能进入正式 GWAS。请先提交数据治理申请。")
    mappings = session.execute(text("""
        SELECT mapping.fid, mapping.iid, material.material_code, material.material_name
        FROM genotype_sample_mapping mapping
        JOIN breeding_material material ON material.id = mapping.material_id
        WHERE mapping.version_id = :version_id AND mapping.status = 'mapped'
        ORDER BY mapping.fid, mapping.iid
    """), {"version_id": version_id}).mappings().all()
    expected_count = int(summary.get("qc_sample_count") or 0)
    if not mappings or len(mappings) != expected_count:
        raise GenomicsError("该版本仍有未映射 QC 样本，不能开始 GWAS。请返回基因型数据工作台完成材料映射。")
    material_codes = [str(item["material_code"]) for item in mappings]
    if len(material_codes) != len(set(material_codes)):
        raise GenomicsError("发现一个材料对应多个 DNA 样本。首版不自动合并技术重复，请提交数据治理申请后再使用。")
    plink_directory = Path(str(version["plink_directory"] or ""))
    prefix = str(version["plink_prefix"] or "qc")
    source_files = [plink_directory / f"{prefix}{suffix}" for suffix in (".bed", ".bim", ".fam")]
    if not all(path.is_file() for path in source_files):
        raise GenomicsError("分析就绪版本的内部 PLINK 文件不完整，请重新执行该版本质控。")
    directory = _private_dir(storage_dir, owner_id, plan_id)
    archive_path = directory / "genotype.zip"
    with zipfile.ZipFile(archive_path, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for source in source_files:
            archive.write(source, arcname=source.name)
    saved = {
        "file_name": archive_path.name,
        "storage_path": str(archive_path),
        "size_bytes": archive_path.stat().st_size,
        "sha256": hashlib.sha256(archive_path.read_bytes()).hexdigest(),
    }
    manifest = dict(plan["input_manifest"] or {})
    manifest["genotype"] = {
        **saved,
        "source_archive_name": f"{version['asset_title']} · QC v{version['version_number']}",
        "prefix": prefix,
        "sample_count": len(mappings),
        "variant_count": int(summary.get("qc_variant_count") or 0),
        "sample_pairs": [f"{item['fid']}\t{item['iid']}" for item in mappings],
        "format": "PLINK_BED_BIM_FAM",
        "source_asset_id": asset_id,
        "source_version_id": version_id,
        "source_qc_template": summary.get("template_name"),
        "source_qc_summary": summary,
        "internal_copy_only": True,
    }
    parameters = dict(plan["parameters"] or DEFAULT_PARAMETERS)
    parameters["reference_assembly"] = str(plan["reference_assembly"])
    return _row_to_dict(_update_plan(session, plan_id, manifest=manifest, preflight=_preflight(manifest, parameters), status="collecting"))


def _table_rows(content: bytes, filename: str) -> tuple[list[dict[str, str]], list[str]]:
    if filename.lower().endswith(".xlsx"):
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
        except Exception as exc:
            raise GenomicsError("无法读取 XLSX 表型文件，请使用系统下载的模板重新保存后上传。") from exc
        iterator = sheet.iter_rows(values_only=True)
        raw_headers = next(iterator, None)
        headers = [str(value).strip() if value is not None else "" for value in (raw_headers or [])]
        if not headers or not any(headers):
            raise GenomicsError("表型文件缺少表头。")
        rows: list[dict[str, str]] = []
        for values in iterator:
            row = {
                headers[index]: "" if value is None else str(value).strip()
                for index, value in enumerate(values[:len(headers)])
                if headers[index]
            }
            if any(value != "" for value in row.values()):
                rows.append(row)
        return rows, headers
    if filename.lower().endswith(".xls"):
        raise GenomicsError("首版请将 .xls 转存为 .xlsx、CSV 或 TSV 后上传，避免旧格式解析差异。")
    text_content = _read_text(content)
    delimiter = "\t" if filename.lower().endswith((".tsv", ".txt")) else ","
    reader = csv.DictReader(io.StringIO(text_content), delimiter=delimiter)
    headers = reader.fieldnames or []
    if not headers:
        raise GenomicsError("表格缺少表头。")
    return [dict(row) for row in reader], headers


def _rows_as_tsv(rows: list[dict[str, str]], headers: list[str]) -> bytes:
    """The local runner consumes one stable TSV form regardless of uploaded spreadsheet format."""
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers, delimiter="\t", lineterminator="\n", extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({header: row.get(header, "") for header in headers})
    return buffer.getvalue().encode("utf-8")


def upload_phenotype(session: Session, owner_id: str, plan_id: str, filename: str, trait_column: str, content: bytes, storage_dir: Path) -> dict[str, Any]:
    plan = _plan_row(session, plan_id)
    _require_collecting_plan(plan)
    if len(content) > MAX_TABULAR_BYTES:
        raise GenomicsError("表型文件超过 32 MB；请先使用专用数据导入流程治理。")
    rows, headers = _table_rows(content, filename)
    normalized = {name.lower(): name for name in headers}
    environment_column = normalized.get("analysis_environment")
    required = [normalized.get("fid"), normalized.get("iid"), environment_column, trait_column]
    if not all(required) or trait_column not in headers:
        raise GenomicsError("表型文件必须包含 FID、IID、analysis_environment 和指定的连续性状列。")
    genotype = _manifest_value(plan, "genotype", {})
    if not genotype:
        raise GenomicsError("请先上传并校验 PLINK 基因型 ZIP，再上传表型文件。")
    genotype_pairs = {tuple(item.split("\t", 1)) for item in genotype.get("sample_pairs", [])}
    valid_rows = 0
    matched_pairs: set[tuple[str, str]] = set()
    environments: set[str] = set()
    observed_pairs: set[tuple[str, str]] = set()
    for number, row in enumerate(rows, start=2):
        value = str(row.get(trait_column, "")).strip()
        if value in {"", "NA", "N/A", ".", "-9"}:
            continue
        try:
            float(value)
        except ValueError as exc:
            raise GenomicsError(f"表型文件第 {number} 行的 {trait_column} 不是连续数值。") from exc
        valid_rows += 1
        pair = (str(row[required[0]]).strip(), str(row[required[1]]).strip())
        environment = str(row.get(environment_column, "")).strip()
        if not environment:
            raise GenomicsError(f"表型文件第 {number} 行缺少 analysis_environment。首版 GWAS 每次只接受一个明确分析环境。")
        if pair in observed_pairs:
            raise GenomicsError(f"表型文件第 {number} 行的 FID/IID 与前一行重复。请对同一样本只保留一个分析环境的一条性状值。")
        observed_pairs.add(pair)
        environments.add(environment)
        if pair in genotype_pairs:
            matched_pairs.add(pair)
    if not valid_rows:
        raise GenomicsError("表型文件中没有可用于连续性状 GWAS 的数值。")
    if len(environments) != 1:
        raise GenomicsError("首版 GWAS 每个表型文件只能包含一个 analysis_environment。多年多点或多重复数据请先通过区域试验资料包治理或计算 BLUP 后再上传。")
    directory = _private_dir(storage_dir, owner_id, plan_id)
    # Preserve the original filename in the manifest, but execute only against a normalized TSV.
    saved = _write_private_file(directory, "phenotype.tsv", _rows_as_tsv(rows, headers))
    manifest = dict(plan["input_manifest"] or {})
    manifest["phenotype"] = {
        **saved,
        "source_file_name": _safe_name(filename, "phenotype.csv"),
        "trait_column": trait_column,
        "row_count": len(rows),
        "numeric_trait_count": valid_rows,
        "matched_sample_count": len(matched_pairs),
        "unmatched_sample_count": max(0, valid_rows - len(matched_pairs)),
        "analysis_environment": next(iter(environments)),
        "headers": headers,
    }
    parameters = dict(plan["parameters"] or DEFAULT_PARAMETERS)
    return _row_to_dict(_update_plan(session, plan_id, manifest=manifest, preflight=_preflight(manifest, parameters), status="collecting"))


def upload_covariates(session: Session, owner_id: str, plan_id: str, filename: str, content: bytes, storage_dir: Path) -> dict[str, Any]:
    plan = _plan_row(session, plan_id)
    _require_collecting_plan(plan)
    if len(content) > MAX_TABULAR_BYTES:
        raise GenomicsError("协变量文件超过 32 MB。")
    rows, headers = _table_rows(content, filename)
    lower = {name.lower() for name in headers}
    if not {"fid", "iid"}.issubset(lower) or len(headers) < 3:
        raise GenomicsError("协变量文件必须包含 FID、IID 及至少一个协变量列。")
    directory = _private_dir(storage_dir, owner_id, plan_id)
    saved = _write_private_file(directory, "covariates.tsv", _rows_as_tsv(rows, headers))
    manifest = dict(plan["input_manifest"] or {})
    manifest["covariates"] = {**saved, "source_file_name": _safe_name(filename, "covariates.csv"), "row_count": len(rows), "headers": headers}
    parameters = dict(plan["parameters"] or DEFAULT_PARAMETERS)
    return _row_to_dict(_update_plan(session, plan_id, manifest=manifest, preflight=_preflight(manifest, parameters), status="collecting"))


def confirm_plan(session: Session, owner_id: str, plan_id: str) -> dict[str, Any]:
    plan = _plan_row(session, plan_id)
    _require_collecting_plan(plan)
    parameters = dict(plan["parameters"] or DEFAULT_PARAMETERS)
    manifest = dict(plan["input_manifest"] or {})
    preflight = _preflight(manifest, parameters)
    if preflight["status"] != "ready":
        raise GenomicsError("输入信息尚未完整或没有可匹配样本，不能确认并执行 GWAS。")
    confirmation = {
        "confirmed_by": owner_id,
        "confirmed_at": _utcnow().isoformat(),
        "confirmed_workflow": WORKFLOW_CODE,
        "confirmed_parameters": parameters,
    }
    return _row_to_dict(_update_plan(session, plan_id, manifest=manifest, preflight=preflight, status="confirmed", confirmation=confirmation))


def request_execution(session: Session, plan_id: str) -> dict[str, Any]:
    plan = _plan_row(session, plan_id)
    if plan["status"] != "confirmed":
        raise GenomicsError("请先确认分析计划；未确认的输入不允许提交到生信计算环境。")
    manifest = dict(plan["input_manifest"] or {})
    parameters = dict(plan["parameters"] or DEFAULT_PARAMETERS)
    preflight = _preflight(manifest, parameters)
    confirmation = dict(plan["confirmation"] or {})
    confirmation["execution_requested_at"] = _utcnow().isoformat()
    confirmation["execution_mode"] = "local_fixed_p3d_runner"
    confirmation["execution_started_at"] = _utcnow().isoformat()
    result = _update_plan(session, plan_id, manifest=manifest, preflight=preflight, status="running", confirmation=confirmation)
    return _row_to_dict(result)


def run_requested_local_execution(session: Session, plan_id: str, storage_dir: Path) -> dict[str, Any]:
    """Run the locked plan locally and persist only deterministic result files."""
    plan = _plan_row(session, plan_id)
    if plan["status"] != "running":
        raise GenomicsError("本地 GWAS 只能执行已确认并已启动的计划。")
    manifest = dict(plan["input_manifest"] or {})
    genotype = manifest.get("genotype") or {}
    phenotype = manifest.get("phenotype") or {}
    covariates = manifest.get("covariates") or {}
    owner_id = str(plan["owner_id"])
    private_dir = _private_dir(storage_dir, owner_id, plan_id)
    genotype_path = private_dir / str(genotype.get("file_name", "genotype.zip"))
    phenotype_path = private_dir / str(phenotype.get("file_name", "phenotype.csv"))
    covariate_path = private_dir / str(covariates.get("file_name")) if covariates else None
    if not genotype_path.is_file() or not phenotype_path.is_file() or (covariate_path and not covariate_path.is_file()):
        raise GenomicsError("已锁定的输入文件不完整，无法启动本地 GWAS。")
    parameters = dict(plan["parameters"] or DEFAULT_PARAMETERS)
    parameters["reference_assembly"] = str(plan["reference_assembly"])
    confirmation = dict(plan["confirmation"] or {})
    try:
        results = run_local_gwas(
            genotype_zip=genotype_path,
            phenotype_path=phenotype_path,
            trait_column=str(phenotype["trait_column"]),
            covariate_path=covariate_path,
            parameters=parameters,
            output_dir=private_dir / "results",
        )
    except Exception as exc:
        confirmation["execution_failed_at"] = _utcnow().isoformat()
        confirmation["execution_error"] = str(exc)
        failed = _update_plan(session, plan_id, manifest=manifest, preflight=_preflight(manifest, parameters), status="failed", confirmation=confirmation, results={})
        payload = _row_to_dict(failed)
        payload["execution_note"] = f"本地 GWAS 未完成：{exc}"
        return payload
    confirmation["execution_completed_at"] = _utcnow().isoformat()
    completed = _update_plan(session, plan_id, manifest=manifest, preflight=_preflight(manifest, parameters), status="completed", confirmation=confirmation, results=results)
    payload = _row_to_dict(completed)
    payload["execution_note"] = "本地 GWAS 已完成；可下载质控摘要、PCA、Manhattan、QQ 和候选位点结果。"
    return payload


def local_result_file(session: Session, plan_id: str, file_key: str, storage_dir: Path) -> tuple[Path, str]:
    plan = _plan_row(session, plan_id)
    results = dict(plan["result_manifest"] or {})
    files = {str(item.get("key")): item for item in results.get("files", [])}
    item = files.get(file_key)
    if not item:
        raise GenomicsError("未找到该 GWAS 结果文件。")
    path = _private_dir(storage_dir, str(plan["owner_id"]), plan_id) / "results" / str(item.get("file_name", ""))
    if not path.is_file():
        raise GenomicsError("GWAS 结果文件不存在或已被清理。")
    return path, str(item.get("mime_type") or "application/octet-stream")


def local_result_bundle(session: Session, plan_id: str, storage_dir: Path) -> tuple[bytes, str, dict[str, Any]]:
    """Create one ZIP containing exactly the completed plan's allow-listed outputs."""
    plan = _plan_row(session, plan_id)
    if plan["status"] != "completed":
        raise GenomicsError("只有已完成的 GWAS 分析才能归档到结果库。")
    results = dict(plan["result_manifest"] or {})
    files = list(results.get("files") or [])
    if not files:
        raise GenomicsError("该 GWAS 计划没有可归档的结果文件。")
    plan_directory = _private_dir(storage_dir, str(plan["owner_id"]), plan_id) / "results"
    folder_name = f"rice_gwas_{plan_id[:8]}"
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for item in files:
            file_name = Path(str(item.get("file_name") or "")).name
            source = plan_directory / file_name
            if not file_name or not source.is_file():
                raise GenomicsError(f"GWAS 结果文件缺失：{file_name or '未知文件'}")
            archive.writestr(f"{folder_name}/{file_name}", source.read_bytes())
    file_name = f"rice_gwas_{plan_id[:8]}_results.zip"
    metadata = {
        "plan_id": plan_id,
        "trait_name": plan["trait_name"],
        "reference_assembly": plan["reference_assembly"],
        "file_count": len(files),
        "folder_name": folder_name,
        "summary": results.get("summary") or {},
    }
    return buffer.getvalue(), file_name, metadata
