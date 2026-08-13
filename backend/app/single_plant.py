"""Single-plant identity, phenotype, genotype mapping and selection services.

The existing platform already stores governed material-level trial facts.  This
module adds the missing identity layer between a breeding material and one
physical plant.  The migration is deliberately compatible with the current
schema: existing rows remain valid because every new foreign-key column is
nullable, and the legacy field-survey uniqueness rules are left intact.
"""

from __future__ import annotations

import hashlib
import io
import json
import re
import uuid
from datetime import datetime, timezone
from itertools import islice
from typing import Any, Literal

from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session


MAX_IMPORT_ROWS = 5000
MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_PHENOTYPE_IMPORT_ROWS = 20000
SAMPLE_TYPES = {"individual_plant", "line", "plot_composite", "tissue", "dna_sample"}
DATA_STATUSES = {"draft", "published", "archived"}
SELECTION_STATUSES = {"candidate", "retained", "observed", "eliminated", "promoted"}
OBSERVATION_SOURCE_TYPES = {"manual", "device", "import", "derived"}
OBSERVATION_QUALITY_STATUSES = {"pending", "passed", "warning", "rejected"}
BASE_SHOWCASE_PROGRAM_CODE = "JX-RICE-DEMO-2021"
BASE_SHOWCASE_PACKAGE_CODE = "RICE-MET-2023-2025-DEMO"

VARIETY_EVALUATION_TRAITS = [
    {"code": "plant_height", "name": "株高", "unit": "cm", "direction": "target", "weight": 0.05, "target": 105.0, "penalty": 4.0},
    {"code": "panicle_length", "name": "穗长", "unit": "cm", "direction": "higher", "weight": 0.05},
    {"code": "effective_panicles", "name": "有效穗数", "unit": "穗/株", "direction": "higher", "weight": 0.10},
    {"code": "grains_per_panicle", "name": "每穗粒数", "unit": "粒/穗", "direction": "higher", "weight": 0.10},
    {"code": "seed_setting_rate", "name": "结实率", "unit": "%", "direction": "higher", "weight": 0.15},
    {"code": "thousand_grain_weight", "name": "千粒重", "unit": "g", "direction": "higher", "weight": 0.10},
    {"code": "yield_per_plant", "name": "单株产量", "unit": "g/株", "direction": "higher", "weight": 0.30},
    {"code": "lodging_grade", "name": "倒伏等级", "unit": "级", "direction": "lower", "weight": 0.15},
]


class SinglePlantError(ValueError):
    """A user-facing validation or reference error."""


class SinglePlantImportPayload(BaseModel):
    project_id: str = Field(min_length=36, max_length=36)
    records: list[dict[str, Any]] = Field(min_length=1, max_length=MAX_IMPORT_ROWS)
    mode: Literal["create_only", "upsert"] = "create_only"


class SinglePlantObservationRequest(BaseModel):
    observation_stage: str = Field(min_length=1, max_length=100)
    trait_code: str = Field(min_length=1, max_length=100)
    trait_name: str = Field(min_length=1, max_length=200)
    value_numeric: float | None = None
    value_text: str | None = Field(default=None, max_length=2000)
    unit: str = Field(default="", max_length=50)
    source_type: Literal["manual", "device", "import", "derived"] = "manual"
    quality_status: Literal["pending", "passed", "warning", "rejected"] = "passed"
    quality_flags: list[str] = Field(default_factory=list, max_length=100)
    observed_at: datetime | None = None
    device_id: str | None = Field(default=None, max_length=160)


class SinglePlantGenotypeMappingRequest(BaseModel):
    sample_id: str | None = Field(default=None, max_length=36)
    note: str = Field(default="", max_length=1000)


class SinglePlantSelectionRequest(BaseModel):
    decision: Literal["retained", "observed", "eliminated", "promoted"]
    selection_criterion: str = Field(min_length=1, max_length=4000)
    selection_year: int | None = Field(default=None, ge=1900, le=2200)
    selection_site: str | None = Field(default=None, max_length=200)
    retention_reason: str | None = Field(default=None, max_length=4000)
    evidence_summary: str | None = Field(default=None, max_length=4000)
    source_record_no: str | None = Field(default=None, max_length=120)
    source_note: str | None = Field(default=None, max_length=4000)
    generation_record_id: str | None = Field(default=None, max_length=36)


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)


def _as_dict(row: Any) -> dict[str, Any]:
    return dict(row._mapping) if row is not None else {}


def _clean(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _optional_text(value: Any) -> str | None:
    cleaned = _clean(value)
    return cleaned or None


def _constraint(statement: str) -> str:
    """Wrap one ALTER TABLE ADD CONSTRAINT in an idempotent PostgreSQL block."""
    table_name, constraint_name = re.search(
        r"ALTER TABLE\s+(\w+)\s+ADD CONSTRAINT\s+(\w+)", statement, flags=re.IGNORECASE
    ).groups()
    return f"""
    DO $single_plant_migration$
    BEGIN
      IF NOT EXISTS (
        SELECT 1 FROM pg_constraint
        WHERE conname = '{constraint_name}' AND conrelid = '{table_name}'::regclass
      ) THEN
        {statement};
      END IF;
    END
    $single_plant_migration$
    """


def ensure_single_plant_schema(session: Session) -> None:
    """Source-manage field-survey tables and add the single-plant identity layer."""
    statements = (
        """
        CREATE TABLE IF NOT EXISTS field_survey_task (
          id VARCHAR(36) PRIMARY KEY,
          task_code VARCHAR(120) NOT NULL UNIQUE,
          task_name VARCHAR(300) NOT NULL,
          trial_name VARCHAR(300) NOT NULL,
          site_name VARCHAR(200) NOT NULL,
          survey_stage VARCHAR(100) NOT NULL,
          status VARCHAR(40) NOT NULL DEFAULT 'active',
          required_traits JSONB NOT NULL DEFAULT '[]'::jsonb,
          assigned_workers JSONB NOT NULL DEFAULT '[]'::jsonb,
          photo_required BOOLEAN NOT NULL DEFAULT TRUE,
          offline_package_version INTEGER NOT NULL DEFAULT 1,
          is_simulated BOOLEAN NOT NULL DEFAULT FALSE,
          description TEXT,
          created_by VARCHAR(160) NOT NULL,
          created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS field_survey_plot (
          id VARCHAR(36) PRIMARY KEY,
          task_id VARCHAR(36) NOT NULL REFERENCES field_survey_task(id) ON DELETE CASCADE,
          plot_code VARCHAR(160) NOT NULL,
          sequence_no INTEGER NOT NULL,
          block_no INTEGER NOT NULL,
          replicate_no INTEGER NOT NULL,
          material_code VARCHAR(100) NOT NULL,
          material_name VARCHAR(200) NOT NULL,
          qr_token VARCHAR(200) NOT NULL UNIQUE,
          assigned_to VARCHAR(160) NOT NULL,
          status VARCHAR(40) NOT NULL DEFAULT 'assigned',
          quality_status VARCHAR(40) NOT NULL DEFAULT 'pending',
          started_at TIMESTAMPTZ,
          completed_at TIMESTAMPTZ,
          completed_by VARCHAR(160),
          device_id VARCHAR(160),
          data_version INTEGER NOT NULL DEFAULT 1,
          UNIQUE(task_id, plot_code)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS field_survey_observation (
          id VARCHAR(36) PRIMARY KEY,
          task_id VARCHAR(36) NOT NULL REFERENCES field_survey_task(id) ON DELETE CASCADE,
          plot_id VARCHAR(36) NOT NULL REFERENCES field_survey_plot(id) ON DELETE CASCADE,
          trait_code VARCHAR(100) NOT NULL,
          trait_name VARCHAR(200) NOT NULL,
          value_numeric DOUBLE PRECISION,
          value_text TEXT,
          unit VARCHAR(50) NOT NULL,
          source_type VARCHAR(40) NOT NULL DEFAULT 'manual',
          device_id VARCHAR(160),
          quality_status VARCHAR(40) NOT NULL DEFAULT 'passed',
          quality_flags JSONB NOT NULL DEFAULT '[]'::jsonb,
          observed_by VARCHAR(160) NOT NULL,
          observed_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          data_version INTEGER NOT NULL DEFAULT 1,
          UNIQUE(plot_id, trait_code, data_version)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS field_survey_photo (
          id VARCHAR(36) PRIMARY KEY,
          task_id VARCHAR(36) NOT NULL REFERENCES field_survey_task(id) ON DELETE CASCADE,
          plot_id VARCHAR(36) NOT NULL REFERENCES field_survey_plot(id) ON DELETE CASCADE,
          trait_code VARCHAR(100) NOT NULL DEFAULT 'leaf_blast_score',
          original_file_name VARCHAR(500) NOT NULL,
          stored_file_name VARCHAR(500),
          storage_path TEXT,
          content_type VARCHAR(160),
          size_bytes BIGINT NOT NULL DEFAULT 0,
          sha256 VARCHAR(64),
          captured_by VARCHAR(160) NOT NULL,
          captured_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          device_id VARCHAR(160),
          data_version INTEGER NOT NULL DEFAULT 1,
          is_simulated BOOLEAN NOT NULL DEFAULT FALSE
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS field_survey_audit (
          id VARCHAR(36) PRIMARY KEY,
          task_id VARCHAR(36) NOT NULL REFERENCES field_survey_task(id) ON DELETE CASCADE,
          plot_id VARCHAR(36) REFERENCES field_survey_plot(id) ON DELETE CASCADE,
          event_type VARCHAR(80) NOT NULL,
          event_label VARCHAR(300) NOT NULL,
          actor VARCHAR(160) NOT NULL,
          device_id VARCHAR(160),
          data_version INTEGER,
          details JSONB NOT NULL DEFAULT '{}'::jsonb,
          created_at TIMESTAMPTZ NOT NULL DEFAULT now()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS biological_sample (
          id VARCHAR(36) PRIMARY KEY,
          project_id VARCHAR(36) REFERENCES research_project(id),
          program_id VARCHAR(36) NOT NULL REFERENCES breeding_program(id),
          material_id VARCHAR(36) NOT NULL REFERENCES breeding_material(id),
          trial_entry_id VARCHAR(36) REFERENCES trial_entry(id),
          sample_code VARCHAR(160) NOT NULL,
          sample_type VARCHAR(40) NOT NULL DEFAULT 'individual_plant',
          generation_label VARCHAR(100),
          plant_no INTEGER,
          parent_sample_id VARCHAR(36) REFERENCES biological_sample(id),
          selection_status VARCHAR(40) NOT NULL DEFAULT 'candidate',
          data_status VARCHAR(40) NOT NULL DEFAULT 'draft',
          source_note TEXT,
          metadata JSONB NOT NULL DEFAULT '{}'::jsonb,
          created_by VARCHAR(160) NOT NULL,
          created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
          UNIQUE(program_id, sample_code),
          CONSTRAINT ck_biological_sample_type CHECK (
            sample_type IN ('individual_plant','line','plot_composite','tissue','dna_sample')
          ),
          CONSTRAINT ck_biological_sample_status CHECK (
            selection_status IN ('candidate','retained','observed','eliminated','promoted')
          ),
          CONSTRAINT ck_biological_sample_data_status CHECK (
            data_status IN ('draft','published','archived')
          ),
          CONSTRAINT ck_biological_sample_not_own_parent CHECK (parent_sample_id IS NULL OR parent_sample_id <> id)
        )
        """,
        "ALTER TABLE field_survey_plot ADD COLUMN IF NOT EXISTS trial_entry_id VARCHAR(36)",
        "ALTER TABLE biological_sample ADD COLUMN IF NOT EXISTS project_id VARCHAR(36)",
        "ALTER TABLE field_survey_observation ADD COLUMN IF NOT EXISTS sample_id VARCHAR(36)",
        "ALTER TABLE field_survey_photo ADD COLUMN IF NOT EXISTS sample_id VARCHAR(36)",
        "ALTER TABLE genotype_sample_mapping ADD COLUMN IF NOT EXISTS sample_id VARCHAR(36)",
        "ALTER TABLE breeding_selection_record ADD COLUMN IF NOT EXISTS sample_id VARCHAR(36)",
        "ALTER TABLE breeding_selection_record ADD COLUMN IF NOT EXISTS recorded_by VARCHAR(160)",
        "ALTER TABLE breeding_selection_record ADD COLUMN IF NOT EXISTS created_at TIMESTAMPTZ NOT NULL DEFAULT now()",
        "ALTER TABLE biological_sample DROP CONSTRAINT IF EXISTS ck_biological_sample_status",
        "ALTER TABLE biological_sample ADD CONSTRAINT ck_biological_sample_status CHECK "
        "(selection_status IN ('candidate','retained','observed','eliminated','promoted'))",
        _constraint(
            "ALTER TABLE biological_sample ADD CONSTRAINT fk_biological_sample_program_material "
            "FOREIGN KEY (program_id, material_id) REFERENCES breeding_program_material(program_id, material_id)"
        ),
        _constraint(
            "ALTER TABLE field_survey_plot ADD CONSTRAINT fk_field_survey_plot_trial_entry "
            "FOREIGN KEY (trial_entry_id) REFERENCES trial_entry(id)"
        ),
        _constraint(
            "ALTER TABLE field_survey_observation ADD CONSTRAINT fk_field_survey_observation_sample "
            "FOREIGN KEY (sample_id) REFERENCES biological_sample(id)"
        ),
        _constraint(
            "ALTER TABLE field_survey_photo ADD CONSTRAINT fk_field_survey_photo_sample "
            "FOREIGN KEY (sample_id) REFERENCES biological_sample(id)"
        ),
        _constraint(
            "ALTER TABLE genotype_sample_mapping ADD CONSTRAINT fk_genotype_sample_mapping_sample "
            "FOREIGN KEY (sample_id) REFERENCES biological_sample(id)"
        ),
        _constraint(
            "ALTER TABLE breeding_selection_record ADD CONSTRAINT fk_breeding_selection_sample "
            "FOREIGN KEY (sample_id) REFERENCES biological_sample(id)"
        ),
        _constraint(
            "ALTER TABLE biological_sample ADD CONSTRAINT fk_biological_sample_project "
            "FOREIGN KEY (project_id) REFERENCES research_project(id)"
        ),
        "ALTER TABLE biological_sample DROP CONSTRAINT IF EXISTS biological_sample_program_id_sample_code_key",
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_biological_sample_project_program_code "
        "ON biological_sample(project_id, program_id, sample_code) WHERE project_id IS NOT NULL",
        "CREATE INDEX IF NOT EXISTS ix_biological_sample_project_material ON biological_sample(project_id, material_id, data_status)",
        "CREATE INDEX IF NOT EXISTS ix_biological_sample_trial_entry ON biological_sample(trial_entry_id)",
        "CREATE INDEX IF NOT EXISTS ix_biological_sample_parent ON biological_sample(parent_sample_id)",
        "CREATE INDEX IF NOT EXISTS ix_field_survey_plot_trial_entry ON field_survey_plot(trial_entry_id)",
        "CREATE INDEX IF NOT EXISTS ix_field_survey_observation_sample ON field_survey_observation(sample_id, observed_at)",
        "CREATE INDEX IF NOT EXISTS ix_field_survey_photo_sample ON field_survey_photo(sample_id, captured_at)",
        "CREATE INDEX IF NOT EXISTS ix_selection_sample ON breeding_selection_record(sample_id, selection_year)",
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_genotype_mapping_version_sample "
        "ON genotype_sample_mapping(version_id, sample_id) WHERE sample_id IS NOT NULL",
    )
    for statement in statements:
        session.execute(text(statement))
    session.commit()


HEADER_ALIASES: dict[str, set[str]] = {
    "program_code": {"programcode", "breedingprogramcode", "育种项目编号", "育种项目编码", "项目编号", "项目编码"},
    "material_code": {"materialcode", "材料编号", "材料编码", "品种编号", "品种编码"},
    "sample_code": {"samplecode", "plantcode", "单株编号", "单株编码", "样本编号", "样本编码"},
    "sample_type": {"sampletype", "样本类型", "单株类型"},
    "trial_code": {"trialcode", "试验编号", "试验编码"},
    "treatment_code": {"treatmentcode", "处理编号", "处理编码"},
    "plot_no": {"plotno", "plotcode", "小区号", "小区编号", "小区编码"},
    "generation_label": {"generationlabel", "generation", "世代", "世代名称"},
    "plant_no": {"plantno", "株号", "田间株号"},
    "parent_sample_code": {"parentsamplecode", "亲本单株编号", "来源单株编号", "上代单株编号"},
    "selection_status": {"selectionstatus", "选择状态", "选育状态"},
    "data_status": {"datastatus", "数据状态", "发布状态"},
    "source_note": {"sourcenote", "来源说明", "数据来源", "备注"},
    "metadata": {"metadata", "扩展信息", "元数据"},
}

PHENOTYPE_HEADER_ALIASES: dict[str, set[str]] = {
    "program_code": {"programcode", "育种项目编号", "育种项目编码", "项目编号", "项目编码"},
    "sample_code": {"samplecode", "plantcode", "单株编号", "单株编码", "样本编号", "样本编码"},
    "material_code": {"materialcode", "材料编号", "材料编码"},
    "trial_code": {"trialcode", "试验编号", "试验编码"},
    "treatment_code": {"treatmentcode", "处理编号", "处理编码"},
    "plot_no": {"plotno", "plotcode", "小区号", "小区编号", "小区编码"},
    "observation_stage": {"observationstage", "stage", "观测时期", "调查时期", "调查阶段"},
    "trait_code": {"traitcode", "指标编号", "指标编码", "性状编号", "性状编码"},
    "trait_name": {"traitname", "指标名称", "性状名称"},
    "value_numeric": {"valuenumeric", "value", "数值", "观测值"},
    "value_text": {"valuetext", "文本值", "文本观测值"},
    "unit": {"unit", "单位"},
    "source_type": {"sourcetype", "来源类型", "采集方式"},
    "quality_status": {"qualitystatus", "质控状态", "质量状态"},
    "quality_flags": {"qualityflags", "质控标记", "质量标记"},
    "observed_at": {"observedat", "观测时间", "调查时间"},
    "device_id": {"deviceid", "设备编号", "设备id"},
    "source_note": {"sourcenote", "来源说明", "数据来源", "备注"},
}


def _header_key(value: Any) -> str:
    return re.sub(r"[\s_\-（）()【】\[\]/]+", "", _clean(value)).lower()


def parse_single_plant_workbook(content: bytes) -> list[dict[str, Any]]:
    """Read a single-plant master sheet without persisting the uploaded file."""
    if not content:
        raise SinglePlantError("上传文件为空。")
    if len(content) > MAX_IMPORT_BYTES:
        raise SinglePlantError("导入文件不能超过 10 MB。")
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise SinglePlantError("无法读取 Excel，请确认文件为有效的 .xlsx 工作簿。") from exc
    worksheet = next(
        (sheet for sheet in workbook.worksheets if _header_key(sheet.title) in {"单株主表", "单株信息", "singleplant", "samples"}),
        workbook.worksheets[0] if workbook.worksheets else None,
    )
    if worksheet is None:
        raise SinglePlantError("Excel 中没有可读取的工作表。")

    header_row = None
    header_map: dict[int, str] = {}
    alias_lookup = {alias: field for field, aliases in HEADER_ALIASES.items() for alias in aliases}
    # Some standards-compliant XLSX writers omit the optional worksheet
    # dimension metadata. In read-only mode openpyxl then reports max_row=None,
    # even though iter_rows() can still stream every cell normally.
    header_rows = islice(worksheet.iter_rows(min_row=1, values_only=True), 15)
    for row_number, values in enumerate(header_rows, 1):
        candidate: dict[int, str] = {}
        for column_number, value in enumerate(values, 1):
            field = alias_lookup.get(_header_key(value))
            if field:
                candidate[column_number] = field
        if {"program_code", "material_code", "sample_code"}.issubset(candidate.values()):
            header_row, header_map = row_number, candidate
            break
    if header_row is None:
        raise SinglePlantError("未找到表头；至少需要：育种项目编号、材料编号、单株编号。")

    records: list[dict[str, Any]] = []
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1
    ):
        record = {
            field: values[column_number - 1] if column_number <= len(values) else None
            for column_number, field in header_map.items()
        }
        if not any(_clean(value) for value in record.values()):
            continue
        record["_source_row"] = row_number
        records.append(record)
        if len(records) > MAX_IMPORT_ROWS:
            raise SinglePlantError(f"单次最多导入 {MAX_IMPORT_ROWS} 条单株记录。")
    if not records:
        raise SinglePlantError("工作表中没有单株数据。")
    return records


def parse_single_plant_phenotype_workbook(content: bytes) -> list[dict[str, Any]]:
    """Read the optional phenotype sheet from a complete single-plant package."""
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise SinglePlantError("无法读取表型观测工作表。") from exc
    worksheet = next(
        (sheet for sheet in workbook.worksheets if _header_key(sheet.title) in {"表型观测", "单株表型", "phenotype", "observations"}),
        None,
    )
    if worksheet is None:
        return []

    header_row = None
    header_map: dict[int, str] = {}
    alias_lookup = {
        alias: field for field, aliases in PHENOTYPE_HEADER_ALIASES.items() for alias in aliases
    }
    for row_number, values in enumerate(islice(worksheet.iter_rows(min_row=1, values_only=True), 15), 1):
        candidate: dict[int, str] = {}
        for column_number, value in enumerate(values, 1):
            field = alias_lookup.get(_header_key(value))
            if field:
                candidate[column_number] = field
        required = {"program_code", "sample_code", "observation_stage", "trait_code", "trait_name"}
        if required.issubset(candidate.values()):
            header_row, header_map = row_number, candidate
            break
    if header_row is None:
        raise SinglePlantError(
            "表型观测工作表未找到有效表头；至少需要项目、单株、观测时期、指标编号和指标名称。"
        )

    records: list[dict[str, Any]] = []
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1
    ):
        record = {
            field: values[column_number - 1] if column_number <= len(values) else None
            for column_number, field in header_map.items()
        }
        if not any(_clean(value) for value in record.values()):
            continue
        record["_source_row"] = row_number
        records.append(record)
        if len(records) > MAX_PHENOTYPE_IMPORT_ROWS:
            raise SinglePlantError(
                f"单次最多导入 {MAX_PHENOTYPE_IMPORT_ROWS} 条单株表型观测。"
            )
    return records


def parse_single_plant_package(content: bytes) -> dict[str, list[dict[str, Any]]]:
    """Parse a master-only workbook or a complete master + phenotype package."""
    return {
        "single_plants": parse_single_plant_workbook(content),
        "phenotypes": parse_single_plant_phenotype_workbook(content),
    }


def _reference_maps(session: Session, records: list[dict[str, Any]], project_id: str) -> dict[str, Any]:
    program_codes = sorted({_clean(item.get("program_code")) for item in records if _clean(item.get("program_code"))})
    material_codes = sorted({_clean(item.get("material_code")) for item in records if _clean(item.get("material_code"))})
    trial_codes = sorted({_clean(item.get("trial_code")) for item in records if _clean(item.get("trial_code"))})
    programs = {
        row.program_code: _as_dict(row)
        for row in session.execute(
            text(
                "SELECT id, program_code, program_name FROM breeding_program "
                "WHERE project_id=:project_id AND program_code = ANY(:codes)"
            ),
            {"project_id": project_id, "codes": program_codes or [""]},
        )
    }
    materials = {
        row.material_code: _as_dict(row)
        for row in session.execute(
            text(
                "SELECT material.id, material.material_code, material.material_name "
                "FROM breeding_material material "
                "JOIN data_material_project_scope scope ON scope.material_id=material.id "
                "WHERE scope.project_id=:project_id AND material.material_code = ANY(:codes)"
            ),
            {"project_id": project_id, "codes": material_codes or [""]},
        )
    }
    entries: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    if trial_codes:
        for row in session.execute(
            text(
                """
                SELECT te.id, te.material_id, te.plot_no, ft.trial_code, tt.treatment_code
                FROM trial_entry te
                JOIN field_trial ft ON ft.id = te.trial_id
                JOIN trial_treatment tt ON tt.id = te.treatment_id
                WHERE ft.project_id=:project_id AND ft.trial_code = ANY(:codes)
                """
            ),
            {"project_id": project_id, "codes": trial_codes},
        ):
            item = _as_dict(row)
            entries.setdefault((item["trial_code"], item["plot_no"], item["material_id"]), []).append(item)
    program_ids = [item["id"] for item in programs.values()]
    memberships: set[tuple[str, str]] = set()
    if program_ids:
        memberships = {
            (row.program_id, row.material_id)
            for row in session.execute(
                text(
                    "SELECT program_id, material_id FROM breeding_program_material "
                    "WHERE program_id = ANY(:program_ids)"
                ),
                {"program_ids": program_ids},
            )
        }
    existing: dict[tuple[str, str], dict[str, Any]] = {}
    if program_ids:
        for row in session.execute(
            text(
                """
                SELECT id, program_id, material_id, sample_code, parent_sample_id
                FROM biological_sample
                WHERE project_id=:project_id AND program_id = ANY(:program_ids)
                """
            ),
            {"project_id": project_id, "program_ids": program_ids},
        ):
            item = _as_dict(row)
            existing[(item["program_id"], item["sample_code"])] = item
    return {
        "programs": programs,
        "materials": materials,
        "memberships": memberships,
        "entries": entries,
        "existing": existing,
    }


def _integer(value: Any, field: str, issues: list[dict[str, str]]) -> int | None:
    if value in (None, ""):
        return None
    try:
        number = int(value)
        if float(value) != number or number <= 0:
            raise ValueError
        return number
    except (TypeError, ValueError):
        issues.append({"field": field, "code": "invalid_integer", "message": "株号必须是正整数。"})
        return None


def _metadata(value: Any, issues: list[dict[str, str]]) -> dict[str, Any]:
    if value in (None, ""):
        return {}
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(str(value))
        if not isinstance(parsed, dict):
            raise ValueError
        return parsed
    except (TypeError, ValueError, json.JSONDecodeError):
        issues.append({"field": "metadata", "code": "invalid_json", "message": "元数据必须是 JSON 对象。"})
        return {}


def preview_single_plant_import(
    session: Session, records: list[dict[str, Any]], project_id: str, mode: str = "create_only"
) -> dict[str, Any]:
    if mode not in {"create_only", "upsert"}:
        raise SinglePlantError("导入模式只能是 create_only 或 upsert。")
    if not records or len(records) > MAX_IMPORT_ROWS:
        raise SinglePlantError(f"每次需导入 1 至 {MAX_IMPORT_ROWS} 条记录。")
    refs = _reference_maps(session, records, project_id)
    seen: set[tuple[str, str]] = set()
    normalized_rows: list[dict[str, Any]] = []

    for position, source in enumerate(records, 1):
        issues: list[dict[str, str]] = []
        source_row = source.get("_source_row") or position
        program_code = _clean(source.get("program_code"))
        material_code = _clean(source.get("material_code"))
        sample_code = _clean(source.get("sample_code"))
        sample_type = _clean(source.get("sample_type")) or "individual_plant"
        selection_status = _clean(source.get("selection_status")) or "candidate"
        data_status = _clean(source.get("data_status")) or "published"
        program = refs["programs"].get(program_code)
        material = refs["materials"].get(material_code)
        if not program_code:
            issues.append({"field": "program_code", "code": "required", "message": "育种项目编号不能为空。"})
        elif not program:
            issues.append({"field": "program_code", "code": "not_found", "message": f"育种项目 {program_code} 不存在。"})
        if not material_code:
            issues.append({"field": "material_code", "code": "required", "message": "材料编号不能为空。"})
        elif not material:
            issues.append({"field": "material_code", "code": "not_found", "message": f"材料 {material_code} 不存在。"})
        if not sample_code:
            issues.append({"field": "sample_code", "code": "required", "message": "单株编号不能为空。"})
        elif len(sample_code) > 160:
            issues.append({"field": "sample_code", "code": "too_long", "message": "单株编号不能超过 160 个字符。"})
        if sample_type not in SAMPLE_TYPES:
            issues.append({"field": "sample_type", "code": "invalid", "message": f"不支持的样本类型：{sample_type}。"})
        if selection_status not in SELECTION_STATUSES:
            issues.append({"field": "selection_status", "code": "invalid", "message": f"不支持的选择状态：{selection_status}。"})
        if data_status not in DATA_STATUSES:
            issues.append({"field": "data_status", "code": "invalid", "message": f"不支持的数据状态：{data_status}。"})

        program_id = program["id"] if program else ""
        material_id = material["id"] if material else ""
        if program and material and (program_id, material_id) not in refs["memberships"]:
            issues.append({
                "field": "material_code",
                "code": "material_not_in_program",
                "message": f"材料 {material_code} 尚未加入育种项目 {program_code}。",
            })
        identity = (program_id or program_code, sample_code)
        if sample_code and identity in seen:
            issues.append({"field": "sample_code", "code": "duplicate_in_file", "message": "同一项目内的单株编号在文件中重复。"})
        seen.add(identity)
        existing = refs["existing"].get((program_id, sample_code)) if program_id and sample_code else None
        if existing and mode == "create_only":
            issues.append({"field": "sample_code", "code": "already_exists", "message": "该项目内的单株编号已经存在。"})

        trial_code = _clean(source.get("trial_code"))
        plot_no = _clean(source.get("plot_no"))
        treatment_code = _clean(source.get("treatment_code"))
        trial_entry_id = None
        if trial_code or plot_no:
            if not trial_code or not plot_no:
                issues.append({"field": "trial_code", "code": "incomplete_trial_reference", "message": "关联试验时必须同时填写试验编号和小区号。"})
            elif material:
                candidates = refs["entries"].get((trial_code, plot_no, material_id), [])
                if treatment_code:
                    candidates = [item for item in candidates if item["treatment_code"] == treatment_code]
                if not candidates:
                    issues.append({"field": "trial_code", "code": "trial_entry_not_found", "message": "未找到与试验、小区和材料一致的试验条目。"})
                elif len(candidates) > 1:
                    issues.append({"field": "treatment_code", "code": "ambiguous_trial_entry", "message": "该小区匹配到多个处理，请补充处理编号。"})
                else:
                    trial_entry_id = candidates[0]["id"]

        parent_code = _clean(source.get("parent_sample_code"))
        if parent_code and parent_code == sample_code:
            issues.append({"field": "parent_sample_code", "code": "self_parent", "message": "单株不能把自己设为来源单株。"})
        if parent_code and program:
            parent_exists = (program_id, parent_code) in refs["existing"]
            parent_in_file = (program_id, parent_code) in seen or any(
                _clean(item.get("program_code")) == program_code and _clean(item.get("sample_code")) == parent_code
                for item in records
            )
            if not parent_exists and not parent_in_file:
                issues.append({"field": "parent_sample_code", "code": "parent_not_found", "message": f"来源单株 {parent_code} 不存在。"})

        record = {
            "id": existing["id"] if existing else None,
            "program_id": program_id or None,
            "program_code": program_code,
            "program_name": program.get("program_name") if program else None,
            "material_id": material_id or None,
            "material_code": material_code,
            "material_name": material.get("material_name") if material else None,
            "trial_entry_id": trial_entry_id,
            "trial_code": trial_code or None,
            "treatment_code": treatment_code or None,
            "plot_no": plot_no or None,
            "sample_code": sample_code,
            "sample_type": sample_type,
            "generation_label": _optional_text(source.get("generation_label")),
            "plant_no": _integer(source.get("plant_no"), "plant_no", issues),
            "parent_sample_code": parent_code or None,
            "selection_status": selection_status,
            "data_status": data_status,
            "source_note": _optional_text(source.get("source_note")),
            "metadata": _metadata(source.get("metadata"), issues),
        }
        normalized_rows.append({
            "row_number": source_row,
            "valid": not issues,
            "action": "update" if existing else "create",
            "issues": issues,
            "record": record,
        })

    invalid_count = sum(not item["valid"] for item in normalized_rows)
    return {
        "mode": mode,
        "row_count": len(normalized_rows),
        "valid_count": len(normalized_rows) - invalid_count,
        "invalid_count": invalid_count,
        "can_publish": invalid_count == 0,
        "rows": normalized_rows,
    }


def _phenotype_number(value: Any, issues: list[dict[str, str]]) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        issues.append({"field": "value_numeric", "code": "invalid_number", "message": "数值必须是有效数字。"})
        return None


def _phenotype_flags(value: Any, issues: list[dict[str, str]]) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(item) for item in value]
    try:
        parsed = json.loads(str(value))
        if not isinstance(parsed, list):
            raise ValueError
        return [str(item) for item in parsed]
    except (TypeError, ValueError, json.JSONDecodeError):
        issues.append({"field": "quality_flags", "code": "invalid_json", "message": "质控标记必须是 JSON 数组。"})
        return []


def _phenotype_time(value: Any, issues: list[dict[str, str]]) -> datetime:
    if value in (None, ""):
        return datetime.now(timezone.utc)
    if isinstance(value, datetime):
        parsed = value
    else:
        try:
            parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        except ValueError:
            issues.append({"field": "observed_at", "code": "invalid_datetime", "message": "观测时间格式无效。"})
            return datetime.now(timezone.utc)
    return parsed.replace(tzinfo=timezone.utc) if parsed.tzinfo is None else parsed


def _normalize_single_plant_phenotypes(
    phenotype_records: list[dict[str, Any]], master_preview: dict[str, Any], mode: str
) -> list[dict[str, Any]]:
    if phenotype_records and mode != "create_only":
        raise SinglePlantError("包含表型观测的完整数据包目前仅支持 create_only 模式。")
    master_by_identity = {
        (item["record"]["program_code"], item["record"]["sample_code"]): item
        for item in master_preview["rows"]
    }
    seen: set[tuple[str, str, str, str]] = set()
    normalized_rows: list[dict[str, Any]] = []
    for position, source in enumerate(phenotype_records, 1):
        issues: list[dict[str, str]] = []
        source_row = source.get("_source_row") or position
        program_code = _clean(source.get("program_code"))
        sample_code = _clean(source.get("sample_code"))
        stage = _clean(source.get("observation_stage"))
        trait_code = _clean(source.get("trait_code"))
        trait_name = _clean(source.get("trait_name"))
        identity = (program_code, sample_code)
        master_item = master_by_identity.get(identity)
        master = master_item["record"] if master_item else None
        if not program_code:
            issues.append({"field": "program_code", "code": "required", "message": "育种项目编号不能为空。"})
        if not sample_code:
            issues.append({"field": "sample_code", "code": "required", "message": "单株编号不能为空。"})
        if not master:
            issues.append({"field": "sample_code", "code": "sample_not_in_package", "message": "表型观测对应的单株不在本次单株主表中。"})
        elif not master_item["valid"]:
            issues.append({"field": "sample_code", "code": "invalid_master_row", "message": "表型观测对应的单株主记录未通过预检。"})
        if not stage:
            issues.append({"field": "observation_stage", "code": "required", "message": "观测时期不能为空。"})
        if not trait_code:
            issues.append({"field": "trait_code", "code": "required", "message": "指标编号不能为空。"})
        if not trait_name:
            issues.append({"field": "trait_name", "code": "required", "message": "指标名称不能为空。"})
        observation_identity = (program_code, sample_code, stage, trait_code)
        if all(observation_identity) and observation_identity in seen:
            issues.append({"field": "trait_code", "code": "duplicate_in_file", "message": "同一单株、时期和指标在文件中重复。"})
        seen.add(observation_identity)

        value_numeric = _phenotype_number(source.get("value_numeric"), issues)
        value_text = _optional_text(source.get("value_text"))
        if value_numeric is None and not value_text:
            issues.append({"field": "value_numeric", "code": "required", "message": "数值和文本值至少填写一项。"})
        source_type = _clean(source.get("source_type")) or "import"
        quality_status = _clean(source.get("quality_status")) or "passed"
        if source_type not in OBSERVATION_SOURCE_TYPES:
            issues.append({"field": "source_type", "code": "invalid", "message": f"不支持的来源类型：{source_type}。"})
        if quality_status not in OBSERVATION_QUALITY_STATUSES:
            issues.append({"field": "quality_status", "code": "invalid", "message": f"不支持的质控状态：{quality_status}。"})

        if master:
            for field in ("material_code", "trial_code", "treatment_code", "plot_no"):
                supplied = _clean(source.get(field))
                expected = _clean(master.get(field))
                if supplied and supplied != expected:
                    issues.append({"field": field, "code": "master_mismatch", "message": f"{field} 与单株主表不一致。"})

        record = {
            "program_code": program_code,
            "sample_code": sample_code,
            "observation_stage": stage,
            "trait_code": trait_code,
            "trait_name": trait_name,
            "value_numeric": value_numeric,
            "value_text": value_text,
            "unit": _clean(source.get("unit")),
            "source_type": source_type,
            "quality_status": quality_status,
            "quality_flags": _phenotype_flags(source.get("quality_flags"), issues),
            "observed_at": _phenotype_time(source.get("observed_at"), issues),
            "device_id": _optional_text(source.get("device_id")),
            "source_note": _optional_text(source.get("source_note")),
        }
        normalized_rows.append({
            "row_number": source_row,
            "valid": not issues,
            "issues": issues,
            "record": record,
        })
    return normalized_rows


def preview_single_plant_package(
    session: Session,
    master_records: list[dict[str, Any]],
    phenotype_records: list[dict[str, Any]],
    project_id: str,
    mode: str = "create_only",
) -> dict[str, Any]:
    master_preview = preview_single_plant_import(session, master_records, project_id, mode)
    phenotype_rows = _normalize_single_plant_phenotypes(phenotype_records, master_preview, mode)
    invalid_phenotypes = [item for item in phenotype_rows if not item["valid"]]
    result = dict(master_preview)
    result.update({
        "package_kind": "complete" if phenotype_records else "master_only",
        "phenotype_row_count": len(phenotype_rows),
        "phenotype_valid_count": len(phenotype_rows) - len(invalid_phenotypes),
        "phenotype_invalid_count": len(invalid_phenotypes),
        "phenotype_rows": invalid_phenotypes[:100],
        "can_publish": master_preview["can_publish"] and not invalid_phenotypes,
    })
    return result


def publish_single_plant_import(
    session: Session,
    records: list[dict[str, Any]],
    actor: str,
    project_id: str,
    mode: str = "create_only",
    *,
    commit: bool = True,
) -> dict[str, Any]:
    preview = preview_single_plant_import(session, records, project_id, mode)
    if not preview["can_publish"]:
        raise SinglePlantError(f"导入预检发现 {preview['invalid_count']} 条问题，未写入数据库。")
    identities: dict[tuple[str, str], str] = {}
    created_ids: list[str] = []
    updated_ids: list[str] = []
    try:
        for item in preview["rows"]:
            record = item["record"]
            sample_id = record["id"] or str(uuid.uuid4())
            identities[(record["program_id"], record["sample_code"])] = sample_id
            parameters = {
                **record,
                "id": sample_id,
                "project_id": project_id,
                "metadata_json": _json(record["metadata"]),
                "actor": actor,
            }
            if item["action"] == "create":
                session.execute(
                    text(
                        """
                        INSERT INTO biological_sample (
                          id, project_id, program_id, material_id, trial_entry_id, sample_code, sample_type,
                          generation_label, plant_no, selection_status, data_status, source_note,
                          metadata, created_by
                        ) VALUES (
                          :id, :project_id, :program_id, :material_id, :trial_entry_id, :sample_code, :sample_type,
                          :generation_label, :plant_no, :selection_status, :data_status, :source_note,
                          CAST(:metadata_json AS JSONB), :actor
                        )
                        """
                    ),
                    parameters,
                )
                created_ids.append(sample_id)
            else:
                session.execute(
                    text(
                        """
                        UPDATE biological_sample SET
                          material_id=:material_id, trial_entry_id=:trial_entry_id,
                          sample_type=:sample_type, generation_label=:generation_label,
                          plant_no=:plant_no, selection_status=:selection_status,
                          data_status=:data_status, source_note=:source_note,
                          metadata=CAST(:metadata_json AS JSONB), updated_at=now()
                        WHERE id=:id AND project_id=:project_id
                        """
                    ),
                    parameters,
                )
                updated_ids.append(sample_id)

        for item in preview["rows"]:
            record = item["record"]
            parent_code = record.get("parent_sample_code")
            if not parent_code:
                continue
            parent_id = identities.get((record["program_id"], parent_code))
            if not parent_id:
                parent_id = session.scalar(
                    text("SELECT id FROM biological_sample WHERE project_id=:project_id AND program_id=:program_id AND sample_code=:sample_code"),
                    {"project_id": project_id, "program_id": record["program_id"], "sample_code": parent_code},
                )
            if not parent_id:
                raise SinglePlantError(f"来源单株 {parent_code} 在发布时未找到。")
            sample_id = identities[(record["program_id"], record["sample_code"])]
            session.execute(
                text("UPDATE biological_sample SET parent_sample_id=:parent_id, updated_at=now() WHERE id=:id AND project_id=:project_id"),
                {"parent_id": parent_id, "id": sample_id, "project_id": project_id},
            )
        if commit:
            session.commit()
        else:
            session.flush()
    except IntegrityError as exc:
        session.rollback()
        raise SinglePlantError("导入数据违反唯一性或关联约束，未写入数据库。") from exc
    except Exception:
        session.rollback()
        raise
    return {
        "published": True,
        "created_count": len(created_ids),
        "updated_count": len(updated_ids),
        "sample_ids": created_ids + updated_ids,
    }


def _publish_single_plant_phenotypes(
    session: Session,
    phenotype_rows: list[dict[str, Any]],
    actor: str,
    project_id: str,
) -> dict[str, int]:
    records = [item["record"] for item in phenotype_rows]
    if not records:
        return {"phenotype_created_count": 0, "survey_unit_created_count": 0, "survey_task_count": 0}

    program_codes = sorted({item["program_code"] for item in records})
    sample_codes = sorted({item["sample_code"] for item in records})
    contexts = {
        (row.program_code, row.sample_code): _as_dict(row)
        for row in session.execute(
            text(
                """
                SELECT bs.id AS sample_id, bs.sample_code, bp.program_code,
                       bs.material_id, bs.trial_entry_id,
                       bm.material_code, bm.material_name,
                       te.trial_id, te.plot_no, te.block_no, te.replicate_no,
                       ft.trial_code, ft.trial_name, ts.site_name
                FROM biological_sample bs
                JOIN breeding_program bp ON bp.id=bs.program_id
                JOIN breeding_material bm ON bm.id=bs.material_id
                JOIN trial_entry te ON te.id=bs.trial_entry_id
                JOIN field_trial ft ON ft.id=te.trial_id
                JOIN trial_site ts ON ts.id=ft.site_id
                WHERE bp.program_code=ANY(:program_codes)
                  AND bs.sample_code=ANY(:sample_codes)
                  AND bs.project_id=:project_id
                  AND bp.project_id=:project_id
                  AND ft.project_id=:project_id
                """
            ),
            {
                "program_codes": program_codes,
                "sample_codes": sample_codes,
                "project_id": project_id,
            },
        )
    }
    if len(contexts) != len({(item["program_code"], item["sample_code"]) for item in records}):
        raise SinglePlantError("发布表型时未找到全部单株或试验小区关联。")

    task_groups: dict[str, dict[str, Any]] = {}
    for record in records:
        context = contexts[(record["program_code"], record["sample_code"])]
        task_code = _task_code(context["trial_id"], record["observation_stage"])
        group = task_groups.setdefault(task_code, {
            "trial_id": context["trial_id"],
            "trial_name": context["trial_name"],
            "site_name": context["site_name"],
            "stage": record["observation_stage"],
            "traits": set(),
        })
        group["traits"].add(record["trait_code"])

    existing_tasks = {
        row.task_code: _as_dict(row)
        for row in session.execute(
            text("SELECT id, task_code, required_traits FROM field_survey_task WHERE task_code=ANY(:codes)"),
            {"codes": list(task_groups)},
        )
    }
    task_ids: dict[str, str] = {}
    created_tasks = 0
    for task_code, group in task_groups.items():
        existing = existing_tasks.get(task_code)
        traits = sorted(group["traits"])
        if existing:
            task_ids[task_code] = existing["id"]
            existing_traits = existing.get("required_traits") or []
            session.execute(
                text("UPDATE field_survey_task SET required_traits=CAST(:traits AS JSONB), updated_at=now() WHERE id=:id"),
                {"traits": _json(sorted(set(existing_traits) | set(traits))), "id": existing["id"]},
            )
            continue
        task_id = str(uuid.uuid4())
        task_ids[task_code] = task_id
        created_tasks += 1
        session.execute(
            text(
                """
                INSERT INTO field_survey_task (
                  id, task_code, task_name, trial_name, site_name, survey_stage,
                  required_traits, assigned_workers, photo_required, created_by, description
                ) VALUES (
                  :id, :code, :name, :trial_name, :site_name, :stage,
                  CAST(:traits AS JSONB), CAST(:workers AS JSONB), FALSE, :actor, :description
                )
                """
            ),
            {
                "id": task_id,
                "code": task_code,
                "name": f"{group['trial_name']} - {group['stage']}单株调查",
                "trial_name": group["trial_name"],
                "site_name": group["site_name"],
                "stage": group["stage"],
                "traits": _json(traits),
                "workers": _json([actor]),
                "actor": actor,
                "description": "完整产品演示包批量导入的单株表型观测。",
            },
        )

    unit_groups: dict[tuple[str, str], dict[str, Any]] = {}
    for record in records:
        context = contexts[(record["program_code"], record["sample_code"])]
        task_code = _task_code(context["trial_id"], record["observation_stage"])
        task_id = task_ids[task_code]
        plot_code = _survey_unit_code(context["sample_code"], context["sample_id"])
        group = unit_groups.setdefault((task_id, plot_code), {
            "task_id": task_id,
            "plot_code": plot_code,
            "context": context,
            "records": [],
        })
        group["records"].append(record)

    task_id_values = sorted(set(task_ids.values()))
    plot_code_values = sorted({key[1] for key in unit_groups})
    existing_units = {
        (row.task_id, row.plot_code): row.id
        for row in session.execute(
            text(
                "SELECT id, task_id, plot_code FROM field_survey_plot "
                "WHERE task_id=ANY(:task_ids) AND plot_code=ANY(:plot_codes)"
            ),
            {"task_ids": task_id_values, "plot_codes": plot_code_values},
        )
    }
    max_sequences = {task_id: 0 for task_id in task_id_values}
    for row in session.execute(
        text(
            "SELECT task_id, COALESCE(MAX(sequence_no),0) AS max_sequence "
            "FROM field_survey_plot WHERE task_id=ANY(:task_ids) GROUP BY task_id"
        ),
        {"task_ids": task_id_values},
    ):
        max_sequences[row.task_id] = row.max_sequence

    unit_ids: dict[tuple[str, str], str] = dict(existing_units)
    unit_inserts: list[dict[str, Any]] = []
    for key, group in unit_groups.items():
        if key in unit_ids:
            continue
        task_id, plot_code = key
        context = group["context"]
        max_sequences[task_id] += 1
        unit_id = str(uuid.uuid4())
        unit_ids[key] = unit_id
        unit_inserts.append({
            "id": unit_id,
            "task_id": task_id,
            "plot_code": plot_code,
            "sequence_no": max_sequences[task_id],
            "block_no": context["block_no"],
            "replicate_no": context["replicate_no"],
            "material_code": context["material_code"],
            "material_name": context["material_name"],
            "qr_token": str(uuid.uuid4()),
            "actor": actor,
            "trial_entry_id": context["trial_entry_id"],
        })
    if unit_inserts:
        session.execute(
            text(
                """
                INSERT INTO field_survey_plot (
                  id, task_id, plot_code, sequence_no, block_no, replicate_no,
                  material_code, material_name, qr_token, assigned_to,
                  status, quality_status, trial_entry_id
                ) VALUES (
                  :id, :task_id, :plot_code, :sequence_no, :block_no, :replicate_no,
                  :material_code, :material_name, :qr_token, :actor,
                  'in_progress', 'pending', :trial_entry_id
                )
                """
            ),
            unit_inserts,
        )

    observation_inserts: list[dict[str, Any]] = []
    audit_inserts: list[dict[str, Any]] = []
    unit_updates: list[dict[str, Any]] = []
    for key, group in unit_groups.items():
        task_id, _plot_code = key
        plot_id = unit_ids[key]
        context = group["context"]
        quality_status = "passed"
        if any(item["quality_status"] == "rejected" for item in group["records"]):
            quality_status = "rejected"
        elif any(item["quality_status"] == "warning" for item in group["records"]):
            quality_status = "warning"
        completed_at = max(item["observed_at"] for item in group["records"])
        device_id = next((item["device_id"] for item in group["records"] if item["device_id"]), None)
        unit_updates.append({
            "quality_status": quality_status,
            "completed_at": completed_at,
            "actor": actor,
            "device_id": device_id,
            "plot_id": plot_id,
        })
        for record in group["records"]:
            observation_id = str(uuid.uuid4())
            observation_inserts.append({
                "id": observation_id,
                "task_id": task_id,
                "plot_id": plot_id,
                "sample_id": context["sample_id"],
                "trait_code": record["trait_code"],
                "trait_name": record["trait_name"],
                "value_numeric": record["value_numeric"],
                "value_text": record["value_text"],
                "unit": record["unit"],
                "source_type": record["source_type"],
                "device_id": record["device_id"],
                "quality_status": record["quality_status"],
                "quality_flags": _json(record["quality_flags"]),
                "actor": actor,
                "observed_at": record["observed_at"],
            })
            audit_inserts.append({
                "id": str(uuid.uuid4()),
                "task_id": task_id,
                "plot_id": plot_id,
                "label": f"批量导入单株表型：{record['trait_name']}",
                "actor": actor,
                "device_id": record["device_id"],
                "details": _json({
                    "sample_id": context["sample_id"],
                    "sample_code": context["sample_code"],
                    "trait_code": record["trait_code"],
                    "source_note": record["source_note"],
                }),
            })

    session.execute(
        text(
            """
            INSERT INTO field_survey_observation (
              id, task_id, plot_id, sample_id, trait_code, trait_name,
              value_numeric, value_text, unit, source_type, device_id,
              quality_status, quality_flags, observed_by, observed_at, data_version
            ) VALUES (
              :id, :task_id, :plot_id, :sample_id, :trait_code, :trait_name,
              :value_numeric, :value_text, :unit, :source_type, :device_id,
              :quality_status, CAST(:quality_flags AS JSONB), :actor, :observed_at, 1
            )
            """
        ),
        observation_inserts,
    )
    session.execute(
        text(
            """
            INSERT INTO field_survey_audit (
              id, task_id, plot_id, event_type, event_label, actor,
              device_id, data_version, details
            ) VALUES (
              :id, :task_id, :plot_id, 'single_plant_observation_import', :label, :actor,
              :device_id, 1, CAST(:details AS JSONB)
            )
            """
        ),
        audit_inserts,
    )
    session.execute(
        text(
            """
            UPDATE field_survey_plot
            SET status='completed', quality_status=:quality_status,
                completed_at=:completed_at, completed_by=:actor, device_id=:device_id
            WHERE id=:plot_id
            """
        ),
        unit_updates,
    )
    return {
        "phenotype_created_count": len(observation_inserts),
        "survey_unit_created_count": len(unit_inserts),
        "survey_task_count": len(task_groups),
    }


def publish_single_plant_package(
    session: Session,
    master_records: list[dict[str, Any]],
    phenotype_records: list[dict[str, Any]],
    actor: str,
    project_id: str,
    mode: str = "create_only",
) -> dict[str, Any]:
    preview = preview_single_plant_package(session, master_records, phenotype_records, project_id, mode)
    if not preview["can_publish"]:
        total_invalid = preview["invalid_count"] + preview["phenotype_invalid_count"]
        raise SinglePlantError(f"完整数据包预检发现 {total_invalid} 条问题，未写入数据库。")
    phenotype_rows = _normalize_single_plant_phenotypes(phenotype_records, preview, mode)
    try:
        master_result = publish_single_plant_import(
            session, master_records, actor, project_id, mode, commit=False
        )
        phenotype_result = _publish_single_plant_phenotypes(
            session, phenotype_rows, actor, project_id
        )
        session.commit()
    except IntegrityError as exc:
        session.rollback()
        raise SinglePlantError("完整数据包违反唯一性或关联约束，未写入数据库。") from exc
    except Exception:
        session.rollback()
        raise
    return {
        **master_result,
        **phenotype_result,
        "package_kind": "complete" if phenotype_records else "master_only",
    }


def list_material_single_plants(
    session: Session,
    material_id: str,
    project_id: str,
    program_id: str | None = None,
    selection_status: str | None = None,
    include_archived: bool = False,
) -> list[dict[str, Any]]:
    material_exists = session.scalar(text("SELECT 1 FROM breeding_material WHERE id=:id"), {"id": material_id})
    if not material_exists:
        raise SinglePlantError("材料不存在。")
    conditions = ["bs.material_id=:material_id", "bs.project_id=:project_id"]
    parameters: dict[str, Any] = {"material_id": material_id, "project_id": project_id}
    if program_id:
        conditions.append("bs.program_id=:program_id")
        parameters["program_id"] = program_id
    if selection_status:
        conditions.append("bs.selection_status=:selection_status")
        parameters["selection_status"] = selection_status
    if not include_archived:
        conditions.append("bs.data_status <> 'archived'")
    rows = session.execute(
        text(
            """
            SELECT bs.id, bs.sample_code, bs.sample_type, bs.generation_label, bs.plant_no,
                   bs.selection_status, bs.data_status, bs.trial_entry_id, bs.parent_sample_id,
                   bs.created_at, bs.updated_at, bp.id AS program_id, bp.program_code, bp.program_name,
                   te.plot_no, ft.trial_code, ft.trial_name,
                   (SELECT COUNT(*) FROM field_survey_observation o WHERE o.sample_id=bs.id) AS observation_count,
                   (SELECT COUNT(*) FROM genotype_sample_mapping gm WHERE gm.sample_id=bs.id) AS genotype_count,
                   (SELECT COUNT(*) FROM breeding_selection_record sr WHERE sr.sample_id=bs.id) AS selection_count
            FROM biological_sample bs
            JOIN breeding_program bp ON bp.id = bs.program_id
            LEFT JOIN trial_entry te ON te.id = bs.trial_entry_id
            LEFT JOIN field_trial ft ON ft.id = te.trial_id
            WHERE """
            + " AND ".join(conditions)
            + """
            ORDER BY bp.program_code, bs.generation_label NULLS LAST,
                     bs.plant_no NULLS LAST, bs.sample_code
            """
        ),
        parameters,
    )
    return [_as_dict(row) for row in rows]


def lookup_single_plants(
    session: Session, keyword: str, project_id: str, program_code: str | None = None, limit: int = 20
) -> list[dict[str, Any]]:
    """Find a plant by scanned code or identifier for field-data entry."""
    term = _clean(keyword)
    if not term:
        raise SinglePlantError("请输入单株编号或扫描二维码内容。")
    rows = session.execute(
        text(
            """
            SELECT bs.id, bs.sample_code, bs.sample_type, bs.generation_label, bs.plant_no,
                   bs.selection_status, bs.data_status, bs.trial_entry_id,
                   bp.id AS program_id, bp.program_code, bp.program_name,
                   bm.id AS material_id, bm.material_code, bm.material_name,
                   te.plot_no, te.block_no, te.replicate_no,
                   ft.trial_code, ft.trial_name, ft.trial_year,
                   ts.site_code, ts.site_name
            FROM biological_sample bs
            JOIN breeding_program bp ON bp.id=bs.program_id
            JOIN breeding_material bm ON bm.id=bs.material_id
            LEFT JOIN trial_entry te ON te.id=bs.trial_entry_id
            LEFT JOIN field_trial ft ON ft.id=te.trial_id
            LEFT JOIN trial_site ts ON ts.id=ft.site_id
            WHERE (bs.id=:term OR bs.sample_code ILIKE :pattern)
              AND bs.project_id=:project_id
              AND (CAST(:program_code AS VARCHAR) IS NULL OR bp.program_code=:program_code)
              AND bs.data_status <> 'archived'
            ORDER BY CASE WHEN bs.id=:term OR bs.sample_code=:term THEN 0 ELSE 1 END,
                     bp.program_code, bs.sample_code
            LIMIT :limit
            """
        ),
        {
            "term": term,
            "project_id": project_id,
            "pattern": f"%{term.replace('%', '').replace('_', '')}%",
            "program_code": _optional_text(program_code),
            "limit": max(1, min(limit, 50)),
        },
    )
    return [_as_dict(row) for row in rows]


def get_single_plant_detail(session: Session, sample_id: str, project_id: str) -> dict[str, Any]:
    sample = session.execute(
        text(
            """
            SELECT bs.*, bp.program_code, bp.program_name,
                   bm.material_code, bm.material_name, bm.material_type,
                   parent.sample_code AS parent_sample_code,
                   te.plot_no, te.block_no, te.replicate_no,
                   ft.trial_code, ft.trial_name, ft.trial_year,
                   ts.site_code, ts.site_name
            FROM biological_sample bs
            JOIN breeding_program bp ON bp.id=bs.program_id
            JOIN breeding_material bm ON bm.id=bs.material_id
            LEFT JOIN biological_sample parent ON parent.id=bs.parent_sample_id
            LEFT JOIN trial_entry te ON te.id=bs.trial_entry_id
            LEFT JOIN field_trial ft ON ft.id=te.trial_id
            LEFT JOIN trial_site ts ON ts.id=ft.site_id
            WHERE bs.id=:sample_id AND bs.project_id=:project_id
            """
        ),
        {"sample_id": sample_id, "project_id": project_id},
    ).first()
    if not sample:
        raise SinglePlantError("单株不存在。")
    observations = [
        _as_dict(row)
        for row in session.execute(
            text(
                """
                SELECT o.id, o.trait_code, o.trait_name, o.value_numeric, o.value_text, o.unit,
                       t.survey_stage AS observation_stage, o.source_type, o.quality_status,
                       o.quality_flags, o.observed_by, o.observed_at, o.data_version,
                       p.plot_code AS survey_unit_code, p.task_id
                FROM field_survey_observation o
                JOIN field_survey_task t ON t.id=o.task_id
                JOIN field_survey_plot p ON p.id=o.plot_id
                WHERE o.sample_id=:sample_id
                ORDER BY o.observed_at DESC, o.trait_code, o.data_version DESC
                """
            ),
            {"sample_id": sample_id},
        )
    ]
    genotype_mappings = [
        _as_dict(row)
        for row in session.execute(
            text(
                """
                SELECT gsm.id, gsm.fid, gsm.iid, gsm.status, gsm.note, gsm.updated_at,
                       gav.id AS version_id, gav.version_number, gav.status AS version_status,
                       ga.id AS asset_id, ga.title AS asset_title, ga.reference_assembly
                FROM genotype_sample_mapping gsm
                JOIN genotype_asset_version gav ON gav.id=gsm.version_id
                JOIN genotype_asset ga ON ga.id=gav.asset_id
                WHERE gsm.sample_id=:sample_id
                ORDER BY gsm.updated_at DESC
                """
            ),
            {"sample_id": sample_id},
        )
    ]
    selections = [
        _as_dict(row)
        for row in session.execute(
            text(
                """
                SELECT id, selection_year, selection_site, selection_criterion, selection_decision,
                       retention_reason, evidence_summary, source_record_no, source_note,
                       generation_record_id, recorded_by, created_at
                FROM breeding_selection_record
                WHERE sample_id=:sample_id
                ORDER BY created_at DESC, id DESC
                """
            ),
            {"sample_id": sample_id},
        )
    ]
    return {
        "sample": _as_dict(sample),
        "observations": observations,
        "genotype_mappings": genotype_mappings,
        "selection_history": selections,
        "evidence_counts": {
            "observations": len(observations),
            "genotype_mappings": len(genotype_mappings),
            "selection_records": len(selections),
        },
    }


def _average(values: list[float]) -> float | None:
    return sum(values) / len(values) if values else None


def _trait_score(value: float | None, trait: dict[str, Any], low: float, high: float) -> float | None:
    if value is None:
        return None
    if trait["direction"] == "target":
        return max(0.0, 100.0 - abs(value - trait["target"]) * trait["penalty"])
    if high == low:
        return 100.0
    if trait["direction"] == "lower":
        return (high - value) / (high - low) * 100.0
    return (value - low) / (high - low) * 100.0


def get_variety_evaluation(
    session: Session, project_id: str, program_code: str | None = None
) -> dict[str, Any]:
    rows = session.execute(
        text(
            """
            WITH ranked_observation AS (
              SELECT o.*, t.survey_stage,
                     ROW_NUMBER() OVER (
                       PARTITION BY o.sample_id, t.survey_stage, o.trait_code
                       ORDER BY o.data_version DESC, o.observed_at DESC, o.id DESC
                     ) AS rn
              FROM field_survey_observation o
              JOIN field_survey_task t ON t.id=o.task_id
              WHERE o.sample_id IS NOT NULL
            )
            SELECT bs.id AS sample_id, bs.material_id, bs.selection_status,
                   bm.material_code, bm.material_name,
                   bp.program_code, bp.program_name,
                   te.id AS trial_entry_id, te.plot_no,
                   ft.id AS trial_id, ft.trial_code,
                   ts.id AS site_id, ts.site_code, ts.site_name,
                   ro.id AS observation_id, ro.trait_code, ro.trait_name,
                   ro.value_numeric, ro.unit, ro.quality_status,
                   EXISTS (
                     SELECT 1 FROM genotype_sample_mapping gsm WHERE gsm.sample_id=bs.id
                   ) AS has_genotype
            FROM biological_sample bs
            JOIN breeding_program bp ON bp.id=bs.program_id
            JOIN breeding_material bm ON bm.id=bs.material_id
            LEFT JOIN trial_entry te ON te.id=bs.trial_entry_id
            LEFT JOIN field_trial ft ON ft.id=te.trial_id
            LEFT JOIN trial_site ts ON ts.id=ft.site_id
            LEFT JOIN ranked_observation ro ON ro.sample_id=bs.id AND ro.rn=1
            WHERE bs.data_status='published'
              AND bs.project_id=:project_id
              AND (CAST(:program_code AS VARCHAR) IS NULL OR bp.program_code=:program_code)
            ORDER BY bm.material_code, bs.sample_code, ro.trait_code
            """
        ),
        {"project_id": project_id, "program_code": _optional_text(program_code)},
    )
    materials: dict[str, dict[str, Any]] = {}
    all_sites: set[str] = set()
    all_trials: set[str] = set()
    total_observations = 0
    total_warnings = 0
    for row in rows:
        item = _as_dict(row)
        material = materials.setdefault(item["material_id"], {
            "material_id": item["material_id"],
            "material_code": item["material_code"],
            "material_name": item["material_name"],
            "program_code": item["program_code"],
            "program_name": item["program_name"],
            "plant_ids": set(),
            "sites": set(),
            "trials": set(),
            "plots": set(),
            "genotype_plants": set(),
            "selection": {status: set() for status in SELECTION_STATUSES},
            "traits": {trait["code"]: [] for trait in VARIETY_EVALUATION_TRAITS},
            "warnings": 0,
            "environment": {},
        })
        material["plant_ids"].add(item["sample_id"])
        material["selection"].setdefault(item["selection_status"], set()).add(item["sample_id"])
        if item["has_genotype"]:
            material["genotype_plants"].add(item["sample_id"])
        if item["site_id"]:
            material["sites"].add(item["site_id"])
            all_sites.add(item["site_id"])
        if item["trial_id"]:
            material["trials"].add(item["trial_id"])
            all_trials.add(item["trial_id"])
        if item["trial_entry_id"]:
            material["plots"].add(item["trial_entry_id"])
        if not item["observation_id"]:
            continue
        total_observations += 1
        if item["quality_status"] == "warning":
            material["warnings"] += 1
            total_warnings += 1
        if item["value_numeric"] is None or item["quality_status"] == "rejected":
            continue
        if item["trait_code"] in material["traits"]:
            material["traits"][item["trait_code"]].append(float(item["value_numeric"]))
        if item["site_id"]:
            environment = material["environment"].setdefault(item["site_id"], {
                "site_id": item["site_id"],
                "site_code": item["site_code"],
                "site_name": item["site_name"],
                "plant_ids": set(),
                "traits": {trait["code"]: [] for trait in VARIETY_EVALUATION_TRAITS},
            })
            environment["plant_ids"].add(item["sample_id"])
            if item["trait_code"] in environment["traits"]:
                environment["traits"][item["trait_code"]].append(float(item["value_numeric"]))

    varieties: list[dict[str, Any]] = []
    for material in materials.values():
        trait_averages = {
            code: round(value, 2) if (value := _average(values)) is not None else None
            for code, values in material["traits"].items()
        }
        environments = []
        for environment in material["environment"].values():
            environments.append({
                "site_id": environment["site_id"],
                "site_code": environment["site_code"],
                "site_name": environment["site_name"],
                "plant_count": len(environment["plant_ids"]),
                "trait_averages": {
                    code: round(value, 2) if (value := _average(values)) is not None else None
                    for code, values in environment["traits"].items()
                },
            })
        varieties.append({
            "material_id": material["material_id"],
            "material_code": material["material_code"],
            "material_name": material["material_name"],
            "program_code": material["program_code"],
            "program_name": material["program_name"],
            "plant_count": len(material["plant_ids"]),
            "site_count": len(material["sites"]),
            "trial_count": len(material["trials"]),
            "plot_count": len(material["plots"]),
            "genotype_count": len(material["genotype_plants"]),
            "quality_warning_count": material["warnings"],
            "selection_counts": {
                status: len(ids) for status, ids in material["selection"].items()
            },
            "trait_averages": trait_averages,
            "environments": sorted(environments, key=lambda value: value["site_code"] or value["site_name"]),
        })

    bounds: dict[str, tuple[float, float]] = {}
    for trait in VARIETY_EVALUATION_TRAITS:
        values = [
            float(variety["trait_averages"][trait["code"]])
            for variety in varieties
            if variety["trait_averages"].get(trait["code"]) is not None
        ]
        bounds[trait["code"]] = (min(values), max(values)) if values else (0.0, 0.0)

    for variety in varieties:
        scores: dict[str, float | None] = {}
        weighted_total = 0.0
        available_weight = 0.0
        for trait in VARIETY_EVALUATION_TRAITS:
            low, high = bounds[trait["code"]]
            score = _trait_score(variety["trait_averages"].get(trait["code"]), trait, low, high)
            scores[trait["code"]] = round(score, 1) if score is not None else None
            if score is not None:
                weighted_total += score * trait["weight"]
                available_weight += trait["weight"]
        variety["trait_scores"] = scores
        variety["score"] = round(weighted_total / available_weight, 1) if available_weight else None
        scored_traits = sorted(
            [
                {"code": trait["code"], "name": trait["name"], "score": scores[trait["code"]]}
                for trait in VARIETY_EVALUATION_TRAITS
                if scores[trait["code"]] is not None
            ],
            key=lambda value: (-value["score"], value["name"]),
        )
        variety["strengths"] = scored_traits[:2]
        variety["weaknesses"] = list(reversed(scored_traits[-2:]))
        site_yields = [
            environment["trait_averages"].get("yield_per_plant")
            for environment in variety["environments"]
            if environment["trait_averages"].get("yield_per_plant") is not None
        ]
        variety["stability_score"] = round(
            max(0.0, 100.0 - (max(site_yields) - min(site_yields)) * 8.0), 1
        ) if len(site_yields) > 1 else None

    varieties.sort(key=lambda value: (-(value["score"] or -1), value["material_code"]))
    for index, variety in enumerate(varieties, 1):
        variety["rank"] = index
    return {
        "generated_at": datetime.now(timezone.utc),
        "program_code": _optional_text(program_code),
        "summary": {
            "variety_count": len(varieties),
            "plant_count": sum(variety["plant_count"] for variety in varieties),
            "observation_count": total_observations,
            "site_count": len(all_sites),
            "trial_count": len(all_trials),
            "quality_warning_count": total_warnings,
        },
        "traits": VARIETY_EVALUATION_TRAITS,
        "bounds": {code: {"min": low, "max": high} for code, (low, high) in bounds.items()},
        "varieties": varieties,
    }


def get_base_dashboard(
    session: Session, project_id: str, program_code: str | None = None
) -> dict[str, Any]:
    effective_program_code = _optional_text(program_code) or BASE_SHOWCASE_PROGRAM_CODE
    rows = session.execute(
        text(
            """
            WITH ranked_observation AS (
              SELECT o.*, t.survey_stage,
                     ROW_NUMBER() OVER (
                       PARTITION BY o.sample_id, t.survey_stage, o.trait_code
                       ORDER BY o.data_version DESC, o.observed_at DESC, o.id DESC
                     ) AS rn
              FROM field_survey_observation o
              JOIN field_survey_task t ON t.id=o.task_id
              WHERE o.sample_id IS NOT NULL
            )
            SELECT ts.id AS site_id, ts.site_code, ts.site_name, ts.province, ts.county,
                   ts.ecological_zone, ts.soil_type, ts.latitude, ts.longitude,
                   ft.id AS trial_id, ft.trial_code, ft.trial_name, ft.trial_year,
                   te.id AS trial_entry_id, te.plot_no, te.block_no, te.replicate_no,
                   tt.treatment_code, tt.treatment_name,
                   bm.id AS material_id, bm.material_code, bm.material_name,
                   COUNT(DISTINCT bs.id) AS plant_count,
                   COUNT(ro.id) AS observation_count,
                   COUNT(DISTINCT bs.id) FILTER (WHERE bs.selection_status='promoted') AS promoted_count,
                   COUNT(DISTINCT bs.id) FILTER (WHERE bs.selection_status='retained') AS retained_count,
                   COUNT(DISTINCT bs.id) FILTER (WHERE bs.selection_status='eliminated') AS eliminated_count,
                   COUNT(DISTINCT bs.id) FILTER (
                     WHERE EXISTS (SELECT 1 FROM genotype_sample_mapping gsm WHERE gsm.sample_id=bs.id)
                   ) AS genotype_count,
                   COUNT(ro.id) FILTER (WHERE ro.quality_status='warning') AS quality_warning_count,
                   AVG(ro.value_numeric) FILTER (WHERE ro.trait_code='plant_height' AND ro.quality_status<>'rejected') AS plant_height,
                   AVG(ro.value_numeric) FILTER (WHERE ro.trait_code='panicle_length' AND ro.quality_status<>'rejected') AS panicle_length,
                   AVG(ro.value_numeric) FILTER (WHERE ro.trait_code='effective_panicles' AND ro.quality_status<>'rejected') AS effective_panicles,
                   AVG(ro.value_numeric) FILTER (WHERE ro.trait_code='grains_per_panicle' AND ro.quality_status<>'rejected') AS grains_per_panicle,
                   AVG(ro.value_numeric) FILTER (WHERE ro.trait_code='seed_setting_rate' AND ro.quality_status<>'rejected') AS seed_setting_rate,
                   AVG(ro.value_numeric) FILTER (WHERE ro.trait_code='thousand_grain_weight' AND ro.quality_status<>'rejected') AS thousand_grain_weight,
                   AVG(ro.value_numeric) FILTER (WHERE ro.trait_code='yield_per_plant' AND ro.quality_status<>'rejected') AS yield_per_plant,
                   AVG(ro.value_numeric) FILTER (WHERE ro.trait_code='lodging_grade' AND ro.quality_status<>'rejected') AS lodging_grade
            FROM biological_sample bs
            JOIN breeding_program bp ON bp.id=bs.program_id
            JOIN breeding_material bm ON bm.id=bs.material_id
            JOIN trial_entry te ON te.id=bs.trial_entry_id
            JOIN field_trial ft ON ft.id=te.trial_id
            JOIN trial_site ts ON ts.id=ft.site_id
            JOIN trial_treatment tt ON tt.id=te.treatment_id
            LEFT JOIN ranked_observation ro ON ro.sample_id=bs.id AND ro.rn=1
            WHERE bs.data_status='published'
              AND (
                bs.project_id=:project_id
                OR (
                  bs.project_id IS NULL
                  AND bp.program_code=:base_showcase_program_code
                  AND bp.is_simulated=TRUE
                )
              )
              AND bp.program_code=:program_code
            GROUP BY ts.id, ts.site_code, ts.site_name, ts.province, ts.county,
                     ts.ecological_zone, ts.soil_type, ts.latitude, ts.longitude,
                     ft.id, ft.trial_code, ft.trial_name, ft.trial_year,
                     te.id, te.plot_no, te.block_no, te.replicate_no,
                     tt.treatment_code, tt.treatment_name,
                     bm.id, bm.material_code, bm.material_name
            ORDER BY ts.site_code, ft.trial_code, te.plot_no
            """
        ),
        {
            "project_id": project_id,
            "program_code": effective_program_code,
            "base_showcase_program_code": BASE_SHOWCASE_PROGRAM_CODE,
        },
    )
    sites: dict[str, dict[str, Any]] = {}
    total_plants = 0
    total_observations = 0
    total_warnings = 0
    for row in rows:
        item = _as_dict(row)
        plot = {
            key: item[key]
            for key in (
                "trial_entry_id", "plot_no", "block_no", "replicate_no",
                "treatment_code", "treatment_name", "material_id", "material_code", "material_name",
                "plant_count", "observation_count", "promoted_count", "retained_count",
                "eliminated_count", "genotype_count", "quality_warning_count",
            )
        }
        plot["trait_averages"] = {
            trait["code"]: round(float(item[trait["code"]]), 2) if item[trait["code"]] is not None else None
            for trait in VARIETY_EVALUATION_TRAITS
        }
        plot["quality_score"] = round(
            max(0.0, 100.0 - (plot["quality_warning_count"] / max(1, plot["observation_count"])) * 100.0), 1
        )
        site = sites.setdefault(item["site_id"], {
            "site_id": item["site_id"],
            "site_code": item["site_code"],
            "site_name": item["site_name"],
            "province": item["province"],
            "county": item["county"],
            "ecological_zone": item["ecological_zone"],
            "soil_type": item["soil_type"],
            "latitude": float(item["latitude"]) if item["latitude"] is not None else None,
            "longitude": float(item["longitude"]) if item["longitude"] is not None else None,
            "trials": {},
        })
        trial = site["trials"].setdefault(item["trial_id"], {
            "trial_id": item["trial_id"],
            "trial_code": item["trial_code"],
            "trial_name": item["trial_name"],
            "trial_year": item["trial_year"],
            "plots": [],
        })
        trial["plots"].append(plot)
        total_plants += plot["plant_count"]
        total_observations += plot["observation_count"]
        total_warnings += plot["quality_warning_count"]

    serialized_sites = []
    for site in sites.values():
        trials = []
        site_materials: set[str] = set()
        site_plants = 0
        site_observations = 0
        site_warnings = 0
        for trial in site["trials"].values():
            plots = trial["plots"]
            trial["plot_count"] = len(plots)
            trial["plant_count"] = sum(plot["plant_count"] for plot in plots)
            trial["observation_count"] = sum(plot["observation_count"] for plot in plots)
            trial["material_count"] = len({plot["material_id"] for plot in plots})
            trial["quality_warning_count"] = sum(plot["quality_warning_count"] for plot in plots)
            trials.append(trial)
            site_materials.update(plot["material_id"] for plot in plots)
            site_plants += trial["plant_count"]
            site_observations += trial["observation_count"]
            site_warnings += trial["quality_warning_count"]
        serialized_sites.append({
            **{key: value for key, value in site.items() if key != "trials"},
            "trial_count": len(trials),
            "plot_count": sum(trial["plot_count"] for trial in trials),
            "material_count": len(site_materials),
            "plant_count": site_plants,
            "observation_count": site_observations,
            "quality_warning_count": site_warnings,
            "trials": sorted(trials, key=lambda value: value["trial_code"]),
        })
    serialized_sites.sort(key=lambda value: value["site_code"] or value["site_name"])
    return {
        "generated_at": datetime.now(timezone.utc),
        "program_code": effective_program_code,
        "summary": {
            "site_count": len(serialized_sites),
            "trial_count": sum(site["trial_count"] for site in serialized_sites),
            "plot_count": sum(site["plot_count"] for site in serialized_sites),
            "material_count": len({
                plot["material_id"]
                for site in serialized_sites
                for trial in site["trials"]
                for plot in trial["plots"]
            }),
            "plant_count": total_plants,
            "observation_count": total_observations,
            "quality_warning_count": total_warnings,
        },
        "traits": VARIETY_EVALUATION_TRAITS,
        "sites": serialized_sites,
    }


def get_base_plot_plants(session: Session, trial_entry_id: str, project_id: str) -> dict[str, Any]:
    context_row = session.execute(
        text(
            """
            SELECT te.id AS trial_entry_id, te.plot_no, te.block_no, te.replicate_no,
                   ft.id AS trial_id, ft.trial_code, ft.trial_name,
                   ts.id AS site_id, ts.site_code, ts.site_name,
                   tt.treatment_code, tt.treatment_name,
                   bm.id AS material_id, bm.material_code, bm.material_name
            FROM trial_entry te
            JOIN field_trial ft ON ft.id=te.trial_id
            JOIN trial_site ts ON ts.id=ft.site_id
            JOIN trial_treatment tt ON tt.id=te.treatment_id
            JOIN breeding_material bm ON bm.id=te.material_id
            WHERE te.id=:trial_entry_id
              AND (
                ft.project_id=:project_id
                OR EXISTS (
                  SELECT 1 FROM trial_data_package package
                  WHERE package.id=ft.package_id
                    AND package.package_code=:base_showcase_package_code
                    AND package.is_simulated=TRUE
                )
              )
            """
        ),
        {
            "trial_entry_id": trial_entry_id,
            "project_id": project_id,
            "base_showcase_package_code": BASE_SHOWCASE_PACKAGE_CODE,
        },
    ).first()
    if not context_row:
        raise SinglePlantError("试验小区不存在。")
    plant_rows = [
        _as_dict(row)
        for row in session.execute(
            text(
                """
                SELECT bs.id, bs.sample_code, bs.generation_label, bs.plant_no,
                       bs.selection_status, bs.data_status,
                       (SELECT COUNT(*) FROM field_survey_observation o WHERE o.sample_id=bs.id) AS observation_count,
                       (SELECT COUNT(*) FROM genotype_sample_mapping gm WHERE gm.sample_id=bs.id) AS genotype_count,
                       (SELECT COUNT(*) FROM breeding_selection_record sr WHERE sr.sample_id=bs.id) AS selection_count
                FROM biological_sample bs
                JOIN breeding_program bp ON bp.id=bs.program_id
                WHERE bs.trial_entry_id=:trial_entry_id
                  AND (
                    bs.project_id=:project_id
                    OR (
                      bs.project_id IS NULL
                      AND bp.program_code=:base_showcase_program_code
                      AND bp.is_simulated=TRUE
                    )
                  )
                  AND bs.data_status<>'archived'
                ORDER BY bs.plant_no NULLS LAST, bs.sample_code
                """
            ),
            {
                "trial_entry_id": trial_entry_id,
                "project_id": project_id,
                "base_showcase_program_code": BASE_SHOWCASE_PROGRAM_CODE,
            },
        )
    ]
    sample_ids = [plant["id"] for plant in plant_rows]
    latest_traits: dict[str, dict[str, Any]] = {sample_id: {} for sample_id in sample_ids}
    if sample_ids:
        for row in session.execute(
            text(
                """
                WITH ranked AS (
                  SELECT o.sample_id, o.trait_code, o.trait_name, o.value_numeric, o.value_text,
                         o.unit, o.quality_status, t.survey_stage,
                         ROW_NUMBER() OVER (
                           PARTITION BY o.sample_id, t.survey_stage, o.trait_code
                           ORDER BY o.data_version DESC, o.observed_at DESC, o.id DESC
                         ) AS rn
                  FROM field_survey_observation o
                  JOIN field_survey_task t ON t.id=o.task_id
                  WHERE o.sample_id=ANY(:sample_ids)
                )
                SELECT * FROM ranked WHERE rn=1 ORDER BY sample_id, trait_code
                """
            ),
            {"sample_ids": sample_ids},
        ):
            item = _as_dict(row)
            latest_traits[item["sample_id"]][item["trait_code"]] = {
                "trait_name": item["trait_name"],
                "value": item["value_numeric"] if item["value_numeric"] is not None else item["value_text"],
                "unit": item["unit"],
                "quality_status": item["quality_status"],
                "observation_stage": item["survey_stage"],
            }
    for plant in plant_rows:
        plant["traits"] = latest_traits.get(plant["id"], {})
    return {"plot": _as_dict(context_row), "plants": plant_rows}


def _task_code(trial_id: str, stage: str) -> str:
    digest = hashlib.sha256(f"{trial_id}|{stage}".encode("utf-8")).hexdigest()[:20]
    return f"SP-{digest}"


def _survey_unit_code(sample_code: str, sample_id: str) -> str:
    clean_code = re.sub(r"[^A-Za-z0-9._\-]+", "-", sample_code).strip("-") or "plant"
    return f"SP-{clean_code[:100]}-{sample_id[:8]}"


def record_single_plant_observation(
    session: Session, sample_id: str, payload: SinglePlantObservationRequest, actor: str, project_id: str
) -> dict[str, Any]:
    if payload.value_numeric is None and not _clean(payload.value_text):
        raise SinglePlantError("数值和文本表型至少填写一项。")
    context_row = session.execute(
        text(
            """
            SELECT bs.id, bs.sample_code, bs.material_id, bs.trial_entry_id,
                   bm.material_code, bm.material_name,
                   te.trial_id, te.plot_no, te.block_no, te.replicate_no,
                   ft.trial_code, ft.trial_name, ts.site_name
            FROM biological_sample bs
            JOIN breeding_material bm ON bm.id=bs.material_id
            LEFT JOIN trial_entry te ON te.id=bs.trial_entry_id
            LEFT JOIN field_trial ft ON ft.id=te.trial_id
            LEFT JOIN trial_site ts ON ts.id=ft.site_id
            WHERE bs.id=:sample_id AND bs.project_id=:project_id
            """
        ),
        {"sample_id": sample_id, "project_id": project_id},
    ).first()
    if not context_row:
        raise SinglePlantError("单株不存在。")
    context = _as_dict(context_row)
    if not context["trial_entry_id"]:
        raise SinglePlantError("该单株尚未关联试验小区，不能录入田间表型。")

    task_code = _task_code(context["trial_id"], payload.observation_stage)
    task_id = session.scalar(text("SELECT id FROM field_survey_task WHERE task_code=:code"), {"code": task_code})
    if not task_id:
        task_id = str(uuid.uuid4())
        session.execute(
            text(
                """
                INSERT INTO field_survey_task (
                  id, task_code, task_name, trial_name, site_name, survey_stage,
                  required_traits, assigned_workers, photo_required, created_by, description
                ) VALUES (
                  :id, :code, :name, :trial_name, :site_name, :stage,
                  CAST(:traits AS JSONB), CAST(:workers AS JSONB), FALSE, :actor, :description
                )
                """
            ),
            {
                "id": task_id,
                "code": task_code,
                "name": f"{context['trial_name']} - {payload.observation_stage}单株调查",
                "trial_name": context["trial_name"],
                "site_name": context["site_name"],
                "stage": payload.observation_stage,
                "traits": _json([payload.trait_code]),
                "workers": _json([actor]),
                "actor": actor,
                "description": "由单株表型录入接口自动建立，与原试验条目保持关联。",
            },
        )
    else:
        session.execute(
            text(
                """
                UPDATE field_survey_task
                SET required_traits=(SELECT jsonb_agg(DISTINCT value) FROM jsonb_array_elements(
                    required_traits || CAST(:traits AS JSONB)
                )), updated_at=now()
                WHERE id=:id
                """
            ),
            {"traits": _json([payload.trait_code]), "id": task_id},
        )

    survey_unit_code = _survey_unit_code(context["sample_code"], sample_id)
    plot_id = session.scalar(
        text("SELECT id FROM field_survey_plot WHERE task_id=:task_id AND plot_code=:plot_code"),
        {"task_id": task_id, "plot_code": survey_unit_code},
    )
    if not plot_id:
        plot_id = str(uuid.uuid4())
        sequence_no = session.scalar(
            text("SELECT COALESCE(MAX(sequence_no), 0) + 1 FROM field_survey_plot WHERE task_id=:task_id"),
            {"task_id": task_id},
        )
        session.execute(
            text(
                """
                INSERT INTO field_survey_plot (
                  id, task_id, plot_code, sequence_no, block_no, replicate_no,
                  material_code, material_name, qr_token, assigned_to,
                  status, quality_status, trial_entry_id
                ) VALUES (
                  :id, :task_id, :plot_code, :sequence_no, :block_no, :replicate_no,
                  :material_code, :material_name, :qr_token, :actor,
                  'in_progress', 'pending', :trial_entry_id
                )
                """
            ),
            {
                "id": plot_id,
                "task_id": task_id,
                "plot_code": survey_unit_code,
                "sequence_no": sequence_no,
                "block_no": context["block_no"],
                "replicate_no": context["replicate_no"],
                "material_code": context["material_code"],
                "material_name": context["material_name"],
                "qr_token": str(uuid.uuid4()),
                "actor": actor,
                "trial_entry_id": context["trial_entry_id"],
            },
        )

    version = session.scalar(
        text(
            "SELECT COALESCE(MAX(data_version),0)+1 FROM field_survey_observation "
            "WHERE plot_id=:plot_id AND trait_code=:trait_code"
        ),
        {"plot_id": plot_id, "trait_code": payload.trait_code},
    )
    observation_id = str(uuid.uuid4())
    observed_at = payload.observed_at or datetime.now(timezone.utc)
    session.execute(
        text(
            """
            INSERT INTO field_survey_observation (
              id, task_id, plot_id, sample_id, trait_code, trait_name,
              value_numeric, value_text, unit, source_type, device_id,
              quality_status, quality_flags, observed_by, observed_at, data_version
            ) VALUES (
              :id, :task_id, :plot_id, :sample_id, :trait_code, :trait_name,
              :value_numeric, :value_text, :unit, :source_type, :device_id,
              :quality_status, CAST(:quality_flags AS JSONB), :actor, :observed_at, :data_version
            )
            """
        ),
        {
            "id": observation_id,
            "task_id": task_id,
            "plot_id": plot_id,
            "sample_id": sample_id,
            "trait_code": payload.trait_code,
            "trait_name": payload.trait_name,
            "value_numeric": payload.value_numeric,
            "value_text": _optional_text(payload.value_text),
            "unit": payload.unit,
            "source_type": payload.source_type,
            "device_id": payload.device_id,
            "quality_status": payload.quality_status,
            "quality_flags": _json(payload.quality_flags),
            "actor": actor,
            "observed_at": observed_at,
            "data_version": version,
        },
    )
    session.execute(
        text(
            """
            INSERT INTO field_survey_audit (
              id, task_id, plot_id, event_type, event_label, actor, device_id, data_version, details
            ) VALUES (
              :id, :task_id, :plot_id, 'single_plant_observation', :label,
              :actor, :device_id, :data_version, CAST(:details AS JSONB)
            )
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "task_id": task_id,
            "plot_id": plot_id,
            "label": f"录入单株表型：{payload.trait_name}",
            "actor": actor,
            "device_id": payload.device_id,
            "data_version": version,
            "details": _json({"sample_id": sample_id, "trait_code": payload.trait_code}),
        },
    )
    session.execute(
        text(
            """
            UPDATE field_survey_plot SET status='completed', quality_status=:quality_status,
              completed_at=:completed_at, completed_by=:actor, device_id=:device_id
            WHERE id=:plot_id
            """
        ),
        {
            "quality_status": payload.quality_status,
            "completed_at": observed_at,
            "actor": actor,
            "device_id": payload.device_id,
            "plot_id": plot_id,
        },
    )
    session.commit()
    return {
        "id": observation_id,
        "sample_id": sample_id,
        "task_id": task_id,
        "survey_unit_id": plot_id,
        "survey_unit_code": survey_unit_code,
        "observation_stage": payload.observation_stage,
        "trait_code": payload.trait_code,
        "data_version": version,
        "observed_at": observed_at,
    }


def map_genotype_to_single_plant(
    session: Session,
    owner_id: str,
    version_id: str,
    fid: str,
    iid: str,
    payload: SinglePlantGenotypeMappingRequest,
    project_id: str,
) -> dict[str, Any]:
    mapping_row = session.execute(
        text(
            """
            SELECT gsm.id, gsm.owner_id, gsm.material_id, gsm.sample_id,
                   gav.status AS version_status
            FROM genotype_sample_mapping gsm
            JOIN genotype_asset_version gav ON gav.id=gsm.version_id
            WHERE gsm.version_id=:version_id AND gsm.fid=:fid AND gsm.iid=:iid
            """
        ),
        {"version_id": version_id, "fid": fid, "iid": iid},
    ).first()
    if not mapping_row:
        raise SinglePlantError("基因型版本中不存在该 FID/IID。")
    mapping = _as_dict(mapping_row)
    if mapping["owner_id"] != owner_id:
        raise SinglePlantError("无权修改该基因型样本映射。")
    if mapping["version_status"] == "analysis_ready":
        raise SinglePlantError("已发布的分析就绪版本不可修改；请先创建映射修订版本。")
    if not payload.sample_id:
        session.execute(
            text(
                "UPDATE genotype_sample_mapping SET sample_id=NULL, note=:note, updated_at=now() "
                "WHERE id=:mapping_id"
            ),
            {"note": payload.note or None, "mapping_id": mapping["id"]},
        )
        session.commit()
        return {
            "id": mapping["id"], "version_id": version_id, "fid": fid, "iid": iid,
            "sample_id": None, "sample_code": None, "material_id": mapping["material_id"],
            "status": "mapped" if mapping["material_id"] else "unmapped",
        }
    sample_row = session.execute(
        text("SELECT id, sample_code, material_id FROM biological_sample WHERE id=:id AND project_id=:project_id"),
        {"id": payload.sample_id, "project_id": project_id},
    ).first()
    if not sample_row:
        raise SinglePlantError("目标单株不存在。")
    sample = _as_dict(sample_row)
    if mapping["material_id"] and mapping["material_id"] != sample["material_id"]:
        raise SinglePlantError("该 FID/IID 现有材料与目标单株所属材料不一致。")
    duplicate = session.execute(
        text(
            """
            SELECT fid, iid FROM genotype_sample_mapping
            WHERE version_id=:version_id AND sample_id=:sample_id AND id<>:mapping_id
            """
        ),
        {"version_id": version_id, "sample_id": payload.sample_id, "mapping_id": mapping["id"]},
    ).first()
    if duplicate:
        raise SinglePlantError(f"该单株已映射到 {duplicate.fid}/{duplicate.iid}；同一版本不能重复映射。")
    session.execute(
        text(
            """
            UPDATE genotype_sample_mapping
            SET sample_id=:sample_id, material_id=:material_id, status='mapped', note=:note, updated_at=now()
            WHERE id=:mapping_id
            """
        ),
        {
            "sample_id": payload.sample_id,
            "material_id": sample["material_id"],
            "note": payload.note or None,
            "mapping_id": mapping["id"],
        },
    )
    session.commit()
    return {
        "id": mapping["id"],
        "version_id": version_id,
        "fid": fid,
        "iid": iid,
        "sample_id": payload.sample_id,
        "sample_code": sample["sample_code"],
        "material_id": sample["material_id"],
        "status": "mapped",
    }


def record_single_plant_selection(
    session: Session, sample_id: str, payload: SinglePlantSelectionRequest, actor: str, project_id: str
) -> dict[str, Any]:
    sample_row = session.execute(
        text("SELECT id, program_id, material_id FROM biological_sample WHERE id=:id AND project_id=:project_id"),
        {"id": sample_id, "project_id": project_id},
    ).first()
    if not sample_row:
        raise SinglePlantError("单株不存在。")
    sample = _as_dict(sample_row)
    if payload.generation_record_id:
        generation = session.execute(
            text(
                """
                SELECT id FROM breeding_generation_record
                WHERE id=:id AND program_id=:program_id AND material_id=:material_id
                """
            ),
            {
                "id": payload.generation_record_id,
                "program_id": sample["program_id"],
                "material_id": sample["material_id"],
            },
        ).first()
        if not generation:
            raise SinglePlantError("世代记录与该单株所属项目或材料不一致。")
    record_id = str(uuid.uuid4())
    session.execute(
        text(
            """
            INSERT INTO breeding_selection_record (
              id, program_id, material_id, sample_id, generation_record_id,
              selection_year, selection_site, selection_criterion, selection_decision,
              retention_reason, source_record_no, evidence_summary, source_note,
              is_simulated, recorded_by
            ) VALUES (
              :id, :program_id, :material_id, :sample_id, :generation_record_id,
              :selection_year, :selection_site, :selection_criterion, :selection_decision,
              :retention_reason, :source_record_no, :evidence_summary, :source_note,
              FALSE, :actor
            )
            """
        ),
        {
            "id": record_id,
            "program_id": sample["program_id"],
            "material_id": sample["material_id"],
            "sample_id": sample_id,
            "generation_record_id": payload.generation_record_id,
            "selection_year": payload.selection_year,
            "selection_site": payload.selection_site,
            "selection_criterion": payload.selection_criterion,
            "selection_decision": payload.decision,
            "retention_reason": payload.retention_reason,
            "source_record_no": payload.source_record_no,
            "evidence_summary": payload.evidence_summary,
            "source_note": payload.source_note,
            "actor": actor,
        },
    )
    session.execute(
        text("UPDATE biological_sample SET selection_status=:decision, updated_at=now() WHERE id=:id"),
        {"decision": payload.decision, "id": sample_id},
    )
    session.commit()
    return {
        "id": record_id,
        "sample_id": sample_id,
        "decision": payload.decision,
        "recorded_by": actor,
        "created_at": datetime.now(timezone.utc),
    }
