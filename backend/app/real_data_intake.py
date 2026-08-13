"""Institution-facing structured data intake.

The intake layer deliberately separates raw files, mappings, staging rows and
published business records.  Institution spreadsheets may use arbitrary
headers; only explicitly mapped semantic fields participate in analysis.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Iterator
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from .data_spine import get_import_batch, record_entity_lineage, transition_import_batch


STRUCTURED_SUFFIXES = {".csv", ".tsv", ".xlsx", ".xls", ".json"}
MAX_STAGING_ROWS = int(os.getenv("DATA_INTAKE_MAX_STAGING_ROWS", "200000"))
PROFILE_SAMPLE_ROWS = 20


class IntakeError(ValueError):
    """A safe error that may be shown to a data processor."""


@dataclass(frozen=True)
class ParsedTable:
    sheet_name: str | None
    columns: tuple[str, ...]
    rows: tuple[tuple[int, dict[str, Any]], ...]


def _clean_header(value: Any, index: int) -> str:
    text_value = str(value).strip() if value is not None else ""
    return text_value or f"未命名字段_{index + 1}"


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _normalized_record(columns: Iterable[str], values: Iterable[Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for column, value in zip(columns, values):
        result[column] = _json_value(value)
    return result


def _deduplicate_headers(values: Iterable[Any]) -> tuple[str, ...]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        base = _clean_header(value, index)
        seen[base] = seen.get(base, 0) + 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return tuple(headers)


def _text_encoding(path: Path) -> str:
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            path.read_text(encoding=encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    raise IntakeError("CSV/TSV 文件编码无法识别，请另存为 UTF-8 或 GB18030。")


def _parse_delimited(path: Path, suffix: str) -> list[ParsedTable]:
    encoding = _text_encoding(path)
    with path.open("r", encoding=encoding, newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        delimiter = "\t" if suffix == ".tsv" else ","
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
        except csv.Error:
            pass
        reader = csv.reader(handle, delimiter=delimiter)
        try:
            columns = _deduplicate_headers(next(reader))
        except StopIteration as exc:
            raise IntakeError("文件没有表头。") from exc
        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in enumerate(reader, 2):
            if not any(str(value).strip() for value in values if value is not None):
                continue
            rows.append((row_number, _normalized_record(columns, values)))
            if len(rows) > MAX_STAGING_ROWS:
                raise IntakeError(f"单文件超过 {MAX_STAGING_ROWS} 行，请拆分后导入。")
    return [ParsedTable(None, columns, tuple(rows))]


def _parse_xlsx(path: Path) -> list[ParsedTable]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: list[ParsedTable] = []
    total_rows = 0
    try:
        for worksheet in workbook.worksheets:
            iterator = worksheet.iter_rows(values_only=True)
            try:
                first = next(iterator)
            except StopIteration:
                continue
            columns = _deduplicate_headers(first)
            rows: list[tuple[int, dict[str, Any]]] = []
            for row_number, values in enumerate(iterator, 2):
                if not any(value is not None and str(value).strip() for value in values):
                    continue
                rows.append((row_number, _normalized_record(columns, values)))
                total_rows += 1
                if total_rows > MAX_STAGING_ROWS:
                    raise IntakeError(f"工作簿超过 {MAX_STAGING_ROWS} 行，请按数据类型或年份拆分。")
            if rows:
                tables.append(ParsedTable(worksheet.title, columns, tuple(rows)))
    finally:
        workbook.close()
    if not tables:
        raise IntakeError("工作簿中没有可导入的数据行。")
    return tables


def _parse_xls(path: Path) -> list[ParsedTable]:
    try:
        import pandas as pd
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise IntakeError("旧版 XLS 解析组件未安装，请另存为 XLSX。") from exc
    try:
        sheets = pd.read_excel(path, sheet_name=None, dtype=object)
    except Exception as exc:
        raise IntakeError("无法读取 XLS，请确认文件未加密或另存为 XLSX。") from exc
    tables: list[ParsedTable] = []
    total_rows = 0
    for sheet_name, frame in sheets.items():
        columns = _deduplicate_headers(frame.columns)
        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in enumerate(frame.itertuples(index=False, name=None), 2):
            cleaned = [None if str(value) == "nan" else value for value in values]
            if not any(value is not None and str(value).strip() for value in cleaned):
                continue
            rows.append((row_number, _normalized_record(columns, cleaned)))
            total_rows += 1
            if total_rows > MAX_STAGING_ROWS:
                raise IntakeError(f"工作簿超过 {MAX_STAGING_ROWS} 行，请拆分后导入。")
        if rows:
            tables.append(ParsedTable(str(sheet_name), columns, tuple(rows)))
    if not tables:
        raise IntakeError("工作簿中没有可导入的数据行。")
    return tables


def _parse_json(path: Path) -> list[ParsedTable]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise IntakeError("JSON 文件必须为 UTF-8 且语法有效。") from exc
    if isinstance(payload, dict):
        records = payload.get("records") or payload.get("data")
        if records is None:
            records = [payload]
    else:
        records = payload
    if not isinstance(records, list) or not all(isinstance(item, dict) for item in records):
        raise IntakeError("JSON 顶层必须是对象数组，或包含 records/data 对象数组。")
    if len(records) > MAX_STAGING_ROWS:
        raise IntakeError(f"JSON 超过 {MAX_STAGING_ROWS} 行，请拆分后导入。")
    columns = _deduplicate_headers(dict.fromkeys(key for item in records for key in item).keys())
    rows = tuple(
        (index, {column: _json_value(item.get(column)) for column in columns})
        for index, item in enumerate(records, 1)
    )
    return [ParsedTable(None, columns, rows)]


def parse_structured_file(path: Path, original_file_name: str) -> list[ParsedTable]:
    lowered = original_file_name.lower()
    suffix = ".xls" if lowered.endswith(".xls") else Path(lowered).suffix
    if suffix not in STRUCTURED_SUFFIXES:
        return []
    if suffix in {".csv", ".tsv"}:
        return _parse_delimited(path, suffix)
    if suffix == ".xlsx":
        return _parse_xlsx(path)
    if suffix == ".xls":
        return _parse_xls(path)
    return _parse_json(path)


def _row_hash(record: dict[str, Any]) -> str:
    return hashlib.sha256(
        json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")
    ).hexdigest()


def stage_parsed_tables(
    session: Session,
    *,
    import_batch_id: str,
    import_file_id: str,
    tables: list[ParsedTable],
) -> dict[str, Any]:
    """Persist raw rows exactly once; no semantic inference happens here."""
    session.execute(
        text("DELETE FROM data_import_staging_row WHERE import_file_id=:file_id"),
        {"file_id": import_file_id},
    )
    inserts: list[dict[str, Any]] = []
    columns: list[str] = []
    samples: list[dict[str, Any]] = []
    for table in tables:
        for column in table.columns:
            if column not in columns:
                columns.append(column)
        for row_number, record in table.rows:
            if len(samples) < PROFILE_SAMPLE_ROWS:
                samples.append({"sheet": table.sheet_name, "row_number": row_number, "record": record})
            inserts.append({
                "id": str(uuid4()),
                "batch_id": import_batch_id,
                "file_id": import_file_id,
                "sheet": table.sheet_name,
                "row_number": row_number,
                "row_hash": _row_hash(record),
                "raw_record": json.dumps(record, ensure_ascii=False, default=str),
            })
            if len(inserts) >= 1000:
                _insert_staging_chunk(session, inserts)
                inserts.clear()
    if inserts:
        _insert_staging_chunk(session, inserts)
    row_count = sum(len(table.rows) for table in tables)
    session.execute(
        text(
            "UPDATE data_import_file SET parse_status='profiled', detected_columns=CAST(:columns AS jsonb), "
            "row_count=:row_count WHERE id=:file_id AND import_batch_id=:batch_id"
        ),
        {
            "columns": json.dumps(columns, ensure_ascii=False),
            "row_count": row_count,
            "file_id": import_file_id,
            "batch_id": import_batch_id,
        },
    )
    session.execute(
        text("UPDATE data_import_batch SET row_count=(SELECT count(*) FROM data_import_staging_row WHERE import_batch_id=:batch_id), updated_at=now() WHERE id=:batch_id"),
        {"batch_id": import_batch_id},
    )
    session.commit()
    return {"structured": True, "columns": columns, "row_count": row_count, "sample_rows": samples}


def _insert_staging_chunk(session: Session, rows: list[dict[str, Any]]) -> None:
    session.execute(
        text(
            """
            INSERT INTO data_import_staging_row(
                id, import_batch_id, import_file_id, source_sheet,
                source_row_number, row_hash, raw_record
            ) VALUES (
                :id, :batch_id, :file_id, :sheet,
                :row_number, :row_hash, CAST(:raw_record AS jsonb)
            )
            """
        ),
        rows,
    )


def list_semantic_fields(session: Session, data_domain: str | None = None) -> list[dict[str, Any]]:
    params: dict[str, Any] = {}
    clause = ""
    if data_domain:
        params["domain"] = data_domain
        clause = " AND data_domain=:domain"
    return [dict(row) for row in session.execute(
        text("SELECT * FROM semantic_field_definition WHERE status='active'" + clause + " ORDER BY data_domain, scope, field_code"),
        params,
    ).mappings().all()]


def create_extension_field(
    session: Session,
    *,
    field_code: str,
    data_domain: str,
    target_entity: str,
    target_field: str,
    field_name: str,
    value_type: str,
    unit: str | None,
    aliases: list[str],
    description: str | None,
    created_by: str,
) -> dict[str, Any]:
    normalized = re.sub(r"[^a-z0-9_.-]+", "_", field_code.strip().lower()).strip("_.")
    if not normalized.startswith("ext."):
        normalized = f"ext.{data_domain}.{normalized}"
    if not normalized or len(normalized) > 180:
        raise IntakeError("扩展字段代码无效。")
    row = session.execute(text(
        """
        INSERT INTO semantic_field_definition(
            id, field_code, data_domain, target_entity, target_field,
            field_name, value_type, unit, scope, aliases, description, created_by
        ) VALUES (
            :id, :field_code, :domain, :entity, :target, :name, :value_type,
            :unit, 'institution_extension', CAST(:aliases AS jsonb), :description, :created_by
        ) RETURNING *
        """
    ), {
        "id": str(uuid4()), "field_code": normalized, "domain": data_domain,
        "entity": target_entity, "target": target_field, "name": field_name.strip(),
        "value_type": value_type, "unit": unit, "aliases": json.dumps(aliases, ensure_ascii=False),
        "description": description, "created_by": created_by,
    }).mappings().one()
    session.commit()
    return dict(row)


def _source_signature(domain: str, columns: Iterable[str]) -> str:
    return hashlib.sha256(
        (domain + "|" + "|".join(sorted(item.strip() for item in columns))).encode("utf-8")
    ).hexdigest()


def save_batch_mapping(
    session: Session,
    *,
    batch_id: str,
    profile_name: str,
    mappings: list[dict[str, Any]],
    binding_context: dict[str, Any],
    actor_id: str,
) -> dict[str, Any]:
    batch = get_import_batch(session, batch_id, for_update=True)
    if not batch:
        raise IntakeError("导入批次不存在。")
    if batch["status"] not in {"uploading", "validating", "failed"}:
        raise IntakeError("当前批次状态不能修改字段映射。")
    source_columns = {
        column
        for row in session.execute(
            text("SELECT detected_columns FROM data_import_file WHERE import_batch_id=:batch_id"),
            {"batch_id": batch_id},
        )
        for column in (row.detected_columns or [])
    }
    if not source_columns:
        raise IntakeError("批次中没有已解析的结构化文件。")
    semantic_rows = {
        row["id"]: dict(row)
        for row in session.execute(
            text("SELECT * FROM semantic_field_definition WHERE status='active' AND data_domain=:domain"),
            {"domain": batch["data_domain"]},
        ).mappings().all()
    }
    seen: set[str] = set()
    normalized_mappings: list[dict[str, Any]] = []
    for item in mappings:
        source = str(item.get("source_column") or "").strip()
        action = str(item.get("mapping_action") or "map").strip()
        semantic_id = item.get("semantic_field_id")
        if source not in source_columns:
            raise IntakeError(f"源字段不存在：{source}")
        if source in seen:
            raise IntakeError(f"源字段重复映射：{source}")
        if action not in {"map", "preserve", "ignore"}:
            raise IntakeError(f"不支持的映射动作：{action}")
        semantic = semantic_rows.get(semantic_id) if semantic_id else None
        if action == "map" and semantic is None:
            raise IntakeError(f"字段 {source} 未选择有效语义字段。")
        seen.add(source)
        normalized_mappings.append({
            "source_column": source,
            "mapping_action": action,
            "semantic": semantic,
            "transform_rule": item.get("transform_rule") or {},
        })
    signature = _source_signature(batch["data_domain"], source_columns)
    version = int(session.scalar(text(
        "SELECT COALESCE(max(version),0)+1 FROM data_mapping_profile WHERE data_domain=:domain AND source_signature=:signature"
    ), {"domain": batch["data_domain"], "signature": signature}) or 1)
    profile_id = str(uuid4())
    session.execute(text(
        """
        INSERT INTO data_mapping_profile(
            id, project_id, profile_name, data_domain, source_signature, version, created_by
        ) VALUES (:id, :project_id, :name, :domain, :signature, :version, :created_by)
        """
    ), {
        "id": profile_id, "project_id": batch["project_id"], "name": profile_name.strip(),
        "domain": batch["data_domain"], "signature": signature, "version": version, "created_by": actor_id,
    })
    session.execute(text("DELETE FROM data_field_mapping WHERE import_batch_id=:batch_id"), {"batch_id": batch_id})
    for item in normalized_mappings:
        semantic = item["semantic"]
        session.execute(text(
            """
            INSERT INTO data_mapping_profile_field(
                id, mapping_profile_id, source_column, semantic_field_id,
                mapping_action, transform_rule
            ) VALUES (:id, :profile_id, :source, :semantic_id, :action, CAST(:rule AS jsonb))
            """
        ), {
            "id": str(uuid4()), "profile_id": profile_id, "source": item["source_column"],
            "semantic_id": semantic["id"] if semantic else None, "action": item["mapping_action"],
            "rule": json.dumps(item["transform_rule"], ensure_ascii=False),
        })
        session.execute(text(
            """
            INSERT INTO data_field_mapping(
                id, import_batch_id, source_column, target_entity, target_field,
                transform_rule, match_confidence, status, confirmed_by, confirmed_at
            ) VALUES (
                :id, :batch_id, :source, :entity, :field,
                CAST(:rule AS jsonb), 1.0, 'confirmed', :actor, now()
            )
            """
        ), {
            "id": str(uuid4()), "batch_id": batch_id, "source": item["source_column"],
            "entity": semantic["target_entity"] if semantic else item["mapping_action"],
            "field": semantic["field_code"] if semantic else item["source_column"],
            "rule": json.dumps(item["transform_rule"], ensure_ascii=False), "actor": actor_id,
        })
    session.execute(text(
        "UPDATE data_import_batch SET mapping_profile_id=:profile_id, binding_context=CAST(:context AS jsonb), status='validating', updated_at=now() WHERE id=:batch_id"
    ), {
        "profile_id": profile_id, "context": json.dumps(binding_context or {}, ensure_ascii=False), "batch_id": batch_id,
    })
    session.commit()
    return {"mapping_profile_id": profile_id, "version": version, "source_signature": signature}


def _transform(value: Any, value_type: str, rule: dict[str, Any]) -> Any:
    if value is None:
        return None
    result = value.strip() if isinstance(value, str) else value
    if isinstance(result, str) and result in set(rule.get("null_values") or []):
        return None
    replacements = rule.get("value_map") or {}
    if str(result) in replacements:
        result = replacements[str(result)]
    if value_type in {"number", "number_or_text"}:
        try:
            cleaned = re.sub(r"[,，%％\s]", "", str(result))
            return float(cleaned)
        except (TypeError, ValueError):
            if value_type == "number":
                raise IntakeError(f"无法转换为数值：{result}")
    if value_type == "integer":
        try:
            return int(float(str(result).strip()))
        except (TypeError, ValueError) as exc:
            raise IntakeError(f"无法转换为整数：{result}") from exc
    if value_type == "text_list":
        if isinstance(result, list):
            return [str(item).strip() for item in result if str(item).strip()]
        separator = rule.get("separator") or r"[;,，、|]"
        return [item.strip() for item in re.split(separator, str(result)) if item.strip()]
    return result


def _required_fields(session: Session, domain: str) -> set[str]:
    return set(session.scalars(text(
        "SELECT field_code FROM semantic_field_definition WHERE data_domain=:domain AND status='active' AND is_required=true"
    ), {"domain": domain}).all())


def _resolve_reference(session: Session, domain: str, mapped: dict[str, Any]) -> tuple[bool, list[dict[str, Any]]]:
    references: list[dict[str, Any]] = []
    if domain == "germplasm":
        return True, references
    material_code = mapped.get(f"{domain}.material_code")
    if material_code:
        material_id = session.scalar(text("SELECT id FROM breeding_material WHERE material_code=:code"), {"code": str(material_code)})
        if not material_id:
            return False, [{"entity_type": "breeding_material", "identifier": str(material_code), "reason": "材料编号尚未建立，请先导入种质主档。"}]
        references.append({"entity_type": "breeding_material", "entity_id": str(material_id), "identifier": str(material_code)})
    trial_code = mapped.get(f"{domain}.trial_code")
    trial_id = None
    if trial_code and domain != "trial":
        trial_id = session.scalar(text("SELECT id FROM field_trial WHERE trial_code=:code"), {"code": str(trial_code)})
        if not trial_id:
            return False, [*references, {"entity_type": "field_trial", "identifier": str(trial_code), "reason": "试验编号尚未建立，请先导入试验与小区。"}]
        references.append({"entity_type": "field_trial", "entity_id": str(trial_id), "identifier": str(trial_code)})
    if domain == "phenotype" and trial_id:
        entry_id = session.scalar(text(
            """
            SELECT entry.id FROM trial_entry entry
            JOIN breeding_material material ON material.id=entry.material_id
            JOIN trial_treatment treatment ON treatment.id=entry.treatment_id
            WHERE entry.trial_id=:trial_id AND material.material_code=:material_code
              AND treatment.treatment_code=:treatment_code
              AND entry.replicate_no=:replicate_no AND entry.plot_no=:plot_no
            """
        ), {
            "trial_id": trial_id, "material_code": str(material_code),
            "treatment_code": str(mapped.get("phenotype.treatment_code")),
            "replicate_no": mapped.get("phenotype.replicate_no"), "plot_no": str(mapped.get("phenotype.plot_no")),
        })
        if not entry_id:
            return False, [*references, {"entity_type": "trial_entry", "identifier": f"{trial_code}/{mapped.get('phenotype.plot_no')}", "reason": "没有匹配到同试验、材料、处理、重复和小区的试验单元。"}]
        references.append({"entity_type": "trial_entry", "entity_id": str(entry_id), "identifier": str(mapped.get("phenotype.plot_no"))})
    return True, references


def validate_staging_batch(session: Session, *, batch_id: str, actor_id: str) -> dict[str, Any]:
    batch = get_import_batch(session, batch_id, for_update=True)
    if not batch or not batch.get("mapping_profile_id"):
        raise IntakeError("请先完成字段映射。")
    mappings = [dict(row) for row in session.execute(text(
        """
        SELECT field.source_column, field.mapping_action, field.transform_rule,
               semantic.field_code, semantic.value_type
        FROM data_mapping_profile_field field
        LEFT JOIN semantic_field_definition semantic ON semantic.id=field.semantic_field_id
        WHERE field.mapping_profile_id=:profile_id
        """
    ), {"profile_id": batch["mapping_profile_id"]}).mappings().all()]
    required = _required_fields(session, batch["data_domain"])
    context = batch.get("binding_context") or {}
    session.execute(text("DELETE FROM data_import_row_error WHERE import_batch_id=:batch_id"), {"batch_id": batch_id})
    session.execute(text("DELETE FROM data_entity_match_candidate WHERE import_batch_id=:batch_id"), {"batch_id": batch_id})
    rows = session.execute(text(
        "SELECT id, source_row_number, raw_record FROM data_import_staging_row WHERE import_batch_id=:batch_id ORDER BY import_file_id, source_row_number"
    ), {"batch_id": batch_id}).mappings().all()
    passed = warnings = failed = 0
    for row in rows:
        mapped = dict(context)
        extensions: dict[str, Any] = {}
        errors: list[tuple[str, str | None, Any, str]] = []
        for mapping in mappings:
            source = mapping["source_column"]
            action = mapping["mapping_action"]
            value = (row["raw_record"] or {}).get(source)
            if action == "ignore":
                continue
            if action == "preserve":
                extensions[source] = value
                continue
            try:
                mapped[mapping["field_code"]] = _transform(value, mapping["value_type"], mapping["transform_rule"] or {})
            except IntakeError as exc:
                errors.append(("VALUE_CONVERSION", source, value, str(exc)))
        if extensions:
            mapped["_extensions"] = extensions
        for field_code in sorted(required):
            if mapped.get(field_code) in {None, ""}:
                errors.append(("REQUIRED_FIELD_MISSING", None, None, f"缺少必需语义字段：{field_code}"))
        resolved = False
        references: list[dict[str, Any]] = []
        if not errors:
            resolved, references = _resolve_reference(session, batch["data_domain"], mapped)
            if not resolved:
                for reference in references:
                    errors.append(("ENTITY_NOT_RESOLVED", None, reference.get("identifier"), reference["reason"]))
        status = "failed" if errors else "passed"
        resolution_status = "unresolved" if errors else "resolved"
        session.execute(text(
            "UPDATE data_import_staging_row SET mapped_record=CAST(:mapped AS jsonb), validation_status=:status, resolution_status=:resolution_status WHERE id=:id"
        ), {
            "mapped": json.dumps(mapped, ensure_ascii=False, default=str), "status": status,
            "resolution_status": resolution_status, "id": row["id"],
        })
        for error_code, source_column, raw_value, detail in errors:
            session.execute(text(
                """
                INSERT INTO data_import_row_error(
                    id, import_batch_id, source_row_number, severity, error_code,
                    source_column, raw_value, detail
                ) VALUES (:id, :batch_id, :row_number, 'error', :code, :column, :value, :detail)
                """
            ), {
                "id": str(uuid4()), "batch_id": batch_id, "row_number": row["source_row_number"],
                "code": error_code, "column": source_column,
                "value": None if raw_value is None else str(raw_value), "detail": detail,
            })
        if errors:
            failed += 1
        else:
            passed += 1
            for reference in references:
                session.execute(text(
                    """
                    INSERT INTO data_entity_match_candidate(
                        id, import_batch_id, staging_row_id, target_entity_type,
                        target_entity_id, source_identifier, match_method, confidence,
                        status, resolved_by, resolved_at
                    ) VALUES (
                        :id, :batch_id, :row_id, :entity_type, :entity_id,
                        :identifier, 'stable_identifier_exact', 1.0,
                        'confirmed', :actor, now()
                    )
                    """
                ), {
                    "id": str(uuid4()), "batch_id": batch_id, "row_id": row["id"],
                    "entity_type": reference["entity_type"], "entity_id": reference["entity_id"],
                    "identifier": reference["identifier"], "actor": actor_id,
                })
    summary = {
        "staging_row_count": len(rows), "passed_count": passed, "failed_count": failed,
        "warning_count": warnings, "uses_real_entity_resolution": True,
    }
    session.execute(text(
        """
        UPDATE data_import_batch SET status=:status, row_count=:row_count,
            accepted_count=:accepted, rejected_count=:rejected, warning_count=:warnings,
            summary=CAST(:summary AS jsonb), validated_at=now(), updated_at=now()
        WHERE id=:batch_id
        """
    ), {
        "status": "ready" if passed and not failed else "failed", "row_count": len(rows),
        "accepted": passed, "rejected": failed, "warnings": warnings,
        "summary": json.dumps(summary, ensure_ascii=False), "batch_id": batch_id,
    })
    session.commit()
    return summary


def list_batch_issues(session: Session, batch_id: str, limit: int = 500) -> list[dict[str, Any]]:
    return [dict(row) for row in session.execute(text(
        """
        SELECT id, source_row_number, severity, error_code, source_column,
               raw_value, detail, resolution_status, created_at
        FROM data_import_row_error WHERE import_batch_id=:batch_id
        ORDER BY source_row_number NULLS LAST, created_at LIMIT :limit
        """
    ), {"batch_id": batch_id, "limit": min(max(limit, 1), 2000)}).mappings().all()]


def _semantic_value(record: dict[str, Any], domain: str, field: str, default: Any = None) -> Any:
    return record.get(f"{domain}.{field}", default)


def _record_revision(
    session: Session, *, project_id: str, batch_id: str, entity_type: str,
    entity_id: str, change_type: str, before: dict[str, Any], after: dict[str, Any], actor_id: str,
) -> None:
    revision_no = int(session.scalar(text(
        "SELECT COALESCE(max(revision_no),0)+1 FROM data_record_revision WHERE entity_type=:entity_type AND entity_id=:entity_id"
    ), {"entity_type": entity_type, "entity_id": entity_id}) or 1)
    session.execute(text(
        """
        INSERT INTO data_record_revision(
            id, project_id, import_batch_id, entity_type, entity_id,
            revision_no, change_type, before_data, after_data, changed_by
        ) VALUES (
            :id, :project_id, :batch_id, :entity_type, :entity_id,
            :revision_no, :change_type, CAST(:before AS jsonb), CAST(:after AS jsonb), :actor
        )
        """
    ), {
        "id": str(uuid4()), "project_id": project_id, "batch_id": batch_id,
        "entity_type": entity_type, "entity_id": entity_id, "revision_no": revision_no,
        "change_type": change_type, "before": json.dumps(before, ensure_ascii=False, default=str),
        "after": json.dumps(after, ensure_ascii=False, default=str), "actor": actor_id,
    })


def _publish_germplasm(session: Session, batch: dict[str, Any], row: dict[str, Any], actor_id: str) -> list[tuple[str, str]]:
    record = row["mapped_record"]
    code = str(_semantic_value(record, "germplasm", "material_code")).strip()
    existing = session.execute(text("SELECT * FROM breeding_material WHERE material_code=:code"), {"code": code}).mappings().one_or_none()
    payload = {
        "material_code": code,
        "material_name": str(_semantic_value(record, "germplasm", "material_name")).strip(),
        "material_type": str(_semantic_value(record, "germplasm", "material_type", "水稻育种材料") or "水稻育种材料").strip(),
        "aliases": _semantic_value(record, "germplasm", "aliases", []) or [],
        "pedigree_summary": _semantic_value(record, "germplasm", "pedigree_summary"),
    }
    if existing:
        material_id = str(existing["id"])
        _record_revision(
            session, project_id=batch["project_id"], batch_id=batch["id"], entity_type="breeding_material",
            entity_id=material_id, change_type="update", before=dict(existing), after=payload, actor_id=actor_id,
        )
        session.execute(text(
            """
            UPDATE breeding_material SET material_name=:material_name, material_type=:material_type,
                aliases=CAST(:aliases AS jsonb), pedigree_summary=:pedigree_summary
            WHERE id=:id
            """
        ), {**payload, "aliases": json.dumps(payload["aliases"], ensure_ascii=False), "id": material_id})
    else:
        material_id = str(uuid4())
        session.execute(text(
            """
            INSERT INTO breeding_material(
                id, material_code, material_name, material_type, is_check, aliases, pedigree_summary
            ) VALUES (:id, :material_code, :material_name, :material_type, false, CAST(:aliases AS jsonb), :pedigree_summary)
            """
        ), {**payload, "aliases": json.dumps(payload["aliases"], ensure_ascii=False), "id": material_id})
        _record_revision(
            session, project_id=batch["project_id"], batch_id=batch["id"], entity_type="breeding_material",
            entity_id=material_id, change_type="create", before={}, after=payload, actor_id=actor_id,
        )
    session.execute(text(
        """
        INSERT INTO data_material_project_scope(project_id, material_id, access_level, source, created_by)
        VALUES (:project_id, :material_id, 'project', 'real_data_intake', :actor)
        ON CONFLICT (project_id, material_id) DO NOTHING
        """
    ), {"project_id": batch["project_id"], "material_id": material_id, "actor": actor_id})
    session.execute(text(
        """
        INSERT INTO data_entity_identifier(
            id, entity_type, entity_id, identifier_namespace, identifier_type,
            identifier_value, normalized_value, is_primary
        ) VALUES (
            :id, 'breeding_material', :entity_id, 'institution', 'material_code',
            :value, :normalized, true
        ) ON CONFLICT (entity_type, identifier_namespace, identifier_type, normalized_value)
          DO UPDATE SET entity_id=excluded.entity_id, identifier_value=excluded.identifier_value
        """
    ), {"id": str(uuid4()), "entity_id": material_id, "value": code, "normalized": _normalize_identifier(code)})
    return [("breeding_material", material_id)]


def _normalize_identifier(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


def _publish_trial(session: Session, batch: dict[str, Any], row: dict[str, Any], actor_id: str) -> list[tuple[str, str]]:
    record = row["mapped_record"]
    trial_code = str(_semantic_value(record, "trial", "trial_code")).strip()
    site_code = str(_semantic_value(record, "trial", "site_code")).strip()
    material_code = str(_semantic_value(record, "trial", "material_code")).strip()
    material_id = session.scalar(text("SELECT id FROM breeding_material WHERE material_code=:code"), {"code": material_code})
    if not material_id:
        raise IntakeError(f"材料 {material_code} 不存在。")
    site_id = session.scalar(text("SELECT id FROM trial_site WHERE site_code=:code"), {"code": site_code})
    if not site_id:
        site_id = str(uuid4())
        session.execute(text(
            """
            INSERT INTO trial_site(id, site_code, site_name, ecological_zone)
            VALUES (:id, :code, :name, '待补充生态区')
            """
        ), {"id": site_id, "code": site_code, "name": str(_semantic_value(record, "trial", "site_name"))})
    package_id = session.scalar(text("SELECT id FROM trial_data_package WHERE project_id=:project_id AND package_code=:code"), {
        "project_id": batch["project_id"], "code": f"INTAKE-{batch['id']}",
    })
    if not package_id:
        package_id = str(uuid4())
        session.execute(text(
            """
            INSERT INTO trial_data_package(
                id, project_id, package_code, package_name, dataset_type,
                governance_status, description, is_simulated
            ) VALUES (:id, :project_id, :code, :name, '机构真实试验数据', 'governed', :description, false)
            """
        ), {
            "id": package_id, "project_id": batch["project_id"], "code": f"INTAKE-{batch['id']}",
            "name": batch["display_name"], "description": "由统一真实数据接入流程创建。",
        })
    trial_id = session.scalar(text("SELECT id FROM field_trial WHERE trial_code=:code"), {"code": trial_code})
    if not trial_id:
        trial_id = str(uuid4())
        session.execute(text(
            """
            INSERT INTO field_trial(
                id, project_id, trial_code, package_id, site_id, trial_year,
                trial_name, design_type, replicate_count, data_status, source_note
            ) VALUES (
                :id, :project_id, :trial_code, :package_id, :site_id, :trial_year,
                :trial_name, :design_type, :replicate_count, 'published', '统一真实数据接入'
            )
            """
        ), {
            "id": trial_id, "project_id": batch["project_id"], "trial_code": trial_code,
            "package_id": package_id, "site_id": site_id, "trial_year": _semantic_value(record, "trial", "trial_year"),
            "trial_name": str(_semantic_value(record, "trial", "trial_name")),
            "design_type": str(_semantic_value(record, "trial", "design_type", "未说明设计")),
            "replicate_count": int(_semantic_value(record, "trial", "replicate_count", _semantic_value(record, "trial", "replicate_no", 1)) or 1),
        })
    treatment_code = str(_semantic_value(record, "trial", "treatment_code")).strip()
    treatment_id = session.scalar(text(
        "SELECT id FROM trial_treatment WHERE trial_id=:trial_id AND treatment_code=:code"
    ), {"trial_id": trial_id, "code": treatment_code})
    if not treatment_id:
        treatment_id = str(uuid4())
        session.execute(text(
            "INSERT INTO trial_treatment(id, trial_id, treatment_code, treatment_name) VALUES (:id, :trial_id, :code, :name)"
        ), {
            "id": treatment_id, "trial_id": trial_id, "code": treatment_code,
            "name": str(_semantic_value(record, "trial", "treatment_name", treatment_code) or treatment_code),
        })
    plot_no = str(_semantic_value(record, "trial", "plot_no")).strip()
    replicate_no = int(_semantic_value(record, "trial", "replicate_no"))
    block_no = int(_semantic_value(record, "trial", "block_no", replicate_no) or replicate_no)
    entry_id = session.scalar(text(
        """
        SELECT id FROM trial_entry WHERE trial_id=:trial_id AND treatment_id=:treatment_id
          AND material_id=:material_id AND replicate_no=:replicate_no AND plot_no=:plot_no
        """
    ), {
        "trial_id": trial_id, "treatment_id": treatment_id, "material_id": material_id,
        "replicate_no": replicate_no, "plot_no": plot_no,
    })
    if not entry_id:
        entry_id = str(uuid4())
        session.execute(text(
            """
            INSERT INTO trial_entry(
                id, trial_id, treatment_id, material_id, replicate_no, block_no,
                plot_no, raw_material_name, source_locator
            ) VALUES (
                :id, :trial_id, :treatment_id, :material_id, :replicate_no, :block_no,
                :plot_no, :material_code, :source_locator
            )
            """
        ), {
            "id": entry_id, "trial_id": trial_id, "treatment_id": treatment_id,
            "material_id": material_id, "replicate_no": replicate_no, "block_no": block_no,
            "plot_no": plot_no, "material_code": material_code,
            "source_locator": f"batch:{batch['id']}/row:{row['source_row_number']}",
        })
    session.execute(text(
        """
        INSERT INTO data_entity_identifier(
            id, entity_type, entity_id, identifier_namespace, identifier_type,
            identifier_value, normalized_value, is_primary
        ) VALUES (:id, 'field_trial', :entity_id, 'institution', 'trial_code', :value, :normalized, true)
        ON CONFLICT (entity_type, identifier_namespace, identifier_type, normalized_value)
          DO UPDATE SET entity_id=excluded.entity_id
        """
    ), {"id": str(uuid4()), "entity_id": trial_id, "value": trial_code, "normalized": _normalize_identifier(trial_code)})
    return [("field_trial", str(trial_id)), ("trial_entry", str(entry_id))]


def _resolved_entry(session: Session, record: dict[str, Any], domain: str) -> str:
    return str(session.scalar(text(
        """
        SELECT entry.id FROM trial_entry entry
        JOIN field_trial trial ON trial.id=entry.trial_id
        JOIN breeding_material material ON material.id=entry.material_id
        JOIN trial_treatment treatment ON treatment.id=entry.treatment_id
        WHERE trial.trial_code=:trial_code AND material.material_code=:material_code
          AND treatment.treatment_code=:treatment_code
          AND entry.replicate_no=:replicate_no AND entry.plot_no=:plot_no
        """
    ), {
        "trial_code": str(_semantic_value(record, domain, "trial_code")),
        "material_code": str(_semantic_value(record, domain, "material_code")),
        "treatment_code": str(_semantic_value(record, domain, "treatment_code")),
        "replicate_no": _semantic_value(record, domain, "replicate_no"),
        "plot_no": str(_semantic_value(record, domain, "plot_no")),
    }))


def _publish_phenotype(session: Session, batch: dict[str, Any], row: dict[str, Any], actor_id: str) -> list[tuple[str, str]]:
    del actor_id
    record = row["mapped_record"]
    entry_id = _resolved_entry(session, record, "phenotype")
    value = _semantic_value(record, "phenotype", "value")
    numeric = value if isinstance(value, (int, float)) else None
    text_value = None if numeric is not None else str(value)
    observation_id = str(uuid4())
    try:
        session.execute(text(
            """
            INSERT INTO trial_phenotype_observation(
                id, entry_id, trait_code, trait_name, trait_category,
                value_numeric, value_text, unit, original_value,
                observation_stage, evaluation_method, source_locator,
                quality_status, publish_status
            ) VALUES (
                :id, :entry_id, :trait_code, :trait_name, '机构映射性状',
                :numeric, :text_value, :unit, :original,
                :stage, :method, :locator, 'passed', 'published'
            )
            """
        ), {
            "id": observation_id, "entry_id": entry_id,
            "trait_code": str(_semantic_value(record, "phenotype", "trait_code")),
            "trait_name": str(_semantic_value(record, "phenotype", "trait_name")),
            "numeric": numeric, "text_value": text_value,
            "unit": str(_semantic_value(record, "phenotype", "unit", "") or ""),
            "original": str(value), "stage": str(_semantic_value(record, "phenotype", "observation_stage")),
            "method": _semantic_value(record, "phenotype", "evaluation_method"),
            "locator": f"batch:{batch['id']}/row:{row['source_row_number']}",
        })
    except Exception as exc:
        raise IntakeError("同一小区、性状和生育期已经存在观测；请通过更正流程生成修订版本。") from exc
    return [("trial_phenotype_observation", observation_id)]


def _publish_environment(session: Session, batch: dict[str, Any], row: dict[str, Any], actor_id: str) -> list[tuple[str, str]]:
    del actor_id
    record = row["mapped_record"]
    trial_id = session.scalar(text("SELECT id FROM field_trial WHERE trial_code=:code"), {"code": str(_semantic_value(record, "environment", "trial_code"))})
    entity_id = str(uuid4())
    session.execute(text(
        """
        INSERT INTO trial_environment_metric(
            id, trial_id, metric_code, metric_name, value_numeric, unit,
            original_value, collection_method, source_locator
        ) VALUES (:id, :trial_id, :code, :name, :value, :unit, :original, :method, :locator)
        ON CONFLICT (trial_id, metric_code) DO UPDATE SET
            metric_name=excluded.metric_name, value_numeric=excluded.value_numeric,
            unit=excluded.unit, original_value=excluded.original_value,
            collection_method=excluded.collection_method, source_locator=excluded.source_locator
        RETURNING id
        """
    ), {
        "id": entity_id, "trial_id": trial_id,
        "code": str(_semantic_value(record, "environment", "metric_code")),
        "name": str(_semantic_value(record, "environment", "metric_name")),
        "value": float(_semantic_value(record, "environment", "value")),
        "unit": str(_semantic_value(record, "environment", "unit")),
        "original": str(_semantic_value(record, "environment", "value")),
        "method": _semantic_value(record, "environment", "collection_method"),
        "locator": f"batch:{batch['id']}/row:{row['source_row_number']}",
    })
    actual_id = session.scalar(text("SELECT id FROM trial_environment_metric WHERE trial_id=:trial_id AND metric_code=:code"), {
        "trial_id": trial_id, "code": str(_semantic_value(record, "environment", "metric_code")),
    })
    return [("trial_environment_metric", str(actual_id))]


def _publish_management(session: Session, batch: dict[str, Any], row: dict[str, Any], actor_id: str) -> list[tuple[str, str]]:
    del actor_id
    record = row["mapped_record"]
    treatment_id = session.scalar(text(
        """
        SELECT treatment.id FROM trial_treatment treatment
        JOIN field_trial trial ON trial.id=treatment.trial_id
        WHERE trial.trial_code=:trial_code AND treatment.treatment_code=:treatment_code
        """
    ), {
        "trial_code": str(_semantic_value(record, "management", "trial_code")),
        "treatment_code": str(_semantic_value(record, "management", "treatment_code")),
    })
    if not treatment_id:
        raise IntakeError("未找到管理记录对应的试验处理。")
    entity_id = str(uuid4())
    session.execute(text(
        """
        INSERT INTO trial_management_event(
            id, treatment_id, event_type, input_name, rate_per_mu, unit, event_stage, notes
        ) VALUES (:id, :treatment_id, :event_type, :input_name, :rate, :unit, :stage, :notes)
        """
    ), {
        "id": entity_id, "treatment_id": treatment_id,
        "event_type": str(_semantic_value(record, "management", "event_type")),
        "input_name": str(_semantic_value(record, "management", "input_name")),
        "rate": _semantic_value(record, "management", "rate_per_mu"),
        "unit": _semantic_value(record, "management", "unit"),
        "stage": _semantic_value(record, "management", "event_stage"),
        "notes": f"统一导入批次 {batch['id']}，源行 {row['source_row_number']}",
    })
    return [("trial_management_event", entity_id)]


PUBLISHERS = {
    "germplasm": _publish_germplasm,
    "trial": _publish_trial,
    "phenotype": _publish_phenotype,
    "environment": _publish_environment,
    "management": _publish_management,
}


def publish_staging_batch(session: Session, *, batch_id: str, actor_id: str) -> dict[str, Any]:
    batch = get_import_batch(session, batch_id, for_update=True)
    if not batch or batch["status"] != "ready":
        raise IntakeError("只有通过校验的批次才能发布。")
    publisher = PUBLISHERS.get(batch["data_domain"])
    if publisher is None:
        raise IntakeError("该数据类型已支持安全存储和映射，但业务发布适配器尚未开放。")
    rows = session.execute(text(
        """
        SELECT staging.*, import_file.file_asset_id
        FROM data_import_staging_row staging
        JOIN data_import_file import_file ON import_file.id=staging.import_file_id
        WHERE staging.import_batch_id=:batch_id
          AND staging.validation_status='passed' AND staging.resolution_status='resolved'
          AND staging.publish_status='staged'
        ORDER BY staging.import_file_id, staging.source_row_number
        """
    ), {"batch_id": batch_id}).mappings().all()
    if not rows:
        raise IntakeError("没有可发布的已解析数据行。")
    published = 0
    try:
        for raw_row in rows:
            row = dict(raw_row)
            refs = publisher(session, batch, row, actor_id)
            for entity_type, entity_id in refs:
                record_entity_lineage(
                    session, project_id=batch["project_id"], import_batch_id=batch_id,
                    file_asset_id=str(row["file_asset_id"]), source_row_number=int(row["source_row_number"]),
                    entity_type=entity_type, entity_id=entity_id,
                    locator={"sheet": row.get("source_sheet"), "row": row["source_row_number"]},
                )
            session.execute(text(
                "UPDATE data_import_staging_row SET publish_status='published', published_entity_refs=CAST(:refs AS jsonb) WHERE id=:id"
            ), {
                "refs": json.dumps([{"entity_type": a, "entity_id": b} for a, b in refs], ensure_ascii=False),
                "id": row["id"],
            })
            published += 1
        session.flush()
        transition_import_batch(
            session, batch_id=batch_id, target_status="published",
            row_count=len(rows), accepted_count=len(rows), rejected_count=0, warning_count=0,
            summary={**(batch.get("summary") or {}), "published_rows": published, "real_entity_lineage": True},
        )
    except Exception:
        session.rollback()
        raise
    return {"batch_id": batch_id, "published_rows": published, "status": "published"}
