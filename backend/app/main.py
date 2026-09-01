import asyncio
import base64
import csv
import hashlib
import io
import json
import logging
import os
import re
import threading
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Generator, Literal
from urllib.parse import quote, urljoin

import fitz
import httpx
from bs4 import BeautifulSoup
from fastapi import BackgroundTasks, Depends, FastAPI, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.responses import FileResponse, Response, StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import Boolean, DateTime, Float, ForeignKey, Integer, JSON, String, Text, UniqueConstraint, create_engine, event, func, select, text
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column, relationship, sessionmaker, with_loader_criteria
from pgvector.sqlalchemy import Vector

from .acps_adapter import (
    AcpsExecutionResult,
    AcpsLeaderDispatchRequest,
    AcpsSettings,
    dispatch_partner_task,
    mount_acps_routes,
)
from .auth import (
    CurrentUser,
    audit_actor,
    require_business_user,
    require_data_platform_user,
    require_data_processor,
    require_field_admin,
    require_knowledge_admin,
    require_knowledge_user,
    require_published_data_reader,
    require_researcher,
)
from .document_parser import SUPPORTED_SUFFIXES as DOCUMENT_SUPPORTED_SUFFIXES, VISION_IMAGE_SUFFIXES, parse_local_document
from .conversation_title import DEFAULT_RESEARCH_SESSION_TITLE, auto_title_for_first_message
from .genomics import (
    AttachGenotypeAssetRequest,
    CreateGwasPlanRequest,
    GenomicsError,
    attach_analysis_ready_genotype,
    confirm_plan as confirm_gwas_plan,
    create_plan as create_gwas_plan,
    ensure_genomics_schema,
    get_plan as get_gwas_plan,
    local_result_bundle as build_local_gwas_result_bundle,
    list_plans as list_gwas_plans,
    local_result_file as get_local_gwas_result_file,
    request_execution as request_gwas_execution,
    run_requested_local_execution,
    upload_covariates as upload_gwas_covariates,
    upload_genotype as upload_gwas_genotype,
    upload_phenotype as upload_gwas_phenotype,
)
from .genotype_assets import (
    CreateGenotypeAssetRequest,
    GenotypeAssetError,
    GovernanceRequestCreate,
    GovernanceRequestResolution,
    MappingUpdateRequest,
    UploadInitRequest,
    analysis_ready_versions,
    artifact_path as genotype_artifact_path,
    batch_update_mapping as batch_update_genotype_mapping,
    build_mapping_template as build_genotype_mapping_template,
    build_phenotype_template as build_genotype_phenotype_template,
    complete_upload as complete_genotype_upload,
    create_asset as create_genotype_asset,
    create_governance_request as create_genotype_governance_request,
    create_mapping_revision as create_genotype_mapping_revision,
    create_upload_session as create_genotype_upload_session,
    ensure_genotype_asset_schema,
    get_asset_version as get_genotype_asset_version,
    get_upload_session as get_genotype_upload_session,
    list_assets as list_genotype_assets,
    list_governance_requests as list_genotype_governance_requests,
    list_material_suggestions as list_genotype_material_suggestions,
    publish_analysis_ready as publish_genotype_analysis_ready,
    record_upload_chunk as record_genotype_upload_chunk,
    resolve_governance_request as resolve_genotype_governance_request,
    upload_chunk_path as genotype_upload_chunk_path,
    update_mapping as update_genotype_mapping,
)
from .knowledge_index import (
    EMBEDDING_DIMENSION,
    KnowledgeIndexUnavailable,
    embed_texts,
    split_markdown,
)
from .published_data_query import (
    SQL_TEMPLATES,
    NumericFilter,
    StructuredQueryRequest,
    execute_published_data_query,
    field_catalog_for_planner,
    is_likely_data_query,
    plan_query_from_question,
    plan_query_from_structured_request,
    template_catalog,
)
from .research_agent import (
    EmptyResearchAnswerError,
    ResearchAgentError,
    infer_controlled_query_request,
    prepare_working_memory_state,
    stream_research_reply,
)
from .research_report import build_analysis_chart_png, build_research_report_pdf, is_report_request
from .research_search import search_public_references
from .breeding_dossier import (
    BreedingDossierError,
    build_breeding_report_context,
    build_breeding_report_evidence_context,
    build_breeding_report_pdf,
    ensure_breeding_dossier_schema,
    is_breeding_report_request,
    seed_mock_breeding_dossiers,
)
from .breeding_intelligence import (
    DEFAULT_RECOMMENDATION_FILTERS,
    DEFAULT_RECOMMENDATION_WEIGHTS,
    _json_safe as intelligence_json_safe,
    build_intelligence_pdf,
    build_material_analysis,
    ensure_breeding_intelligence_schema,
    get_material_analysis_run,
    get_parent_recommendation_run,
    get_recommendation_rule,
    list_materials as list_intelligence_materials,
    recommendation_csv,
    run_parent_recommendation,
    save_material_analysis,
    save_parent_recommendation,
    update_recommendation_rule,
)
from .institution_data import (
    DATASET_LABELS,
    FIELD_ALIASES,
    SUPPORTED_SUFFIXES as INSTITUTION_SUPPORTED_SUFFIXES,
    InstitutionDataError,
    InstitutionDataSettings,
    InstitutionDatabaseManager,
    MinioInstitutionStore,
    default_access_policy,
    import_into_institution_database,
    list_batches as list_institution_batches,
    object_key_for,
    parse_field_mapping,
    safe_identifier,
    stage_upload,
    trace_entity as trace_institution_entity,
)
from .trial_package import (
    build_published_trial_evidence,
    ensure_trial_package_schema,
    get_trial_import_batch,
    list_trial_import_batches,
    publish_trial_package,
    retire_legacy_seeded_trial_demo,
    upload_trial_package,
)
from .trial_statistics import TrialStatisticsError, run_controlled_trial_analysis


# Keep background parser failures in the normal Uvicorn container log stream.
logger = logging.getLogger("uvicorn.error")

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+psycopg://rice:rice_demo_password@localhost:54329/rice_demo")
MIGRATION_DATABASE_URL = os.getenv("MIGRATION_DATABASE_URL", DATABASE_URL)
APP_DATABASE_ROLE = os.getenv("APP_DATABASE_ROLE", "rice_app")
APP_DATABASE_PASSWORD = os.getenv("APP_DATABASE_PASSWORD", "rice_app_demo_password")
DEPLOYMENT_ENV = os.getenv("DEPLOYMENT_ENV", "development").strip().lower()
CORS_ALLOWED_ORIGINS = [
    origin.strip()
    for origin in os.getenv("CORS_ALLOWED_ORIGINS", "http://localhost:5173,http://localhost:5183").split(",")
    if origin.strip()
]
TRUSTED_HOSTS = [
    host.strip()
    for host in os.getenv("TRUSTED_HOSTS", "localhost,127.0.0.1").split(",")
    if host.strip()
]
RAW_STORAGE_DIR = Path(os.getenv("RAW_STORAGE_DIR", "./data/raw"))
RESEARCH_STORAGE_DIR = Path(os.getenv("RESEARCH_STORAGE_DIR", "./data/research"))

# 海南南繁 is the server-owned default institution. Accounts can be
# pre-provisioned into another institution, but users never create, select, or
# switch institutions from the browser or from an access-token claim.
INSTITUTION_ID = "hainan-nanfan"
INSTITUTION_CODE = "HNNF"
INSTITUTION_NAME = "海南南繁"
DEFAULT_PROJECT_ID = "00000000-0000-4000-8000-000000000001"
DEFAULT_PROJECT_CODE = "HNNF-DEFAULT"
DEFAULT_PROJECT_NAME = "海南南繁水稻育种研究"
BUSINESS_ROLES = ("data_processor", "field_admin", "researcher")
INSTITUTION_DATA_ENABLED = os.getenv("INSTITUTION_DATA_ENABLED", "false").strip().lower() in {
    "1", "true", "yes", "on",
}
INSTITUTION_DATA_SETTINGS = InstitutionDataSettings.from_env()
INSTITUTION_OBJECT_STORE = MinioInstitutionStore(INSTITUTION_DATA_SETTINGS)
INSTITUTION_DATABASES = InstitutionDatabaseManager(INSTITUTION_DATA_SETTINGS)


def _unsafe_production_value(value: str) -> bool:
    """Identify values that are acceptable in a demo but unsafe for deployment."""
    normalized = (value or "").strip().lower()
    return not normalized or any(marker in normalized for marker in (
        "replace-with",
        "change-me",
        "demo_password",
        "rice_demo_password",
        "rice_app_demo_password",
    ))


def validate_runtime_configuration() -> None:
    """Fail closed when a production container is accidentally started as a demo.

    Docker Compose can validate that an environment variable is present, but it
    cannot determine whether a copied example password or an HTTP callback URL
    was left in place.  A startup check gives the operator a clear failure
    before any data service begins accepting requests.
    """
    if DEPLOYMENT_ENV not in {"production", "staging"}:
        return

    issuer = os.getenv("KEYCLOAK_ISSUER", "")
    jwks_url = os.getenv("KEYCLOAK_JWKS_URL", "")
    checks = {
        "DATABASE_URL": DATABASE_URL,
        "MIGRATION_DATABASE_URL": MIGRATION_DATABASE_URL,
        "APP_DATABASE_PASSWORD": APP_DATABASE_PASSWORD,
        "KEYCLOAK_ISSUER": issuer,
        "KEYCLOAK_JWKS_URL": jwks_url,
        "CORS_ALLOWED_ORIGINS": ",".join(CORS_ALLOWED_ORIGINS),
        "TRUSTED_HOSTS": ",".join(TRUSTED_HOSTS),
    }
    if INSTITUTION_DATA_ENABLED:
        checks.update({
            "MINIO_ROOT_USER": INSTITUTION_DATA_SETTINGS.minio_access_key,
            "MINIO_ROOT_PASSWORD": INSTITUTION_DATA_SETTINGS.minio_secret_key,
        })
    invalid = [key for key, value in checks.items() if _unsafe_production_value(value)]
    if invalid:
        raise RuntimeError(f"生产环境配置缺失或仍是演示值: {', '.join(invalid)}")
    if any(origin == "*" or not origin.startswith("https://") for origin in CORS_ALLOWED_ORIGINS):
        raise RuntimeError("生产环境 CORS_ALLOWED_ORIGINS 只能包含明确的 https:// 来源")
    if not issuer.startswith("https://"):
        raise RuntimeError("生产环境 KEYCLOAK_ISSUER 必须使用 HTTPS 公网/内网地址")
    if "*" in TRUSTED_HOSTS:
        raise RuntimeError("生产环境 TRUSTED_HOSTS 不允许使用通配符")


def _download_content_disposition(filename: str, fallback_filename: str = "download") -> str:
    """Build a browser-safe attachment header while retaining a UTF-8 display name.

    Starlette encodes response headers as Latin-1.  A Chinese filename therefore
    cannot be placed directly in ``filename=`` and would otherwise turn a valid
    download into a 500 response.  RFC 5987's ``filename*`` form keeps the name
    available to modern browsers, with an ASCII fallback for older clients.
    """
    requested = Path(str(filename or fallback_filename)).name.replace("\r", "").replace("\n", "").strip()
    if not requested:
        requested = fallback_filename

    requested_suffix = Path(requested).suffix
    fallback_stem = re.sub(r"[^A-Za-z0-9._-]+", "-", Path(fallback_filename).stem).strip("-._") or "download"
    fallback_suffix = requested_suffix or Path(fallback_filename).suffix
    fallback = f"{fallback_stem}{fallback_suffix}"
    return f"attachment; filename=\"{fallback}\"; filename*=UTF-8''{quote(requested, safe='')}"
ENABLE_SOURCE_DEDUPLICATION = os.getenv("ENABLE_SOURCE_DEDUPLICATION", "false").strip().lower() in {"1", "true", "yes", "on"}
MAX_RESEARCH_ATTACHMENT_BYTES = 10 * 1024 * 1024
MAX_RESEARCH_CONTEXT_CHARS = 90000
MAX_NATIVE_VISION_IMAGES = 4
MAX_KNOWLEDGE_COSINE_DISTANCE = 0.34
MAX_EXPLICIT_DOCUMENT_COSINE_DISTANCE = 0.48
MAX_EXPLICIT_DOCUMENT_DISTANCE_GAP = 0.045
MAX_KNOWLEDGE_DOCUMENTS = 3
MAX_KNOWLEDGE_CHUNKS_PER_DOCUMENT = 2
KNOWLEDGE_CROP_TERMS = ("水稻", "小麦", "玉米", "大豆", "棉花", "油菜", "马铃薯")
MAX_NATIVE_VISION_BYTES = 20 * 1024 * 1024
MAX_NATIVE_VISION_EDGE = 1600
NATIVE_VISION_JPEG_QUALITY = 82
MAX_KNOWLEDGE_FILE_BYTES = 100 * 1024 * 1024
MAX_KNOWLEDGE_BATCH_FILES = 10
MAX_KNOWLEDGE_BATCH_BYTES = 500 * 1024 * 1024


def _knowledge_document_concurrency() -> int:
    """Return the bounded number of CPU-heavy document jobs for this API."""
    raw_value = os.getenv("KNOWLEDGE_DOCUMENT_CONCURRENCY", "1").strip()
    try:
        value = int(raw_value)
    except ValueError:
        logger.warning(
            "Ignoring invalid KNOWLEDGE_DOCUMENT_CONCURRENCY=%r; using 1",
            raw_value,
        )
        return 1
    if value < 1:
        logger.warning(
            "Ignoring non-positive KNOWLEDGE_DOCUMENT_CONCURRENCY=%r; using 1",
            raw_value,
        )
        return 1
    # This shared-server profile has one MinerU CLI slot and CPU-bound
    # Docling/PaddleOCR/BGE stages.  A higher value only makes different
    # stages contend for the same host CPU, so production stays serial.
    if value > 1:
        logger.warning(
            "KNOWLEDGE_DOCUMENT_CONCURRENCY=%r is not supported by the local "
            "CPU parser pipeline; using 1",
            raw_value,
        )
    return 1


KNOWLEDGE_DOCUMENT_CONCURRENCY = _knowledge_document_concurrency()
KNOWLEDGE_DOCUMENT_SEMAPHORE = threading.BoundedSemaphore(KNOWLEDGE_DOCUMENT_CONCURRENCY)
KNOWLEDGE_STORAGE_DIR = RESEARCH_STORAGE_DIR / "knowledge"
RESULT_STORAGE_DIR = RESEARCH_STORAGE_DIR / "results"
RAW_STORAGE_DIR.mkdir(parents=True, exist_ok=True)
RESEARCH_STORAGE_DIR.mkdir(parents=True, exist_ok=True)
KNOWLEDGE_STORAGE_DIR.mkdir(parents=True, exist_ok=True)
RESULT_STORAGE_DIR.mkdir(parents=True, exist_ok=True)

# RiceData renders protected digits as tiny GIFs. The filenames rotate per page,
# while the binary glyph for each digit is stable. Matching the image SHA-256 is
# more reliable than OCRing a full page and remains fail-safe for unknown glyphs.
RICE_DATA_GLYPH_HASH_TO_DIGIT = {
    "4022dcae2922ed5bfa952bbcf1b33c7d99d33e065739bcd2b5d0298df67765fd": "0",
    "b956aa941142ba9a2aad0f5999b271116c17f31ca891905340b71f3282ec0420": "1",
    "11a7a3862223d0e7b1aeb154b59dfd9d7351dac5d09d75e94147aff3b25e4d6f": "2",
    "cd16f6a7405f4dd83471ca2867789d0693769c3739ab07eda28f35e9e694406d": "3",
    "6c080f166b2aba3423008bd44ee6b4596e4e1827e44b037ae7c38e48ac7708d2": "4",
    "d9b3b643730ed48642a41839a6894577d339d06a56d9890c73a5d0390a28e38a": "5",
    "53a482d2fc3ac08a4033e183db813bfc708d41baf90fc53e9ee38e0ef1e38cf5": "6",
    "de3b1e78c21a9dd6ce8dfc3f887f3b5066a3d3d4d78ee93043281a1fad89d88e": "7",
    "516ae310ccbf277ebc076e5fa07d40cbe0eb6236043ea8f5e39ee0cae2efd7e0": "8",
    "47d3be245110793b44eb85baeb629fcf4a3dc4b6cc2a5db7a715054b52ed540f": "9",
}


class Base(DeclarativeBase):
    pass


class Institution(Base):
    __tablename__ = "institution"

    id: Mapped[str] = mapped_column(String(80), primary_key=True)
    institution_code: Mapped[str] = mapped_column(String(40), unique=True)
    institution_name: Mapped[str] = mapped_column(String(200))
    status: Mapped[str] = mapped_column(String(30), default="active")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class InstitutionDataConfig(Base):
    """Control-plane routing for one institution's isolated data plane."""

    __tablename__ = "institution_data_config"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    institution_id: Mapped[str] = mapped_column(
        ForeignKey("institution.id", ondelete="CASCADE"), unique=True, index=True
    )
    minio_bucket: Mapped[str] = mapped_column(String(200), unique=True)
    business_database: Mapped[str] = mapped_column(String(63), unique=True)
    access_policy: Mapped[dict] = mapped_column(JSON, default=dict)
    status: Mapped[str] = mapped_column(String(30), default="active")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc)
    )


class PlatformAccount(Base):
    """Application directory entry backed by a Keycloak account."""

    __tablename__ = "platform_account"

    username: Mapped[str] = mapped_column(String(120), primary_key=True)
    keycloak_subject: Mapped[str | None] = mapped_column(String(120), nullable=True, unique=True)
    display_name: Mapped[str] = mapped_column(String(200))
    business_role: Mapped[str] = mapped_column(String(40), index=True)
    institution_id: Mapped[str] = mapped_column(String(80), default=INSTITUTION_ID, index=True)
    active: Mapped[bool] = mapped_column(Boolean, default=True)
    last_login_at: Mapped[datetime | None] = mapped_column(DateTime(timezone=True), nullable=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))


class ResearchProject(Base):
    """A project boundary inside the one Hainan NanFan institution."""

    __tablename__ = "research_project"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_code: Mapped[str] = mapped_column(String(80), unique=True, index=True)
    project_name: Mapped[str] = mapped_column(String(300))
    description: Mapped[str | None] = mapped_column(Text, nullable=True)
    institution_id: Mapped[str] = mapped_column(String(80), default=INSTITUTION_ID, index=True)
    status: Mapped[str] = mapped_column(String(30), default="active", index=True)
    created_by: Mapped[str] = mapped_column(String(120))
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))


class ProjectMember(Base):
    __tablename__ = "project_member"
    __table_args__ = (UniqueConstraint("project_id", "username", name="uq_project_member_project_user"),)

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id", ondelete="CASCADE"), index=True)
    username: Mapped[str] = mapped_column(ForeignKey("platform_account.username", ondelete="CASCADE"), index=True)
    member_role: Mapped[str] = mapped_column(String(40), default="researcher")
    created_by: Mapped[str] = mapped_column(String(120))
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class PermissionAudit(Base):
    __tablename__ = "permission_audit"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    actor_id: Mapped[str] = mapped_column(String(120), index=True)
    actor_name: Mapped[str] = mapped_column(String(200))
    action: Mapped[str] = mapped_column(String(100), index=True)
    target_type: Mapped[str] = mapped_column(String(80), index=True)
    target_id: Mapped[str] = mapped_column(String(160), index=True)
    institution_id: Mapped[str] = mapped_column(String(80), default=INSTITUTION_ID, index=True)
    project_id: Mapped[str | None] = mapped_column(String(36), nullable=True, index=True)
    before_state: Mapped[dict] = mapped_column(JSON, default=dict)
    after_state: Mapped[dict] = mapped_column(JSON, default=dict)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), index=True)


class Variety(Base):
    __tablename__ = "variety_basic"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    variety_name: Mapped[str] = mapped_column(String(200), index=True)
    normalized_name: Mapped[str] = mapped_column(String(200), index=True)
    alias_names: Mapped[list] = mapped_column(JSON, default=list)
    raw_variety_title: Mapped[str | None] = mapped_column(Text, nullable=True)
    variety_type: Mapped[str | None] = mapped_column(String(100), nullable=True)
    female_parent: Mapped[str | None] = mapped_column(String(200), nullable=True)
    male_parent: Mapped[str | None] = mapped_column(String(200), nullable=True)
    breeding_unit: Mapped[str | None] = mapped_column(String(300), nullable=True)
    approval_number: Mapped[str | None] = mapped_column(String(100), nullable=True)
    approval_year: Mapped[str | None] = mapped_column(String(20), nullable=True)
    suitable_region: Mapped[str | None] = mapped_column(Text, nullable=True)
    source_review_id: Mapped[str | None] = mapped_column(ForeignKey("source_review.id"), nullable=True)
    data_status: Mapped[str] = mapped_column(String(30), default="draft")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))

    observations: Mapped[list["PhenotypeObservation"]] = relationship(back_populates="variety", cascade="all, delete-orphan")


class SourceReview(Base):
    __tablename__ = "source_review"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    source_type: Mapped[str] = mapped_column(String(30))
    source_name: Mapped[str] = mapped_column(String(500))
    source_url: Mapped[str | None] = mapped_column(Text, nullable=True)
    file_path: Mapped[str | None] = mapped_column(Text, nullable=True)
    file_hash: Mapped[str | None] = mapped_column(String(128), nullable=True)
    raw_text: Mapped[str | None] = mapped_column(Text, nullable=True)
    page_or_locator: Mapped[str | None] = mapped_column(String(300), nullable=True)
    parsing_status: Mapped[str] = mapped_column(String(30), default="parsed")
    quality_status: Mapped[str] = mapped_column(String(30), default="pending")
    template_version_id: Mapped[str | None] = mapped_column(String(36), nullable=True, index=True)
    review_history: Mapped[list] = mapped_column(JSON, default=list)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class PhenotypeObservation(Base):
    __tablename__ = "phenotype_observation"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    variety_id: Mapped[str] = mapped_column(ForeignKey("variety_basic.id"), index=True)
    source_review_id: Mapped[str | None] = mapped_column(ForeignKey("source_review.id"), nullable=True)
    trait_code: Mapped[str] = mapped_column(String(100), index=True)
    trait_name: Mapped[str] = mapped_column(String(100))
    trait_category: Mapped[str] = mapped_column(String(50))
    observation_type: Mapped[str] = mapped_column(String(30), default="numeric")
    value_numeric: Mapped[float | None] = mapped_column(Float, nullable=True)
    value_min: Mapped[float | None] = mapped_column(Float, nullable=True)
    value_max: Mapped[float | None] = mapped_column(Float, nullable=True)
    value_text: Mapped[str | None] = mapped_column(Text, nullable=True)
    unit: Mapped[str | None] = mapped_column(String(50), nullable=True)
    original_value: Mapped[str] = mapped_column(Text)
    original_field: Mapped[str | None] = mapped_column(String(150), nullable=True)
    source_locator: Mapped[str | None] = mapped_column(String(300), nullable=True)
    trial_year: Mapped[str | None] = mapped_column(String(30), nullable=True)
    trial_location: Mapped[str | None] = mapped_column(String(200), nullable=True)
    evaluation_method: Mapped[str | None] = mapped_column(String(200), nullable=True)
    rule_version: Mapped[str] = mapped_column(String(50), default="v1.0")
    quality_status: Mapped[str] = mapped_column(String(30), default="pending")
    publish_status: Mapped[str] = mapped_column(String(30), default="pending")
    review_comment: Mapped[str | None] = mapped_column(Text, nullable=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))

    variety: Mapped[Variety] = relationship(back_populates="observations")


class DataRule(Base):
    __tablename__ = "data_rule"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    rule_code: Mapped[str] = mapped_column(String(50), index=True)
    rule_name: Mapped[str] = mapped_column(String(200))
    rule_type: Mapped[str] = mapped_column(String(50))
    version: Mapped[str] = mapped_column(String(30))
    severity: Mapped[str] = mapped_column(String(30), default="info")
    config: Mapped[dict] = mapped_column(JSON, default=dict)
    status: Mapped[str] = mapped_column(String(30), default="published")
    change_reason: Mapped[str] = mapped_column(Text, default="初始规则")
    created_by: Mapped[str] = mapped_column(String(100), default="数据处理员-张三")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class DataTemplate(Base):
    __tablename__ = "data_template"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    template_code: Mapped[str] = mapped_column(String(80), unique=True, index=True)
    template_name: Mapped[str] = mapped_column(String(200))
    data_scope: Mapped[str] = mapped_column(String(100))
    target_table: Mapped[str] = mapped_column(String(100))
    description: Mapped[str] = mapped_column(Text)
    current_version_id: Mapped[str | None] = mapped_column(String(36), nullable=True)
    status: Mapped[str] = mapped_column(String(30), default="published")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class TemplateVersion(Base):
    __tablename__ = "template_version"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    template_id: Mapped[str] = mapped_column(ForeignKey("data_template.id"), index=True)
    version: Mapped[str] = mapped_column(String(30))
    change_summary: Mapped[str] = mapped_column(Text)
    field_definitions: Mapped[list] = mapped_column(JSON, default=list)
    status: Mapped[str] = mapped_column(String(30), default="published")
    created_by: Mapped[str] = mapped_column(String(100), default="字段管理员")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class FieldChangeRequest(Base):
    __tablename__ = "field_change_request"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    source_review_id: Mapped[str] = mapped_column(ForeignKey("source_review.id"), index=True)
    template_id: Mapped[str] = mapped_column(ForeignKey("data_template.id"), index=True)
    source_field: Mapped[str] = mapped_column(String(200))
    sample_value: Mapped[str | None] = mapped_column(Text, nullable=True)
    request_note: Mapped[str | None] = mapped_column(Text, nullable=True)
    status: Mapped[str] = mapped_column(String(30), default="pending")
    submitted_by: Mapped[str] = mapped_column(String(100))
    resolved_by: Mapped[str | None] = mapped_column(String(100), nullable=True)
    resolution_note: Mapped[str | None] = mapped_column(Text, nullable=True)
    resolved_version_id: Mapped[str | None] = mapped_column(String(36), nullable=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class RootPhenotypeObservation(Base):
    __tablename__ = "root_phenotype_observation"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    variety_id: Mapped[str] = mapped_column(ForeignKey("variety_basic.id"), index=True)
    source_review_id: Mapped[str | None] = mapped_column(ForeignKey("source_review.id"), nullable=True)
    trait_code: Mapped[str] = mapped_column(String(100), index=True)
    trait_name: Mapped[str] = mapped_column(String(100))
    trait_category: Mapped[str] = mapped_column(String(80))
    value_numeric: Mapped[float | None] = mapped_column(Float, nullable=True)
    value_text: Mapped[str | None] = mapped_column(Text, nullable=True)
    unit: Mapped[str | None] = mapped_column(String(50), nullable=True)
    original_value: Mapped[str] = mapped_column(Text)
    original_field: Mapped[str | None] = mapped_column(String(150), nullable=True)
    source_locator: Mapped[str | None] = mapped_column(String(300), nullable=True)
    template_version: Mapped[str] = mapped_column(String(30))
    published_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class ResearchSession(Base):
    """A private conversation owned by one authenticated researcher."""

    __tablename__ = "research_session"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    owner_id: Mapped[str] = mapped_column(String(120), index=True)
    title: Mapped[str] = mapped_column(String(200), default=DEFAULT_RESEARCH_SESSION_TITLE)
    memory_summary: Mapped[str | None] = mapped_column(Text, nullable=True)
    memory_state: Mapped[dict] = mapped_column(JSON, default=dict)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))

    messages: Mapped[list["ResearchMessage"]] = relationship(back_populates="session", cascade="all, delete-orphan")
    attachments: Mapped[list["ResearchAttachment"]] = relationship(back_populates="session", cascade="all, delete-orphan")


class ResearchMessage(Base):
    __tablename__ = "research_message"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    session_id: Mapped[str] = mapped_column(ForeignKey("research_session.id", ondelete="CASCADE"), index=True)
    owner_id: Mapped[str] = mapped_column(String(120), index=True)
    role: Mapped[str] = mapped_column(String(20))
    content: Mapped[str] = mapped_column(Text)
    evidence: Mapped[list] = mapped_column(JSON, default=list)
    operation_state: Mapped[list] = mapped_column(JSON, default=list)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))

    session: Mapped[ResearchSession] = relationship(back_populates="messages")


class ResearchAttachment(Base):
    __tablename__ = "research_attachment"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    session_id: Mapped[str] = mapped_column(ForeignKey("research_session.id", ondelete="CASCADE"), index=True)
    owner_id: Mapped[str] = mapped_column(String(120), index=True)
    file_name: Mapped[str] = mapped_column(String(500))
    content_type: Mapped[str | None] = mapped_column(String(120), nullable=True)
    size_bytes: Mapped[int] = mapped_column()
    storage_path: Mapped[str] = mapped_column(Text)
    parser_name: Mapped[str | None] = mapped_column(String(80), nullable=True)
    parsing_status: Mapped[str] = mapped_column(String(30), default="queued")
    parsed_markdown: Mapped[str | None] = mapped_column(Text, nullable=True)
    parsed_metadata: Mapped[dict] = mapped_column(JSON, default=dict)
    parser_warnings: Mapped[list] = mapped_column(JSON, default=list)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))

    session: Mapped[ResearchSession] = relationship(back_populates="attachments")


class ResearchAudit(Base):
    __tablename__ = "research_audit"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    owner_id: Mapped[str] = mapped_column(String(120), index=True)
    session_id: Mapped[str | None] = mapped_column(String(36), nullable=True, index=True)
    action: Mapped[str] = mapped_column(String(100))
    audit_metadata: Mapped[dict] = mapped_column("metadata", JSON, default=dict)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class ResearchResult(Base):
    """A durable, private research product generated by the assistant.

    Conversation messages can be deleted without removing a deliberately saved
    research output.  Every result is still scoped by the Keycloak subject at
    the database layer and may retain a nullable pointer to its source message.
    """

    __tablename__ = "research_result"
    __table_args__ = (
        UniqueConstraint("owner_id", "source_message_id", "result_type", name="uq_research_result_source_type"),
    )

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    owner_id: Mapped[str] = mapped_column(String(120), index=True)
    session_id: Mapped[str | None] = mapped_column(ForeignKey("research_session.id", ondelete="SET NULL"), nullable=True, index=True)
    source_message_id: Mapped[str | None] = mapped_column(ForeignKey("research_message.id", ondelete="SET NULL"), nullable=True, index=True)
    analysis_run_id: Mapped[str | None] = mapped_column(String(36), nullable=True, index=True)
    result_type: Mapped[str] = mapped_column(String(40), index=True)
    title: Mapped[str] = mapped_column(String(500))
    content_type: Mapped[str] = mapped_column(String(120))
    file_name: Mapped[str] = mapped_column(String(500))
    size_bytes: Mapped[int] = mapped_column(Integer, default=0)
    storage_path: Mapped[str] = mapped_column(Text)
    summary: Mapped[str | None] = mapped_column(Text, nullable=True)
    result_metadata: Mapped[dict] = mapped_column("metadata", JSON, default=dict)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), index=True)


class KnowledgeFolder(Base):
    """A private folder tree or an administrator-managed public category tree."""

    __tablename__ = "knowledge_folder"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    scope: Mapped[str] = mapped_column(String(20), index=True)
    owner_id: Mapped[str] = mapped_column(String(120), index=True)
    parent_id: Mapped[str | None] = mapped_column(ForeignKey("knowledge_folder.id", ondelete="CASCADE"), nullable=True, index=True)
    folder_name: Mapped[str] = mapped_column(String(200))
    description: Mapped[str | None] = mapped_column(Text, nullable=True)
    created_by: Mapped[str] = mapped_column(String(120))
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))


class KnowledgeDocument(Base):
    """Non-structured reference material. It never writes phenotype tables."""

    __tablename__ = "knowledge_document"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    scope: Mapped[str] = mapped_column(String(20), index=True)
    owner_id: Mapped[str] = mapped_column(String(120), index=True)
    folder_id: Mapped[str | None] = mapped_column(ForeignKey("knowledge_folder.id", ondelete="SET NULL"), nullable=True, index=True)
    original_file_name: Mapped[str] = mapped_column(String(500))
    display_title: Mapped[str] = mapped_column(String(500), index=True)
    content_type: Mapped[str | None] = mapped_column(String(160), nullable=True)
    size_bytes: Mapped[int] = mapped_column(Integer)
    content_hash: Mapped[str] = mapped_column(String(64), index=True)
    storage_path: Mapped[str] = mapped_column(Text)
    source_organization: Mapped[str | None] = mapped_column(String(300), nullable=True)
    author: Mapped[str | None] = mapped_column(String(300), nullable=True)
    publication_year: Mapped[str | None] = mapped_column(String(20), nullable=True)
    source_url: Mapped[str | None] = mapped_column(Text, nullable=True)
    short_description: Mapped[str | None] = mapped_column(Text, nullable=True)
    authorization_basis: Mapped[str | None] = mapped_column(Text, nullable=True)
    license_scope: Mapped[str | None] = mapped_column(String(100), nullable=True)
    topic_tags: Mapped[list] = mapped_column(JSON, default=list)
    parsing_status: Mapped[str] = mapped_column(String(30), default="processing")
    indexing_status: Mapped[str] = mapped_column(String(30), default="pending")
    parser_name: Mapped[str | None] = mapped_column(String(80), nullable=True)
    parsed_characters: Mapped[int] = mapped_column(Integer, default=0)
    parser_warnings: Mapped[list] = mapped_column(JSON, default=list)
    status: Mapped[str] = mapped_column(String(30), default="processing", index=True)
    version_number: Mapped[int] = mapped_column(Integer, default=1)
    supersedes_document_id: Mapped[str | None] = mapped_column(String(36), nullable=True, index=True)
    version_change_summary: Mapped[str | None] = mapped_column(Text, nullable=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))
    published_at: Mapped[datetime | None] = mapped_column(DateTime(timezone=True), nullable=True)
    withdrawn_at: Mapped[datetime | None] = mapped_column(DateTime(timezone=True), nullable=True)


class KnowledgeChunk(Base):
    """A local vector-search unit. Full text is never exposed as a download."""

    __tablename__ = "knowledge_chunk"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(ForeignKey("research_project.id"), default=DEFAULT_PROJECT_ID, index=True)
    document_id: Mapped[str] = mapped_column(ForeignKey("knowledge_document.id", ondelete="CASCADE"), index=True)
    scope: Mapped[str] = mapped_column(String(20), index=True)
    owner_id: Mapped[str] = mapped_column(String(120), index=True)
    folder_id: Mapped[str | None] = mapped_column(String(36), nullable=True, index=True)
    document_status: Mapped[str] = mapped_column(String(30), default="processing", index=True)
    ordinal: Mapped[int] = mapped_column(Integer)
    content: Mapped[str] = mapped_column(Text)
    source_locator: Mapped[str] = mapped_column(String(200))
    embedding: Mapped[list[float] | None] = mapped_column(Vector(EMBEDDING_DIMENSION), nullable=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


engine = create_engine(DATABASE_URL, pool_pre_ping=True)
migration_engine = create_engine(MIGRATION_DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, expire_on_commit=False)
MigrationSessionLocal = sessionmaker(bind=migration_engine, autoflush=False, autocommit=False, expire_on_commit=False)


def get_session() -> Generator[Session, None, None]:
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


def _business_role(user: CurrentUser) -> str:
    for role in ("field_admin", "data_processor", "researcher"):
        if role in user.roles:
            return role
    raise HTTPException(403, "当前账号未配置海南南繁平台业务角色。")


def sync_platform_account(session: Session, user: CurrentUser) -> PlatformAccount:
    """Keep the application directory aligned with the verified Keycloak identity."""
    role = _business_role(user)
    account = session.get(PlatformAccount, user.username)
    if not account:
        account = PlatformAccount(
            username=user.username,
            keycloak_subject=user.id,
            display_name=user.display_name or user.username,
            business_role=role,
            institution_id=INSTITUTION_ID,
            active=True,
        )
        session.add(account)
    else:
        account.keycloak_subject = user.id
        account.display_name = user.display_name or user.username
        account.business_role = role
        # Institution assignment is provisioned by the platform directory and
        # must not be overwritten on every login. New accounts still enter the
        # default Hainan NanFan institution automatically.
    account.last_login_at = datetime.now(timezone.utc)
    session.flush()
    if not account.active:
        raise HTTPException(403, "当前业务账号已停用，请联系字段管理员。")
    return account


def accessible_projects(session: Session, user: CurrentUser) -> list[ResearchProject]:
    account = sync_platform_account(session, user)
    statement = select(ResearchProject).where(
        ResearchProject.institution_id == account.institution_id,
        ResearchProject.status == "active",
    )
    if "researcher" in user.roles and not {"field_admin", "data_processor"}.intersection(user.roles):
        statement = statement.join(ProjectMember, ProjectMember.project_id == ResearchProject.id).where(
            ProjectMember.username == user.username
        )
    return list(session.scalars(statement.order_by(ResearchProject.created_at, ResearchProject.project_name)).all())


def resolve_project_access(session: Session, user: CurrentUser, requested_project_id: str | None = None) -> ResearchProject:
    projects = accessible_projects(session, user)
    if not projects:
        raise HTTPException(403, "当前账号尚未加入任何课题，请联系字段管理员分配课题。")
    project_id = (requested_project_id or "").strip()
    if project_id:
        project = next((item for item in projects if item.id == project_id), None)
        if not project:
            raise HTTPException(403, "当前账号无权访问所选课题。")
        return project
    return next((item for item in projects if item.id == DEFAULT_PROJECT_ID), projects[0])


def record_permission_audit(
    session: Session,
    user: CurrentUser,
    action: str,
    target_type: str,
    target_id: str,
    *,
    project_id: str | None = None,
    before: dict[str, Any] | None = None,
    after: dict[str, Any] | None = None,
) -> None:
    account = session.get(PlatformAccount, user.username)
    session.add(PermissionAudit(
        actor_id=user.id,
        actor_name=audit_actor(user),
        action=action,
        target_type=target_type,
        target_id=target_id,
        institution_id=account.institution_id if account else INSTITUTION_ID,
        project_id=project_id,
        before_state=before or {},
        after_state=after or {},
    ))


def serialize_project(project: ResearchProject, member_count: int | None = None) -> dict[str, Any]:
    result = {
        "id": project.id,
        "project_code": project.project_code,
        "project_name": project.project_name,
        "description": project.description or "",
        "institution_id": project.institution_id,
        "institution_name": INSTITUTION_NAME if project.institution_id == INSTITUTION_ID else project.institution_id,
        "status": project.status,
        "created_by": project.created_by,
        "created_at": project.created_at.isoformat(),
        "updated_at": project.updated_at.isoformat(),
    }
    if member_count is not None:
        result["member_count"] = member_count
    return result


@event.listens_for(Session, "after_begin")
def apply_request_rls_context(session: Session, _transaction: Any, connection: Any) -> None:
    """Restore request-scoped RLS variables whenever a pooled connection begins work.

    PostgreSQL settings belong to a physical connection, while SQLAlchemy can
    release and later reacquire a connection after ``commit`` or ``rollback``.
    Keeping the verified identity on the SQLAlchemy Session and applying it as
    a transaction-local value prevents both false RLS rejections and identity
    leakage from one pooled connection to another user.
    """
    owner_id = session.info.get("research_owner_id")
    if owner_id:
        connection.execute(
            text("SELECT set_config('app.research_user_id', :owner_id, true)"),
            {"owner_id": owner_id},
        )
    if "knowledge_is_admin" in session.info:
        connection.execute(
            text("SELECT set_config('app.knowledge_is_admin', :is_admin, true)"),
            {"is_admin": session.info["knowledge_is_admin"]},
        )
    project_id = session.info.get("active_project_id")
    if project_id:
        connection.execute(
            text("SELECT set_config('app.project_id', :project_id, true)"),
            {"project_id": project_id},
        )


def _set_research_owner(session: Session, owner_id: str) -> None:
    """Scope PostgreSQL RLS to one verified Keycloak subject for this request."""
    session.info["research_owner_id"] = owner_id
    session.execute(
        text("SELECT set_config('app.research_user_id', :owner_id, true)"),
        {"owner_id": owner_id},
    )


def _set_active_project(session: Session, project_id: str) -> None:
    session.info["active_project_id"] = project_id
    session.execute(
        text("SELECT set_config('app.project_id', :project_id, true)"),
        {"project_id": project_id},
    )


def get_research_session(
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_researcher),
) -> Generator[Session, None, None]:
    session = SessionLocal()
    try:
        _set_research_owner(session, user.id)
        session.info["research_username"] = user.username
        session.info["research_roles"] = tuple(user.roles)
        project = resolve_project_access(session, user, x_project_id)
        _set_active_project(session, project.id)
        yield session
    finally:
        session.close()


def _set_knowledge_context(session: Session, user: CurrentUser) -> None:
    """Set RLS variables for both the caller identity and the public-KB admin gate."""
    _set_research_owner(session, user.id)
    session.info["research_username"] = user.username
    session.info["research_roles"] = tuple(user.roles)
    session.info["knowledge_is_admin"] = "true" if "field_admin" in user.roles else "false"
    session.execute(
        text("SELECT set_config('app.knowledge_is_admin', :is_admin, true)"),
        {"is_admin": session.info["knowledge_is_admin"]},
    )


def get_knowledge_session(
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_knowledge_user),
) -> Generator[Session, None, None]:
    session = SessionLocal()
    try:
        _set_knowledge_context(session, user)
        project = resolve_project_access(session, user, x_project_id)
        _set_active_project(session, project.id)
        yield session
    finally:
        session.close()


def active_project_id(session: Session) -> str:
    return str(session.info.get("active_project_id") or DEFAULT_PROJECT_ID)


def get_business_project_session(
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_business_user),
) -> Generator[Session, None, None]:
    session = SessionLocal()
    try:
        project = resolve_project_access(session, user, x_project_id)
        _set_active_project(session, project.id)
        yield session
    finally:
        session.close()


PROJECT_SCOPED_MODELS = (
    Variety,
    SourceReview,
    PhenotypeObservation,
    RootPhenotypeObservation,
    ResearchSession,
    ResearchMessage,
    ResearchAttachment,
    ResearchAudit,
    ResearchResult,
    KnowledgeFolder,
    KnowledgeDocument,
    KnowledgeChunk,
    FieldChangeRequest,
)


@event.listens_for(Session, "before_flush")
def assign_active_project_to_new_rows(session: Session, _flush_context: Any, _instances: Any) -> None:
    project_id = session.info.get("active_project_id")
    if not project_id:
        return
    for item in session.new:
        if isinstance(item, PROJECT_SCOPED_MODELS) and not getattr(item, "project_id", None):
            item.project_id = str(project_id)


@event.listens_for(Session, "do_orm_execute")
def restrict_orm_selects_to_active_project(execute_state: Any) -> None:
    if not execute_state.is_select:
        return
    project_id = execute_state.session.info.get("active_project_id")
    if not project_id:
        return
    statement = execute_state.statement
    for model in PROJECT_SCOPED_MODELS:
        statement = statement.options(with_loader_criteria(
            model,
            lambda cls: cls.project_id == project_id,
            include_aliases=True,
        ))
    execute_state.statement = statement


TRAITS: dict[str, dict[str, Any]] = {
    "growth_duration": {"name": "全生育期", "category": "生育期", "unit": "天", "aliases": ["全生育期", "生育期"], "patterns": [r"全生育期(?:为|平均)?(?P<value>\d+(?:\.\d+)?)天"]},
    "plant_height": {"name": "株高", "category": "农艺性状", "unit": "cm", "aliases": ["株高"], "patterns": [r"株高(?P<value>\d+(?:\.\d+)?)(?:厘米|cm)"]},
    "panicle_length": {"name": "穗长", "category": "农艺性状", "unit": "cm", "aliases": ["穗长"], "patterns": [r"穗长(?P<value>\d+(?:\.\d+)?)(?:厘米|cm)"]},
    "effective_panicles_per_mu": {"name": "每亩有效穗", "category": "产量构成", "unit": "万穗/亩", "aliases": ["亩有效穗", "每亩有效穗", "有效穗数"], "patterns": [r"(?:每亩|亩)有效穗(?:数)?(?P<value>\d+(?:\.\d+)?)万"]},
    "total_grains_per_panicle": {"name": "每穗总粒数", "category": "产量构成", "unit": "粒/穗", "aliases": ["每穗总粒数"], "patterns": [r"每穗总粒数(?P<value>\d+(?:\.\d+)?)粒"]},
    "grains_per_panicle": {"name": "每穗粒数", "category": "产量构成", "unit": "粒/穗", "aliases": ["每穗粒数"], "patterns": [r"每穗粒数(?P<value>\d+(?:\.\d+)?)粒"]},
    "filled_grains_per_panicle": {"name": "每穗实粒数", "category": "产量构成", "unit": "粒/穗", "aliases": ["每穗实粒数", "实粒数"], "patterns": [r"(?:每穗)?实粒数(?P<value>\d+(?:\.\d+)?)粒"]},
    "seed_setting_rate": {"name": "结实率", "category": "产量构成", "unit": "%", "aliases": ["结实率"], "patterns": [r"结实率(?P<value>\d+(?:\.\d+)?)%"]},
    "thousand_grain_weight": {"name": "千粒重", "category": "产量构成", "unit": "g", "aliases": ["千粒重", "千粒质量"], "patterns": [r"千粒(?:重|质量)(?P<value>\d+(?:\.\d+)?)(?:克|g)"]},
    "yield_per_mu": {"name": "亩产", "category": "产量表现", "unit": "kg/亩", "aliases": ["亩产", "平均亩产"], "patterns": [r"(?:平均)?亩产(?P<value>\d+(?:\.\d+)?)(?:公斤|千克|kg)"]},
    "brown_rice_rate": {"name": "出糙率/糙米率", "category": "加工品质", "unit": "%", "aliases": ["出糙率", "糙米率"], "patterns": [r"(?:出糙率|糙米率)(?P<value>\d+(?:\.\d+)?)%"]},
    "milled_rice_rate": {"name": "精米率", "category": "加工品质", "unit": "%", "aliases": ["精米率"], "patterns": [r"(?<!整)精米率(?P<value>\d+(?:\.\d+)?)%"]},
    "head_rice_rate": {"name": "整精米率", "category": "加工品质", "unit": "%", "aliases": ["整精米率"], "patterns": [r"整精米率(?P<value>\d+(?:\.\d+)?)%"]},
    "chalky_grain_rate": {"name": "垩白粒率", "category": "外观品质", "unit": "%", "aliases": ["垩白粒率"], "patterns": [r"垩白粒率(?P<value>\d+(?:\.\d+)?)%"]},
    "chalkiness_degree": {"name": "垩白度", "category": "外观品质", "unit": "%", "aliases": ["垩白度"], "patterns": [r"垩白度(?P<value>\d+(?:\.\d+)?)%"]},
    "amylose_content": {"name": "直链淀粉含量", "category": "蒸煮品质", "unit": "%", "aliases": ["直链淀粉含量"], "patterns": [r"直链淀粉含量(?P<value>\d+(?:\.\d+)?)%"]},
    "gel_consistency": {"name": "胶稠度", "category": "蒸煮品质", "unit": "mm", "aliases": ["胶稠度"], "patterns": [r"胶稠度(?P<value>\d+(?:\.\d+)?)(?:毫米|mm)"]},
    "grain_length": {"name": "粒长", "category": "籽粒形态", "unit": "mm", "aliases": ["粒长"], "patterns": [r"粒长(?P<value>\d+(?:\.\d+)?)(?:毫米|mm)"]},
    "grain_length_width_ratio": {"name": "长宽比", "category": "籽粒形态", "unit": "", "aliases": ["长宽比"], "patterns": [r"长宽比(?P<value>\d+(?:\.\d+)?)"]},
    "seedling_blast_score": {"name": "苗瘟等级", "category": "抗病性", "unit": "级", "aliases": ["苗瘟"], "patterns": [r"苗瘟(?P<value>\d+(?:\.\d+)?)级"]},
    "leaf_blast_score": {"name": "叶瘟等级", "category": "抗病性", "unit": "级", "aliases": ["叶瘟"], "patterns": [r"叶瘟(?P<value>\d+(?:\.\d+)?)级"]},
    "panicle_blast_score": {"name": "穗瘟等级", "category": "抗病性", "unit": "级", "aliases": ["穗瘟"], "patterns": [r"穗瘟(?P<value>\d+(?:\.\d+)?)级"]},
}

ROOT_TRAITS: dict[str, dict[str, Any]] = {
    "root_length": {"name": "根长", "category": "根系形态", "unit": "cm", "aliases": ["根长", "单根长度", "主根长"]},
    "total_root_length": {"name": "总根长", "category": "根系形态", "unit": "cm", "aliases": ["总根长", "总根系长度", "根系总长"]},
    "root_count": {"name": "根数", "category": "根系形态", "unit": "条", "aliases": ["根数", "总根数"]},
    "root_surface_area": {"name": "根表面积", "category": "根系形态", "unit": "cm²", "aliases": ["根表面积", "根系表面积"]},
    "root_volume": {"name": "根体积", "category": "根系形态", "unit": "cm³", "aliases": ["根体积", "根系体积"]},
    "average_root_diameter": {"name": "平均根径", "category": "根系形态", "unit": "mm", "aliases": ["平均根径", "根平均直径", "根径"]},
    "root_tip_count": {"name": "根尖数", "category": "根系形态", "unit": "个", "aliases": ["根尖数", "根尖数量"]},
    "root_dry_weight": {"name": "根干重", "category": "根系生物量", "unit": "g", "aliases": ["根干重", "根系干重"]},
    "root_angle": {"name": "根系角度", "category": "根系构型", "unit": "°", "aliases": ["根系角度", "根角"]},
    "root_shoot_ratio": {"name": "根冠比", "category": "根系生物量", "unit": "", "aliases": ["根冠比", "根冠质量比"]},
}


def national_template_fields() -> list[dict[str, Any]]:
    fields = [{"code": "variety_name", "name": "品种名称", "category": "基础信息", "unit": "", "aliases": ["品种名称", "原始材料名称", "材料名"], "required": True, "kind": "basic"}]
    for code, trait in TRAITS.items():
        fields.append({"code": code, "name": trait["name"], "category": trait["category"], "unit": trait["unit"], "aliases": trait["aliases"], "required": False, "kind": "trait"})
    return fields


def root_template_fields() -> list[dict[str, Any]]:
    fields = [{"code": "variety_name", "name": "材料/品种名称", "category": "基础信息", "unit": "", "aliases": ["品种名称", "材料名称", "样品名称", "材料编号"], "required": True, "kind": "basic"}]
    for code, trait in ROOT_TRAITS.items():
        fields.append({"code": code, "name": trait["name"], "category": trait["category"], "unit": trait["unit"], "aliases": trait["aliases"], "required": False, "kind": "trait"})
    return fields

# First-phase spreadsheet mappings. Institute-specific mappings can later be
# persisted as versioned templates rather than changing parser code.
EXCEL_HEADER_TRAIT_CODES = {
    "植高记录": "plant_height",
    "植株高度": "plant_height",
    "株高记录": "plant_height",
    "1000粒质量": "thousand_grain_weight",
    "千粒质量": "thousand_grain_weight",
    "每穗总粒": "total_grains_per_panicle",
    "总粒数": "total_grains_per_panicle",
    "饱粒数穗": "filled_grains_per_panicle",
    "饱粒数": "filled_grains_per_panicle",
    "结实百分比": "seed_setting_rate",
    "结实百分率": "seed_setting_rate",
    "亩收获量": "yield_per_mu",
    "有效穗": "effective_panicles_per_mu",
    "叶瘟病级": "leaf_blast_score",
    "叶瘟等级": "leaf_blast_score",
}

EXCEL_NAME_HEADER_TOKENS = ("品种", "材料名称", "材料名", "原始材料", "样品名称", "样本名称", "材料编号")
EXCEL_ALIAS_HEADER_TOKENS = ("别名", "代号")


SEED_RULES = [
    ("R001", "品种名称不能为空", "publish", "block", {"field": "variety_name", "required": True}),
    ("R003", "发布数据必须有来源", "publish", "block", {"field": "source_review_id", "required": True}),
    ("R004", "发布表型必须保留原始值", "publish", "block", {"field": "original_value", "required": True}),
    ("R101", "品种标题括号别名识别", "name", "info", {"pattern": "名称（别名）"}),
    ("R201", "百分比必须在0至100之间", "quality", "block", {"unit": "%", "min": 0, "max": 100}),
    ("R204", "每穗实粒数不得大于总粒数", "quality", "block", {"relation": "filled_grains_per_panicle <= total_grains_per_panicle"}),
    ("R211", "株高异常预警", "quality", "warning", {"trait_code": "plant_height", "min": 30, "max": 250}),
    ("R212", "千粒重异常预警", "quality", "warning", {"trait_code": "thousand_grain_weight", "min": 5, "max": 60}),
    ("R213", "全生育期异常预警", "quality", "warning", {"trait_code": "growth_duration", "min": 30, "max": 250}),
    ("R301", "稻瘟病等级不自动推导抗性分类", "semantic", "info", {"requires_scale": True}),
    ("R401", "已发布规则仅允许新增版本", "governance", "block", {"immutable": True}),
]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_name(value: str) -> str:
    return re.sub(r"[\s()（）\-_/]", "", value or "").lower()


def append_history(source: SourceReview | None, actor: str, action: str, details: dict[str, Any]) -> None:
    if not source:
        return
    history = list(source.review_history or [])
    history.append({"at": now_iso(), "actor": actor, "action": action, "details": details})
    source.review_history = history


def trait_by_alias(value: str, trait_catalog: dict[str, dict[str, Any]] = TRAITS, header_mappings: dict[str, str] | None = None) -> dict[str, Any] | None:
    cleaned = normalize_name(value)
    mapped_code = (header_mappings or EXCEL_HEADER_TRAIT_CODES).get(cleaned)
    if mapped_code:
        return {"code": mapped_code, **trait_catalog[mapped_code]}
    for code, trait in trait_catalog.items():
        if cleaned == normalize_name(trait["name"]) or any(cleaned == normalize_name(alias) for alias in trait["aliases"]):
            return {"code": code, **trait}
    for code, trait in trait_catalog.items():
        names = [trait["name"], *trait["aliases"]]
        if any(len(normalize_name(name)) >= 2 and normalize_name(name) in cleaned for name in names):
            return {"code": code, **trait}
    return None


def normalize_spreadsheet_value(trait_code: str, raw_value: Any, trait_catalog: dict[str, dict[str, Any]] = TRAITS) -> dict[str, Any]:
    """Parse common institute spreadsheet unit variants into the platform unit."""
    raw = str(raw_value).strip()
    number_match = re.search(r"-?\d+(?:\.\d+)?", raw.replace(",", ""))
    if not number_match:
        return {"observation_type": "text", "value_text": raw, "value_numeric": None, "unit": trait_catalog[trait_code]["unit"], "requires_confirmation": True, "conversion_suggestion": "未提取到数值，请人工确认。"}

    value = float(number_match.group(0))
    normalized = raw.lower().replace(" ", "").replace("²", "2")
    unit = trait_catalog[trait_code]["unit"]
    suggestion = ""
    requires_confirmation = False

    if trait_code == "plant_height":
        if "mm" in normalized or "毫米" in normalized:
            value /= 10
            suggestion = "单位换算：mm -> cm，数值除以10。"
        elif "cm" in normalized or "厘米" in normalized:
            suggestion = "单位已标准化：cm。"
        elif re.search(r"(?<![a-z])m(?![a-z])", normalized) or ("米" in normalized and "厘米" not in normalized and "毫米" not in normalized):
            value *= 100
            suggestion = "单位换算：m -> cm，数值乘以100。"
        else:
            requires_confirmation = True
            suggestion = "未识别株高单位；根据数值推测可能为 cm，需人工确认。"
    elif trait_code == "thousand_grain_weight":
        if "mg" in normalized or "毫克" in normalized:
            value /= 1000
            suggestion = "单位换算：mg -> g，数值除以1000。"
        elif "kg" in normalized or "千克" in normalized:
            value *= 1000
            suggestion = "单位换算：kg -> g，数值乘以1000。"
        elif "g" in normalized or "克" in normalized:
            suggestion = "单位已标准化：g。"
        else:
            requires_confirmation = True
            suggestion = "未识别千粒重单位；根据数值推测可能为 g，需人工确认。"
    elif trait_code == "seed_setting_rate":
        if "%" in normalized or "百分" in normalized:
            suggestion = "单位已标准化：%。"
        elif 0 <= value <= 1:
            value *= 100
            suggestion = "比例换算：0-1 小数 -> %，数值乘以100。"
        else:
            suggestion = "未带百分号；按百分比数值处理，建议人工核对。"
    elif trait_code == "yield_per_mu":
        if "t/ha" in normalized or "吨/公顷" in normalized:
            value = value * 1000 / 15
            suggestion = "单位换算：t/ha -> kg/亩，数值乘以1000后除以15。"
        elif "kg/hm2" in normalized or "kg/ha" in normalized or "公斤/公顷" in normalized or "千克/公顷" in normalized:
            value /= 15
            suggestion = "单位换算：kg/hm2 -> kg/亩，数值除以15。"
        elif "kg/亩" in normalized or "公斤/亩" in normalized or "千克/亩" in normalized:
            suggestion = "单位已标准化：kg/亩。"
        else:
            requires_confirmation = True
            suggestion = "未识别产量单位；需人工确认后才能作为 kg/亩 入库。"
    elif trait_code == "effective_panicles_per_mu":
        if "万" in raw:
            suggestion = "单位已标准化：万穗/亩。"
        elif "穗/亩" in raw:
            value /= 10000
            suggestion = "单位换算：穗/亩 -> 万穗/亩，数值除以10000。"
        elif value >= 1000:
            value /= 10000
            requires_confirmation = True
            suggestion = "未带单位；按穗/亩推测并除以10000，需人工确认。"
        else:
            requires_confirmation = True
            suggestion = "未识别有效穗单位，需人工确认。"
    elif trait_code in {"seedling_blast_score", "leaf_blast_score", "panicle_blast_score"}:
        suggestion = "提取病害等级；第一版不自动换算为抗性分类。"
    elif trait_code in {"total_grains_per_panicle", "grains_per_panicle", "filled_grains_per_panicle"}:
        suggestion = "已提取粒数；将结合总粒数和实粒数关系进行质量校验。"
    elif trait_code in {"root_length", "total_root_length"}:
        if "mm" in normalized or "毫米" in normalized:
            value /= 10
            suggestion = "单位换算：mm -> cm，数值除以10。"
        elif "cm" in normalized or "厘米" in normalized:
            suggestion = "单位已标准化：cm。"
        elif re.search(r"(?<![a-z])m(?![a-z])", normalized) or ("米" in normalized and "厘米" not in normalized and "毫米" not in normalized):
            value *= 100
            suggestion = "单位换算：m -> cm，数值乘以100。"
        else:
            requires_confirmation = True
            suggestion = "未识别根长单位，需人工确认。"
    elif trait_code == "average_root_diameter":
        if "mm" in normalized or "毫米" in normalized:
            suggestion = "单位已标准化：mm。"
        elif "cm" in normalized or "厘米" in normalized:
            value *= 10
            suggestion = "单位换算：cm -> mm，数值乘以10。"
        else:
            requires_confirmation = True
            suggestion = "未识别根径单位，需人工确认。"
    elif trait_code == "root_dry_weight":
        if "mg" in normalized or "毫克" in normalized:
            value /= 1000
            suggestion = "单位换算：mg -> g，数值除以1000。"
        elif "g" in normalized or "克" in normalized:
            suggestion = "单位已标准化：g。"
        else:
            requires_confirmation = True
            suggestion = "未识别根干重单位，需人工确认。"
    elif trait_code in {"root_count", "root_tip_count", "root_angle", "root_shoot_ratio", "root_surface_area", "root_volume"}:
        suggestion = f"已提取{trait_catalog[trait_code]['name']}；请核对测定方法和单位。"

    return {
        "observation_type": "numeric",
        "value_numeric": round(value, 6),
        "value_text": None,
        "unit": unit,
        "requires_confirmation": requires_confirmation,
        "conversion_suggestion": suggestion,
    }


def active_custom_quality_rules(session: Session) -> list[DataRule]:
    return [
        rule for rule in session.scalars(
            select(DataRule).where(DataRule.rule_type == "quality", DataRule.status == "published")
        ).all()
        if (rule.config or {}).get("trait_code")
    ]


def validate_observation(
    observation: PhenotypeObservation,
    all_observations: list[PhenotypeObservation] | None = None,
    custom_quality_rules: list[DataRule] | None = None,
) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    if not observation.original_value:
        issues.append({"rule": "R004", "severity": "block", "message": "缺少原始值，不能发布。"})
    if observation.observation_type == "numeric" and observation.value_numeric is None:
        issues.append({"rule": "R206", "severity": "block", "message": "数值无法解析，不能作为标准化数值发布。"})
    if observation.unit == "%" and observation.value_numeric is not None and not 0 <= observation.value_numeric <= 100:
        issues.append({"rule": "R201", "severity": "block", "message": "百分比必须在0至100之间。"})
    if observation.value_numeric is not None and observation.trait_code in {"plant_height", "panicle_length", "thousand_grain_weight", "grain_length"} and observation.value_numeric <= 0:
        issues.append({"rule": "R202", "severity": "block", "message": "该数值必须大于0。"})
    warning_ranges = {"plant_height": (30, 250, "R211"), "thousand_grain_weight": (5, 60, "R212"), "growth_duration": (30, 250, "R213"), "grains_per_panicle": (0, 500, "R214"), "total_grains_per_panicle": (0, 500, "R214"), "grain_length": (0, 15, "R215"), "gel_consistency": (0, 150, "R216")}
    if observation.trait_code in warning_ranges and observation.value_numeric is not None:
        low, high, rule = warning_ranges[observation.trait_code]
        if not low <= observation.value_numeric <= high:
            issues.append({"rule": rule, "severity": "warning", "message": f"超出第一版预警范围 {low}–{high}，需人工确认。"})
    if observation.value_numeric is not None:
        for rule in custom_quality_rules or []:
            config = rule.config or {}
            if config.get("trait_code") != observation.trait_code:
                continue
            low = config.get("min")
            high = config.get("max")
            if (low is not None and observation.value_numeric < float(low)) or (high is not None and observation.value_numeric > float(high)):
                bounds = f"{low if low is not None else '-∞'}–{high if high is not None else '∞'}"
                issues.append({"rule": rule.rule_code, "severity": rule.severity, "message": f"{rule.rule_name}：标准值超出允许范围 {bounds}。"})
    if all_observations and observation.trait_code in {"filled_grains_per_panicle", "total_grains_per_panicle", "grains_per_panicle"}:
        values = {item.trait_code: item.value_numeric for item in all_observations}
        filled = values.get("filled_grains_per_panicle")
        total = values.get("total_grains_per_panicle") or values.get("grains_per_panicle")
        if filled is not None and total is not None and filled > total:
            issues.append({"rule": "R204", "severity": "block", "message": "每穗实粒数不能大于每穗总粒数。"})
    return issues


def parse_title(value: str) -> tuple[str, list[str]]:
    match = re.match(r"\s*([^（(\n]{1,100}?)(?:[（(]([^）)\n]+)[）)])?\s*$", value or "")
    if not match:
        return value.strip(), []
    name = match.group(1).strip()
    aliases = [item.strip() for item in (match.group(2) or "").split("、") if item.strip()]
    return name, aliases


def decode_html_bytes(content: bytes) -> str:
    """Decode legacy agricultural websites that often omit an HTTP charset header."""
    probe = content[:2048].decode("ascii", errors="ignore")
    declared = re.search(r"charset\s*=\s*['\"]?([\w-]+)", probe, flags=re.I)
    encodings = [declared.group(1) if declared else "", "utf-8", "gb18030", "gbk", "gb2312"]
    candidates: list[tuple[int, str]] = []
    for encoding in dict.fromkeys(item.lower() for item in encodings if item):
        try:
            decoded = content.decode(encoding, errors="strict")
        except (LookupError, UnicodeDecodeError):
            continue
        chinese_count = len(re.findall(r"[\u4e00-\u9fff]", decoded))
        score = chinese_count * 10 - decoded.count("�") * 1000
        candidates.append((score, decoded))
    if candidates:
        return max(candidates, key=lambda item: item[0])[1]
    return content.decode("utf-8", errors="replace")


def extract_variety_title(soup: BeautifulSoup) -> str:
    title_text = soup.title.get_text(" ", strip=True) if soup.title else ""
    # RiceData uses "品种名>>>RiceData==..." as the page title.
    title_text = re.split(r">>>|RiceData", title_text, maxsplit=1, flags=re.I)[0].strip()
    if title_text and len(title_text) <= 100:
        return title_text
    for node in soup.find_all(["h1", "h2", "h3", "strong", "b"]):
        text = node.get_text(" ", strip=True)
        if 2 <= len(text) <= 100 and "审定" not in text:
            return text.lstrip("·• ")
    return ""


def is_rice_data_digit_glyph(image: Any) -> bool:
    source = image.get("src", "")
    return bool(image.find_parent(["p", "td", "li"]) and (
        "/figures/" in source or ("_files/" in source and source.lower().endswith(".gif"))
    ))


def rice_data_glyph_url(page_url: str, image_source: str) -> str:
    if "/figures/" in image_source:
        return urljoin(page_url, image_source)
    # Chrome saved pages rewrite GIF paths to ./<page>_files/<name>.gif.
    # The original page URL in the saved-page comment lets us recover only
    # the matching single-character image from the public page's image folder.
    return urljoin(page_url, f"/images/figures/{image_source.rsplit('/', 1)[-1]}")


def extract_saved_page_url(html: str) -> str | None:
    match = re.search(r"<!--\s*saved from url=\(\d+\)(https?://[^\s>]+)\s*-->", html, flags=re.I)
    return match.group(1) if match else None


async def resolve_rice_data_digit_glyphs(soup: BeautifulSoup, page_url: str, client: httpx.AsyncClient) -> int:
    """Replace known RiceData inline digit GIFs before extracting the visible text."""
    glyph_images = [
        image for image in soup.find_all("img") if is_rice_data_digit_glyph(image)
    ]
    image_bytes: dict[str, bytes] = {}
    for source_url in {rice_data_glyph_url(page_url, image.get("src", "")) for image in glyph_images}:
        try:
            response = await client.get(source_url)
            response.raise_for_status()
            image_bytes[source_url] = response.content
        except httpx.HTTPError:
            continue

    resolved = 0
    for image in glyph_images:
        data = image_bytes.get(rice_data_glyph_url(page_url, image.get("src", "")))
        digit = RICE_DATA_GLYPH_HASH_TO_DIGIT.get(hashlib.sha256(data).hexdigest()) if data else None
        if digit is not None:
            image.replace_with(digit)
            resolved += 1
    return resolved


def extract_html_text_with_markers(soup: BeautifulSoup) -> tuple[str, int]:
    """Keep anti-scraping image glyphs visible instead of guessing the hidden digit."""
    masked_digits = 0
    for image in list(soup.find_all("img")):
        if is_rice_data_digit_glyph(image):
            image.replace_with("【图形字符】")
            masked_digits += 1
    return soup.get_text(" ", strip=True), masked_digits


def extract_text_observations(text: str) -> list[dict[str, Any]]:
    # Legacy pages often place each label, number and unit in separate HTML cells.
    # Remove layout whitespace only for matching; retain the full raw text in source_review.
    match_text = re.sub(r"\s+", "", text)
    results: list[dict[str, Any]] = []
    used: set[tuple[str, str]] = set()
    for code, trait in TRAITS.items():
        for pattern in trait["patterns"]:
            for match in re.finditer(pattern, match_text, flags=re.I):
                raw = match.group(0)
                marker = (code, raw)
                if marker in used:
                    continue
                used.add(marker)
                results.append({
                    "trait_code": code,
                    "trait_name": trait["name"],
                    "trait_category": trait["category"],
                    "observation_type": "numeric",
                    "value_numeric": float(match.group("value")),
                    "unit": trait["unit"],
                    "original_value": raw,
                    "original_field": trait["name"],
                    "source_locator": f"字符位置 {match.start()}",
                })
    plant_type = re.search(r"株型([^。；]{1,60})", match_text)
    if plant_type:
        results.append({"trait_code": "plant_type_description", "trait_name": "株型描述", "trait_category": "农艺性状", "observation_type": "text", "value_text": plant_type.group(1).strip("，： "), "unit": "", "original_value": plant_type.group(0), "original_field": "株型", "source_locator": f"字符位置 {plant_type.start()}"})
    blast = re.search(r"(?:综合评价为|鉴定为)([^。；]{0,30}(?:稻瘟病|苗瘟|叶瘟|穗瘟)[^。；]{0,30})", match_text)
    if blast:
        results.append({"trait_code": "rice_blast_assessment_raw", "trait_name": "稻瘟病评价原文", "trait_category": "抗病性", "observation_type": "text", "value_text": blast.group(1).strip(), "unit": "", "original_value": blast.group(0), "original_field": "稻瘟病抗性", "source_locator": f"字符位置 {blast.start()}"})
    return results


VARIETY_BASIC_FIELDS = (
    "variety_type", "female_parent", "male_parent", "breeding_unit",
    "approval_number", "approval_year", "suitable_region",
)

APPROVAL_SECTION_START = re.compile(
    r"(?:19|20)\d{2}年[^。；\n]{0,30}?(?:审定|认定|登记)[，,、:：]*编号[:：]?",
    flags=re.I,
)


def keep_first_approval_section(text: str) -> tuple[str, dict[str, Any]]:
    """First-phase rule: parse only the first approval section from a multi-approval source."""
    matches = list(APPROVAL_SECTION_START.finditer(text))
    if len(matches) < 2:
        return text, {"approval_count": len(matches), "ignored_approval_count": 0}
    first_header = re.split(r"[。；\n]", text[matches[0].start():matches[0].start() + 120], maxsplit=1)[0].strip()
    return text[:matches[1].start()].rstrip(), {
        "approval_count": len(matches),
        "ignored_approval_count": len(matches) - 1,
        "first_approval_header": first_header,
    }


def extract_variety_basic_info(text: str) -> dict[str, str | None]:
    """Extract only explicit basic facts; never infer a breeding conclusion."""
    compact = re.sub(r"\s+", "", text)

    def capture(pattern: str) -> str | None:
        match = re.search(pattern, compact, flags=re.I)
        return match.group("value").strip("：:；;，,。 ") if match else None

    parents = capture(r"亲本来源[:：](?P<value>.*?)(?=选育单位[:：]|品种类型[:：]|适(?:种|宜)地区[:：]|温馨提示|$)")
    female = None
    male = None
    if parents:
        female_match = re.search(r"(?P<value>.+?)[（(]♀[）)]", parents)
        male_match = re.search(r"[）)](?P<value>[^（()]{1,120})[（(]♂[）)]", parents)
        female = female_match.group("value").strip() if female_match else None
        male = male_match.group("value").strip() if male_match else None

    approval_match = re.search(
        r"(?P<year>(?:19|20)\d{2})年[^。；]{0,30}?(?:审定|认定|登记)[，,、:：]*编号[:：]?(?P<number>[^。；]{1,100}?)(?=品种来源[:：]|特征特性[:：]|产量表现[:：]|适(?:种|宜)地区[:：]|$)",
        compact,
    )
    return {
        "variety_type": capture(r"品种类型[:：](?P<value>.*?)(?=适(?:种|宜)地区[:：]|亲本来源[:：]|左边有品种系谱树|温馨提示|$)"),
        "female_parent": female,
        "male_parent": male,
        "breeding_unit": capture(r"选育单位[:：](?P<value>.*?)(?=品种类型[:：]|适(?:种|宜)地区[:：]|亲本来源[:：]|温馨提示|$)"),
        "approval_number": approval_match.group("number") if approval_match else None,
        "approval_year": approval_match.group("year") if approval_match else None,
        "suitable_region": capture(r"适(?:种|宜)地区[:：](?P<value>[^。；]{1,500})"),
    }


def table_to_candidates(headers: list[str], rows: list[dict[str, Any]], trait_catalog: dict[str, dict[str, Any]] = TRAITS, header_mappings: dict[str, str] | None = None) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    name_key = next((header for header in headers if any(token in normalize_name(header) for token in EXCEL_NAME_HEADER_TOKENS)), None)
    alias_key = next((header for header in headers if any(token in normalize_name(header) for token in EXCEL_ALIAS_HEADER_TOKENS)), None)
    mapped_headers = {header: trait_by_alias(str(header), trait_catalog, header_mappings) for header in headers}
    unmapped_headers = [header for header in headers if header != name_key and header != alias_key and mapped_headers[header] is None]
    for index, row in enumerate(rows):
        title = str(row.get(name_key, "") if name_key else "")
        name, aliases = parse_title(title)
        if alias_key and row.get(alias_key) not in (None, ""):
            aliases = list(dict.fromkeys([*aliases, *[item.strip() for item in re.split(r"[、,，;/]", str(row[alias_key])) if item.strip()]]))
        observations = []
        for header, value in row.items():
            if value in (None, ""):
                continue
            trait = mapped_headers.get(header)
            if not trait:
                continue
            raw = str(value)
            normalized_value = normalize_spreadsheet_value(trait["code"], raw, trait_catalog)
            observations.append({
                "trait_code": trait["code"], "trait_name": trait["name"], "trait_category": trait["category"],
                "observation_type": normalized_value["observation_type"], "value_numeric": normalized_value["value_numeric"],
                "value_text": normalized_value["value_text"], "unit": normalized_value["unit"], "original_value": raw,
                "original_field": str(header), "source_locator": f"第 {index + 2} 行 / {header}",
                "conversion_suggestion": f"字段映射：{header} -> {trait['name']}。{normalized_value['conversion_suggestion']}",
                "requires_confirmation": normalized_value["requires_confirmation"],
            })
        warnings = [f"已识别 {len(observations)} 个标准字段；{sum(item['requires_confirmation'] for item in observations)} 个字段需要人工确认单位或换算。"]
        if unmapped_headers:
            warnings.append(f"未映射字段保留在原始来源中：{'、'.join(unmapped_headers)}。")
        candidates.append({"variety_name": name, "aliases": aliases, "raw_title": title, "observations": observations, "source_locator": f"第 {index + 2} 行", "parser_warnings": warnings, "unmapped_fields": [{"field": header, "sample_value": str(row.get(header, ""))} for header in unmapped_headers if row.get(header) not in (None, "")]})
    return candidates


async def parse_uploaded_content(filename: str, content: bytes, template: DataTemplate | None = None, template_version: TemplateVersion | None = None) -> tuple[str, str, list[dict[str, Any]]]:
    suffix = Path(filename).suffix.lower()
    is_root_template = template and template.template_code == "rice_root_phenotype"
    trait_catalog, header_mappings = template_parsing_catalog(template, template_version) if template and template_version else (ROOT_TRAITS if is_root_template else TRAITS, {} if is_root_template else EXCEL_HEADER_TRAIT_CODES)
    if suffix == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        rows = list(csv.DictReader(io.StringIO(text)))
        return "csv", text, table_to_candidates(list(rows[0].keys()) if rows else [], rows, trait_catalog, header_mappings)
    if suffix in {".xlsx", ".xls"}:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return "excel", "", []
        headers = [str(item or "") for item in values[0]]
        rows = [dict(zip(headers, row)) for row in values[1:] if any(cell is not None for cell in row)]
        raw = "\n".join(" | ".join(str(cell or "") for cell in row) for row in values)
        return "excel", raw, table_to_candidates(headers, rows, trait_catalog, header_mappings)
    if suffix == ".pdf":
        document = fitz.open(stream=content, filetype="pdf")
        text = "\n\n".join(f"[第 {index + 1} 页]\n{page.get_text()}" for index, page in enumerate(document))
        if not text.strip():
            raise HTTPException(422, "未提取到可复制文字。第一版暂不支持扫描图片型 PDF。")
        if is_root_template:
            raise HTTPException(422, "根系表型模板第一版仅支持 Excel 或 CSV；请保留根系 PDF 为原始来源后再补充结构化表格。")
        return "pdf", text, [{"variety_name": "", "aliases": [], "raw_title": "", "observations": extract_text_observations(text), "source_locator": "PDF 正文"}]
    if suffix in {".html", ".htm"}:
        decoded = decode_html_bytes(content)
        soup = BeautifulSoup(decoded, "html.parser")
        saved_page_url = extract_saved_page_url(decoded)
        resolved_digits = 0
        if saved_page_url:
            async with httpx.AsyncClient(timeout=15, follow_redirects=True, headers={"User-Agent": "RiceDataGovernanceDemo/1.0 (saved-page import)"}) as glyph_client:
                resolved_digits = await resolve_rice_data_digit_glyphs(soup, saved_page_url, glyph_client)
        raw_title = extract_variety_title(soup)
        text, masked_digits = extract_html_text_with_markers(soup)
        text, approval_context = keep_first_approval_section(text)
        name, aliases = parse_title(raw_title)
        if is_root_template:
            raise HTTPException(422, "根系表型模板第一版仅支持 Excel 或 CSV。")
        candidate = {"variety_name": name, "aliases": aliases, "raw_title": raw_title, "observations": extract_text_observations(text), "source_locator": "网页正文", **extract_variety_basic_info(text)}
        parser_warnings: list[str] = []
        if resolved_digits:
            parser_warnings.append(f"已根据保存网页中的原始链接还原 {resolved_digits} 个图形数字。")
        if masked_digits:
            parser_warnings.append(f"网页正文仍有 {masked_digits} 个未识别图形字符，相关字段未自动标准化，请以原始网页人工核对。")
        if approval_context["ignored_approval_count"]:
            parser_warnings.append(f"检测到 {approval_context['approval_count']} 条审定记录。第一版仅处理第一条，后续 {approval_context['ignored_approval_count']} 条未解析、未写入数据库。")
        if parser_warnings:
            candidate["parser_warnings"] = parser_warnings
        candidate["_saved_page_url"] = saved_page_url
        candidate["_parsing_status"] = "partial" if masked_digits else "parsed"
        candidate["_quality_status"] = "requires_manual_check" if masked_digits else "pending"
        candidate["_approval_context"] = approval_context
        return "html", text, [candidate]
    raise HTTPException(422, "仅支持 HTML、PDF、Excel 和 CSV 文件。")


def serialize_source(source: SourceReview) -> dict[str, Any]:
    return {"id": source.id, "project_id": source.project_id, "source_type": source.source_type, "source_name": source.source_name, "source_url": source.source_url, "raw_text": source.raw_text or "", "page_or_locator": source.page_or_locator, "parsing_status": source.parsing_status, "quality_status": source.quality_status, "template_version_id": source.template_version_id, "review_history": source.review_history or [], "created_at": source.created_at.isoformat()}


def serialize_observation(observation: PhenotypeObservation) -> dict[str, Any]:
    return {"id": observation.id, "project_id": observation.project_id, "variety_id": observation.variety_id, "source_review_id": observation.source_review_id, "trait_code": observation.trait_code, "trait_name": observation.trait_name, "trait_category": observation.trait_category, "observation_type": observation.observation_type, "value_numeric": observation.value_numeric, "value_text": observation.value_text, "unit": observation.unit, "original_value": observation.original_value, "original_field": observation.original_field, "source_locator": observation.source_locator, "trial_year": observation.trial_year, "trial_location": observation.trial_location, "evaluation_method": observation.evaluation_method, "rule_version": observation.rule_version, "quality_status": observation.quality_status, "publish_status": observation.publish_status, "review_comment": observation.review_comment, "created_at": observation.created_at.isoformat()}


def serialize_variety(variety: Variety, include_observations: bool = False) -> dict[str, Any]:
    data = {"id": variety.id, "project_id": variety.project_id, "variety_name": variety.variety_name, "alias_names": variety.alias_names or [], "raw_variety_title": variety.raw_variety_title, "variety_type": variety.variety_type, "female_parent": variety.female_parent, "male_parent": variety.male_parent, "breeding_unit": variety.breeding_unit, "approval_number": variety.approval_number, "approval_year": variety.approval_year, "suitable_region": variety.suitable_region, "data_status": variety.data_status}
    if include_observations:
        data["observations"] = [serialize_observation(item) for item in variety.observations]
    return data


def seed_data(session: Session) -> None:
    if session.scalar(select(func.count(Variety.id))) or session.scalar(select(func.count(DataRule.id))):
        return
    for code, name, rule_type, severity, config in SEED_RULES:
        session.add(DataRule(rule_code=code, rule_name=name, rule_type=rule_type, version="v1.0", severity=severity, config=config, status="published", change_reason="第一版预置规则"))
    source = SourceReview(source_type="webpage", source_name="公开样例：水稻品种详情页", source_url="https://www.ricedata.cn/", raw_text="公开品种资料样例。用于演示原始来源、字段识别与审核流程。", page_or_locator="网页详情页", parsing_status="parsed", quality_status="reviewed")
    append_history(source, "数据处理员-张三", "创建公开样例来源", {"note": "仅用于本地演示"})
    session.add(source)
    session.flush()
    samples = [
        ("田两优9号", ["田两优佳99"], "籼型两系杂交水稻", "田丰S-2", "桂99", "江西省赣州市农业科学研究所", "赣审稻20040026", "2004", "江西省稻瘟病轻发区", {"growth_duration": 116.1, "plant_height": 103.3, "effective_panicles_per_mu": 19.5, "total_grains_per_panicle": 114.7, "filled_grains_per_panicle": 88.6, "seed_setting_rate": 77.2, "thousand_grain_weight": 24.3, "yield_per_mu": 427.76, "brown_rice_rate": 81.4, "head_rice_rate": 60.3, "chalky_grain_rate": 12, "chalkiness_degree": 2.4, "amylose_content": 19.64, "gel_consistency": 50, "grain_length": 7.2, "grain_length_width_ratio": 3.6, "leaf_blast_score": 3, "panicle_blast_score": 5}),
        ("江四优992", [], "籼型三系杂交水稻", "G4A", "R99257", "江西省农业科学院水稻研究所", "赣审稻2004021", "2004", "江西省各地均可种植", {"growth_duration": 111.4, "plant_height": 109.0, "effective_panicles_per_mu": 21.6, "total_grains_per_panicle": 123.8, "filled_grains_per_panicle": 90.6, "seed_setting_rate": 73.2, "thousand_grain_weight": 25.7, "brown_rice_rate": 77.5, "head_rice_rate": 64.9, "chalky_grain_rate": 10.0, "chalkiness_degree": 1.0, "amylose_content": 22.4, "gel_consistency": 60, "grain_length": 7.0, "grain_length_width_ratio": 3.4, "leaf_blast_score": 3, "panicle_blast_score": 0}),
        ("闽优示范1号", [], "籼型杂交水稻", None, None, "演示农业技术团队", "闽审稻演示001", "2024", "福建省中低海拔稻区", {"growth_duration": 122.1, "plant_height": 113.2, "panicle_length": 24.3, "effective_panicles_per_mu": 26.6, "grains_per_panicle": 158.9, "seed_setting_rate": 83.7, "thousand_grain_weight": 26.4, "brown_rice_rate": 80.7, "milled_rice_rate": 71.4, "head_rice_rate": 60.0, "chalky_grain_rate": 34, "chalkiness_degree": 4.6, "amylose_content": 18.0, "gel_consistency": 61, "grain_length": 6.3, "grain_length_width_ratio": 2.7}),
    ]
    for name, aliases, variety_type, female, male, unit, approval, year, region, trait_values in samples:
        variety = Variety(variety_name=name, normalized_name=normalize_name(name), alias_names=aliases, raw_variety_title=name if not aliases else f"{name}（{aliases[0]}）", variety_type=variety_type, female_parent=female, male_parent=male, breeding_unit=unit, approval_number=approval, approval_year=year, suitable_region=region, source_review_id=source.id, data_status="published")
        session.add(variety)
        session.flush()
        for code, value in trait_values.items():
            trait = TRAITS[code]
            observation = PhenotypeObservation(variety_id=variety.id, source_review_id=source.id, trait_code=code, trait_name=trait["name"], trait_category=trait["category"], observation_type="numeric", value_numeric=float(value), unit=trait["unit"], original_value=f"{trait['name']}{value}{trait['unit']}", original_field=trait["name"], source_locator="公开样例详情页", rule_version="v1.0", quality_status="passed", publish_status="published")
            session.add(observation)
    pending_variety = Variety(variety_name="待审核示例稻", normalized_name=normalize_name("待审核示例稻"), alias_names=[], raw_variety_title="待审核示例稻", variety_type="籼型常规稻", source_review_id=source.id, data_status="pending")
    session.add(pending_variety)
    session.flush()
    pending = PhenotypeObservation(variety_id=pending_variety.id, source_review_id=source.id, trait_code="seed_setting_rate", trait_name="结实率", trait_category="产量构成", observation_type="numeric", value_numeric=110.0, unit="%", original_value="结实率110%", original_field="结实率", source_locator="演示异常记录", rule_version="v1.0", quality_status="blocked", publish_status="pending", review_comment="演示异常值：需核对原始来源")
    session.add(pending)
    append_history(source, "数据处理员-张三", "录入演示数据", {"published_varieties": 3, "pending_records": 1})
    session.commit()


LEGACY_INTAKE_TEMPLATE_CODES = frozenset({"rice_data_center", "rice_root_phenotype"})
STRUCTURED_GOVERNANCE_TEMPLATE_CODES = frozenset({
    "germplasm_master",
    "pedigree_relationship",
    "field_trial_package",
    "single_plant_master",
    "genotype_dataset",
    "knowledge_document",
})


def serialize_template(template: DataTemplate, version: TemplateVersion | None) -> dict[str, Any]:
    is_structured_governance = template.template_code in STRUCTURED_GOVERNANCE_TEMPLATE_CODES
    return {
        "id": template.id,
        "template_code": template.template_code,
        "template_name": template.template_name,
        "data_scope": template.data_scope,
        "target_table": template.target_table,
        "description": template.description,
        "status": template.status,
        "current_version": version.version if version else "-",
        "current_version_id": version.id if version else None,
        "change_summary": version.change_summary if version else "",
        "fields": version.field_definitions if version else [],
        "template_group": "structured_governance" if is_structured_governance else "legacy_intake",
        "intake_supported": template.template_code in LEGACY_INTAKE_TEMPLATE_CODES,
    }


def seed_templates(session: Session) -> None:
    germplasm_fields = [
        {"code": "material_code", "name": "材料编码", "kind": "identifier", "required": False, "aliases": ["种质编号", "材料编号", "品系编号"]},
        {"code": "variety_name", "name": "材料名称", "kind": "identifier", "required": True, "aliases": ["种质名称", "品种名称", "品系名称"]},
        {"code": "variety_type", "name": "材料类型", "kind": "attribute", "required": False, "aliases": ["种质类型", "品种类型"]},
        {"code": "aliases", "name": "别名", "kind": "attribute", "required": False, "aliases": ["曾用名", "其他名称"]},
        {"code": "breeding_unit", "name": "选育单位", "kind": "attribute", "required": False, "aliases": ["育种单位", "来源单位"]},
        {"code": "approval_number", "name": "审定编号", "kind": "attribute", "required": False, "aliases": ["审定号"]},
    ]
    pedigree_fields = [
        {"code": "material_code", "name": "子代材料编码", "kind": "identifier", "required": False, "aliases": ["后代编号", "子代编号"]},
        {"code": "variety_name", "name": "子代材料名称", "kind": "basic", "required": True, "aliases": ["子代", "子代材料"]},
        {"code": "female_parent", "name": "母本名称或编码", "kind": "basic", "required": False, "aliases": ["母本", "母本编号"]},
        {"code": "male_parent", "name": "父本名称或编码", "kind": "basic", "required": False, "aliases": ["父本", "父本编号"]},
    ]
    trial_fields = [
        {"code": "trial_code", "name": "试验编号", "kind": "identifier", "required": True, "aliases": ["试验编码"]},
        {"code": "site_code", "name": "试点编号", "kind": "identifier", "required": True, "aliases": ["地点编号"]},
        {"code": "material_code", "name": "材料编码", "kind": "identifier", "required": True, "aliases": ["种质编号"]},
        {"code": "plot_no", "name": "小区号", "kind": "identifier", "required": True, "aliases": ["小区编号"]},
        {"code": "trial_year", "name": "试验年份", "kind": "attribute", "required": True, "aliases": ["年份"]},
    ]
    single_plant_fields = [
        {"code": "sample_code", "name": "单株编号", "kind": "identifier", "required": True, "aliases": ["样本编号"]},
        {"code": "material_code", "name": "材料编码", "kind": "identifier", "required": True, "aliases": ["种质编号"]},
        {"code": "trial_code", "name": "试验编号", "kind": "identifier", "required": False, "aliases": ["试验编码"]},
        {"code": "plot_no", "name": "小区号", "kind": "identifier", "required": False, "aliases": ["小区编号"]},
    ]
    genotype_fields = [
        {"code": "sample_id", "name": "基因型样本编号", "kind": "identifier", "required": True, "aliases": ["IID", "样本号"]},
        {"code": "material_code", "name": "材料编码", "kind": "identifier", "required": True, "aliases": ["种质编号"]},
        {"code": "reference_assembly", "name": "参考基因组版本", "kind": "attribute", "required": True, "aliases": ["参考版本", "GenomeBuild"]},
    ]
    knowledge_fields = [
        {"code": "display_title", "name": "资料标题", "kind": "attribute", "required": True, "aliases": ["标题", "文献标题"]},
        {"code": "source_organization", "name": "来源单位", "kind": "attribute", "required": True, "aliases": ["发布单位", "机构"]},
        {"code": "publication_year", "name": "发布年份", "kind": "attribute", "required": False, "aliases": ["年份"]},
        {"code": "source_url", "name": "来源链接", "kind": "attribute", "required": False, "aliases": ["URL", "DOI"]},
    ]
    seeds = [
        ("rice_data_center", "国家水稻数据中心信息标准", "水稻品种与地上部表型", "phenotype_observation", "用于国家水稻数据中心网页、审定资料及同类品种表型数据的归集。", national_template_fields()),
        ("rice_root_phenotype", "水稻根系表型数据标准", "水稻根系表型", "root_phenotype_observation", "用于根系扫描、根系成像和人工测量等根系表型数据。", root_template_fields()),
        ("germplasm_master", "种质主数据模板", "种质资源主档", "breeding_material", "用于建立平台内稳定的材料编码、名称和别名，是跨文件关联的首要基础。", germplasm_fields),
        ("pedigree_relationship", "材料系谱关系模板", "亲本与后代关系", "variety_basic", "使用子代、母本和父本语义字段表达系谱关系，并兼容不同来源的原始列名。", pedigree_fields),
        ("field_trial_package", "多年多点试验资料包模板", "试验、环境、管理与表型", "trial_data_package", "用于治理试验设计、小区布局、环境管理和表型观测组成的关联资料包。", trial_fields),
        ("single_plant_master", "单株主数据模板", "材料下的单株与样本", "biological_sample", "用于把单株、试验小区、表型照片和后续基因型样本关联起来。", single_plant_fields),
        ("genotype_dataset", "基因型数据与样本映射模板", "VCF/PLINK 与材料映射", "genotype_asset", "用于登记参考版本，并将基因型样本编号映射到种质材料或单株。", genotype_fields),
        ("knowledge_document", "育种文献与情报元数据模板", "公共文献与行业情报", "knowledge_document", "用于维护资料来源、版本和可追溯元数据；正文由本地解析与索引流程处理。", knowledge_fields),
    ]
    for code, name, scope, target, description, fields in seeds:
        template = session.scalar(select(DataTemplate).where(DataTemplate.template_code == code))
        if template:
            continue
        template = DataTemplate(template_code=code, template_name=name, data_scope=scope, target_table=target, description=description)
        session.add(template)
        session.flush()
        version = TemplateVersion(template_id=template.id, version="v1.0", change_summary="第一版预置标准模板", field_definitions=fields, created_by="字段管理员")
        session.add(version)
        session.flush()
        template.current_version_id = version.id
    session.commit()


def get_template_version(session: Session, version_id: str | None) -> tuple[DataTemplate, TemplateVersion]:
    version = session.get(TemplateVersion, version_id) if version_id else None
    if not version:
        raise HTTPException(422, "请选择一套已发布的处理标准模板。")
    template = session.get(DataTemplate, version.template_id)
    if not template or template.template_code not in LEGACY_INTAKE_TEMPLATE_CODES or template.status != "published" or version.status != "published":
        raise HTTPException(422, "所选标准模板不可用，请选择已发布版本。")
    return template, version


def template_parsing_catalog(template: DataTemplate, version: TemplateVersion) -> tuple[dict[str, dict[str, Any]], dict[str, str]]:
    base = ROOT_TRAITS if template.template_code == "rice_root_phenotype" else TRAITS
    catalog = {code: dict(value) for code, value in base.items()}
    header_mappings: dict[str, str] = {} if template.template_code == "rice_root_phenotype" else dict(EXCEL_HEADER_TRAIT_CODES)
    for field in version.field_definitions or []:
        if field.get("kind") != "trait":
            continue
        code = field["code"]
        catalog[code] = {**catalog.get(code, {}), "name": field["name"], "category": field.get("category", "扩展性状"), "unit": field.get("unit", ""), "aliases": field.get("aliases") or []}
        for alias in [field["name"], *(field.get("aliases") or [])]:
            header_mappings[normalize_name(alias)] = code
    return catalog, header_mappings


def template_quality_issues(observation: PhenotypeObservation, version: TemplateVersion | None) -> list[dict[str, str]]:
    if observation.value_numeric is None or not version:
        return []
    field = next((item for item in (version.field_definitions or []) if item.get("code") == observation.trait_code), None)
    if not field:
        return []
    low, high = field.get("min"), field.get("max")
    if (low is not None and observation.value_numeric < float(low)) or (high is not None and observation.value_numeric > float(high)):
        bounds = f"{low if low is not None else '-∞'}–{high if high is not None else '∞'}"
        return [{"rule": f"{version.version}:{observation.trait_code}", "severity": field.get("severity", "warning"), "message": f"模板规则：{field['name']}超出允许范围 {bounds}。"}]
    return []


def spreadsheet_headers(filename: str, content: bytes) -> list[str]:
    suffix = Path(filename).suffix.lower()
    try:
        if suffix == ".csv":
            reader = csv.reader(io.StringIO(content.decode("utf-8-sig", errors="replace")))
            return [str(value or "") for value in next(reader, [])]
        if suffix in {".xlsx", ".xls"}:
            workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            return [str(value or "") for value in next(workbook[workbook.sheetnames[0]].iter_rows(values_only=True), [])]
    except Exception:
        return []
    return []


def infer_template_from_spreadsheet(session: Session, filename: str, content: bytes, selected_template: DataTemplate, selected_version: TemplateVersion) -> tuple[DataTemplate, TemplateVersion, bool]:
    headers = spreadsheet_headers(filename, content)
    if not headers:
        return selected_template, selected_version, False
    candidates: list[tuple[DataTemplate, TemplateVersion]] = []
    for template in session.scalars(select(DataTemplate).where(DataTemplate.status == "published", DataTemplate.template_code.in_(LEGACY_INTAKE_TEMPLATE_CODES))).all():
        version = session.get(TemplateVersion, template.current_version_id)
        if version and version.status == "published":
            candidates.append((template, version))
    scores: list[tuple[int, DataTemplate, TemplateVersion]] = []
    for template, version in candidates:
        catalog, mappings = template_parsing_catalog(template, version)
        score = sum(1 for header in headers if trait_by_alias(header, catalog, mappings))
        scores.append((score, template, version))
    if not scores:
        return selected_template, selected_version, False
    best_score, best_template, best_version = max(scores, key=lambda item: item[0])
    selected_score = next((score for score, template, _ in scores if template.id == selected_template.id), 0)
    if best_score >= 3 and best_template.id != selected_template.id and best_score > selected_score:
        return best_template, best_version, True
    return selected_template, selected_version, False


def backfill_missing_variety_basic_info(session: Session) -> None:
    """Upgrade already-imported demo records from their preserved source text."""
    changed = False
    for variety in session.scalars(select(Variety)).all():
        if not variety.source_review_id or all(getattr(variety, field) for field in VARIETY_BASIC_FIELDS):
            continue
        source = session.get(SourceReview, variety.source_review_id)
        if not source or not source.raw_text:
            continue
        parsed = extract_variety_basic_info(source.raw_text)
        updated_fields = []
        for field, value in parsed.items():
            if value and not getattr(variety, field):
                setattr(variety, field, value)
                updated_fields.append(field)
        if updated_fields:
            append_history(source, "系统", "从已保存原始内容回填基础信息", {"variety": variety.variety_name, "fields": updated_fields})
            changed = True
    if changed:
        session.commit()


def consolidate_duplicate_sources(session: Session) -> None:
    """Merge legacy duplicate imports before enforcing a content-hash unique index."""
    sources_by_hash: dict[str, list[SourceReview]] = {}
    for source in session.scalars(select(SourceReview).where(SourceReview.file_hash.is_not(None)).order_by(SourceReview.created_at.asc())).all():
        sources_by_hash.setdefault(source.file_hash, []).append(source)

    changed = False
    for file_hash, sources in sources_by_hash.items():
        if len(sources) < 2:
            continue
        scores = {}
        for source in sources:
            observation_count = session.scalar(select(func.count(PhenotypeObservation.id)).where(PhenotypeObservation.source_review_id == source.id)) or 0
            variety_count = session.scalar(select(func.count(Variety.id)).where(Variety.source_review_id == source.id)) or 0
            scores[source.id] = variety_count * 1000 + observation_count
        canonical = max(sources, key=lambda source: (scores[source.id], -source.created_at.timestamp()))
        canonical_observations = session.scalars(select(PhenotypeObservation).where(PhenotypeObservation.source_review_id == canonical.id)).all()
        observation_keys = {
            (item.variety_id, item.trait_code, item.observation_type, item.value_numeric, item.value_text, item.unit, item.original_value)
            for item in canonical_observations
        }
        merged_source_ids = []
        dropped_observations = 0
        for duplicate in (source for source in sources if source.id != canonical.id):
            merged_source_ids.append(duplicate.id)
            for observation in session.scalars(select(PhenotypeObservation).where(PhenotypeObservation.source_review_id == duplicate.id)).all():
                key = (observation.variety_id, observation.trait_code, observation.observation_type, observation.value_numeric, observation.value_text, observation.unit, observation.original_value)
                if key in observation_keys:
                    session.delete(observation)
                    dropped_observations += 1
                else:
                    observation.source_review_id = canonical.id
                    observation_keys.add(key)
            for variety in session.scalars(select(Variety).where(Variety.source_review_id == duplicate.id)).all():
                variety.source_review_id = canonical.id
            session.delete(duplicate)
            changed = True
        append_history(canonical, "系统", "合并重复导入来源", {"file_hash": file_hash, "merged_source_ids": merged_source_ids, "dropped_duplicate_observations": dropped_observations})

    if changed:
        session.commit()
    session.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_source_review_file_hash ON source_review (file_hash) WHERE file_hash IS NOT NULL"))
    session.commit()


def find_duplicate_source(session: Session, file_hash: str, source_url: str | None = None) -> tuple[SourceReview | None, str | None]:
    existing = session.scalar(select(SourceReview).where(SourceReview.file_hash == file_hash))
    if existing:
        return existing, "相同文件内容"
    if source_url:
        existing = session.scalar(select(SourceReview).where(SourceReview.source_url == source_url).order_by(SourceReview.created_at.asc()))
        if existing:
            return existing, "相同原始网页地址"
    return None, None


def reject_duplicate_source(existing: SourceReview, reason: str) -> None:
    created_at = existing.created_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    raise HTTPException(409, f"{reason}已导入：{existing.source_name}（{created_at}）。系统未创建重复来源记录。")


def trait_identity(observation: PhenotypeObservation | RootPhenotypeObservation) -> tuple[str, str]:
    """Current demo identity: a variety has one value for one standardized trait.

    Trial/year/location are not yet modeled as separate experimental contexts. Until
    they are, accepting another value would make research queries ambiguous, so the
    formal database deliberately keeps one record per variety and trait.
    """
    return observation.variety_id, observation.trait_code


def observation_retention_rank(observation: PhenotypeObservation) -> tuple[int, int, datetime]:
    """Prefer a published and higher-quality record when consolidating old data."""
    quality_rank = {"passed": 0, "warning": 1, "pending": 2, "blocked": 3}
    return (
        0 if observation.publish_status == "published" else 1,
        quality_rank.get(observation.quality_status, 9),
        observation.created_at,
    )


def consolidate_duplicate_traits(session: Session) -> None:
    """Consolidate legacy duplicates before creating database-level constraints.

    Raw files and source records are retained. Only repeated standardized field
    records are removed, and their source history records which value was retained.
    """
    observation_groups: dict[tuple[str, str], list[PhenotypeObservation]] = {}
    for observation in session.scalars(select(PhenotypeObservation).order_by(PhenotypeObservation.created_at.asc())).all():
        observation_groups.setdefault(trait_identity(observation), []).append(observation)

    for records in observation_groups.values():
        if len(records) < 2:
            continue
        retained = min(records, key=observation_retention_rank)
        for duplicate in (item for item in records if item.id != retained.id):
            source = session.get(SourceReview, duplicate.source_review_id)
            append_history(source, "系统", "清理重复标准字段", {
                "duplicate_observation_id": duplicate.id,
                "retained_observation_id": retained.id,
                "variety_id": duplicate.variety_id,
                "trait_code": duplicate.trait_code,
                "rule": "同一品种 + 同一标准字段仅保留一条",
            })
            session.delete(duplicate)

    root_groups: dict[tuple[str, str], list[RootPhenotypeObservation]] = {}
    for observation in session.scalars(select(RootPhenotypeObservation).order_by(RootPhenotypeObservation.published_at.asc())).all():
        root_groups.setdefault(trait_identity(observation), []).append(observation)
    for records in root_groups.values():
        if len(records) < 2:
            continue
        retained = min(records, key=lambda item: item.published_at)
        for duplicate in (item for item in records if item.id != retained.id):
            source = session.get(SourceReview, duplicate.source_review_id)
            append_history(source, "系统", "清理重复根系标准字段", {
                "duplicate_root_observation_id": duplicate.id,
                "retained_root_observation_id": retained.id,
                "variety_id": duplicate.variety_id,
                "trait_code": duplicate.trait_code,
                "rule": "同一品种 + 同一标准字段仅保留一条",
            })
            session.delete(duplicate)
    session.commit()


def ensure_trait_uniqueness_constraints(session: Session) -> None:
    """Make the deduplication rule a PostgreSQL constraint, not a UI convention."""
    session.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_phenotype_variety_trait ON phenotype_observation (variety_id, trait_code)"))
    session.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_root_phenotype_variety_trait ON root_phenotype_observation (variety_id, trait_code)"))
    session.commit()


class ImportCommit(BaseModel):
    variety_name: str = ""
    aliases: list[str] = Field(default_factory=list)
    raw_title: str = ""
    variety_type: str | None = None
    female_parent: str | None = None
    male_parent: str | None = None
    breeding_unit: str | None = None
    approval_number: str | None = None
    approval_year: str | None = None
    suitable_region: str | None = None
    observations: list[dict[str, Any]] = Field(default_factory=list)
    actor: str = "数据处理员-张三"


class VarietyCreate(BaseModel):
    variety_name: str
    aliases: list[str] = Field(default_factory=list)
    variety_type: str | None = None
    source_review_id: str | None = None
    actor: str = "数据处理员-张三"


class ObservationCreate(BaseModel):
    variety_id: str
    trait_code: str
    value_numeric: float | None = None
    value_text: str | None = None
    unit: str | None = None
    original_value: str
    source_review_id: str | None = None
    source_locator: str | None = None
    actor: str = "数据处理员-张三"


class ObservationUpdate(BaseModel):
    value_numeric: float | None = None
    value_text: str | None = None
    unit: str | None = None
    original_value: str | None = None
    review_comment: str
    actor: str = "数据处理员-张三"


class PublishRequest(BaseModel):
    observation_ids: list[str]
    actor: str = "数据处理员-张三"


class RuleCreate(BaseModel):
    rule_code: str
    rule_name: str
    rule_type: str
    severity: str = "info"
    config: dict[str, Any] = Field(default_factory=dict)
    change_reason: str
    created_by: str = "数据处理员-张三"


class UrlImport(BaseModel):
    url: str
    template_version_id: str
    actor: str = "数据处理员-张三"


class PdfReport(BaseModel):
    filters: dict[str, Any] = Field(default_factory=dict)
    rows: list[dict[str, Any]] = Field(default_factory=list)


class ManualRecord(BaseModel):
    variety_id: str | None = None
    variety_name: str = ""
    aliases: list[str] = Field(default_factory=list)
    variety_type: str | None = None
    trait_code: str
    value_numeric: float | None = None
    value_text: str | None = None
    unit: str | None = None
    original_value: str
    source_reference: str
    source_note: str = ""
    actor: str = "数据处理员-张三"


class TemplateVersionCreate(BaseModel):
    change_summary: str
    action: str = "add_field"
    field_code: str = ""
    field_name: str
    field_kind: Literal["identifier", "basic", "attribute", "trait"] = "attribute"
    category: str = "扩展性状"
    unit: str = ""
    aliases: list[str] = Field(default_factory=list)
    required: bool = False
    min_value: float | None = None
    max_value: float | None = None
    severity: str = "warning"
    request_id: str | None = None
    actor: str = "字段管理员"


class FieldChangeRequestCreate(BaseModel):
    source_review_id: str
    source_field: str
    sample_value: str = ""
    request_note: str = ""
    actor: str = "数据处理员-张三"


class ProjectCreate(BaseModel):
    project_code: str = Field(min_length=2, max_length=80, pattern=r"^[A-Za-z0-9][A-Za-z0-9._-]*$")
    project_name: str = Field(min_length=2, max_length=300)
    description: str = Field(default="", max_length=2000)


class ProjectUpdate(BaseModel):
    project_name: str | None = Field(default=None, min_length=2, max_length=300)
    description: str | None = Field(default=None, max_length=2000)
    status: Literal["active", "archived"] | None = None


class ProjectMemberUpdate(BaseModel):
    member_role: Literal["researcher"] = "researcher"


class PlatformAccountUpdate(BaseModel):
    active: bool


class ResearchSessionCreate(BaseModel):
    title: str = Field(default=DEFAULT_RESEARCH_SESSION_TITLE, min_length=1, max_length=200)


class ResearchSessionRename(BaseModel):
    title: str = Field(min_length=1, max_length=200)


class ResearchChatRequest(BaseModel):
    content: str = Field(min_length=1, max_length=12000)
    knowledge_scope: Literal["private", "public", "both"] = "both"
    attachment_ids: list[str] = Field(default_factory=list, max_length=20)


class ResearchStructuredQueryRequest(BaseModel):
    """User-built query parameters for the public, governed data panel."""

    scope: Literal["rice_phenotype", "root_phenotype"] = "rice_phenotype"
    variety_names: list[str] = Field(default_factory=list, max_length=20)
    trait_codes: list[str] = Field(default_factory=list, max_length=64)
    filters: list[NumericFilter] = Field(default_factory=list, max_length=6)
    limit: int = Field(default=100, ge=1, le=100)


class KnowledgeFolderCreate(BaseModel):
    folder_name: str = Field(min_length=1, max_length=200)
    parent_id: str | None = None
    description: str = Field(default="", max_length=1000)


class KnowledgeFolderUpdate(BaseModel):
    folder_name: str = Field(min_length=1, max_length=200)
    parent_id: str | None = None
    description: str = Field(default="", max_length=1000)


class KnowledgeDocumentMetadataUpdate(BaseModel):
    display_title: str = Field(min_length=1, max_length=500)
    folder_id: str | None = None
    source_organization: str = Field(default="", max_length=300)
    author: str = Field(default="", max_length=300)
    publication_year: str = Field(default="", max_length=20)
    source_url: str = Field(default="", max_length=2000)
    short_description: str = Field(default="", max_length=2000)
    authorization_basis: str = Field(default="", max_length=2000)
    license_scope: str = Field(default="", max_length=100)
    topic_tags: list[str] = Field(default_factory=list, max_length=30)


class KnowledgeDocumentVersionCreate(KnowledgeDocumentMetadataUpdate):
    change_summary: str = Field(min_length=1, max_length=2000)


class GermplasmAnalysisRequest(BaseModel):
    material_key: str = Field(min_length=1, max_length=300)


class ParentRecommendationFilterSettings(BaseModel):
    exclude_common_parent: bool = False
    minimum_evidence_coverage: float = Field(default=0.0, ge=0, le=1)
    minimum_confidence: Literal["低", "中", "高"] = "低"
    required_dimensions: list[Literal[
        "yield", "stability", "lodging", "disease", "complementarity", "quality", "pedigree", "genotype",
    ]] = Field(default_factory=list, max_length=8)


class ParentRecommendationRequest(BaseModel):
    candidate_keys: list[str] = Field(min_length=2, max_length=30)
    breeding_goal: str = Field(min_length=2, max_length=1000)
    constraints: list[str] = Field(default_factory=list, max_length=30)
    weights: dict[str, float] | None = None
    filter_settings: ParentRecommendationFilterSettings | None = None
    sort_mode: Literal["score", "evidence_coverage", "confidence"] | None = None


class ParentRecommendationRuleUpdate(BaseModel):
    weights: dict[str, float] = Field(default_factory=lambda: dict(DEFAULT_RECOMMENDATION_WEIGHTS))
    filter_settings: ParentRecommendationFilterSettings = Field(
        default_factory=lambda: ParentRecommendationFilterSettings(**DEFAULT_RECOMMENDATION_FILTERS)
    )
    sort_mode: Literal["score", "evidence_coverage", "confidence"] = "score"
    rule_note: str = Field(default="", max_length=2000)


class ControlledTrialAnalysisRequest(BaseModel):
    package_id: str = Field(min_length=1, max_length=36)
    analysis_type: Literal["same_trial", "stability", "environment", "management", "tradeoff", "decline"]
    year: int | None = Field(default=None, ge=1900, le=2200)
    site_name: str = Field(default="", max_length=200)
    material_code: str = Field(default="", max_length=200)
    treatment_code: str = Field(default="", max_length=100)


RESEARCH_SKILL_CATALOG: tuple[dict[str, Any], ...] = (
    {
        "code": "regional_trial_statistics",
        "name": "区域试验统计与高产稳产筛选",
        "status": "available",
        "status_label": "已接入",
        "workspace": "assistant",
        "category": "试验统计",
        "summary": "面向已发布的多年多点随机区组区域试验资料，按受控统计结果回答材料比较、稳定性、环境与管理效应、性状权衡和风险拆解问题。",
        "inputs": ["已发布区域试验资料包", "材料、地点、年份、处理与重复记录", "科研人员的自然语言问题"],
        "outputs": ["同试验材料比较与 Tukey 多重比较", "多年多点平均表现、波动和有效环境数", "环境/管理影响与可追溯 PDF 报告"],
        "traceability": ["资料包与发布批次", "统计运行 ID（trial_analysis_run）", "原始记录定位与标准模板版本", "本次问答、证据卡片和研究产物"],
        "suggested_question": "候选材料 A-08 与对照 CK-01 在 3 年 4 点的平均产量、相对增产、波动和有效环境数如何？请生成 PDF 报告。",
    },
    {
        "code": "continuous_trait_gwas",
        "name": "水稻连续性状 GWAS",
        "status": "available",
        "status_label": "已接入",
        "workspace": "gwas",
        "category": "基因组分析",
        "summary": "以受控的 PLINK 基因型、连续性状表型和协变量为输入，完成样本交集预检、固定参数分析和结果归档，避免由大模型直接编造计算结论。",
        "inputs": ["PLINK 三件套基因型压缩包", "含 FID、IID 的连续性状表", "可选协变量表与参考基因组版本"],
        "outputs": ["样本交集与质量预检", "PCA、Manhattan 图、QQ 图和候选位点", "可下载结果包与结果库归档"],
        "traceability": ["上传文件名与校验信息", "样本匹配结果", "已确认的分析计划和参数", "本地运行状态与结果文件清单"],
        "suggested_question": "",
    },
    {
        "code": "genotype_import_qc",
        "name": "基因型导入与水稻专用质控",
        "status": "available",
        "status_label": "已接入",
        "workspace": "genotype",
        "category": "基因型数据治理",
        "summary": "将 VCF/VCF.GZ 或 PLINK 三件套统一为可追溯的 PLINK 分析版本，执行水稻常规育种材料 QC、材料映射和人工发布；合格版本可直接供连续性状 GWAS 选择。",
        "inputs": ["VCF / VCF.GZ", "或包含 .bed/.bim/.fam 的单一 ZIP", "参考基因组版本、材料群体类型"],
        "outputs": ["统一 PLINK 二进制版本", "样本/SNP 质控报告与材料映射表", "可由 GWAS 直接调用的分析就绪版本"],
        "traceability": ["原始文件哈希和分片上传记录", "参考基因组与染色体命名检查", "QC 模板、阈值、运行版本和结果包", "材料映射修订与人工发布记录"],
        "suggested_question": "",
    },
    {
        "code": "field_image_phenotyping",
        "name": "田间图像表型提取",
        "status": "planned",
        "status_label": "规划中",
        "workspace": "",
        "category": "图像表型",
        "summary": "计划以本地 PlantCV 图像处理服务为基础，对带有拍摄协议和标尺信息的田间、穗部或籽粒图像提取可复核表型指标；当前不在平台内执行。",
        "inputs": ["原始图像与拍摄批次", "标尺/校准信息", "明确的目标性状与人工复核规则"],
        "outputs": ["可量化图像表型候选值", "图像质量标记与人工核验队列", "算法版本、校准参数与来源留痕"],
        "traceability": ["原图文件哈希", "采集批次与拍摄协议", "算法/模型版本和参数", "人工核验人与修改记录"],
        "suggested_question": "",
    },
)


def _research_skill_by_code(skill_code: str) -> dict[str, Any] | None:
    return next((item for item in RESEARCH_SKILL_CATALOG if item["code"] == skill_code), None)


def _serialize_research_skill(skill: dict[str, Any], last_opened_at: datetime | None = None) -> dict[str, Any]:
    return {
        **skill,
        "last_opened_at": last_opened_at.isoformat() if last_opened_at else None,
    }


def serialize_research_session(item: ResearchSession) -> dict[str, Any]:
    return {
        "id": item.id,
        "project_id": item.project_id,
        "title": item.title,
        "created_at": item.created_at.isoformat(),
        "updated_at": item.updated_at.isoformat(),
    }


def serialize_research_message(item: ResearchMessage) -> dict[str, Any]:
    operation_state = item.operation_state or []
    return {
        "id": item.id,
        "project_id": item.project_id,
        "role": item.role,
        "content": item.content,
        "evidence": item.evidence or [],
        "operation_state": operation_state,
        "report_available": any(item.get("state") == "report_ready" for item in operation_state if isinstance(item, dict)),
        "created_at": item.created_at.isoformat(),
    }


RESEARCH_RESULT_LABELS = {
    "pdf_report": "PDF 研究报告",
    "chart_png": "分析图表",
    "statistics_json": "结构化统计结果",
    "genotype_qc_package": "基因型质控结果包",
    "gwas_result_zip": "GWAS 分析结果",
}


def serialize_research_result(item: ResearchResult) -> dict[str, Any]:
    return {
        "id": item.id,
        "project_id": item.project_id,
        "session_id": item.session_id,
        "source_message_id": item.source_message_id,
        "analysis_run_id": item.analysis_run_id,
        "result_type": item.result_type,
        "result_type_label": RESEARCH_RESULT_LABELS.get(item.result_type, item.result_type),
        "title": item.title,
        "content_type": item.content_type,
        "file_name": item.file_name,
        "size_bytes": item.size_bytes,
        "summary": item.summary or "",
        "metadata": item.result_metadata or {},
        "created_at": item.created_at.isoformat(),
    }


def _safe_result_filename(name: str, fallback: str) -> str:
    suffix = Path(name).suffix.lower()
    stem = re.sub(r"[^\w.\-()（）]+", "_", Path(name).stem).strip("._") or fallback
    return f"{stem}{suffix}"


def _result_summary(analysis: dict[str, Any] | None, result_type: str) -> str:
    if not analysis:
        return "基于本轮科研助手回答和实际引用证据生成。"
    analysis_type = str(analysis.get("analysis_type") or "")
    title = str(analysis.get("title") or "已发布区域试验统计分析")
    source_count = analysis.get("source_record_count") or analysis.get("sample_size")
    suffix = f"纳入 {source_count} 条可追溯记录。" if source_count else "仅使用本轮可追溯结构化统计结果。"
    if result_type == "chart_png":
        return f"{title}的可下载图表。{suffix}"
    if result_type == "statistics_json":
        return f"{title}的原始结构化统计结果，统计类型：{analysis_type or '受控分析'}。{suffix}"
    return f"{title}的可追溯 PDF 研究报告。{suffix}"


def _store_research_result(
    session: Session,
    *,
    owner_id: str,
    project_id: str,
    source_message_id: str,
    session_id: str | None,
    analysis_run_id: str | None,
    result_type: str,
    title: str,
    file_name: str,
    content_type: str,
    content: bytes,
    analysis: dict[str, Any] | None = None,
) -> ResearchResult:
    """Upsert one private output and write it under the local research volume."""
    existing = session.scalar(select(ResearchResult).where(
        ResearchResult.owner_id == owner_id,
        ResearchResult.source_message_id == source_message_id,
        ResearchResult.result_type == result_type,
    ))
    result = existing or ResearchResult(
        owner_id=owner_id,
        project_id=project_id,
        source_message_id=source_message_id,
        session_id=session_id,
        analysis_run_id=analysis_run_id,
        result_type=result_type,
        title=title,
        content_type=content_type,
        file_name=_safe_result_filename(file_name, "research_result"),
        storage_path="",
    )
    if not existing:
        session.add(result)
        session.flush()

    result.project_id = project_id
    directory = RESULT_STORAGE_DIR / project_id / owner_id / result.id
    directory.mkdir(parents=True, exist_ok=True)
    target_path = directory / _safe_result_filename(file_name, "research_result")
    temporary_path = target_path.with_suffix(f"{target_path.suffix}.tmp")
    temporary_path.write_bytes(content)
    temporary_path.replace(target_path)

    result.session_id = session_id
    result.analysis_run_id = analysis_run_id
    result.title = title
    result.content_type = content_type
    result.file_name = target_path.name
    result.size_bytes = len(content)
    result.storage_path = str(target_path)
    result.summary = _result_summary(analysis, result_type)
    result.result_metadata = {
        "analysis_type": (analysis or {}).get("analysis_type"),
        "analysis_title": (analysis or {}).get("title"),
        "source_record_count": (analysis or {}).get("source_record_count"),
        "generated_from": "agricultural_research_assistant",
    }
    return result


def _store_gwas_result_bundle(
    session: Session,
    *,
    owner_id: str,
    project_id: str,
    plan_id: str,
    content: bytes,
    file_name: str,
    metadata: dict[str, Any],
) -> ResearchResult:
    """Upsert one downloadable, private ZIP for one completed GWAS plan."""
    result = session.scalar(select(ResearchResult).where(
        ResearchResult.owner_id == owner_id,
        ResearchResult.analysis_run_id == plan_id,
        ResearchResult.result_type == "gwas_result_zip",
    ))
    if not result:
        result = ResearchResult(
            owner_id=owner_id,
            project_id=project_id,
            session_id=None,
            source_message_id=None,
            analysis_run_id=plan_id,
            result_type="gwas_result_zip",
            title=f"{metadata.get('trait_name') or '水稻性状'} GWAS 分析结果包",
            content_type="application/zip",
            file_name=_safe_result_filename(file_name, "rice_gwas_results"),
            storage_path="",
        )
        session.add(result)
        session.flush()
    result.project_id = project_id
    directory = RESULT_STORAGE_DIR / project_id / owner_id / result.id
    directory.mkdir(parents=True, exist_ok=True)
    target = directory / _safe_result_filename(file_name, "rice_gwas_results")
    temporary = target.with_suffix(f"{target.suffix}.tmp")
    temporary.write_bytes(content)
    temporary.replace(target)
    result.title = f"{metadata.get('trait_name') or '水稻性状'} GWAS 分析结果包"
    result.content_type = "application/zip"
    result.file_name = target.name
    result.size_bytes = len(content)
    result.storage_path = str(target)
    result.summary = f"包含 {metadata.get('file_count', 0)} 个已完成的 GWAS 输出：质控、PCA、Manhattan、QQ、候选位点及可复核数据。"
    result.result_metadata = {**metadata, "generated_from": "local_rice_gwas"}
    return result


def _load_trial_analysis(session: Session, analysis_run_id: str | None) -> dict[str, Any] | None:
    if not analysis_run_id:
        return None
    row = session.execute(text("""
        SELECT result_json
        FROM trial_analysis_run
        WHERE id = :run_id
        LIMIT 1
    """), {"run_id": analysis_run_id}).scalar_one_or_none()
    if isinstance(row, str):
        try:
            row = json.loads(row)
        except (TypeError, ValueError, json.JSONDecodeError):
            return None
    return row if isinstance(row, dict) else None


def _save_structured_result_artifacts(
    session: Session,
    *,
    owner_id: str,
    message: ResearchMessage,
    question: str,
    analysis_run_id: str | None,
    analysis: dict[str, Any] | None,
) -> list[ResearchResult]:
    """Archive stable JSON and PNG products whenever formal statistics ran."""
    if not analysis:
        return []
    analysis_title = str(analysis.get("title") or "区域试验统计分析")
    payload = json.dumps({
        "question": question,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "analysis": analysis,
    }, ensure_ascii=False, indent=2).encode("utf-8")
    saved = [_store_research_result(
        session,
        owner_id=owner_id,
        project_id=message.project_id,
        source_message_id=message.id,
        session_id=message.session_id,
        analysis_run_id=analysis_run_id,
        result_type="statistics_json",
        title=f"{analysis_title} · 结构化统计结果",
        file_name=f"{analysis_title}-结构化统计结果.json",
        content_type="application/json",
        content=payload,
        analysis=analysis,
    )]
    try:
        chart_png = build_analysis_chart_png(analysis)
    except Exception:
        logger.exception("Could not render result-library chart for analysis_run_id=%s", analysis_run_id)
        chart_png = None
    if chart_png:
        saved.append(_store_research_result(
            session,
            owner_id=owner_id,
            project_id=message.project_id,
            source_message_id=message.id,
            session_id=message.session_id,
            analysis_run_id=analysis_run_id,
            result_type="chart_png",
            title=f"{analysis_title} · 分析图表",
            file_name=f"{analysis_title}-分析图表.png",
            content_type="image/png",
            content=chart_png,
            analysis=analysis,
        ))
    return saved


def serialize_research_attachment(item: ResearchAttachment, include_preview: bool = False) -> dict[str, Any]:
    result = {
        "id": item.id,
        "project_id": item.project_id,
        "file_name": item.file_name,
        "content_type": item.content_type,
        "size_bytes": item.size_bytes,
        "parser_name": item.parser_name,
        "parsing_status": item.parsing_status,
        "parser_warnings": item.parser_warnings or [],
        "created_at": item.created_at.isoformat(),
    }
    if include_preview:
        text_preview = item.parsed_markdown or ""
        if item.parsing_status == "image_ready":
            text_preview = "图片附件不进行本地 Docling 或 OCR 文字解析。提问时，原图会直接提交给神农进行多模态视觉分析。"
        result.update({
            "preview": text_preview[:60000],
            "preview_truncated": len(text_preview) > 60000,
        })
    return result


def serialize_knowledge_folder(item: KnowledgeFolder) -> dict[str, Any]:
    return {
        "id": item.id,
        "project_id": item.project_id,
        "scope": item.scope,
        "owner_id": item.owner_id,
        "parent_id": item.parent_id,
        "folder_name": item.folder_name,
        "description": item.description or "",
        "created_at": item.created_at.isoformat(),
        "updated_at": item.updated_at.isoformat(),
    }


def serialize_knowledge_document(item: KnowledgeDocument, folder: KnowledgeFolder | None = None) -> dict[str, Any]:
    return {
        "id": item.id,
        "project_id": item.project_id,
        "scope": item.scope,
        "folder_id": item.folder_id,
        "folder_name": folder.folder_name if folder else "未分类",
        "original_file_name": item.original_file_name,
        "display_title": item.display_title,
        "content_type": item.content_type,
        "size_bytes": item.size_bytes,
        "source_organization": item.source_organization or "",
        "author": item.author or "",
        "publication_year": item.publication_year or "",
        "source_url": item.source_url or "",
        "short_description": item.short_description or "",
        "authorization_basis": item.authorization_basis or "",
        "license_scope": item.license_scope or "",
        "topic_tags": item.topic_tags or [],
        "parsing_status": item.parsing_status,
        "indexing_status": item.indexing_status,
        "parser_name": item.parser_name,
        "parsed_characters": item.parsed_characters,
        "parser_warnings": item.parser_warnings or [],
        "status": item.status,
        "version_number": item.version_number,
        "supersedes_document_id": item.supersedes_document_id,
        "version_change_summary": item.version_change_summary or "",
        "created_at": item.created_at.isoformat(),
        "updated_at": item.updated_at.isoformat(),
        "published_at": item.published_at.isoformat() if item.published_at else None,
    }


def public_standard_field_catalog() -> dict[str, Any]:
    """The two queryable public templates exposed by the research assistant."""
    return {
        "datasets": [
            {
                "scope": "rice_phenotype",
                "title": "国家水稻数据中心信息标准",
                "description": "品种基础名称加 22 个已发布水稻表型标准字段。",
                "fields": national_template_fields(),
            },
            {
                "scope": "root_phenotype",
                "title": "水稻根系表型数据标准",
                "description": "材料/品种名称加 10 个已发布根系表型标准字段。",
                "fields": root_template_fields(),
            },
        ]
    }


def serialize_public_query_execution(execution: Any) -> dict[str, Any]:
    """Turn long-table rows into a stable field-oriented result for the UI."""
    records_by_variety: dict[str, dict[str, Any]] = {}
    for row in execution.records:
        variety_id = str(row["variety_id"])
        record = records_by_variety.setdefault(variety_id, {
            "id": variety_id,
            "variety_name": row["variety_name"],
            "aliases": row["alias_names"] or [],
            "variety_type": row["variety_type"],
            "approval_number": row["approval_number"],
            "approval_year": row["approval_year"],
            "suitable_region": row["suitable_region"],
            "traits": {},
        })
        record["traits"][row["trait_code"]] = {
            "name": row["trait_name"],
            "value": row["value_numeric"] if row["value_numeric"] is not None else row["value_text"],
            "unit": row["unit"] or "",
            "trial_year": row["trial_year"],
            "trial_location": row["trial_location"],
            "evaluation_method": row["evaluation_method"],
        }
    records = list(records_by_variety.values())
    return {
        "template_code": execution.template_code,
        "parameters": execution.parameters,
        "record_count": len(records),
        "observation_count": len(execution.records),
        "matched_variety_names": execution.matched_variety_names,
        "unresolved_variety_names": execution.unresolved_variety_names,
        "records": records,
    }


app = FastAPI(title="隆耘 Agent 育种智能体", version="1.5.0")
ACPS_SETTINGS = AcpsSettings.from_env()
app.add_middleware(TrustedHostMiddleware, allowed_hosts=TRUSTED_HOSTS)
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def ensure_application_database_role(session: Session) -> None:
    """Create the least-privileged API role before application tables exist.

    Database migrations run through the bootstrap account. The web API never
    connects as that superuser, otherwise PostgreSQL would bypass RLS.
    """
    if not re.fullmatch(r"[a-z_][a-z0-9_]{0,62}", APP_DATABASE_ROLE):
        raise RuntimeError("APP_DATABASE_ROLE must be a simple PostgreSQL role name")
    escaped_password = APP_DATABASE_PASSWORD.replace("'", "''")
    role = APP_DATABASE_ROLE
    exists = session.scalar(text("SELECT 1 FROM pg_roles WHERE rolname = :role"), {"role": role})
    if not exists:
        session.execute(text(
            f"CREATE ROLE {role} LOGIN NOSUPERUSER NOCREATEDB NOCREATEROLE "
            f"NOREPLICATION NOBYPASSRLS PASSWORD '{escaped_password}'"
        ))
    else:
        session.execute(text(f"ALTER ROLE {role} NOSUPERUSER NOCREATEDB NOCREATEROLE NOREPLICATION NOBYPASSRLS PASSWORD '{escaped_password}'"))
    session.execute(text(f"GRANT USAGE ON SCHEMA public TO {role}"))
    session.execute(text(f"GRANT SELECT, INSERT, UPDATE, DELETE ON ALL TABLES IN SCHEMA public TO {role}"))
    session.execute(text(f"GRANT USAGE, SELECT ON ALL SEQUENCES IN SCHEMA public TO {role}"))
    session.execute(text(f"ALTER DEFAULT PRIVILEGES IN SCHEMA public GRANT SELECT, INSERT, UPDATE, DELETE ON TABLES TO {role}"))
    session.execute(text(f"ALTER DEFAULT PRIVILEGES IN SCHEMA public GRANT USAGE, SELECT ON SEQUENCES TO {role}"))
    session.commit()


def backfill_image_ready_research_attachments(session: Session) -> None:
    """Move all earlier image attachments to the native-vision-only workflow."""
    attachments = session.scalars(select(ResearchAttachment)).all()
    changed = False
    for attachment in attachments:
        if Path(attachment.file_name).suffix.lower() not in VISION_IMAGE_SUFFIXES:
            continue
        if (
            attachment.parser_name != "native_image"
            or attachment.parsing_status != "image_ready"
            or attachment.parsed_markdown
        ):
            attachment.parser_name = "native_image"
            attachment.parsing_status = "image_ready"
            attachment.parsed_markdown = None
            attachment.parsed_metadata = {}
            attachment.parser_warnings = ["图片不进行本地 Docling 或 OCR 文字解析；提问时将原图直接提交给神农进行多模态分析。"]
            changed = True
    if changed:
        session.commit()


def normalize_legacy_range_markers(session: Session) -> None:
    """Repair escaped range markers stored before parser normalization was added."""
    escaped_marker = "chr(92) || '~'"
    statements = (
        "UPDATE knowledge_chunk SET content = replace(content, " + escaped_marker + ", '～') "
        "WHERE position(" + escaped_marker + " in content) > 0",
        "UPDATE research_attachment SET parsed_markdown = replace(parsed_markdown, " + escaped_marker + ", '～') "
        "WHERE parsed_markdown IS NOT NULL AND position(" + escaped_marker + " in parsed_markdown) > 0",
        "UPDATE research_message SET content = replace(content, " + escaped_marker + ", '～') "
        "WHERE position(" + escaped_marker + " in content) > 0",
    )
    changed = False
    for statement in statements:
        changed = session.execute(text(statement)).rowcount > 0 or changed
    if changed:
        session.commit()


def ensure_single_institution_schema(session: Session) -> None:
    """Install the one-institution/project boundary and backfill legacy rows."""
    institution = session.get(Institution, INSTITUTION_ID)
    if not institution:
        session.add(Institution(
            id=INSTITUTION_ID,
            institution_code=INSTITUTION_CODE,
            institution_name=INSTITUTION_NAME,
            status="active",
        ))

    default_accounts = (
        ("wang.researcher", "王研究员", "researcher"),
        ("li.researcher", "李研究员", "researcher"),
        ("zhang.processor", "张数据处理员", "data_processor"),
        ("chen.fieldadmin", "陈字段管理员", "field_admin"),
    )
    for username, display_name, role in default_accounts:
        account = session.get(PlatformAccount, username)
        if not account:
            session.add(PlatformAccount(
                username=username,
                display_name=display_name,
                business_role=role,
                institution_id=INSTITUTION_ID,
                active=True,
            ))
        # Existing directory assignments may belong to another provisioned
        # institution; startup must never silently move those accounts.

    project = session.get(ResearchProject, DEFAULT_PROJECT_ID)
    if not project:
        project = ResearchProject(
            id=DEFAULT_PROJECT_ID,
            project_code=DEFAULT_PROJECT_CODE,
            project_name=DEFAULT_PROJECT_NAME,
            description="海南南繁统一默认课题；历史业务数据已自动归入此课题。",
            institution_id=INSTITUTION_ID,
            status="active",
            created_by="system-bootstrap",
        )
        session.add(project)
    else:
        project.institution_id = INSTITUTION_ID
    session.flush()

    for username in ("wang.researcher", "li.researcher"):
        membership = session.scalar(select(ProjectMember).where(
            ProjectMember.project_id == DEFAULT_PROJECT_ID,
            ProjectMember.username == username,
        ))
        if not membership:
            session.add(ProjectMember(
                project_id=DEFAULT_PROJECT_ID,
                username=username,
                member_role="researcher",
                created_by="system-bootstrap",
            ))
    session.commit()

    orm_project_tables = (
        "variety_basic",
        "source_review",
        "phenotype_observation",
        "root_phenotype_observation",
        "research_session",
        "research_message",
        "research_attachment",
        "research_audit",
        "research_result",
        "knowledge_folder",
        "knowledge_document",
        "knowledge_chunk",
        "field_change_request",
    )
    raw_project_tables = (
        "trial_import_batch",
        "trial_data_package",
        "field_trial",
        "gwas_analysis_plan",
        "genotype_asset",
        "genotype_asset_version",
        "genotype_upload_session",
        "genotype_processing_job",
        "genotype_sample_mapping",
        "genotype_governance_request",
    )
    for table_name in (*orm_project_tables, *raw_project_tables):
        exists = session.scalar(text("SELECT to_regclass(:table_name)"), {"table_name": f"public.{table_name}"})
        if not exists:
            continue
        session.execute(text(
            f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS project_id VARCHAR(36) "
            f"NOT NULL DEFAULT '{DEFAULT_PROJECT_ID}'"
        ))
        session.execute(text(f"UPDATE {table_name} SET project_id = :project_id WHERE project_id IS NULL"), {
            "project_id": DEFAULT_PROJECT_ID,
        })
        session.execute(text(f"CREATE INDEX IF NOT EXISTS ix_{table_name}_project_id ON {table_name}(project_id)"))
    session.commit()


def ensure_institution_data_planes(session: Session) -> None:
    """Provision one private bucket and one business database per institution."""
    if not INSTITUTION_DATA_ENABLED:
        return
    institutions = session.scalars(select(Institution).where(Institution.status == "active")).all()
    for institution in institutions:
        config = session.scalar(select(InstitutionDataConfig).where(
            InstitutionDataConfig.institution_id == institution.id
        ))
        bucket_name = safe_identifier(institution.institution_code, "longyun").replace("_", "-")
        database_name = INSTITUTION_DATABASES.database_name(institution.institution_code)
        if not config:
            config = InstitutionDataConfig(
                institution_id=institution.id,
                minio_bucket=bucket_name,
                business_database=database_name,
                access_policy=default_access_policy(institution.id, bucket_name),
                status="provisioning",
            )
            session.add(config)
            session.commit()
        try:
            INSTITUTION_OBJECT_STORE.ensure_private_bucket(config.minio_bucket)
            INSTITUTION_DATABASES.ensure_database(config.business_database)
            config.status = "active"
            config.access_policy = default_access_policy(institution.id, config.minio_bucket)
            session.commit()
        except Exception:
            config.status = "error"
            session.commit()
            raise


def institution_data_config_for_user(
    session: Session,
    user: CurrentUser,
) -> tuple[PlatformAccount, InstitutionDataConfig]:
    if not INSTITUTION_DATA_ENABLED:
        raise HTTPException(503, "机构数据接入层尚未启用。")
    account = sync_platform_account(session, user)
    config = session.scalar(select(InstitutionDataConfig).where(
        InstitutionDataConfig.institution_id == account.institution_id
    ))
    if not config:
        raise HTTPException(503, "当前机构尚未配置独立 MinIO Bucket 和业务数据库。")
    if config.status != "active":
        raise HTTPException(503, f"当前机构数据平面状态为 {config.status}，请联系管理员。")
    return account, config


@app.on_event("startup")
def startup() -> None:
    validate_runtime_configuration()
    with MigrationSessionLocal() as session:
        ensure_application_database_role(session)
        try:
            session.execute(text("CREATE EXTENSION IF NOT EXISTS vector"))
            session.commit()
        except Exception as exc:
            raise RuntimeError("PostgreSQL pgvector 扩展不可用，请使用包含 pgvector 的数据库镜像。") from exc
    Base.metadata.create_all(migration_engine)
    with MigrationSessionLocal() as session:
        # Trial data remains separate from legacy variety-level records.  A
        # measurement is meaningful only together with trial, treatment,
        # replicate, environment and raw-source location.
        ensure_trial_package_schema(session)
        ensure_breeding_dossier_schema(session)
        ensure_breeding_intelligence_schema(session)
        ensure_genomics_schema(session)
        ensure_genotype_asset_schema(session)
        ensure_single_institution_schema(session)
        session.execute(text(
            "ALTER TABLE permission_audit ADD COLUMN IF NOT EXISTS institution_id VARCHAR(80) "
            f"NOT NULL DEFAULT '{INSTITUTION_ID}'"
        ))
        session.execute(text(
            "CREATE INDEX IF NOT EXISTS ix_permission_audit_institution_id "
            "ON permission_audit(institution_id)"
        ))
        ensure_institution_data_planes(session)
        retire_legacy_seeded_trial_demo(session)
        session.execute(text("ALTER TABLE source_review ADD COLUMN IF NOT EXISTS template_version_id VARCHAR(36)"))
        session.execute(text("ALTER TABLE research_session ADD COLUMN IF NOT EXISTS memory_state JSONB NOT NULL DEFAULT '{}'::jsonb"))
        session.execute(text("ALTER TABLE knowledge_document ADD COLUMN IF NOT EXISTS version_change_summary TEXT"))
        session.execute(text("ALTER TABLE knowledge_document ADD COLUMN IF NOT EXISTS authorization_basis TEXT"))
        session.execute(text("ALTER TABLE knowledge_document ADD COLUMN IF NOT EXISTS license_scope VARCHAR(100)"))
        session.execute(text("ALTER TABLE knowledge_document ADD COLUMN IF NOT EXISTS topic_tags JSONB NOT NULL DEFAULT '[]'::jsonb"))
        session.commit()
        backfill_image_ready_research_attachments(session)
        normalize_legacy_range_markers(session)
        seed_templates(session)
        seed_data(session)
        # The dossier seed is intentionally dependent on the material table.
        # It becomes active once the regional-trial package has been published.
        seed_mock_breeding_dossiers(session)
        backfill_missing_variety_basic_info(session)
        # Demo defaults to retaining every import. Future deployments can enable
        # request-level source de-duplication without changing the import flow.
        session.execute(text("DROP INDEX IF EXISTS uq_source_review_file_hash"))
        session.commit()
        consolidate_duplicate_traits(session)
        ensure_trait_uniqueness_constraints(session)
        ensure_research_rls(session)
        seed_public_knowledge_folders(session)
        ensure_knowledge_rls(session)
        ensure_application_database_role(session)
    resume_interrupted_knowledge_documents()


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


INSTITUTION_TEMPLATE_DATASETS = {
    "germplasm_master": {"germplasm"},
    "pedigree_relationship": {"pedigree"},
    "field_trial_package": {"phenotype", "environment"},
    "genotype_dataset": {"genotype"},
    "knowledge_document": {"literature"},
}


@app.get("/api/institution-data/contracts")
def institution_data_contracts(
    user: CurrentUser = Depends(require_business_user),
    session: Session = Depends(get_business_project_session),
) -> dict[str, Any]:
    account, config = institution_data_config_for_user(session, user)
    return {
        "institution_id": account.institution_id,
        "project_id": active_project_id(session),
        "bucket": config.minio_bucket,
        "business_database": config.business_database,
        "limits": {"regular_bytes": 200 * 1024 * 1024, "genotype_archive_bytes": 2 * 1024 * 1024 * 1024},
        "datasets": [
            {
                "id": dataset_type,
                "label": label,
                "formats": sorted({suffix.lstrip(".").upper() for suffix in INSTITUTION_SUPPORTED_SUFFIXES[dataset_type]}),
                "standard_fields": sorted(FIELD_ALIASES.get(dataset_type, {})),
            }
            for dataset_type, label in DATASET_LABELS.items()
        ],
        "field_mapping_contract": "source-column -> standard-field",
    }


@app.get("/api/institution-data/config")
def institution_data_config(
    user: CurrentUser = Depends(require_business_user),
    session: Session = Depends(get_business_project_session),
) -> dict[str, Any]:
    account, config = institution_data_config_for_user(session, user)
    institution = session.get(Institution, account.institution_id)
    return {
        "institution_id": account.institution_id,
        "institution_name": institution.institution_name if institution else account.institution_id,
        "minio_bucket": config.minio_bucket,
        "business_database": config.business_database,
        "access_policy": config.access_policy,
        "status": config.status,
    }


@app.post("/api/institution-data/imports")
async def import_institution_data(
    dataset_type: str = Form(...),
    field_mapping: str | None = Form(default=None),
    template_version_id: str | None = Form(default=None),
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_data_processor),
    session: Session = Depends(get_business_project_session),
) -> dict[str, Any]:
    account, config = institution_data_config_for_user(session, user)
    project_id = active_project_id(session)
    mapping: dict[str, str] = {}
    resolved_template_version_id: str | None = None
    staged = None
    raw_stored = False
    try:
        mapping = parse_field_mapping(field_mapping)
        if template_version_id:
            version = session.get(TemplateVersion, template_version_id)
            template = session.get(DataTemplate, version.template_id) if version else None
            if not version or not template or version.status != "published":
                raise HTTPException(422, "所选标准模板版本不存在或尚未发布。")
            allowed_datasets = INSTITUTION_TEMPLATE_DATASETS.get(template.template_code, set())
            if dataset_type not in allowed_datasets:
                raise HTTPException(422, f"模板“{template.template_name}”不适用于 {DATASET_LABELS.get(dataset_type, dataset_type)}。")
            resolved_template_version_id = version.id
        staged = await stage_upload(file, dataset_type, INSTITUTION_DATA_SETTINGS.staging_dir)
        batch_id = str(uuid.uuid4())
        object_key = object_key_for(account.institution_id, project_id, dataset_type, batch_id, staged.file_name)
        INSTITUTION_OBJECT_STORE.put_file(
            config.minio_bucket,
            object_key,
            staged,
            file.content_type or "application/octet-stream",
        )
        raw_stored = True
        result = import_into_institution_database(
            INSTITUTION_DATABASES.engine(config.business_database),
            batch_id=batch_id,
            institution_id=account.institution_id,
            project_id=project_id,
            dataset_type=dataset_type,
            upload=staged,
            bucket_name=config.minio_bucket,
            object_key=object_key,
            field_mapping=mapping,
            template_version_id=resolved_template_version_id,
        )
        record_permission_audit(
            session,
            user,
            "institution_data_imported",
            "institution_import_batch",
            batch_id,
            project_id=project_id,
            after={
                "institution_id": account.institution_id,
                "dataset_type": dataset_type,
                "bucket": config.minio_bucket,
                "object_key": object_key,
                "entity_count": result["entity_count"],
                "issue_count": result["issue_count"],
            },
        )
        session.commit()
        return {
            "batch_id": batch_id,
            "institution_id": account.institution_id,
            "project_id": project_id,
            "dataset_type": dataset_type,
            "dataset_label": DATASET_LABELS[dataset_type],
            "raw_object": {"bucket": config.minio_bucket, "key": object_key, "sha256": staged.sha256, "size_bytes": staged.size_bytes},
            "structured_database": config.business_database,
            "template_version_id": resolved_template_version_id,
            "mapping_mode": "field_mapping" if mapping else "standard_template",
            **result,
        }
    except InstitutionDataError as exc:
        logger.warning(
            "Institution data import rejected institution=%s project=%s dataset=%s file=%s reason=%s",
            account.institution_id,
            project_id,
            dataset_type,
            getattr(file, "filename", "upload"),
            str(exc),
        )
        suffix = "；原始文件已保存到机构 Bucket，可修正数据或映射后重新导入" if raw_stored else ""
        raise HTTPException(422, f"{exc}{suffix}") from exc
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Institution data import failed")
        suffix = "；原始文件已安全保存到机构 Bucket，可修复后重新处理" if raw_stored else ""
        raise HTTPException(502, f"机构数据导入失败：{str(exc)[:180]}{suffix}") from exc
    finally:
        if staged:
            staged.path.unlink(missing_ok=True)


@app.get("/api/institution-data/imports")
def get_institution_imports(
    limit: int = Query(default=100, ge=1, le=500),
    user: CurrentUser = Depends(require_business_user),
    session: Session = Depends(get_business_project_session),
) -> list[dict[str, Any]]:
    account, config = institution_data_config_for_user(session, user)
    return list_institution_batches(
        INSTITUTION_DATABASES.engine(config.business_database),
        account.institution_id,
        active_project_id(session),
        limit,
    )


@app.get("/api/institution-data/trace/{entity_key}")
def get_institution_entity_trace(
    entity_key: str,
    user: CurrentUser = Depends(require_business_user),
    session: Session = Depends(get_business_project_session),
) -> dict[str, Any]:
    account, config = institution_data_config_for_user(session, user)
    trace = trace_institution_entity(
        INSTITUTION_DATABASES.engine(config.business_database),
        account.institution_id,
        active_project_id(session),
        entity_key,
    )
    if not trace["entities"] and not trace["relations"] and not trace["issues"]:
        raise HTTPException(404, "当前机构和课题下未找到该实体标识。")
    return {
        "institution_id": account.institution_id,
        "project_id": active_project_id(session),
        **trace,
    }


@app.get("/api/research/intelligence/materials")
def get_intelligence_materials(
    q: str = Query(default="", max_length=200),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> list[dict[str, Any]]:
    account, config = institution_data_config_for_user(session, user)
    return list_intelligence_materials(
        INSTITUTION_DATABASES.engine(config.business_database),
        account.institution_id,
        active_project_id(session),
        q,
    )


@app.post("/api/research/intelligence/material-analysis")
def create_material_analysis(
    payload: GermplasmAnalysisRequest,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    account, config = institution_data_config_for_user(session, user)
    try:
        result = build_material_analysis(
            INSTITUTION_DATABASES.engine(config.business_database),
            account.institution_id,
            active_project_id(session),
            payload.material_key.strip(),
        )
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    run_id = save_material_analysis(session, user.id, result)
    record_permission_audit(
        session, user, "germplasm_analysis_created", "germplasm_analysis_run", run_id,
        project_id=active_project_id(session),
        after={"material_key": payload.material_key, "evidence_counts": result["evidence_counts"]},
    )
    session.commit()
    return {"run_id": run_id, **result}


@app.get("/api/research/intelligence/material-analysis/{run_id}/report.pdf")
def material_analysis_report(
    run_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> Response:
    result = get_material_analysis_run(session, run_id, user.id, active_project_id(session))
    if not result:
        raise HTTPException(404, "未找到当前账号在本课题创建的种质解析记录。")
    return Response(
        content=build_intelligence_pdf("种质资源综合解析报告", result),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="germplasm-analysis-{run_id}.pdf"'},
    )


@app.get("/api/research/intelligence/recommendation-rules")
def get_parent_recommendation_rules(
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    return get_recommendation_rule(session, active_project_id(session))


@app.put("/api/research/intelligence/recommendation-rules")
def put_parent_recommendation_rules(
    payload: ParentRecommendationRuleUpdate,
    user: CurrentUser = Depends(require_field_admin),
    session: Session = Depends(get_business_project_session),
) -> dict[str, Any]:
    try:
        result = update_recommendation_rule(
            session, active_project_id(session), payload.weights, payload.rule_note, user.id,
            filter_settings=payload.filter_settings.model_dump(), sort_mode=payload.sort_mode,
        )
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    record_permission_audit(
        session, user, "parent_recommendation_rule_updated", "parent_recommendation_rule",
        active_project_id(session), project_id=active_project_id(session), after=result,
    )
    session.commit()
    return result


@app.post("/api/research/intelligence/parent-recommendations")
def create_parent_recommendations(
    payload: ParentRecommendationRequest,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    candidate_keys = list(dict.fromkeys(item.strip() for item in payload.candidate_keys if item.strip()))
    if len(candidate_keys) < 2:
        raise HTTPException(422, "至少选择 2 个不同的候选亲本。")
    account, config = institution_data_config_for_user(session, user)
    rule = get_recommendation_rule(session, active_project_id(session))
    weights = dict(rule["weights"])
    if payload.weights is not None:
        unknown = sorted(set(payload.weights) - set(DEFAULT_RECOMMENDATION_WEIGHTS))
        if unknown:
            raise HTTPException(422, f"未知权重维度：{'、'.join(unknown)}。")
        weights.update({key: float(value) for key, value in payload.weights.items()})
    filter_settings = (
        payload.filter_settings.model_dump()
        if payload.filter_settings is not None
        else dict(rule.get("filter_settings") or DEFAULT_RECOMMENDATION_FILTERS)
    )
    sort_mode = payload.sort_mode or str(rule.get("sort_mode") or "score")
    try:
        result = run_parent_recommendation(
            INSTITUTION_DATABASES.engine(config.business_database),
            account.institution_id,
            active_project_id(session),
            candidate_keys,
            weights,
            payload.breeding_goal,
            payload.constraints,
            filter_settings,
            sort_mode,
        )
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    request_json = payload.model_dump()
    request_json["candidate_keys"] = candidate_keys
    run_id = save_parent_recommendation(
        session,
        owner_id=user.id,
        institution_id=account.institution_id,
        project_id=active_project_id(session),
        request_payload=request_json,
        result=result,
        rule_version=int(rule["version"]),
    )
    record_permission_audit(
        session, user, "parent_auxiliary_recommendation_created", "parent_recommendation_run", run_id,
        project_id=active_project_id(session),
        after={"candidate_count": len(candidate_keys), "combination_count": len(result["recommendations"])},
    )
    session.commit()
    return {"run_id": run_id, "rule_version": rule["version"], **result}


def _parent_recommendation_for_current_user(session: Session, run_id: str, user: CurrentUser) -> dict[str, Any]:
    result = get_parent_recommendation_run(session, run_id, user.id, active_project_id(session))
    if not result:
        raise HTTPException(404, "未找到当前账号在本课题创建的亲本辅助推荐记录。")
    return result


@app.get("/api/research/intelligence/parent-recommendations/{run_id}/report.pdf")
def parent_recommendation_report(
    run_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> Response:
    result = _parent_recommendation_for_current_user(session, run_id, user)
    return Response(
        content=build_intelligence_pdf("亲本组合辅助推荐报告", result),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="parent-recommendation-{run_id}.pdf"'},
    )


@app.get("/api/research/intelligence/parent-recommendations/{run_id}/result.csv")
def parent_recommendation_csv(
    run_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> Response:
    result = _parent_recommendation_for_current_user(session, run_id, user)
    return Response(
        content=recommendation_csv(result),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="parent-recommendation-{run_id}.csv"'},
    )


@app.get("/api/research/trial-analysis/packages")
def get_trial_analysis_packages(
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> list[dict[str, Any]]:
    rows = session.execute(text("""
        SELECT package.id, package.package_code, package.package_name, package.created_at,
               COUNT(DISTINCT trial.id) AS trial_count,
               COUNT(DISTINCT entry.material_id) AS material_count,
               COUNT(observation.id) AS observation_count
        FROM trial_data_package package
        LEFT JOIN field_trial trial ON trial.package_id=package.id AND trial.data_status='published'
        LEFT JOIN trial_entry entry ON entry.trial_id=trial.id
        LEFT JOIN trial_phenotype_observation observation ON observation.entry_id=entry.id AND observation.publish_status='published'
        WHERE package.project_id=:project_id AND package.governance_status='published'
        GROUP BY package.id, package.package_code, package.package_name, package.created_at
        ORDER BY package.created_at DESC
    """), {"project_id": active_project_id(session)}).mappings().all()
    return [intelligence_json_safe(dict(row)) for row in rows]


@app.get("/api/research/trial-analysis/packages/{package_id}/options")
def get_trial_analysis_package_options(
    package_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    package_exists = session.execute(text("""
        SELECT 1 FROM trial_data_package
        WHERE id=:package_id AND project_id=:project_id AND governance_status='published'
    """), {"package_id": package_id, "project_id": active_project_id(session)}).scalar_one_or_none()
    if not package_exists:
        raise HTTPException(404, "所选区域试验资料包不存在、未发布或不属于当前课题。")
    trials = session.execute(text("""
        SELECT DISTINCT trial.trial_year, site.site_code, site.site_name
        FROM field_trial trial JOIN trial_site site ON site.id=trial.site_id
        WHERE trial.package_id=:package_id AND trial.data_status='published'
        ORDER BY trial.trial_year, site.site_name
    """), {"package_id": package_id}).mappings().all()
    materials = session.execute(text("""
        SELECT DISTINCT material.material_code, material.material_name, material.is_check
        FROM field_trial trial
        JOIN trial_entry entry ON entry.trial_id=trial.id
        JOIN breeding_material material ON material.id=entry.material_id
        WHERE trial.package_id=:package_id AND trial.data_status='published'
        ORDER BY material.material_code
    """), {"package_id": package_id}).mappings().all()
    treatments = session.execute(text("""
        SELECT DISTINCT treatment.treatment_code, treatment.treatment_name
        FROM field_trial trial JOIN trial_treatment treatment ON treatment.trial_id=trial.id
        WHERE trial.package_id=:package_id AND trial.data_status='published'
        ORDER BY treatment.treatment_code
    """), {"package_id": package_id}).mappings().all()
    return intelligence_json_safe({
        "years": sorted({int(item["trial_year"]) for item in trials}),
        "sites": [dict(item) for item in trials],
        "materials": [dict(item) for item in materials],
        "treatments": [dict(item) for item in treatments],
    })


def _controlled_trial_question(payload: ControlledTrialAnalysisRequest) -> str:
    context = " ".join(
        str(value).strip() for value in (payload.year, payload.site_name, payload.material_code, payload.treatment_code)
        if value not in (None, "")
    )
    prompts = {
        "same_trial": f"同一试验 {context} 标准施氮全部材料方差分析和 Tukey 多重比较",
        "stability": f"{context} 多年多点平均产量、相对增产、波动和有效环境稳定性分析",
        "environment": f"{context} 土壤 pH、有效磷、降雨和环境影响关联分析",
        "management": f"{context} 不同施氮管理措施与材料施氮交互效应分析",
        "tradeoff": f"{context} 高产材料的产量、抗病、株高和米质权衡分析",
        "decline": f"{context} 材料表现下降或异常的环境、病害和性状证据拆解",
    }
    return prompts[payload.analysis_type]


@app.post("/api/research/trial-analysis/run")
def create_controlled_trial_analysis(
    payload: ControlledTrialAnalysisRequest,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    package = session.execute(text("""
        SELECT id, package_code, package_name, created_at FROM trial_data_package
        WHERE id=:package_id AND project_id=:project_id AND governance_status='published'
    """), {"package_id": payload.package_id, "project_id": active_project_id(session)}).mappings().first()
    if not package:
        raise HTTPException(404, "所选区域试验资料包不存在、未发布或不属于当前课题。")
    if payload.analysis_type == "same_trial" and (not payload.year or not payload.site_name.strip()):
        raise HTTPException(422, "同一试验材料比较必须选择年份和地点。")
    if payload.analysis_type == "decline" and not payload.material_code.strip():
        raise HTTPException(422, "表现异常证据拆解必须选择材料。")
    question = _controlled_trial_question(payload)
    try:
        analysis = run_controlled_trial_analysis(session, dict(package), question, user.id)
    except TrialStatisticsError as exc:
        raise HTTPException(422, str(exc)) from exc
    if not analysis:
        raise HTTPException(422, "当前参数未匹配到受控分析方法。")
    session.commit()
    record_permission_audit(
        session, user, "controlled_trial_analysis_created", "trial_analysis_run",
        str(analysis["analysis_run_id"]), project_id=active_project_id(session),
        after={"package_id": payload.package_id, "analysis_type": analysis.get("analysis_type")},
    )
    session.commit()
    return {"question": question, "package": intelligence_json_safe(dict(package)), **intelligence_json_safe(analysis)}


def _trial_run_for_current_user(session: Session, run_id: str, user: CurrentUser) -> dict[str, Any]:
    row = session.execute(text("""
        SELECT run.id, run.result_json, run.request_question, run.completed_at,
               package.package_code, package.package_name
        FROM trial_analysis_run run
        JOIN trial_data_package package ON package.id=run.package_id
        WHERE run.id=:run_id AND run.requested_by=:owner_id AND package.project_id=:project_id
    """), {"run_id": run_id, "owner_id": user.id, "project_id": active_project_id(session)}).mappings().first()
    if not row:
        raise HTTPException(404, "未找到当前账号在本课题创建的试验分析记录。")
    return intelligence_json_safe(dict(row))


@app.get("/api/research/trial-analysis/runs/{run_id}/chart.png")
def controlled_trial_chart(
    run_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> Response:
    run = _trial_run_for_current_user(session, run_id, user)
    content = build_analysis_chart_png(run["result_json"])
    if not content:
        raise HTTPException(422, "该分析结果没有可安全绘制的数值序列。")
    return Response(content=content, media_type="image/png")


@app.get("/api/research/trial-analysis/runs/{run_id}/report.pdf")
def controlled_trial_report(
    run_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> Response:
    run = _trial_run_for_current_user(session, run_id, user)
    analysis = run["result_json"]
    evidence = [{
        "priority": "P0",
        "title": f"已发布试验资料包 {run['package_name']}",
        "detail": f"资料包编号 {run['package_code']}；分析运行 {run_id}；原始记录 {analysis.get('source_record_count', 0)} 条",
    }]
    answer = f"已按 {analysis.get('model_formula') or '受控统计方法'} 完成分析。{analysis.get('limitations') or ''}"
    content = build_research_report_pdf(
        question=run["request_question"], answer=answer, evidence=evidence, analysis=analysis,
    )
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="trial-analysis-{run_id}.pdf"'},
    )


@app.get("/api/context")
def platform_context(
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_business_user),
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    project = resolve_project_access(session, user, x_project_id)
    projects = accessible_projects(session, user)
    account = session.get(PlatformAccount, user.username)
    institution = session.get(Institution, account.institution_id) if account else None
    session.commit()
    return {
        "institution": {
            "id": institution.id if institution else project.institution_id,
            "code": institution.institution_code if institution else project.institution_id,
            "name": institution.institution_name if institution else project.institution_id,
        },
        "user": {
            "id": user.id,
            "username": user.username,
            "display_name": user.display_name,
            "business_role": _business_role(user),
            "roles": sorted(user.roles),
        },
        "active_project_id": project.id,
        "projects": [serialize_project(item) for item in projects],
    }


@app.get("/api/projects")
def list_projects(
    user: CurrentUser = Depends(require_business_user),
    session: Session = Depends(get_session),
) -> list[dict[str, Any]]:
    projects = accessible_projects(session, user)
    result = []
    for project in projects:
        member_count = session.scalar(select(func.count(ProjectMember.id)).where(ProjectMember.project_id == project.id)) or 0
        result.append(serialize_project(project, member_count))
    session.commit()
    return result


@app.post("/api/projects")
def create_project(
    payload: ProjectCreate,
    user: CurrentUser = Depends(require_field_admin),
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    account = sync_platform_account(session, user)
    code = payload.project_code.strip().upper()
    if session.scalar(select(ResearchProject.id).where(ResearchProject.project_code == code)):
        raise HTTPException(409, "课题编号已存在。")
    project = ResearchProject(
        project_code=code,
        project_name=payload.project_name.strip(),
        description=payload.description.strip() or None,
        institution_id=account.institution_id,
        status="active",
        created_by=audit_actor(user),
    )
    session.add(project)
    session.flush()
    # Public knowledge folders are protected by project-aware RLS.  A newly
    # created project cannot come from the request header yet, so establish
    # the verified field-admin and project context before seeding its folders.
    _set_knowledge_context(session, user)
    _set_active_project(session, project.id)
    seed_public_knowledge_folders(session, project.id)
    record_permission_audit(
        session,
        user,
        "project_created",
        "research_project",
        project.id,
        project_id=project.id,
        after={"project_code": project.project_code, "project_name": project.project_name, "status": project.status},
    )
    result = serialize_project(project, 0)
    session.commit()
    return result


@app.patch("/api/projects/{project_id}")
def update_project(
    project_id: str,
    payload: ProjectUpdate,
    user: CurrentUser = Depends(require_field_admin),
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    account = sync_platform_account(session, user)
    project = session.get(ResearchProject, project_id)
    if not project or project.institution_id != account.institution_id:
        raise HTTPException(404, "课题不存在。")
    if project.id == DEFAULT_PROJECT_ID and payload.status == "archived":
        raise HTTPException(409, "海南南繁默认课题不能停用。")
    before = {"project_name": project.project_name, "description": project.description or "", "status": project.status}
    if payload.project_name is not None:
        project.project_name = payload.project_name.strip()
    if payload.description is not None:
        project.description = payload.description.strip() or None
    if payload.status is not None:
        project.status = payload.status
    after = {"project_name": project.project_name, "description": project.description or "", "status": project.status}
    record_permission_audit(session, user, "project_updated", "research_project", project.id, project_id=project.id, before=before, after=after)
    session.flush()
    result = serialize_project(project, session.scalar(select(func.count(ProjectMember.id)).where(ProjectMember.project_id == project.id)) or 0)
    session.commit()
    return result


@app.get("/api/accounts")
def list_platform_accounts(
    user: CurrentUser = Depends(require_field_admin),
    session: Session = Depends(get_session),
) -> list[dict[str, Any]]:
    current_account = sync_platform_account(session, user)
    accounts = session.scalars(select(PlatformAccount).where(
        PlatformAccount.institution_id == current_account.institution_id
    ).order_by(PlatformAccount.business_role, PlatformAccount.username)).all()
    session.commit()
    return [{
        "username": account.username,
        "display_name": account.display_name,
        "business_role": account.business_role,
        "active": account.active,
        "identity_bound": bool(account.keycloak_subject),
        "last_login_at": account.last_login_at.isoformat() if account.last_login_at else None,
    } for account in accounts]


@app.patch("/api/accounts/{username}")
def update_platform_account(
    username: str,
    payload: PlatformAccountUpdate,
    user: CurrentUser = Depends(require_field_admin),
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    current_account = sync_platform_account(session, user)
    account = session.get(PlatformAccount, username)
    if not account or account.institution_id != current_account.institution_id:
        raise HTTPException(404, "当前机构账号目录中不存在该账号。")
    if username == user.username and not payload.active:
        raise HTTPException(409, "不能停用当前登录的字段管理员账号。")
    before = {"active": account.active, "business_role": account.business_role}
    account.active = payload.active
    record_permission_audit(
        session,
        user,
        "account_activated" if account.active else "account_deactivated",
        "platform_account",
        account.username,
        before=before,
        after={"active": account.active, "business_role": account.business_role},
    )
    result = {
        "username": account.username,
        "display_name": account.display_name,
        "business_role": account.business_role,
        "active": account.active,
        "identity_bound": bool(account.keycloak_subject),
    }
    session.commit()
    return result


@app.get("/api/projects/{project_id}/members")
def list_project_members(
    project_id: str,
    user: CurrentUser = Depends(require_field_admin),
    session: Session = Depends(get_session),
) -> list[dict[str, Any]]:
    account = sync_platform_account(session, user)
    project = session.get(ResearchProject, project_id)
    if not project or project.institution_id != account.institution_id:
        raise HTTPException(404, "课题不存在。")
    rows = session.execute(
        select(ProjectMember, PlatformAccount)
        .join(PlatformAccount, PlatformAccount.username == ProjectMember.username)
        .where(ProjectMember.project_id == project_id)
        .order_by(PlatformAccount.display_name)
    ).all()
    session.commit()
    return [{
        "id": member.id,
        "username": account.username,
        "display_name": account.display_name,
        "business_role": account.business_role,
        "member_role": member.member_role,
        "created_by": member.created_by,
        "created_at": member.created_at.isoformat(),
    } for member, account in rows]


@app.put("/api/projects/{project_id}/members/{username}")
def upsert_project_member(
    project_id: str,
    username: str,
    payload: ProjectMemberUpdate,
    user: CurrentUser = Depends(require_field_admin),
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    current_account = sync_platform_account(session, user)
    project = session.get(ResearchProject, project_id)
    account = session.get(PlatformAccount, username)
    if not project or project.institution_id != current_account.institution_id:
        raise HTTPException(404, "课题不存在。")
    if not account or account.institution_id != current_account.institution_id:
        raise HTTPException(404, "当前机构账号目录中不存在该账号。")
    if account.business_role != "researcher":
        raise HTTPException(422, "只有科研人员需要配置课题成员关系；数据处理员和字段管理员按岗位访问课题。")
    member = session.scalar(select(ProjectMember).where(
        ProjectMember.project_id == project_id,
        ProjectMember.username == username,
    ))
    created = member is None
    if not member:
        member = ProjectMember(
            project_id=project_id,
            username=username,
            member_role=payload.member_role,
            created_by=audit_actor(user),
        )
        session.add(member)
        session.flush()
    else:
        member.member_role = payload.member_role
    record_permission_audit(
        session,
        user,
        "project_member_added" if created else "project_member_updated",
        "project_member",
        member.id,
        project_id=project_id,
        after={"username": username, "member_role": member.member_role},
    )
    result = {"id": member.id, "project_id": project_id, "username": username, "member_role": member.member_role}
    session.commit()
    return result


@app.delete("/api/projects/{project_id}/members/{username}")
def remove_project_member(
    project_id: str,
    username: str,
    user: CurrentUser = Depends(require_field_admin),
    session: Session = Depends(get_session),
) -> dict[str, bool]:
    account = sync_platform_account(session, user)
    project = session.get(ResearchProject, project_id)
    if not project or project.institution_id != account.institution_id:
        raise HTTPException(404, "课题不存在。")
    member = session.scalar(select(ProjectMember).where(
        ProjectMember.project_id == project_id,
        ProjectMember.username == username,
    ))
    if not member:
        raise HTTPException(404, "课题成员关系不存在。")
    member_id = member.id
    record_permission_audit(
        session,
        user,
        "project_member_removed",
        "project_member",
        member_id,
        project_id=project_id,
        before={"username": username, "member_role": member.member_role},
    )
    session.delete(member)
    session.commit()
    return {"deleted": True}


@app.get("/api/permission-audits")
def list_permission_audits(
    project_id: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
    user: CurrentUser = Depends(require_field_admin),
    session: Session = Depends(get_session),
) -> list[dict[str, Any]]:
    account = sync_platform_account(session, user)
    statement = select(PermissionAudit).where(PermissionAudit.institution_id == account.institution_id)
    if project_id:
        project = session.get(ResearchProject, project_id)
        if not project or project.institution_id != account.institution_id:
            raise HTTPException(404, "课题不存在。")
        statement = statement.where(PermissionAudit.project_id == project_id)
    rows = session.scalars(statement.order_by(PermissionAudit.created_at.desc()).limit(limit)).all()
    session.commit()
    return [{
        "id": item.id,
        "actor_name": item.actor_name,
        "action": item.action,
        "target_type": item.target_type,
        "target_id": item.target_id,
        "project_id": item.project_id,
        "before": item.before_state,
        "after": item.after_state,
        "created_at": item.created_at.isoformat(),
    } for item in rows]


@app.get("/api/trial-packages")
def get_trial_packages(
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_data_processor),
    session: Session = Depends(get_session),
) -> list[dict[str, Any]]:
    """List regional-trial packages visible to the data processor."""
    project = resolve_project_access(session, user, x_project_id)
    return list_trial_import_batches(session, project.id)


@app.get("/api/trial-packages/{batch_id}")
def get_trial_package(
    batch_id: str,
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_data_processor),
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    try:
        project = resolve_project_access(session, user, x_project_id)
        return get_trial_import_batch(session, batch_id, project.id)
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.post("/api/trial-packages/upload")
async def upload_trial_package_endpoint(
    file: UploadFile = File(...),
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_data_processor),
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    try:
        project = resolve_project_access(session, user, x_project_id)
        return upload_trial_package(
            session,
            file.filename or "regional-trial-package.zip",
            await file.read(),
            audit_actor(user),
            RAW_STORAGE_DIR,
            project.id,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/trial-packages/{batch_id}/publish")
def publish_trial_package_endpoint(
    batch_id: str,
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_data_processor),
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    try:
        project = resolve_project_access(session, user, x_project_id)
        result = publish_trial_package(session, batch_id, audit_actor(user), project.id)
        dossier_count = seed_mock_breeding_dossiers(session, project.id)
        session.commit()
        if dossier_count:
            result["simulated_breeding_dossiers"] = dossier_count
        return result
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@app.get("/api/research/me")
def research_current_user(user: CurrentUser = Depends(require_researcher)) -> dict[str, Any]:
    """Return the verified Keycloak identity used by the research assistant."""
    return {
        "id": user.id,
        "username": user.username,
        "display_name": user.display_name,
        "roles": sorted(user.roles),
    }


@app.get("/api/research/skills")
def list_research_skills(
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    """Return the controlled platform-skill catalog and this user's audit trail."""
    audits = session.scalars(
        select(ResearchAudit)
        .where(ResearchAudit.action == "research_skill_opened")
        .order_by(ResearchAudit.created_at.desc())
        .limit(30)
    ).all()
    last_opened: dict[str, datetime] = {}
    recent_runs: list[dict[str, Any]] = []
    for audit in audits:
        metadata = audit.audit_metadata or {}
        skill_code = str(metadata.get("skill_code") or "")
        if skill_code and skill_code not in last_opened:
            last_opened[skill_code] = audit.created_at
        skill = _research_skill_by_code(skill_code)
        recent_runs.append({
            "id": audit.id,
            "skill_code": skill_code,
            "skill_name": skill["name"] if skill else skill_code or "未识别技能",
            "action": "打开技能工作台",
            "workspace": metadata.get("workspace") or "",
            "created_at": audit.created_at.isoformat(),
        })
    return {
        "skills": [_serialize_research_skill(skill, last_opened.get(skill["code"])) for skill in RESEARCH_SKILL_CATALOG],
        "recent_runs": recent_runs,
    }


@app.post("/api/research/skills/{skill_code}/launch")
def launch_research_skill(
    skill_code: str,
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    """Audit a controlled skill launch before sending the user to its workspace."""
    skill = _research_skill_by_code(skill_code)
    if not skill:
        raise HTTPException(404, "未找到该科研技能。")
    if skill["status"] != "available":
        raise HTTPException(409, "该技能尚在规划中，当前版本未接入执行能力。")
    project = resolve_project_access(session, user, x_project_id)
    audit = ResearchAudit(
        owner_id=user.id,
        project_id=project.id,
        session_id=None,
        action="research_skill_opened",
        audit_metadata={
            "skill_code": skill["code"],
            "skill_name": skill["name"],
            "workspace": skill["workspace"],
            "skill_status": skill["status"],
        },
    )
    session.add(audit)
    session.flush()
    session.commit()
    return {
        "skill": _serialize_research_skill(skill, audit.created_at),
        "audit": {
            "id": audit.id,
            "action": "打开技能工作台",
            "created_at": audit.created_at.isoformat(),
        },
    }


@app.get("/api/gwas/plans")
def list_gwas_analysis_plans(
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> list[dict[str, Any]]:
    """List only the authenticated researcher's controlled GWAS plans."""
    return list_gwas_plans(session)


@app.get("/api/genotype-assets")
def list_private_genotype_assets(
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> list[dict[str, Any]]:
    """Only list genotype assets owned by the authenticated researcher."""
    return list_genotype_assets(session)


@app.get("/api/genotype-assets/analysis-ready")
def list_analysis_ready_genotype_assets(
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> list[dict[str, Any]]:
    return analysis_ready_versions(session)


@app.get("/api/genotype-assets/materials")
def search_genotype_material_master(
    keyword: str = Query(default="", max_length=200),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> list[dict[str, Any]]:
    return list_genotype_material_suggestions(session, keyword)


@app.post("/api/genotype-assets")
def create_private_genotype_asset(
    payload: CreateGenotypeAssetRequest,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        result = create_genotype_asset(session, user.id, payload, active_project_id(session))
        session.commit()
        return result
    except GenotypeAssetError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/genotype-assets/{asset_id}/uploads")
def initialize_genotype_asset_upload(
    asset_id: str,
    payload: UploadInitRequest,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        result = create_genotype_upload_session(session, user.id, asset_id, payload, RESEARCH_STORAGE_DIR)
        session.commit()
        return result
    except GenotypeAssetError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.get("/api/genotype-assets/{asset_id}/uploads/{upload_id}")
def get_private_genotype_upload(
    asset_id: str,
    upload_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        return get_genotype_upload_session(session, asset_id, upload_id)
    except GenotypeAssetError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.put("/api/genotype-assets/{asset_id}/uploads/{upload_id}/chunks/{chunk_index}")
async def upload_genotype_asset_chunk(
    asset_id: str,
    upload_id: str,
    chunk_index: int,
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    if chunk_index < 0:
        raise HTTPException(400, "上传分片编号不能为负数。")
    try:
        destination = genotype_upload_chunk_path(RESEARCH_STORAGE_DIR, user.id, asset_id, upload_id, chunk_index)
        written = 0
        with destination.open("wb") as sink:
            while block := await file.read(1024 * 1024):
                written += len(block)
                if written > 10 * 1024 * 1024:
                    destination.unlink(missing_ok=True)
                    raise HTTPException(400, "单个上传分片超过 10 MB，请按客户端分片大小重新上传。")
                sink.write(block)
        record_genotype_upload_chunk(session, upload_id, chunk_index)
        session.commit()
        return {"uploaded": True, "chunk_index": chunk_index, "size_bytes": written}
    except GenotypeAssetError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/genotype-assets/{asset_id}/uploads/{upload_id}/complete")
def complete_genotype_asset_upload(
    asset_id: str,
    upload_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        result = complete_genotype_upload(session, user.id, asset_id, upload_id, RESEARCH_STORAGE_DIR)
        session.commit()
        return result
    except GenotypeAssetError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.get("/api/genotype-assets/{asset_id}")
def get_private_genotype_asset(
    asset_id: str,
    version_id: str | None = Query(default=None),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        return get_genotype_asset_version(session, asset_id, version_id)
    except GenotypeAssetError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.patch("/api/genotype-assets/{asset_id}/versions/{version_id}/mappings/{fid}/{iid}")
def save_genotype_sample_mapping(
    asset_id: str,
    version_id: str,
    fid: str,
    iid: str,
    payload: MappingUpdateRequest,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        get_genotype_asset_version(session, asset_id, version_id)
        result = update_genotype_mapping(session, user.id, version_id, fid, iid, payload)
        session.commit()
        return result
    except GenotypeAssetError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/genotype-assets/{asset_id}/versions/{version_id}/mappings/import")
async def import_genotype_sample_mapping(
    asset_id: str,
    version_id: str,
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        get_genotype_asset_version(session, asset_id, version_id)
        result = batch_update_genotype_mapping(session, user.id, version_id, await file.read())
        session.commit()
        return result
    except GenotypeAssetError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/genotype-assets/{asset_id}/versions/{version_id}/mapping-revision")
def create_genotype_mapping_revision_endpoint(
    asset_id: str,
    version_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        result = create_genotype_mapping_revision(session, user.id, asset_id, version_id)
        session.commit()
        return result
    except GenotypeAssetError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/genotype-assets/{asset_id}/versions/{version_id}/publish")
def publish_genotype_analysis_version(
    asset_id: str,
    version_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        get_genotype_asset_version(session, asset_id, version_id)
        result = publish_genotype_analysis_ready(session, user.id, version_id)
        session.commit()
        return result
    except GenotypeAssetError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/genotype-assets/{asset_id}/versions/{version_id}/governance-requests")
def submit_genotype_governance_request(
    asset_id: str,
    version_id: str,
    payload: GovernanceRequestCreate,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        get_genotype_asset_version(session, asset_id, version_id)
        result = create_genotype_governance_request(session, user.id, asset_id, version_id, payload)
        session.commit()
        return result
    except GenotypeAssetError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.get("/api/genotype-governance-requests")
def get_genotype_governance_requests(
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_data_processor),
    session: Session = Depends(get_session),
) -> list[dict[str, Any]]:
    """Metadata-only queue for data processors; raw genotype files stay private."""
    project = resolve_project_access(session, user, x_project_id)
    _set_active_project(session, project.id)
    session.execute(text("SELECT set_config('app.genotype_governance_processor', 'true', true)"))
    return list_genotype_governance_requests(session)


@app.patch("/api/genotype-governance-requests/{request_id}")
def resolve_genotype_governance_request_endpoint(
    request_id: str,
    payload: GovernanceRequestResolution,
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_data_processor),
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    try:
        project = resolve_project_access(session, user, x_project_id)
        _set_active_project(session, project.id)
        session.execute(text("SELECT set_config('app.genotype_governance_processor', 'true', true)"))
        result = resolve_genotype_governance_request(session, request_id, audit_actor(user), payload)
        session.commit()
        return result
    except GenotypeAssetError as exc:
        session.rollback()
        raise HTTPException(404, str(exc)) from exc


@app.get("/api/genotype-assets/{asset_id}/versions/{version_id}/phenotype-template")
def download_genotype_phenotype_template(
    asset_id: str,
    version_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> Response:
    try:
        get_genotype_asset_version(session, asset_id, version_id)
        content = build_genotype_phenotype_template(session, version_id)
        filename = quote(f"{asset_id}-v{version_id[:8]}-连续性状表型模板.xlsx")
        return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})
    except GenotypeAssetError as exc:
        raise HTTPException(400, str(exc)) from exc


@app.get("/api/genotype-assets/{asset_id}/versions/{version_id}/mapping-template")
def download_genotype_mapping_template(
    asset_id: str,
    version_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> Response:
    try:
        get_genotype_asset_version(session, asset_id, version_id)
        content = build_genotype_mapping_template(session, version_id)
        filename = quote(f"{asset_id}-v{version_id[:8]}-样本材料映射模板.csv")
        return Response(content=content, media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})
    except GenotypeAssetError as exc:
        raise HTTPException(400, str(exc)) from exc


@app.get("/api/genotype-assets/{asset_id}/versions/{version_id}/artifacts/{kind}")
def download_genotype_qc_artifact(
    asset_id: str,
    version_id: str,
    kind: Literal["report", "package"],
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> FileResponse:
    try:
        get_genotype_asset_version(session, asset_id, version_id)
        path, media_type = genotype_artifact_path(session, version_id, kind)
        return FileResponse(path, media_type=media_type, filename=path.name)
    except GenotypeAssetError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.post("/api/gwas/plans")
def create_gwas_analysis_plan(
    payload: CreateGwasPlanRequest,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        result = create_gwas_plan(session, user.id, payload, active_project_id(session))
        session.commit()
        return result
    except GenomicsError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.get("/api/gwas/plans/{plan_id}")
def get_gwas_analysis_plan(
    plan_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        return get_gwas_plan(session, plan_id)
    except GenomicsError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.post("/api/gwas/plans/{plan_id}/genotype-asset")
def attach_genotype_asset_to_gwas_plan(
    plan_id: str,
    payload: AttachGenotypeAssetRequest,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    """Attach one immutable, analysis-ready genotype version to a collecting GWAS plan."""
    try:
        result = attach_analysis_ready_genotype(
            session,
            user.id,
            plan_id,
            payload.asset_id,
            payload.version_id,
            RESEARCH_STORAGE_DIR,
        )
        session.commit()
        return result
    except (GenomicsError, GenotypeAssetError) as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/gwas/plans/{plan_id}/genotype")
async def upload_gwas_genotype_package(
    plan_id: str,
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        result = upload_gwas_genotype(session, user.id, plan_id, file.filename or "genotype.zip", await file.read(), RESEARCH_STORAGE_DIR)
        session.commit()
        return result
    except GenomicsError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/gwas/plans/{plan_id}/phenotype")
async def upload_gwas_phenotype_table(
    plan_id: str,
    trait_column: str = Query(..., min_length=1, max_length=120),
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        result = upload_gwas_phenotype(session, user.id, plan_id, file.filename or "phenotype.csv", trait_column, await file.read(), RESEARCH_STORAGE_DIR)
        session.commit()
        return result
    except GenomicsError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/gwas/plans/{plan_id}/covariates")
async def upload_gwas_covariate_table(
    plan_id: str,
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        result = upload_gwas_covariates(session, user.id, plan_id, file.filename or "covariates.csv", await file.read(), RESEARCH_STORAGE_DIR)
        session.commit()
        return result
    except GenomicsError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/gwas/plans/{plan_id}/confirm")
def confirm_gwas_analysis_plan(
    plan_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        result = confirm_gwas_plan(session, user.id, plan_id)
        session.commit()
        return result
    except GenomicsError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/gwas/plans/{plan_id}/run")
def queue_gwas_analysis_plan(
    plan_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    try:
        request_gwas_execution(session, plan_id)
        session.commit()
        result = run_requested_local_execution(session, plan_id, RESEARCH_STORAGE_DIR)
        if result["status"] == "completed":
            content, file_name, metadata = build_local_gwas_result_bundle(session, plan_id, RESEARCH_STORAGE_DIR)
            project_id = session.execute(text("SELECT project_id FROM gwas_analysis_plan WHERE id = :id"), {"id": plan_id}).scalar_one()
            saved = _store_gwas_result_bundle(
                session,
                owner_id=user.id,
                project_id=str(project_id),
                plan_id=plan_id,
                content=content,
                file_name=file_name,
                metadata=metadata,
            )
            result["result_library_item"] = serialize_research_result(saved)
            result["execution_note"] = "本地 GWAS 已完成，七个输出已打包并保存到结果库的“GWAS 分析结果”分类。"
        session.commit()
        return result
    except GenomicsError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.get("/api/gwas/plans/{plan_id}/results/{file_key}")
def download_local_gwas_result(
    plan_id: str,
    file_key: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> FileResponse:
    try:
        path, media_type = get_local_gwas_result_file(session, plan_id, file_key, RESEARCH_STORAGE_DIR)
        return FileResponse(path, media_type=media_type, filename=path.name)
    except GenomicsError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.post("/api/gwas/plans/{plan_id}/archive")
def archive_completed_gwas_result(
    plan_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    """Archive an already-completed plan, including runs completed before auto-archive existed."""
    try:
        content, file_name, metadata = build_local_gwas_result_bundle(session, plan_id, RESEARCH_STORAGE_DIR)
        project_id = session.execute(text("SELECT project_id FROM gwas_analysis_plan WHERE id = :id"), {"id": plan_id}).scalar_one()
        saved = _store_gwas_result_bundle(
            session,
            owner_id=user.id,
            project_id=str(project_id),
            plan_id=plan_id,
            content=content,
            file_name=file_name,
            metadata=metadata,
        )
        session.commit()
        return {"plan": get_gwas_plan(session, plan_id), "result_library_item": serialize_research_result(saved)}
    except GenomicsError as exc:
        session.rollback()
        raise HTTPException(400, str(exc)) from exc


@app.get("/api/research/query-templates")
def research_query_templates(user: CurrentUser = Depends(require_researcher)) -> list[dict[str, Any]]:
    """Expose supported templates for audit and future controlled tool configuration."""
    return template_catalog()


@app.get("/api/research/standard-fields")
def research_standard_fields(user: CurrentUser = Depends(require_researcher)) -> dict[str, Any]:
    """Return every field available in the two public structured-query templates."""
    return public_standard_field_catalog()


@app.get("/api/research/published-data/varieties")
def research_published_variety_options(
    q: str = Query(default="", max_length=100),
    scope: Literal["rice_phenotype", "root_phenotype"] = Query(default="rice_phenotype"),
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> list[dict[str, Any]]:
    """Search only the published varieties that belong to the active standard template."""
    project = resolve_project_access(session, user, x_project_id)
    keyword = q.strip()
    if scope == "root_phenotype":
        observation_filter = """
            EXISTS (
                SELECT 1
                FROM root_phenotype_observation observation
                WHERE observation.variety_id = v.id
                  AND observation.project_id = :project_id
            )
        """
    else:
        observation_filter = """
            EXISTS (
                SELECT 1
                FROM phenotype_observation observation
                WHERE observation.variety_id = v.id
                  AND observation.project_id = :project_id
                  AND observation.publish_status = 'published'
            )
        """
    rows = session.execute(text(f"""
        SELECT v.id, v.variety_name, v.alias_names, v.variety_type, v.approval_number
        FROM variety_basic v
        WHERE v.data_status = 'published'
          AND v.project_id = :project_id
          AND {observation_filter}
          AND (
              :keyword = ''
              OR v.variety_name ILIKE :pattern
              OR v.normalized_name ILIKE :pattern
              OR COALESCE(v.alias_names::text, '') ILIKE :pattern
          )
        ORDER BY
            CASE WHEN :keyword <> '' AND v.variety_name ILIKE :exact THEN 0 ELSE 1 END,
            v.variety_name
        LIMIT 20
    """), {
        "keyword": keyword,
        "project_id": project.id,
        "pattern": f"%{keyword}%",
        "exact": keyword,
    }).mappings().all()
    return [{
        "id": str(row["id"]),
        "variety_name": row["variety_name"],
        "aliases": row["alias_names"] or [],
        "variety_type": row["variety_type"],
        "approval_number": row["approval_number"],
    } for row in rows]


@app.post("/api/research/published-data/query")
def research_published_data_query(
    payload: ResearchStructuredQueryRequest,
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    """Run a user-built query plan through the same guarded templates as the agent."""
    project = resolve_project_access(session, user, x_project_id)
    normalized_names = [name.strip() for name in payload.variety_names if name.strip()]
    request = StructuredQueryRequest(
        query_needed=True,
        scope=payload.scope,
        variety_names=normalized_names,
        trait_codes=payload.trait_codes,
        filters=payload.filters,
    )
    plan, unresolved_names = plan_query_from_structured_request(
        session,
        request,
        TRAITS,
        ROOT_TRAITS,
        project.id,
    )
    if not plan:
        if unresolved_names:
            raise HTTPException(404, f"未在已发布标准数据中找到：{'、'.join(unresolved_names)}")
        raise HTTPException(422, "请至少选择一个标准字段、填写一个品种/材料名称，或添加一个数值筛选条件。")
    plan = plan.model_copy(update={"limit": payload.limit})
    execution = execute_published_data_query(session, plan, project.id)
    execution.unresolved_variety_names = unresolved_names
    return serialize_public_query_execution(execution)


@app.get("/api/research/sessions")
def list_research_sessions(
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> list[dict[str, Any]]:
    project = resolve_project_access(session, user, x_project_id)
    items = session.scalars(
        select(ResearchSession)
        .where(ResearchSession.owner_id == user.id, ResearchSession.project_id == project.id)
        .order_by(ResearchSession.updated_at.desc())
    ).all()
    return [serialize_research_session(item) for item in items]


@app.post("/api/research/sessions")
def create_research_session(
    payload: ResearchSessionCreate,
    x_project_id: str | None = Header(default=None, alias="X-Project-Id"),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    project = resolve_project_access(session, user, x_project_id)
    item = ResearchSession(owner_id=user.id, project_id=project.id, title=payload.title.strip())
    session.add(item)
    session.flush()
    result = serialize_research_session(item)
    session.commit()
    return result


def get_owned_research_session(session: Session, session_id: str) -> ResearchSession:
    item = session.get(ResearchSession, session_id)
    if not item:
        raise HTTPException(404, "未找到该会话，或当前账号无权访问。")
    username = str(session.info.get("research_username") or "")
    roles = set(session.info.get("research_roles") or ())
    if username and "researcher" in roles and not {"field_admin", "data_processor"}.intersection(roles):
        membership = session.scalar(select(ProjectMember.id).where(
            ProjectMember.project_id == item.project_id,
            ProjectMember.username == username,
        ))
        if not membership:
            raise HTTPException(404, "未找到该会话，或当前账号已无权访问所属课题。")
    return item


@app.patch("/api/research/sessions/{research_session_id}")
def rename_research_session(
    research_session_id: str,
    payload: ResearchSessionRename,
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    item = get_owned_research_session(session, research_session_id)
    item.title = payload.title.strip()
    item.updated_at = datetime.now(timezone.utc)
    session.flush()
    result = serialize_research_session(item)
    session.commit()
    return result


@app.delete("/api/research/sessions/{research_session_id}")
def delete_research_session(
    research_session_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, bool]:
    item = get_owned_research_session(session, research_session_id)
    attachments = session.scalars(
        select(ResearchAttachment).where(ResearchAttachment.session_id == item.id)
    ).all()
    paths = [Path(attachment.storage_path) for attachment in attachments]
    for attachment in attachments:
        session.delete(attachment)
    for message in session.scalars(
        select(ResearchMessage).where(ResearchMessage.session_id == item.id)
    ).all():
        session.delete(message)
    for audit in session.scalars(
        select(ResearchAudit).where(ResearchAudit.session_id == item.id, ResearchAudit.owner_id == user.id)
    ).all():
        session.delete(audit)
    session.delete(item)
    session.commit()
    for path in paths:
        try:
            path.unlink(missing_ok=True)
        except OSError:
            # The database record is gone. A future local maintenance job can
            # remove an orphaned file without retaining research content.
            pass
    session_dir = RESEARCH_STORAGE_DIR / item.project_id / user.id / research_session_id
    try:
        session_dir.rmdir()
    except OSError:
        pass
    return {"deleted": True}


@app.get("/api/research/sessions/{research_session_id}/messages")
def list_research_messages(
    research_session_id: str,
    session: Session = Depends(get_research_session),
) -> list[dict[str, Any]]:
    research_session = get_owned_research_session(session, research_session_id)
    items = session.scalars(
        select(ResearchMessage)
        .where(ResearchMessage.session_id == research_session_id)
        .order_by(ResearchMessage.created_at)
    ).all()
    return [serialize_research_message(item) for item in items]


@app.get("/api/research/sessions/{research_session_id}/attachments")
def list_research_attachments(
    research_session_id: str,
    session: Session = Depends(get_research_session),
) -> list[dict[str, Any]]:
    research_session = get_owned_research_session(session, research_session_id)
    items = session.scalars(
        select(ResearchAttachment)
        .where(ResearchAttachment.session_id == research_session_id)
        .order_by(ResearchAttachment.created_at)
    ).all()
    return [serialize_research_attachment(item) for item in items]


@app.post("/api/research/sessions/{research_session_id}/attachments")
async def upload_research_attachment(
    research_session_id: str,
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    research_session = get_owned_research_session(session, research_session_id)
    original_name = (file.filename or "附件").strip()
    suffix = Path(original_name).suffix.lower()
    if suffix not in DOCUMENT_SUPPORTED_SUFFIXES and suffix not in VISION_IMAGE_SUFFIXES:
        raise HTTPException(415, "暂不支持该附件格式。请上传 PDF、Office 文档、表格、图片或文本文件。")
    content = await file.read(MAX_RESEARCH_ATTACHMENT_BYTES + 1)
    if len(content) > MAX_RESEARCH_ATTACHMENT_BYTES:
        raise HTTPException(413, "单个附件不能超过 10 MB，请拆分后重新上传。")
    if not content:
        raise HTTPException(400, "上传的附件为空。")

    safe_name = re.sub(r"[^\w.\-()（）]+", "_", original_name).strip("._") or f"attachment{suffix}"
    attachment_id = str(uuid.uuid4())
    directory = RESEARCH_STORAGE_DIR / research_session.project_id / user.id / research_session_id
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / f"{attachment_id}_{safe_name}"
    path.write_bytes(content)

    if suffix in VISION_IMAGE_SUFFIXES:
        attachment = ResearchAttachment(
            id=attachment_id,
            session_id=research_session_id,
            project_id=research_session.project_id,
            owner_id=user.id,
            file_name=original_name,
            content_type=file.content_type,
            size_bytes=len(content),
            storage_path=str(path),
            parser_name="native_image",
            parsing_status="image_ready",
            parser_warnings=["图片不进行本地 Docling 或 OCR 文字解析；提问时将原图直接提交给当前配置的大模型进行多模态分析。"],
        )
    else:
        try:
            parsed = await parse_local_document(path)
            attachment = ResearchAttachment(
                id=attachment_id,
                session_id=research_session_id,
                project_id=research_session.project_id,
                owner_id=user.id,
                file_name=original_name,
                content_type=file.content_type,
                size_bytes=len(content),
                storage_path=str(path),
                parser_name=parsed.parser,
                parsing_status="parsed",
                parsed_markdown=parsed.markdown,
                parsed_metadata=parsed.metadata,
                parser_warnings=parsed.warnings,
            )
        except Exception as exc:
            attachment = ResearchAttachment(
                id=attachment_id,
                session_id=research_session_id,
                project_id=research_session.project_id,
                owner_id=user.id,
                file_name=original_name,
                content_type=file.content_type,
                size_bytes=len(content),
                storage_path=str(path),
                parsing_status="failed",
                parser_warnings=[str(exc)[:500]],
            )
    session.add(attachment)
    session.add(ResearchAudit(
        owner_id=user.id,
        project_id=research_session.project_id,
        session_id=research_session_id,
        action="attachment_uploaded",
        audit_metadata={"file_name": original_name, "size_bytes": len(content), "status": attachment.parsing_status},
    ))
    session.flush()
    result = serialize_research_attachment(attachment)
    session.commit()
    return result


@app.get("/api/research/attachments/{attachment_id}/preview")
def preview_research_attachment(
    attachment_id: str,
    session: Session = Depends(get_research_session),
) -> dict[str, Any]:
    attachment = session.get(ResearchAttachment, attachment_id)
    if not attachment:
        raise HTTPException(404, "未找到该私有附件，或当前账号无权访问。")
    return serialize_research_attachment(attachment, include_preview=True)


@app.get("/api/research/attachments/{attachment_id}/image")
def read_research_attachment_image(
    attachment_id: str,
    session: Session = Depends(get_research_session),
) -> FileResponse:
    """Return an authorized private image for inline assistant preview only."""
    attachment = session.get(ResearchAttachment, attachment_id)
    if not attachment:
        raise HTTPException(404, "未找到该私有图片，或当前账号无权访问。")

    suffix = Path(attachment.file_name).suffix.lower()
    mime_by_suffix = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }
    if suffix not in mime_by_suffix:
        raise HTTPException(415, "该附件不是可预览的图片。")

    path = Path(attachment.storage_path)
    if not path.is_file():
        raise HTTPException(409, "图片原文件已不存在，请重新上传。")
    media_type = attachment.content_type if attachment.content_type in mime_by_suffix.values() else mime_by_suffix[suffix]
    return FileResponse(path, media_type=media_type, headers={"Content-Disposition": "inline"})


@app.delete("/api/research/attachments/{attachment_id}")
def delete_research_attachment(
    attachment_id: str,
    session: Session = Depends(get_research_session),
) -> dict[str, bool]:
    attachment = session.get(ResearchAttachment, attachment_id)
    if not attachment:
        raise HTTPException(404, "未找到该私有附件，或当前账号无权访问。")
    path = Path(attachment.storage_path)
    session.delete(attachment)
    session.commit()
    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass
    return {"deleted": True}


def knowledge_folder_for_document(session: Session, document: KnowledgeDocument) -> KnowledgeFolder | None:
    return session.get(KnowledgeFolder, document.folder_id) if document.folder_id else None


def get_visible_knowledge_document(session: Session, document_id: str) -> KnowledgeDocument:
    document = session.get(KnowledgeDocument, document_id)
    if not document or document.project_id != active_project_id(session):
        raise HTTPException(404, "未找到该知识库资料，或当前账号无权访问。")
    return document


def validate_knowledge_folder(
    session: Session,
    *,
    folder_id: str | None,
    scope: str,
    user: CurrentUser,
) -> KnowledgeFolder:
    if not folder_id and scope == "private":
        return ensure_private_uncategorized_folder(session, user)
    if not folder_id:
        raise HTTPException(422, "公共知识库资料必须选择一个资料分类。")
    folder = session.get(KnowledgeFolder, folder_id)
    if not folder or folder.scope != scope or folder.project_id != active_project_id(session):
        raise HTTPException(422, "所选资料夹不存在，或不属于当前知识库范围。")
    if scope == "private" and folder.owner_id != user.id:
        raise HTTPException(403, "不能将资料保存到其他科研人员的私人文件夹。")
    return folder


def safe_knowledge_filename(name: str) -> str:
    suffix = Path(name).suffix.lower()
    stem = re.sub(r"[^\w.\-()（）]+", "_", Path(name).stem).strip("._") or "knowledge_document"
    return f"{stem}{suffix}"


def delete_knowledge_file(path: str) -> None:
    try:
        Path(path).unlink(missing_ok=True)
    except OSError:
        pass


def process_knowledge_document(document_id: str) -> None:
    """Parse and index an authorized on-host file in a background task."""
    queued_at = time.monotonic()
    logger.info(
        "Knowledge document queued: document_id=%s configured_concurrency=%s",
        document_id,
        KNOWLEDGE_DOCUMENT_CONCURRENCY,
    )
    # MinerU executes one local CLI job at a time, while Docling and PaddleOCR
    # are CPU-intensive.  Start only the number of jobs the process has
    # explicitly been configured to sustain; by default the work is serial.
    with KNOWLEDGE_DOCUMENT_SEMAPHORE:
        queue_wait_seconds = time.monotonic() - queued_at
        with MigrationSessionLocal() as session:
            document = session.get(KnowledgeDocument, document_id)
            if not document or document.status in {"withdrawn", "superseded", "deleted"}:
                logger.info("Knowledge document skipped: document_id=%s", document_id)
                return

            file_name = document.original_file_name
            document.parsing_status = "processing"
            document.indexing_status = "pending"
            session.commit()
            logger.info(
                "Knowledge document processing started: document_id=%s file=%s size_bytes=%s "
                "queue_wait_seconds=%.1f",
                document.id,
                file_name,
                document.size_bytes,
                queue_wait_seconds,
            )
            total_started_at = time.monotonic()
            try:
                parse_started_at = time.monotonic()
                parsed = asyncio.run(parse_local_document(Path(document.storage_path)))
                parse_elapsed_seconds = time.monotonic() - parse_started_at
                chunks = split_markdown(parsed.markdown)
                session.execute(
                    text("DELETE FROM knowledge_chunk WHERE document_id = :document_id"),
                    {"document_id": document.id},
                )
                document.parser_name = parsed.parser
                document.parsed_characters = len(parsed.markdown.strip())
                document.parser_warnings = parsed.warnings
                document.parsing_status = "parsed"
                document.indexing_status = "processing"
                session.commit()
                logger.info(
                    "Knowledge document parsing completed: document_id=%s file=%s parser=%s "
                    "characters=%s chunks=%s elapsed_seconds=%.1f",
                    document.id,
                    file_name,
                    parsed.parser,
                    document.parsed_characters,
                    len(chunks),
                    parse_elapsed_seconds,
                )

                indexing_started_at = time.monotonic()
                vectors = embed_texts([content for content, _locator in chunks])
                for ordinal, ((content, locator), vector) in enumerate(zip(chunks, vectors), start=1):
                    session.add(KnowledgeChunk(
                        document_id=document.id,
                        project_id=document.project_id,
                        scope=document.scope,
                        owner_id=document.owner_id,
                        folder_id=document.folder_id,
                        document_status=document.status,
                        ordinal=ordinal,
                        content=content,
                        source_locator=locator,
                        embedding=vector,
                    ))
                document.indexing_status = "ready"
                document.status = "review" if document.scope == "public" else "ready"
                session.commit()
                logger.info(
                    "Knowledge document processing completed: document_id=%s file=%s parser=%s "
                    "indexing_elapsed_seconds=%.1f total_elapsed_seconds=%.1f",
                    document.id,
                    file_name,
                    parsed.parser,
                    time.monotonic() - indexing_started_at,
                    time.monotonic() - total_started_at,
                )
            except KnowledgeIndexUnavailable as exc:
                # Keep a successfully parsed document visible to its owner/admin,
                # but never let it enter semantic retrieval until BGE-M3 is ready.
                logger.warning(
                    "Knowledge document indexing is waiting for the local model: "
                    "document_id=%s file=%s error=%s",
                    document_id,
                    file_name,
                    str(exc),
                )
                document = session.get(KnowledgeDocument, document_id)
                if document:
                    document.indexing_status = "pending_model"
                    document.parser_warnings = [*(document.parser_warnings or []), str(exc)]
                    document.status = "review" if document.scope == "public" and document.parsing_status == "parsed" else "ready"
                    session.commit()
            except Exception as exc:
                logger.exception(
                    "Knowledge document processing failed: document_id=%s file=%s "
                    "parsing_status=%s indexing_status=%s error_type=%s",
                    document_id,
                    file_name,
                    document.parsing_status,
                    document.indexing_status,
                    exc.__class__.__name__,
                )
                document = session.get(KnowledgeDocument, document_id)
                if document:
                    document.parsing_status = "failed"
                    document.indexing_status = "failed"
                    document.status = "failed"
                    document.parser_warnings = [str(exc)[:800]]
                    session.execute(
                        text("DELETE FROM knowledge_chunk WHERE document_id = :document_id"),
                        {"document_id": document.id},
                    )
                    session.commit()


def resume_interrupted_knowledge_documents() -> None:
    """Resume durable knowledge jobs that were interrupted by an API restart."""
    with MigrationSessionLocal() as session:
        document_ids = list(session.scalars(
            select(KnowledgeDocument.id).where(KnowledgeDocument.status == "processing")
        ))
    for document_id in document_ids:
        threading.Thread(
            target=process_knowledge_document,
            args=(document_id,),
            daemon=True,
            name=f"knowledge-resume-{document_id[:8]}",
        ).start()


@app.get("/api/knowledge/summary")
def knowledge_summary(
    user: CurrentUser = Depends(require_knowledge_user),
    session: Session = Depends(get_knowledge_session),
) -> dict[str, Any]:
    project_id = active_project_id(session)
    private_count = session.scalar(select(func.count(KnowledgeDocument.id)).where(KnowledgeDocument.scope == "private", KnowledgeDocument.project_id == project_id)) or 0
    private_bytes = session.scalar(select(func.coalesce(func.sum(KnowledgeDocument.size_bytes), 0)).where(KnowledgeDocument.scope == "private", KnowledgeDocument.project_id == project_id)) or 0
    public_count = session.scalar(
        select(func.count(KnowledgeDocument.id)).where(KnowledgeDocument.scope == "public", KnowledgeDocument.status == "published", KnowledgeDocument.project_id == project_id)
    ) or 0
    result = {
        "private_document_count": private_count,
        "private_size_bytes": int(private_bytes),
        "public_published_count": public_count,
        "is_public_admin": "field_admin" in user.roles,
    }
    if "field_admin" in user.roles:
        result["public_review_count"] = session.scalar(
            select(func.count(KnowledgeDocument.id)).where(KnowledgeDocument.scope == "public", KnowledgeDocument.status.in_(("processing", "review", "failed")), KnowledgeDocument.project_id == project_id)
        ) or 0
    return result


@app.get("/api/knowledge/folders")
def list_knowledge_folders(
    scope: Literal["private", "public"] = Query(...),
    user: CurrentUser = Depends(require_knowledge_user),
    session: Session = Depends(get_knowledge_session),
) -> list[dict[str, Any]]:
    if scope == "private":
        ensure_private_uncategorized_folder(session, user)
    items = session.scalars(
        select(KnowledgeFolder)
        .where(KnowledgeFolder.scope == scope, KnowledgeFolder.project_id == active_project_id(session))
        .order_by(KnowledgeFolder.parent_id.nullsfirst(), KnowledgeFolder.folder_name)
    ).all()
    return [serialize_knowledge_folder(item) for item in items]


@app.post("/api/knowledge/folders")
def create_knowledge_folder(
    payload: KnowledgeFolderCreate,
    scope: Literal["private", "public"] = Query(...),
    user: CurrentUser = Depends(require_knowledge_user),
    session: Session = Depends(get_knowledge_session),
) -> dict[str, Any]:
    if scope == "public" and "field_admin" not in user.roles:
        raise HTTPException(403, "公共知识库分类仅允许字段管理员维护。")
    parent = None
    if payload.parent_id:
        parent = session.get(KnowledgeFolder, payload.parent_id)
        if not parent or parent.scope != scope or parent.project_id != active_project_id(session):
            raise HTTPException(422, "上级文件夹不存在，或不属于当前知识库范围。")
        if scope == "private" and parent.owner_id != user.id:
            raise HTTPException(403, "不能在其他科研人员的私人文件夹下新建目录。")
    duplicate_conditions = [
        KnowledgeFolder.scope == scope,
        KnowledgeFolder.project_id == active_project_id(session),
        KnowledgeFolder.parent_id == (parent.id if parent else None),
        KnowledgeFolder.folder_name == payload.folder_name.strip(),
    ]
    if scope == "private":
        duplicate_conditions.append(KnowledgeFolder.owner_id == user.id)
    duplicate = session.scalar(select(KnowledgeFolder).where(*duplicate_conditions))
    if duplicate:
        raise HTTPException(409, "同一目录下已存在同名文件夹。")
    item = KnowledgeFolder(
        project_id=active_project_id(session),
        scope=scope,
        owner_id=user.id,
        parent_id=parent.id if parent else None,
        folder_name=payload.folder_name.strip(),
        description=payload.description.strip() or None,
        created_by=user.id,
    )
    session.add(item)
    session.flush()
    result = serialize_knowledge_folder(item)
    session.commit()
    return result


@app.patch("/api/knowledge/folders/{folder_id}")
def update_knowledge_folder(
    folder_id: str,
    payload: KnowledgeFolderUpdate,
    user: CurrentUser = Depends(require_knowledge_user),
    session: Session = Depends(get_knowledge_session),
) -> dict[str, Any]:
    folder = session.get(KnowledgeFolder, folder_id)
    if not folder or folder.project_id != active_project_id(session):
        raise HTTPException(404, "未找到该文件夹。")
    if folder.scope == "public" and "field_admin" not in user.roles:
        raise HTTPException(403, "公共知识库分类仅允许字段管理员维护。")
    if folder.scope == "private" and folder.owner_id != user.id:
        raise HTTPException(403, "不能修改其他科研人员的私人文件夹。")
    if payload.parent_id == folder.id:
        raise HTTPException(422, "文件夹不能移动到自己下面。")
    if payload.parent_id:
        parent = session.get(KnowledgeFolder, payload.parent_id)
        if not parent or parent.scope != folder.scope or parent.project_id != active_project_id(session):
            raise HTTPException(422, "目标文件夹不存在，或不属于当前知识库范围。")
        if folder.scope == "private" and parent.owner_id != user.id:
            raise HTTPException(403, "不能移动到其他科研人员的私人文件夹。")
        ancestor = parent
        while ancestor:
            if ancestor.id == folder.id:
                raise HTTPException(422, "不能将文件夹移动到自己的子文件夹中。")
            ancestor = session.get(KnowledgeFolder, ancestor.parent_id) if ancestor.parent_id else None
    duplicate_conditions = [
        KnowledgeFolder.scope == folder.scope,
        KnowledgeFolder.project_id == active_project_id(session),
        KnowledgeFolder.parent_id == payload.parent_id,
        KnowledgeFolder.folder_name == payload.folder_name.strip(),
        KnowledgeFolder.id != folder.id,
    ]
    if folder.scope == "private":
        duplicate_conditions.append(KnowledgeFolder.owner_id == user.id)
    if session.scalar(select(KnowledgeFolder).where(*duplicate_conditions)):
        raise HTTPException(409, "同一目录下已存在同名文件夹。")
    folder.folder_name = payload.folder_name.strip()
    folder.parent_id = payload.parent_id
    folder.description = payload.description.strip() or None
    session.flush()
    result = serialize_knowledge_folder(folder)
    session.commit()
    return result


@app.delete("/api/knowledge/folders/{folder_id}")
def delete_knowledge_folder(
    folder_id: str,
    user: CurrentUser = Depends(require_knowledge_user),
    session: Session = Depends(get_knowledge_session),
) -> dict[str, bool]:
    folder = session.get(KnowledgeFolder, folder_id)
    if not folder or folder.project_id != active_project_id(session):
        raise HTTPException(404, "未找到该文件夹。")
    if folder.scope == "public" and "field_admin" not in user.roles:
        raise HTTPException(403, "公共知识库分类仅允许字段管理员维护。")
    if folder.scope == "private" and folder.owner_id != user.id:
        raise HTTPException(403, "不能删除其他科研人员的私人文件夹。")
    child_count = session.scalar(select(func.count(KnowledgeFolder.id)).where(KnowledgeFolder.parent_id == folder.id)) or 0
    document_count = session.scalar(select(func.count(KnowledgeDocument.id)).where(KnowledgeDocument.folder_id == folder.id)) or 0
    if child_count or document_count:
        raise HTTPException(409, "文件夹仍包含子文件夹或资料，请先移动或删除其中内容。")
    session.delete(folder)
    session.commit()
    return {"deleted": True}


@app.get("/api/knowledge/documents")
def list_knowledge_documents(
    scope: Literal["private", "public"] = Query(...),
    folder_id: str | None = Query(default=None),
    q: str = Query(default="", max_length=200),
    include_unpublished: bool = Query(default=False),
    user: CurrentUser = Depends(require_knowledge_user),
    session: Session = Depends(get_knowledge_session),
) -> list[dict[str, Any]]:
    project_id = active_project_id(session)
    statement = select(KnowledgeDocument).where(KnowledgeDocument.scope == scope, KnowledgeDocument.project_id == project_id)
    if scope == "public" and not (include_unpublished and "field_admin" in user.roles):
        statement = statement.where(KnowledgeDocument.status == "published")
    if folder_id:
        statement = statement.where(KnowledgeDocument.folder_id == folder_id)
    keyword = q.strip()
    if keyword:
        pattern = f"%{keyword}%"
        statement = statement.where(
            KnowledgeDocument.display_title.ilike(pattern)
            | KnowledgeDocument.original_file_name.ilike(pattern)
            | func.coalesce(KnowledgeDocument.author, "").ilike(pattern)
            | func.coalesce(KnowledgeDocument.source_organization, "").ilike(pattern)
            | KnowledgeDocument.topic_tags.cast(Text).ilike(pattern)
        )
    items = session.scalars(statement.order_by(KnowledgeDocument.updated_at.desc())).all()
    folders = {item.id: item for item in session.scalars(select(KnowledgeFolder).where(KnowledgeFolder.scope == scope, KnowledgeFolder.project_id == project_id)).all()}
    return [serialize_knowledge_document(item, folders.get(item.folder_id)) for item in items]


@app.post("/api/knowledge/documents")
async def upload_knowledge_documents(
    background_tasks: BackgroundTasks,
    files: list[UploadFile] = File(...),
    scope: Literal["private", "public"] = Query(...),
    folder_id: str | None = Query(default=None),
    source_organization: str = Query(default="", max_length=300),
    author: str = Query(default="", max_length=300),
    publication_year: str = Query(default="", max_length=20),
    source_url: str = Query(default="", max_length=2000),
    short_description: str = Query(default="", max_length=2000),
    authorization_basis: str = Query(default="", max_length=2000),
    license_scope: str = Query(default="", max_length=100),
    topic_tags: str = Query(default="", max_length=1000),
    supersedes_document_id: str | None = Query(default=None),
    version_change_summary: str = Query(default="", max_length=2000),
    user: CurrentUser = Depends(require_knowledge_user),
    session: Session = Depends(get_knowledge_session),
) -> list[dict[str, Any]]:
    if not files or len(files) > MAX_KNOWLEDGE_BATCH_FILES:
        raise HTTPException(422, f"每次最多上传 {MAX_KNOWLEDGE_BATCH_FILES} 个知识库文件。")
    if scope == "public" and "field_admin" not in user.roles:
        raise HTTPException(403, "公共知识库仅允许字段管理员上传和维护。")
    if scope == "public" and (not authorization_basis.strip() or not license_scope.strip()):
        raise HTTPException(422, "公共资料必须填写授权依据和授权范围（公开资料或合法授权资料）。")
    parsed_topic_tags = list(dict.fromkeys(
        item.strip() for item in re.split(r"[,，;；]", topic_tags) if item.strip()
    ))
    if len(parsed_topic_tags) > 30:
        raise HTTPException(422, "主题标签最多填写 30 个。")
    if supersedes_document_id and (scope != "public" or "field_admin" not in user.roles or len(files) != 1):
        raise HTTPException(422, "公共资料的新版本一次只能上传一个文件，并且只能由字段管理员操作。")
    folder = validate_knowledge_folder(session, folder_id=folder_id, scope=scope, user=user)
    previous: KnowledgeDocument | None = None
    if supersedes_document_id:
        previous = session.get(KnowledgeDocument, supersedes_document_id)
        if not previous or previous.scope != "public" or previous.project_id != active_project_id(session):
            raise HTTPException(404, "未找到需要替代的公共知识库资料。")
        if not version_change_summary.strip():
            raise HTTPException(422, "发布公共资料新版本时必须说明本次新增或修改内容。")

    created: list[KnowledgeDocument] = []
    total_size = 0
    for file in files:
        original_name = (file.filename or "knowledge-document").strip()
        suffix = Path(original_name).suffix.lower()
        if suffix not in DOCUMENT_SUPPORTED_SUFFIXES:
            raise HTTPException(415, f"{original_name} 不是第一版知识库支持的文档格式。")
        content = await file.read(MAX_KNOWLEDGE_FILE_BYTES + 1)
        if not content:
            raise HTTPException(400, f"{original_name} 为空文件，未保存。")
        if len(content) > MAX_KNOWLEDGE_FILE_BYTES:
            raise HTTPException(413, f"{original_name} 超过单文件 100 MB 上限。")
        total_size += len(content)
        if total_size > MAX_KNOWLEDGE_BATCH_BYTES:
            raise HTTPException(413, "本次上传文件总量超过 500 MB 上限。")
        content_hash = hashlib.sha256(content).hexdigest()
        duplicate_conditions = [
            KnowledgeDocument.scope == scope,
            KnowledgeDocument.project_id == active_project_id(session),
            KnowledgeDocument.content_hash == content_hash,
            KnowledgeDocument.status.not_in(("withdrawn", "superseded", "deleted")),
        ]
        if scope == "private":
            duplicate_conditions.append(KnowledgeDocument.owner_id == user.id)
        duplicate = session.scalar(select(KnowledgeDocument).where(*duplicate_conditions))
        if duplicate:
            raise HTTPException(409, f"{original_name} 的相同内容已存在于当前知识库，未重复保存。")
        document_id = str(uuid.uuid4())
        directory = KNOWLEDGE_STORAGE_DIR / active_project_id(session) / scope / (user.id if scope == "private" else "public") / document_id
        directory.mkdir(parents=True, exist_ok=True)
        path = directory / safe_knowledge_filename(original_name)
        path.write_bytes(content)
        item = KnowledgeDocument(
            id=document_id,
            project_id=active_project_id(session),
            scope=scope,
            owner_id=user.id,
            folder_id=folder.id,
            original_file_name=original_name,
            display_title=Path(original_name).stem or original_name,
            content_type=file.content_type,
            size_bytes=len(content),
            content_hash=content_hash,
            storage_path=str(path),
            source_organization=source_organization.strip() or None,
            author=author.strip() or None,
            publication_year=publication_year.strip() or None,
            source_url=source_url.strip() or None,
            short_description=short_description.strip() or None,
            authorization_basis=authorization_basis.strip() or None,
            license_scope=license_scope.strip() or None,
            topic_tags=parsed_topic_tags,
            status="processing",
            version_number=(previous.version_number + 1) if previous else 1,
            supersedes_document_id=previous.id if previous else None,
            version_change_summary=version_change_summary.strip() or None,
        )
        session.add(item)
        created.append(item)
    session.flush()
    response = [serialize_knowledge_document(item, folder) for item in created]
    session.commit()
    for item in created:
        background_tasks.add_task(process_knowledge_document, item.id)
    return response


@app.post("/api/knowledge/sync-institution-literature")
def sync_institution_literature(
    background_tasks: BackgroundTasks,
    user: CurrentUser = Depends(require_knowledge_admin),
    session: Session = Depends(get_knowledge_session),
) -> dict[str, Any]:
    """Index authorized text extracted by the governed institution import.

    The original PDF/DOCX/TXT remains in the institution MinIO bucket.  The
    knowledge library stores only an extracted text derivative plus its batch,
    object and authorization metadata, so every citation remains traceable.
    """
    account, config = institution_data_config_for_user(session, user)
    project_id = active_project_id(session)
    with INSTITUTION_DATABASES.engine(config.business_database).connect() as connection:
        rows = [dict(row) for row in connection.execute(text("""
            SELECT entity.entity_key, entity.payload, entity.source_batch_id,
                   batch.source_file_name, batch.object_bucket, batch.object_key,
                   batch.file_sha256
            FROM data_entity entity
            JOIN ingest_batch batch ON batch.id=entity.source_batch_id
            WHERE entity.institution_id=:institution_id AND entity.project_id=:project_id
              AND entity.entity_type='literature_document'
            ORDER BY entity.updated_at
        """), {"institution_id": account.institution_id, "project_id": project_id}).mappings()]
    folder = session.scalar(select(KnowledgeFolder).where(
        KnowledgeFolder.project_id == project_id,
        KnowledgeFolder.scope == "public",
        KnowledgeFolder.folder_name == "论文与综述",
    ))
    if not folder:
        raise HTTPException(503, "公共知识库“论文与综述”分类尚未初始化。")
    created: list[KnowledgeDocument] = []
    skipped = 0
    for row in rows:
        payload = row.get("payload") or {}
        extracted = str(payload.get("text") or "").strip()
        if not extracted:
            skipped += 1
            continue
        derivative_hash = hashlib.sha256(extracted.encode("utf-8")).hexdigest()
        duplicate = session.scalar(select(KnowledgeDocument).where(
            KnowledgeDocument.project_id == project_id,
            KnowledgeDocument.scope == "public",
            KnowledgeDocument.content_hash == derivative_hash,
            KnowledgeDocument.status.not_in(("withdrawn", "superseded", "deleted")),
        ))
        if duplicate:
            skipped += 1
            continue
        document_id = str(uuid.uuid4())
        directory = KNOWLEDGE_STORAGE_DIR / project_id / "public" / "public" / document_id
        directory.mkdir(parents=True, exist_ok=True)
        original_name = str(payload.get("file_name") or row.get("source_file_name") or f"literature-{row['entity_key']}")
        path = directory / f"institution-{row['entity_key']}.txt"
        path.write_text(extracted, encoding="utf-8")
        lower = f"{original_name} {extracted[:5000]}".lower()
        tags = [label for label, markers in {
            "品种": ("品种", "种质", "材料"),
            "基因": ("基因", "qtl", "gwas"),
            "性状": ("性状", "表型", "产量", "株高"),
            "育种目标": ("育种", "亲本", "选育"),
            "试验": ("试验", "区试", "田间"),
        }.items() if any(marker in lower for marker in markers)]
        item = KnowledgeDocument(
            id=document_id,
            project_id=project_id,
            scope="public",
            owner_id=user.id,
            folder_id=folder.id,
            original_file_name=original_name,
            display_title=Path(original_name).stem or original_name,
            content_type="text/plain; charset=utf-8",
            size_bytes=len(extracted.encode("utf-8")),
            content_hash=derivative_hash,
            storage_path=str(path),
            source_organization=INSTITUTION_NAME,
            source_url=f"minio://{row['object_bucket']}/{row['object_key']}",
            short_description=f"由机构导入批次 {row['source_batch_id']} 提取的可检索文本；原文件 SHA-256：{row['file_sha256']}",
            authorization_basis="机构数据处理员按公开资料或合法授权资料导入，字段管理员复核后发布。",
            license_scope="公开资料或合法授权资料",
            topic_tags=tags,
            status="processing",
        )
        session.add(item)
        created.append(item)
    session.flush()
    created_ids = [item.id for item in created]
    record_permission_audit(
        session, user, "institution_literature_synced", "knowledge_document", project_id,
        project_id=project_id,
        after={"created_document_ids": created_ids, "skipped_count": skipped},
    )
    session.commit()
    for item in created:
        background_tasks.add_task(process_knowledge_document, item.id)
    return {"created_count": len(created), "skipped_count": skipped, "document_ids": created_ids}


@app.patch("/api/knowledge/documents/{document_id}")
def update_knowledge_document(
    document_id: str,
    payload: KnowledgeDocumentMetadataUpdate,
    user: CurrentUser = Depends(require_knowledge_user),
    session: Session = Depends(get_knowledge_session),
) -> dict[str, Any]:
    document = get_visible_knowledge_document(session, document_id)
    if document.scope == "public" and "field_admin" not in user.roles:
        raise HTTPException(403, "公共知识库资料仅允许字段管理员维护。")
    if document.scope == "private" and document.owner_id != user.id:
        raise HTTPException(403, "不能修改其他科研人员的私人资料。")
    folder = validate_knowledge_folder(session, folder_id=payload.folder_id, scope=document.scope, user=user)
    document.display_title = payload.display_title.strip()
    document.folder_id = folder.id
    document.source_organization = payload.source_organization.strip() or None
    document.author = payload.author.strip() or None
    document.publication_year = payload.publication_year.strip() or None
    document.source_url = payload.source_url.strip() or None
    document.short_description = payload.short_description.strip() or None
    if document.scope == "public" and (not payload.authorization_basis.strip() or not payload.license_scope.strip()):
        raise HTTPException(422, "公共资料必须保留授权依据和授权范围。")
    document.authorization_basis = payload.authorization_basis.strip() or None
    document.license_scope = payload.license_scope.strip() or None
    document.topic_tags = list(dict.fromkeys(item.strip() for item in payload.topic_tags if item.strip()))
    session.flush()
    result = serialize_knowledge_document(document, folder)
    session.commit()
    return result


@app.get("/api/knowledge/documents/{document_id}/preview")
def preview_public_knowledge_document(
    document_id: str,
    user: CurrentUser = Depends(require_knowledge_admin),
    session: Session = Depends(get_knowledge_session),
) -> dict[str, Any]:
    document = get_visible_knowledge_document(session, document_id)
    if document.scope != "public":
        raise HTTPException(403, "为防止知识原文外泄，第一版不提供私人知识库正文预览。")
    chunks = session.scalars(
        select(KnowledgeChunk)
        .where(KnowledgeChunk.document_id == document.id)
        .order_by(KnowledgeChunk.ordinal)
    ).all()
    if chunks:
        preview = "\n\n".join(f"[{item.source_locator}]\n{item.content}" for item in chunks)
    elif document.parsing_status == "parsed":
        preview = (
            f"已由 {document.parser_name or '本地解析器'} 提取 {document.parsed_characters:,} 个字符，"
            f"当前正在建立向量索引（状态：{document.indexing_status}）。"
            "索引完成后会在此显示可核验的分段预览。"
        )
    else:
        preview = "资料尚未完成本地解析，请等待处理完成或查看解析失败原因。"
    return {
        "id": document.id,
        "display_title": document.display_title,
        "parsing_status": document.parsing_status,
        "indexing_status": document.indexing_status,
        "parser_warnings": document.parser_warnings or [],
        "preview": preview[:60000],
        "preview_truncated": len(preview) > 60000,
    }


@app.post("/api/knowledge/documents/{document_id}/reindex")
def reindex_knowledge_document(
    document_id: str,
    background_tasks: BackgroundTasks,
    user: CurrentUser = Depends(require_knowledge_user),
    session: Session = Depends(get_knowledge_session),
) -> dict[str, Any]:
    document = get_visible_knowledge_document(session, document_id)
    if document.scope == "public" and "field_admin" not in user.roles:
        raise HTTPException(403, "公共知识库资料仅允许字段管理员重新处理。")
    if document.scope == "private" and document.owner_id != user.id:
        raise HTTPException(403, "不能重新处理其他科研人员的私人资料。")
    if document.status == "processing" or document.parsing_status == "processing":
        raise HTTPException(409, "该资料正在解析中，请等待当前任务完成后再重试。")
    document.parsing_status = "processing"
    document.indexing_status = "pending"
    document.status = "processing"
    session.commit()
    background_tasks.add_task(process_knowledge_document, document.id)
    return {"queued": True}


@app.post("/api/knowledge/documents/{document_id}/publish")
def publish_knowledge_document(
    document_id: str,
    user: CurrentUser = Depends(require_knowledge_admin),
    session: Session = Depends(get_knowledge_session),
) -> dict[str, Any]:
    document = get_visible_knowledge_document(session, document_id)
    if document.scope != "public":
        raise HTTPException(422, "只有公共知识库资料需要发布。")
    if document.parsing_status != "parsed" or document.indexing_status != "ready":
        raise HTTPException(422, "资料尚未通过本地解析和 bge-m3 索引检查，暂不能发布。")
    if not document.source_organization or not document.folder_id or not document.display_title:
        raise HTTPException(422, "请补全标题、资料分类和来源单位后再发布。")
    if not document.authorization_basis or not document.license_scope:
        raise HTTPException(422, "请补全授权依据和授权范围后再发布。")
    document.status = "published"
    document.published_at = datetime.now(timezone.utc)
    session.execute(
        text("UPDATE knowledge_chunk SET document_status = 'published' WHERE document_id = :document_id"),
        {"document_id": document.id},
    )
    if document.supersedes_document_id:
        previous = session.get(KnowledgeDocument, document.supersedes_document_id)
        if previous:
            previous.status = "superseded"
            session.execute(
                text("UPDATE knowledge_chunk SET document_status = 'superseded' WHERE document_id = :document_id"),
                {"document_id": previous.id},
            )
    session.flush()
    result = serialize_knowledge_document(document, knowledge_folder_for_document(session, document))
    session.commit()
    return result


@app.post("/api/knowledge/documents/{document_id}/withdraw")
def withdraw_public_knowledge_document(
    document_id: str,
    user: CurrentUser = Depends(require_knowledge_admin),
    session: Session = Depends(get_knowledge_session),
) -> dict[str, Any]:
    document = get_visible_knowledge_document(session, document_id)
    if document.scope != "public":
        raise HTTPException(422, "私人知识库资料请使用永久删除。")
    document.status = "withdrawn"
    document.withdrawn_at = datetime.now(timezone.utc)
    session.execute(
        text("UPDATE knowledge_chunk SET document_status = 'withdrawn' WHERE document_id = :document_id"),
        {"document_id": document.id},
    )
    session.commit()
    return {"withdrawn": True}


@app.delete("/api/knowledge/documents/{document_id}")
def delete_private_knowledge_document(
    document_id: str,
    user: CurrentUser = Depends(require_knowledge_user),
    session: Session = Depends(get_knowledge_session),
) -> dict[str, bool]:
    document = get_visible_knowledge_document(session, document_id)
    if document.scope != "private" or document.owner_id != user.id:
        raise HTTPException(403, "公共资料仅能撤回；私人资料仅允许本人永久删除。")
    storage_path = document.storage_path
    session.delete(document)
    session.commit()
    delete_knowledge_file(storage_path)
    try:
        Path(storage_path).parent.rmdir()
    except OSError:
        pass
    return {"deleted": True}


def ensure_research_rls(session: Session) -> None:
    """Apply database-side owner isolation to all private research records.

    The application always sets ``app.research_user_id`` from a verified
    Keycloak token before accessing these tables. FORCE ROW LEVEL SECURITY is
    deliberate: the application database owner must not bypass the policy.
    """
    table_names = [
        "research_session",
        "research_message",
        "research_attachment",
        "research_audit",
        "research_result",
    ]
    for table_name in table_names:
        session.execute(text(f"ALTER TABLE {table_name} ENABLE ROW LEVEL SECURITY"))
        session.execute(text(f"ALTER TABLE {table_name} FORCE ROW LEVEL SECURITY"))
        session.execute(text(f"DROP POLICY IF EXISTS research_owner_only ON {table_name}"))
        session.execute(text(
            f"CREATE POLICY research_owner_only ON {table_name} "
            "FOR ALL USING (owner_id = current_setting('app.research_user_id', true) "
            "AND project_id = current_setting('app.project_id', true)) "
            "WITH CHECK (owner_id = current_setting('app.research_user_id', true) "
            "AND project_id = current_setting('app.project_id', true))"
        ))
    session.commit()


PUBLIC_KNOWLEDGE_FOLDERS = (
    ("论文与综述", "公开或已获授权的论文、综述和学术资料。"),
    ("标准与规范", "数据标准、评价规范、操作规程和管理制度。"),
    ("品种审定与试验资料", "审定公告、区试资料和经确认的试验报告。"),
    ("病虫害与栽培资料", "植保、栽培、施肥和农业生产技术资料。"),
    ("项目方案与研究报告", "项目方案、阶段报告和经批准的研究总结。"),
    ("其他资料", "暂不属于既有分类的公共参考资料。"),
)


def seed_public_knowledge_folders(session: Session, project_id: str = DEFAULT_PROJECT_ID) -> None:
    for folder_name, description in PUBLIC_KNOWLEDGE_FOLDERS:
        exists = session.scalar(
            select(KnowledgeFolder).where(
                KnowledgeFolder.scope == "public",
                KnowledgeFolder.project_id == project_id,
                KnowledgeFolder.parent_id.is_(None),
                KnowledgeFolder.folder_name == folder_name,
            )
        )
        if not exists:
            session.add(KnowledgeFolder(
                project_id=project_id,
                scope="public",
                owner_id="system-public",
                folder_name=folder_name,
                description=description,
                created_by="system",
            ))
    session.commit()


def ensure_private_uncategorized_folder(session: Session, user: CurrentUser) -> KnowledgeFolder:
    project_id = active_project_id(session)
    folder = session.scalar(
        select(KnowledgeFolder).where(
            KnowledgeFolder.scope == "private",
            KnowledgeFolder.project_id == project_id,
            KnowledgeFolder.owner_id == user.id,
            KnowledgeFolder.parent_id.is_(None),
            KnowledgeFolder.folder_name == "未分类",
        )
    )
    if folder:
        return folder
    folder = KnowledgeFolder(
        project_id=project_id,
        scope="private",
        owner_id=user.id,
        folder_name="未分类",
        description="未指定文件夹的私人资料。",
        created_by=user.id,
    )
    session.add(folder)
    session.flush()
    session.commit()
    return folder


def ensure_knowledge_rls(session: Session) -> None:
    """Enforce public/private knowledge visibility in PostgreSQL, not only in APIs."""
    policies = {
        "knowledge_folder": (
            "(project_id = current_setting('app.project_id', true) AND (owner_id = current_setting('app.research_user_id', true) "
            "OR scope = 'public' "
            "OR COALESCE(current_setting('app.knowledge_is_admin', true), 'false') = 'true'))",
            "(project_id = current_setting('app.project_id', true) AND ((scope = 'private' AND owner_id = current_setting('app.research_user_id', true)) "
            "OR COALESCE(current_setting('app.knowledge_is_admin', true), 'false') = 'true'))",
        ),
        "knowledge_document": (
            "(project_id = current_setting('app.project_id', true) AND (owner_id = current_setting('app.research_user_id', true) "
            "OR (scope = 'public' AND status = 'published') "
            "OR COALESCE(current_setting('app.knowledge_is_admin', true), 'false') = 'true'))",
            "(project_id = current_setting('app.project_id', true) AND ((scope = 'private' AND owner_id = current_setting('app.research_user_id', true)) "
            "OR COALESCE(current_setting('app.knowledge_is_admin', true), 'false') = 'true'))",
        ),
        "knowledge_chunk": (
            "(project_id = current_setting('app.project_id', true) AND (owner_id = current_setting('app.research_user_id', true) "
            "OR (scope = 'public' AND document_status = 'published') "
            "OR COALESCE(current_setting('app.knowledge_is_admin', true), 'false') = 'true'))",
            "(project_id = current_setting('app.project_id', true) AND ((scope = 'private' AND owner_id = current_setting('app.research_user_id', true)) "
            "OR COALESCE(current_setting('app.knowledge_is_admin', true), 'false') = 'true'))",
        ),
    }
    for table_name, (using_clause, check_clause) in policies.items():
        session.execute(text(f"ALTER TABLE {table_name} ENABLE ROW LEVEL SECURITY"))
        session.execute(text(f"ALTER TABLE {table_name} FORCE ROW LEVEL SECURITY"))
        session.execute(text(f"DROP POLICY IF EXISTS knowledge_access_policy ON {table_name}"))
        session.execute(text(
            f"CREATE POLICY knowledge_access_policy ON {table_name} "
            f"FOR ALL USING {using_clause} WITH CHECK {check_clause}"
        ))
    session.execute(text(
        "CREATE INDEX IF NOT EXISTS ix_knowledge_chunk_embedding_hnsw "
        "ON knowledge_chunk USING hnsw (embedding vector_cosine_ops) "
        "WHERE embedding IS NOT NULL"
    ))
    session.commit()


async def build_published_evidence_context(
    session: Session,
    question: str,
    requested_by: str = "系统",
    project_id: str = DEFAULT_PROJECT_ID,
) -> tuple[str, list[dict[str, Any]]]:
    """Use controlled templates instead of sending a bulk database snapshot to the model."""
    trial_context, trial_cards = build_published_trial_evidence(session, question, requested_by, project_id)
    if trial_context:
        return trial_context, trial_cards
    query_plan = plan_query_from_question(session, question, TRAITS, ROOT_TRAITS, project_id)
    likely_data_query = is_likely_data_query(question)
    query_planner = "规则解析"
    clarification: str | None = None
    unresolved_variety_names: list[str] = []

    # The model only fills a request form when deterministic matching cannot
    # understand a likely data query. It never receives table data or SQL.
    if not query_plan and likely_data_query:
        planner_result = await infer_controlled_query_request(
            question=question,
            field_catalog=field_catalog_for_planner(TRAITS, ROOT_TRAITS),
        )
        if planner_result:
            try:
                structured_request = StructuredQueryRequest.model_validate(planner_result)
                clarification = structured_request.clarification
                query_plan, unresolved_variety_names = plan_query_from_structured_request(
                    session,
                    structured_request,
                    TRAITS,
                    ROOT_TRAITS,
                    project_id,
                )
                query_planner = "神农受控参数解析"
            except ValueError:
                # Invalid model output is deliberately ignored; no free-form
                # fallback can ever reach the database.
                pass

    if not query_plan:
        if not likely_data_query:
            return "本轮未请求平台已发布标准数据查询。", []
        if clarification:
            detail = f"请先补充查询条件：{clarification}"
        elif unresolved_variety_names:
            detail = f"未能在已发布标准数据中识别品种：{'、'.join(unresolved_variety_names)}。请确认标准品种名或别名。"
        else:
            detail = "请明确品种名称、标准字段或可计算的数值筛选条件。"
        return f"本次问题尚不能执行已发布标准数据查询。{detail}", [{
            "priority": 1,
            "type": "query_clarification",
            "title": "已发布标准数据查询需要补充条件",
            "detail": detail,
            "query_planner": query_planner,
        }]

    execution = execute_published_data_query(session, query_plan, project_id)
    if not execution.records:
        template_title = SQL_TEMPLATES.get(execution.template_code)
        label = template_title.title if template_title else "已发布标准数据查询"
        return f"已执行{label}，但没有命中符合条件的已发布标准数据。", [{
            "priority": 1,
            "type": "published_standard_data",
            "title": f"平台已发布标准数据查询：{label}",
            "detail": "查询已在只读标准数据范围内执行，未返回记录。",
            "query_template": execution.template_code,
            "query_parameters": execution.parameters,
            "query_planner": query_planner,
        }]

    records_by_variety: dict[str, dict[str, Any]] = {}
    for row in execution.records:
        variety_id = str(row["variety_id"])
        record = records_by_variety.setdefault(variety_id, {
            "variety_name": row["variety_name"],
            "aliases": row["alias_names"] or [],
            "variety_type": row["variety_type"],
            "approval_number": row["approval_number"],
            "approval_year": row["approval_year"],
            "suitable_region": row["suitable_region"],
            "traits": {},
        })
        record["traits"][row["trait_name"]] = {
            "value": row["value_numeric"] if row["value_numeric"] is not None else row["value_text"],
            "unit": row["unit"],
            "trait_code": row["trait_code"],
            "trial_year": row["trial_year"],
            "trial_location": row["trial_location"],
            "evaluation_method": row["evaluation_method"],
            "source": "published_standard_data",
        }
    template = SQL_TEMPLATES[execution.template_code]
    records = list(records_by_variety.values())
    cards = [{
        "priority": 1,
        "type": "published_standard_data",
        "title": f"平台已发布标准数据：{record['variety_name']}",
        "detail": f"SQL 模板：{template.title}；返回 {len(record['traits'])} 个标准字段。",
        "query_template": execution.template_code,
        "query_parameters": execution.parameters,
        "query_planner": query_planner,
    } for record in records]
    return "平台已发布标准数据（受控查询结果 JSON）：\n" + json.dumps(records, ensure_ascii=False), cards


def build_attachment_evidence_context(
    attachments: list[ResearchAttachment],
) -> tuple[str, list[dict[str, Any]]]:
    parsed = [item for item in attachments if item.parsing_status == "parsed" and item.parsed_markdown]
    cards = [{
        "priority": 2,
        "type": "private_attachment",
        "title": f"当前会话私有附件：{item.file_name}",
        "detail": f"本地 {item.parser_name or '文档'} 解析；仅当前科研人员可访问。",
    } for item in parsed]
    if not parsed:
        return "当前会话没有已解析的私有附件。", cards
    blocks = [f"### 私有附件：{item.file_name}\n{item.parsed_markdown}" for item in parsed]
    return "\n\n".join(blocks), cards


def knowledge_crop_terms(question: str) -> list[str]:
    """Keep an explicit crop request from retrieving a different crop's document."""
    lowered = question.lower()
    return [term for term in KNOWLEDGE_CROP_TERMS if term in lowered]


KNOWLEDGE_LEXICAL_TOPICS = frozenset({
    "品种", "种质", "材料", "基因", "qtl", "gwas", "性状", "表型", "产量", "株高",
    "育种", "亲本", "选育", "试验", "区试", "田间", "病害", "抗病", "品质", "环境",
})


def knowledge_lexical_matches(question: str, document_text: str) -> list[str]:
    """Return strong exact identifiers or multiple controlled topic matches.

    Vector similarity can be conservative for very short governed records.  A
    material/gene identifier that occurs verbatim in both the question and the
    indexed evidence is stronger than a generic semantic near-match.  Topic
    fallback requires at least two controlled terms to avoid accepting a
    document merely because both sides contain a broad word such as “材料”.
    """
    identifier_pattern = r"(?i)\b[a-z][a-z0-9]*(?:[-_.][a-z0-9]+)+\b"
    question_identifiers = {item.lower() for item in re.findall(identifier_pattern, question)}
    document_identifiers = {item.lower() for item in re.findall(identifier_pattern, document_text)}
    identifier_matches = question_identifiers & document_identifiers
    if identifier_matches:
        return sorted(identifier_matches)
    lowered_question = question.lower()
    lowered_document = document_text.lower()
    topic_matches = {
        term for term in KNOWLEDGE_LEXICAL_TOPICS
        if term in lowered_question and term in lowered_document
    }
    return sorted(topic_matches) if len(topic_matches) >= 2 else []


def explicitly_requested_knowledge_documents(
    question: str,
    documents: list[KnowledgeDocument],
) -> list[KnowledgeDocument]:
    """Find files explicitly named by the researcher, not merely semantic matches."""
    normalized_question = re.sub(r"\s+", "", question).lower()
    return [
        document
        for document in documents
        if len(document.display_title.strip()) >= 8
        and re.sub(r"\s+", "", document.display_title).lower() in normalized_question
    ][:MAX_KNOWLEDGE_DOCUMENTS]


def build_knowledge_evidence_context(
    session: Session,
    user: CurrentUser,
    knowledge_scope: Literal["private", "public", "both"],
    question: str,
) -> tuple[str, list[dict[str, Any]]]:
    """Retrieve only local BGE-M3 chunks that PostgreSQL RLS already authorizes."""
    document_conditions = [
        KnowledgeDocument.parsing_status == "parsed",
        KnowledgeDocument.indexing_status == "ready",
    ]
    if knowledge_scope == "private":
        document_conditions.extend((KnowledgeDocument.scope == "private", KnowledgeDocument.owner_id == user.id, KnowledgeDocument.status == "ready"))
    elif knowledge_scope == "public":
        document_conditions.extend((KnowledgeDocument.scope == "public", KnowledgeDocument.status == "published"))
    else:
        document_conditions.append(
            ((KnowledgeDocument.scope == "private") & (KnowledgeDocument.owner_id == user.id) & (KnowledgeDocument.status == "ready"))
            | ((KnowledgeDocument.scope == "public") & (KnowledgeDocument.status == "published"))
        )

    documents = session.scalars(select(KnowledgeDocument).where(*document_conditions)).all()
    if not documents:
        return "本次问题未检索到可用于支撑结论的本地知识库资料。", []
    explicit_documents = explicitly_requested_knowledge_documents(question, documents)
    explicit_document_ids = {document.id for document in explicit_documents}
    retrieval_question = question
    if explicit_documents:
        # A named file is an explicit retrieval scope, not wording to discard.
        # Removing its title turns a precise question into fragments such as
        # “其中，面临的问题是什么”, which makes vector retrieval miss the
        # document that the researcher explicitly named.
        retrieval_question = "\n".join(
            [question, *(document.display_title for document in explicit_documents)]
        )
    retrieval_question = re.sub(r"\s+", " ", retrieval_question).strip() or question
    conditions = [*document_conditions, KnowledgeChunk.embedding.is_not(None)]

    candidates = session.scalar(
        select(func.count(KnowledgeChunk.id))
        .join(KnowledgeDocument, KnowledgeDocument.id == KnowledgeChunk.document_id)
        .where(*conditions)
    ) or 0
    if not candidates:
        return "本次问题未检索到可用于支撑结论的本地知识库资料。", []
    try:
        question_vector = embed_texts([retrieval_question])[0]
    except KnowledgeIndexUnavailable:
        return "本次问题未检索到可用于支撑结论的本地知识库资料（本地 bge-m3 检索服务尚未就绪）。", []

    distance = KnowledgeChunk.embedding.cosine_distance(question_vector)
    retrieval_query = (
        select(KnowledgeChunk, KnowledgeDocument, KnowledgeFolder, distance.label("distance"))
        .join(KnowledgeDocument, KnowledgeDocument.id == KnowledgeChunk.document_id)
        .outerjoin(KnowledgeFolder, KnowledgeFolder.id == KnowledgeDocument.folder_id)
        .where(*conditions)
        .order_by(distance)
    )
    if explicit_document_ids:
        retrieval_query = retrieval_query.where(KnowledgeChunk.document_id.in_(explicit_document_ids)).limit(12)
    else:
        retrieval_query = retrieval_query.limit(MAX_KNOWLEDGE_DOCUMENTS * MAX_KNOWLEDGE_CHUNKS_PER_DOCUMENT * 3)
    rows = session.execute(retrieval_query).all()
    if not rows:
        return "本次问题未检索到可用于支撑结论的本地知识库资料。", []

    requested_crops = knowledge_crop_terms(question)
    selected_by_document: dict[str, list[tuple[KnowledgeChunk, KnowledgeDocument, KnowledgeFolder | None, float]]] = {}
    for chunk, document, folder, chunk_distance in rows:
        numeric_distance = float(chunk_distance)
        document_text = (
            f"{document.display_title}\n{' '.join(document.topic_tags or [])}\n{chunk.content}"
        ).lower()
        is_explicit_document = document.id in explicit_document_ids
        lexical_matches = knowledge_lexical_matches(retrieval_question, document_text)
        # A named document is an explicit retrieval scope.  Once the user has
        # named it, always retain its best matching segment after crop checks;
        # an embedding-distance threshold must not erase the requested source.
        # Unspecified documents still require a strict semantic threshold.
        if (
            not is_explicit_document
            and numeric_distance > MAX_KNOWLEDGE_COSINE_DISTANCE
            and not lexical_matches
        ):
            continue
        if requested_crops and not all(term in document_text for term in requested_crops):
            continue
        if document.id not in selected_by_document:
            if len(selected_by_document) >= MAX_KNOWLEDGE_DOCUMENTS:
                continue
            selected_by_document[document.id] = []
        selected_rows = selected_by_document[document.id]
        if is_explicit_document and selected_rows:
            # A named document is searched internally. Retain a second segment
            # only when it is nearly as relevant as the primary hit.
            if numeric_distance - selected_rows[0][3] > MAX_EXPLICIT_DOCUMENT_DISTANCE_GAP:
                continue
        if len(selected_rows) < MAX_KNOWLEDGE_CHUNKS_PER_DOCUMENT:
            selected_rows.append((chunk, document, folder, numeric_distance))

    if not selected_by_document:
        return "本轮未检索到与问题直接相关的本地知识库资料。", []

    blocks: list[str] = []
    cards: list[dict[str, Any]] = []
    for document_rows in selected_by_document.values():
        chunk, document, folder, best_distance = document_rows[0]
        scope_label = "我的知识库" if document.scope == "private" else "公共知识库"
        locators = "、".join(item.source_locator for item, _, _, _ in document_rows)
        excerpts = [re.sub(r"\s+", " ", item.content).strip()[:200] for item, _, _, _ in document_rows]
        matched_terms = knowledge_lexical_matches(
            retrieval_question,
            f"{document.display_title}\n{' '.join(document.topic_tags or [])}\n{chunk.content}",
        )
        retrieval_method = (
            "explicit_document"
            if document.id in explicit_document_ids
            else "semantic_vector"
            if best_distance <= MAX_KNOWLEDGE_COSINE_DISTANCE
            else "hybrid_exact_or_topic"
        )
        selected_chunks = "\n\n".join(
            f"片段位置：{item.source_locator}\n{item.content}"
            for item, _, _, _ in document_rows
        )
        blocks.append(
            f"### {scope_label}：{document.display_title}\n"
            f"来源：{document.source_organization or document.author or '未填写'}；命中位置：{locators}\n"
            f"授权范围：{document.license_scope or ('私人资料，仅当前账号' if document.scope == 'private' else '未填写')}；"
            f"主题标签：{'、'.join(document.topic_tags or []) or '未填写'}\n"
            f"{selected_chunks}"
        )
        cards.append({
            "priority": 3,
            "type": "private_knowledge" if document.scope == "private" else "public_knowledge",
            "title": f"{scope_label}：{document.display_title}",
            "detail": f"{folder.folder_name if folder else '未分类'} · {document.source_organization or document.author or '来源待补充'} · {document.license_scope or '私人资料'} · {locators}",
            "excerpt": "\n".join(excerpts),
            "topic_tags": document.topic_tags or [],
            "retrieval_method": retrieval_method,
            "matched_terms": matched_terms,
        })
    return "\n\n".join(blocks), cards


def build_native_vision_blocks(
    attachments: list[ResearchAttachment],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Load local images as compact, temporary multimodal model blocks.

    Original files stay untouched in protected storage. Before an image leaves
    the API process, it is normalized in memory to avoid sending oversized PNG
    screenshots as a Base64 prompt. This keeps visual requests responsive and
    does not change what the researcher can preview locally.
    """
    images = [item for item in attachments if Path(item.file_name).suffix.lower() in VISION_IMAGE_SUFFIXES]
    if len(images) > MAX_NATIVE_VISION_IMAGES:
        raise HTTPException(413, f"当前会话包含 {len(images)} 张图片；单次视觉分析最多支持 {MAX_NATIVE_VISION_IMAGES} 张，请移除部分图片后重试。")
    total_size = sum(item.size_bytes for item in images)
    if total_size > MAX_NATIVE_VISION_BYTES:
        raise HTTPException(413, "当前会话图片合计超过 20 MB，请移除部分图片或拆分后再分析。")

    blocks: list[dict[str, Any]] = []
    cards: list[dict[str, Any]] = []
    for index, item in enumerate(images, start=1):
        path = Path(item.storage_path)
        if not path.is_file():
            raise HTTPException(409, f"图片附件“{item.file_name}”已不在本地存储中，请重新上传后再分析。")
        image_url, model_size, normalized = build_native_vision_image_url(path)
        blocks.append({
            "type": "text",
            "text": f"本轮图片 {index}（文件名：{item.file_name}）。",
        })
        blocks.append({
            "type": "image_url",
            "image_url": {
                "url": image_url,
            },
        })
        transfer_note = (
            f"原图 {item.size_bytes // 1024} KB，在内存中规范化为 {model_size // 1024} KB JPEG 后发送给模型。"
            if normalized
            else "原图在本地受控保存，并以原始格式临时发送给模型。"
        )
        cards.append({
            "priority": 2,
            "type": "private_image",
            "title": f"当前会话私有图片：{item.file_name}",
            "detail": f"{transfer_note} 原图不写入会话记忆。",
        })
    return blocks, cards


def message_attachment_evidence(attachments: list[ResearchAttachment]) -> list[dict[str, Any]]:
    """Record which private files were deliberately sent with one question."""
    return [{
        "type": "message_attachment",
        "attachment_id": item.id,
        "title": item.file_name,
        "detail": "随本轮问题提交的私有图片，将在后续追问中继续作为会话参考。"
        if Path(item.file_name).suffix.lower() in VISION_IMAGE_SUFFIXES
        else "随本轮问题提交的私有附件，将在后续追问中继续作为会话参考。",
        "is_image": Path(item.file_name).suffix.lower() in VISION_IMAGE_SUFFIXES,
        "size_bytes": item.size_bytes,
    } for item in attachments]


def referenced_attachment_ids(messages: list[ResearchMessage]) -> set[str]:
    """Return only attachments that have actually been sent in this conversation."""
    attachment_ids: set[str] = set()
    for message in messages:
        for item in message.evidence or []:
            if isinstance(item, dict) and item.get("type") == "message_attachment" and item.get("attachment_id"):
                attachment_ids.add(str(item["attachment_id"]))
    return attachment_ids


def is_vision_attachment(attachment: ResearchAttachment) -> bool:
    """Return whether an attachment can be sent to the native vision model."""
    return Path(attachment.file_name).suffix.lower() in VISION_IMAGE_SUFFIXES


def select_vision_attachments(
    *,
    current_turn_attachments: list[ResearchAttachment],
    history_items: list[ResearchMessage],
    attachments_by_id: dict[str, ResearchAttachment],
) -> tuple[list[ResearchAttachment], bool]:
    """Choose the image set for one vision request.

    A newly submitted image starts a new visual observation.  It must never be
    mixed with pictures from earlier turns: otherwise a model can describe the
    previous symptom as if it appeared in the new photograph.  A text-only
    follow-up still uses the most recent image-bearing user turn, which keeps
    natural follow-up questions working without retaining every past image.
    """
    previously_sent_ids = referenced_attachment_ids(history_items)
    # The browser normally clears its composer after every send.  Keep the
    # server authoritative as well: if a stale browser state repeats old IDs
    # alongside a newly uploaded photo, send only the newly uploaded photo.
    # This makes an independent image question deterministic even if an older
    # client or a retry mistakenly includes prior attachment IDs.
    new_current_images = [
        item
        for item in current_turn_attachments
        if is_vision_attachment(item) and item.id not in previously_sent_ids
    ]
    if new_current_images:
        return new_current_images, True

    # An explicit request containing only historical image IDs is a deliberate
    # re-analysis request.  It is not a new visual observation, so retain the
    # selected image set and the text history for a follow-up discussion.
    explicit_history_images = [item for item in current_turn_attachments if is_vision_attachment(item)]
    if explicit_history_images:
        return explicit_history_images, False

    # `history_items` is ordered newest first.  Keep all images sent together
    # in the latest applicable user turn, but stop before an older image turn.
    for message in history_items:
        if message.role != "user":
            continue
        image_ids = [
            str(item["attachment_id"])
            for item in message.evidence or []
            if isinstance(item, dict)
            and item.get("type") == "message_attachment"
            and item.get("attachment_id")
            and item.get("is_image")
        ]
        images = [
            attachments_by_id[item_id]
            for item_id in image_ids
            if item_id in attachments_by_id and is_vision_attachment(attachments_by_id[item_id])
        ]
        if images:
            return images, False
    return [], False


def build_native_vision_image_url(path: Path) -> tuple[str, int, bool]:
    """Return a model-ready image data URL without modifying local originals."""
    raw_bytes = path.read_bytes()
    try:
        from PIL import Image, ImageOps

        with Image.open(io.BytesIO(raw_bytes)) as source:
            image = ImageOps.exif_transpose(source)
            image.thumbnail((MAX_NATIVE_VISION_EDGE, MAX_NATIVE_VISION_EDGE))
            if "A" in image.getbands():
                background = Image.new("RGB", image.size, "white")
                background.paste(image, mask=image.getchannel("A"))
                image = background
            else:
                image = image.convert("RGB")
            normalized_bytes = io.BytesIO()
            image.save(
                normalized_bytes,
                format="JPEG",
                quality=NATIVE_VISION_JPEG_QUALITY,
                optimize=True,
                progressive=True,
            )
        model_bytes = normalized_bytes.getvalue()
        return f"data:image/jpeg;base64,{base64.b64encode(model_bytes).decode('ascii')}", len(model_bytes), True
    except Exception:
        mime_by_suffix = {
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".webp": "image/webp",
        }
        media_type = mime_by_suffix.get(path.suffix.lower(), "application/octet-stream")
        return f"data:{media_type};base64,{base64.b64encode(raw_bytes).decode('ascii')}", len(raw_bytes), False


def sse_event(event: str, payload: dict[str, Any]) -> str:
    return f"event: {event}\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"


def _trial_analysis_run_id_from_context(context: str) -> str | None:
    """Keep a reference to the reviewed local analysis behind a report.

    `build_published_trial_evidence` already serializes its controlled result
    as JSON for the assistant.  We extract only the immutable run id here so a
    later PDF download can render the same local statistics without asking an
    LLM to recreate numbers or figures.
    """
    marker = "区域试验已发布数据（本轮可追溯证据 JSON）：\n"
    if not context.startswith(marker):
        return None
    try:
        payload = json.loads(context[len(marker):])
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    analysis = payload.get("analysis") if isinstance(payload, dict) else None
    run_id = analysis.get("analysis_run_id") if isinstance(analysis, dict) else None
    return str(run_id) if run_id else None


def _report_operation(operations: list[dict[str, Any]] | None) -> dict[str, Any] | None:
    for item in operations or []:
        if isinstance(item, dict) and item.get("state") == "report_ready":
            return item
    return None


def _report_fallback_content(
    *,
    breeding_report_requested: bool,
    breeding_report_context: dict[str, Any] | None,
) -> str:
    """Describe a deterministic report without misrepresenting an LLM answer.

    A requested PDF is rendered from controlled local data.  If the model
    layer returns only an unusable placeholder, preserve that valid report
    workflow but make its status message deterministic and auditable instead
    of persisting the provider placeholder as a research conclusion.
    """
    if breeding_report_requested and breeding_report_context:
        material = breeding_report_context.get("material") or {}
        material_name = str(material.get("material_name") or "候选材料")
        material_code = str(material.get("material_code") or "").strip()
        subject = f"{material_name}（{material_code}）" if material_code else material_name
        return (
            f"已依据已发布区域试验数据和育种档案，生成 {subject} 的品种选育报告（审定辅助草稿）。"
            "请通过下方结果卡下载，并按原始档案和正式审定要求复核。"
        )
    return "已依据本轮受控证据生成可下载 PDF 报告。请通过下方结果卡下载并复核来源。"


@app.post("/api/research/sessions/{research_session_id}/chat/stream")
async def research_chat_stream(
    research_session_id: str,
    payload: ResearchChatRequest,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> StreamingResponse:
    research_session = get_owned_research_session(session, research_session_id)
    breeding_report_requested = is_breeding_report_request(payload.content)
    report_requested = is_report_request(payload.content) or breeding_report_requested
    breeding_report_context: dict[str, Any] | None = None
    breeding_context = ""
    breeding_cards: list[dict[str, Any]] = []
    if breeding_report_requested:
        try:
            breeding_report_context = build_breeding_report_context(session, payload.content, research_session.project_id)
            breeding_context, breeding_cards = build_breeding_report_evidence_context(breeding_report_context)
        except BreedingDossierError as exc:
            raise HTTPException(422, str(exc)) from exc
    attachments = session.scalars(
        select(ResearchAttachment)
        .where(ResearchAttachment.session_id == research_session_id)
        .order_by(ResearchAttachment.created_at)
    ).all()
    attachments_by_id = {item.id: item for item in attachments}
    current_turn_attachment_ids = list(dict.fromkeys(payload.attachment_ids))
    unknown_attachment_ids = [item_id for item_id in current_turn_attachment_ids if item_id not in attachments_by_id]
    if unknown_attachment_ids:
        raise HTTPException(422, "存在不属于当前会话的附件，无法随本轮问题提交。请刷新会话后重试。")
    current_turn_attachments = [attachments_by_id[item_id] for item_id in current_turn_attachment_ids]
    history_items = session.scalars(
        select(ResearchMessage)
        .where(ResearchMessage.session_id == research_session_id)
        .order_by(ResearchMessage.created_at.desc())
        .limit(8)
    ).all()
    previous_session_title = research_session.title
    automatic_session_title = auto_title_for_first_message(
        research_session.title,
        payload.content,
        has_messages=bool(history_items),
    )
    if automatic_session_title:
        research_session.title = automatic_session_title
        research_session.updated_at = datetime.now(timezone.utc)
    context_attachment_ids = referenced_attachment_ids(history_items) | set(current_turn_attachment_ids)
    context_attachments = [item for item in attachments if item.id in context_attachment_ids]
    published_context, published_cards = await build_published_evidence_context(
        session,
        payload.content,
        audit_actor(user),
        research_session.project_id,
    )
    analysis_run_id = _trial_analysis_run_id_from_context(published_context)
    attachment_context, attachment_cards = build_attachment_evidence_context(context_attachments)
    knowledge_context, knowledge_cards = build_knowledge_evidence_context(
        session,
        user,
        payload.knowledge_scope,
        payload.content,
    )
    vision_attachments, has_current_vision_images = select_vision_attachments(
        current_turn_attachments=current_turn_attachments,
        history_items=history_items,
        attachments_by_id=attachments_by_id,
    )
    vision_blocks, vision_cards = build_native_vision_blocks(vision_attachments)
    conversation_history = [
        {"role": item.role, "content": item.content}
        for item in reversed(history_items)
        if item.role in {"user", "assistant"}
    ]
    static_evidence_context = f"{published_context}\n\n{breeding_context}\n\n{attachment_context}\n\n{knowledge_context}"
    if len(static_evidence_context) > MAX_RESEARCH_CONTEXT_CHARS:
        raise HTTPException(
            413,
            "当前会话附件与证据材料过长，未向模型截断。请移除部分附件或拆分后再分析。",
        )
    static_evidence = [*published_cards, *breeding_cards, *attachment_cards, *knowledge_cards, *vision_cards]
    user_message = ResearchMessage(
        session_id=research_session_id,
        project_id=research_session.project_id,
        owner_id=user.id,
        role="user",
        content=payload.content.strip(),
        evidence=message_attachment_evidence(current_turn_attachments),
        operation_state=[
            {"state": "accepted", "label": "已接收问题"},
            *([{ "state": "attachments", "label": f"已随本轮提交 {len(current_turn_attachments)} 个附件" }] if current_turn_attachments else []),
            *([{
                "state": "report_requested",
                "label": "已识别品种选育报告请求" if breeding_report_requested else "已识别生成 PDF 报告请求",
            }] if report_requested else []),
        ],
    )
    research_session.updated_at = datetime.now(timezone.utc)
    session.add(user_message)
    if automatic_session_title:
        session.add(ResearchAudit(
            owner_id=user.id,
            project_id=research_session.project_id,
            session_id=research_session_id,
            action="assistant_session_auto_renamed",
            audit_metadata={
                "previous_title": previous_session_title,
                "title": automatic_session_title,
                "source": "first_user_message",
            },
        ))
    session.add(ResearchAudit(
        owner_id=user.id,
        project_id=research_session.project_id,
        session_id=research_session_id,
        action="assistant_question_submitted",
        audit_metadata={
            "current_turn_attachment_count": len(current_turn_attachments),
            "context_attachment_count": len(context_attachments),
            "vision_attachment_count": len(vision_attachments),
            "vision_attachment_source": "current_turn" if has_current_vision_images else ("latest_history_turn" if vision_attachments else "none"),
            "published_evidence_count": len(published_cards),
            "breeding_dossier_evidence_count": len(breeding_cards),
            "knowledge_evidence_count": len(knowledge_cards),
            "knowledge_scope": payload.knowledge_scope,
        },
    ))
    memory_state = research_session.memory_state or {}
    session.commit()

    async def event_stream() -> Any:
        if automatic_session_title:
            yield sse_event("session_title", {
                "session_id": research_session_id,
                "title": automatic_session_title,
            })
        yield sse_event("status", {"label": "正在读取已发布标准数据、当前会话附件和本地知识库证据"})
        full_text = ""
        model_answer_started = False
        try:
            evidence = list(static_evidence)
            evidence_context = static_evidence_context
            public_web_context = ""
            if vision_blocks:
                yield sse_event("status", {"label": f"正在准备 {len(vision_attachments)} 张本地图片供当前模型进行视觉分析"})
            yield sse_event("status", {"label": "正在判断是否需要检索近期可信公开资料"})
            web_results, search_note = await search_public_references(payload.content)
            if web_results:
                web_context = "\n\n".join(
                    f"### 可信公开来源：{item.title}\n链接：{item.url}\n摘要：{item.snippet}"
                    for item in web_results
                )
                public_web_context = web_context
                evidence.extend({
                    "priority": 4,
                    "type": "trusted_public_web",
                    "title": f"公开参考：{item.title}",
                    "detail": "Tavily 公开检索；仅作为补充证据，未写入平台知识库。",
                    "url": item.url,
                } for item in web_results)
                yield sse_event("status", {"label": f"已检索到 {len(web_results)} 条可信公开参考来源"})
            elif search_note:
                yield sse_event("status", {"label": search_note})

            if len(f"{evidence_context}\n\n{public_web_context}") > MAX_RESEARCH_CONTEXT_CHARS:
                yield sse_event("error", {
                    "detail": "当前会话附件、已发布数据与公开资料合计过长，未向模型截断。请移除部分附件或拆分问题后重试。",
                })
                return

            # Yield before compacting so the browser can render the same
            # waiting state users see in mature research assistants.  The
            # current request then continues with the compacted context.
            if len((memory_state or {}).get("content") or []) > 8:
                yield sse_event("status", {"label": "正在自动压缩上下文"})
                # The compaction itself is intentionally lightweight. Keep a
                # tiny yield window so the browser can paint the progress
                # state before the request continues to the model call.
                await asyncio.sleep(0.12)
            working_memory_state, memory_was_compacted = prepare_working_memory_state(memory_state)
            if memory_was_compacted and len((memory_state or {}).get("content") or []) <= 8:
                yield sse_event("status", {"label": "正在自动压缩上下文"})
                await asyncio.sleep(0.12)
            yield sse_event("status", {"label": "正在调用大模型"})
            async for result in stream_research_reply(
                user_prompt=payload.content.strip(),
                evidence_context=evidence_context,
                memory_state=working_memory_state,
                public_web_context=public_web_context,
                vision_images=vision_blocks,
                # A new image is a new observation.  Do not send prior visual
                # diagnoses as chat history, or they can be mistaken for facts
                # about this image.  Pure text follow-ups retain the history.
                conversation_history=[] if has_current_vision_images else conversation_history,
                has_current_vision_images=has_current_vision_images,
            ):
                if result["type"] == "token":
                    if not model_answer_started:
                        model_answer_started = True
                        yield sse_event("status", {"label": "正在接收大模型回答"})
                    full_text += result["text"]
                    yield sse_event("token", {"text": result["text"]})
                    continue

                with SessionLocal() as write_session:
                    _set_research_owner(write_session, user.id)
                    _set_active_project(write_session, research_session.project_id)
                    stored_session = get_owned_research_session(write_session, research_session_id)
                    stored_session.memory_state = result["memory_state"]
                    stored_session.memory_summary = result.get("memory_summary")
                    stored_session.updated_at = datetime.now(timezone.utc)
                    assistant_message = ResearchMessage(
                        session_id=research_session_id,
                        project_id=stored_session.project_id,
                        owner_id=user.id,
                        role="assistant",
                        content=result["content"],
                        evidence=evidence,
                        operation_state=[
                            {"state": "completed", "label": "已完成大模型分析"},
                            *([{ "state": "web_search", "label": f"已补充 {len(web_results)} 条可信公开来源" }] if web_results else []),
                            {"state": "evidence", "label": f"已附带 {len(evidence)} 项证据"},
                            *([{
                                "state": "report_ready",
                                "label": "已按已发布试验表现与育种档案生成可下载品种选育报告" if breeding_report_requested else "已按本轮证据生成可下载 PDF 报告",
                                "analysis_run_id": analysis_run_id,
                                "report_kind": "breeding_dossier" if breeding_report_requested else "research_report",
                                "material_code": breeding_report_context["material"]["material_code"] if breeding_report_context else None,
                            }] if report_requested else []),
                        ],
                    )
                    write_session.add(assistant_message)
                    write_session.add(ResearchAudit(
                        owner_id=user.id,
                        project_id=stored_session.project_id,
                        session_id=research_session_id,
                        action="assistant_answer_completed",
                        audit_metadata={
                            "evidence_count": len(evidence),
                            "knowledge_evidence_count": len(knowledge_cards),
                            "public_web_source_count": len(web_results),
                            "report_requested": report_requested,
                            "report_kind": "breeding_dossier" if breeding_report_requested else None,
                        },
                    ))
                    write_session.flush()
                    analysis = _load_trial_analysis(write_session, analysis_run_id)
                    saved_artifacts = _save_structured_result_artifacts(
                        write_session,
                        owner_id=user.id,
                        message=assistant_message,
                        question=payload.content.strip(),
                        analysis_run_id=analysis_run_id,
                        analysis=analysis,
                    )
                    if saved_artifacts:
                        assistant_message.operation_state = [
                            *(assistant_message.operation_state or []),
                            {
                                "state": "results_archived",
                                "label": f"已保存 {len(saved_artifacts)} 项结构化研究产物到结果库",
                                "analysis_run_id": analysis_run_id,
                            },
                        ]
                    response_message = serialize_research_message(assistant_message)
                    write_session.commit()
                yield sse_event("complete", {"message": response_message})
        except HTTPException as exc:
            detail = exc.detail if isinstance(exc.detail, str) else "请求未能完成，请检查输入内容后重试。"
            yield sse_event("error", {"detail": detail})
        except EmptyResearchAnswerError:
            if not report_requested:
                yield sse_event("error", {
                    "detail": "大模型未返回可展示的研究结论（仅收到占位符或空内容）。本轮未保存回答，请重新提问。",
                })
                return

            # A report request has a deterministic, controlled-data output
            # path. Keep that result available, but never label a provider
            # placeholder as a completed LLM analysis.
            fallback_content = _report_fallback_content(
                breeding_report_requested=breeding_report_requested,
                breeding_report_context=breeding_report_context,
            )
            with SessionLocal() as write_session:
                _set_research_owner(write_session, user.id)
                _set_active_project(write_session, research_session.project_id)
                stored_session = get_owned_research_session(write_session, research_session_id)
                stored_session.updated_at = datetime.now(timezone.utc)
                assistant_message = ResearchMessage(
                    session_id=research_session_id,
                    project_id=stored_session.project_id,
                    owner_id=user.id,
                    role="assistant",
                    content=fallback_content,
                    evidence=evidence,
                    operation_state=[
                        {"state": "completed", "label": "已完成受控数据报告生成"},
                        {
                            "state": "model_output_unavailable",
                            "label": "大模型未返回可展示的说明文本，未将占位内容保存为科研结论",
                        },
                        *([{ "state": "web_search", "label": f"已补充 {len(web_results)} 条可信公开来源" }] if web_results else []),
                        {"state": "evidence", "label": f"已附带 {len(evidence)} 项证据"},
                        {
                            "state": "report_ready",
                            "label": "已按已发布试验表现与育种档案生成可下载品种选育报告" if breeding_report_requested else "已按本轮证据生成可下载 PDF 报告",
                            "analysis_run_id": analysis_run_id,
                            "report_kind": "breeding_dossier" if breeding_report_requested else "research_report",
                            "material_code": breeding_report_context["material"]["material_code"] if breeding_report_context else None,
                        },
                    ],
                )
                write_session.add(assistant_message)
                write_session.add(ResearchAudit(
                    owner_id=user.id,
                    project_id=stored_session.project_id,
                    session_id=research_session_id,
                    action="assistant_report_completed_without_model_text",
                    audit_metadata={
                        "evidence_count": len(evidence),
                        "knowledge_evidence_count": len(knowledge_cards),
                        "public_web_source_count": len(web_results),
                        "report_kind": "breeding_dossier" if breeding_report_requested else "research_report",
                        "model_output": "placeholder_or_empty",
                    },
                ))
                response_message = serialize_research_message(assistant_message)
                write_session.commit()
            logger.warning(
                "Completed deterministic report after rejecting placeholder model output: session_id=%s user_id=%s report_kind=%s",
                research_session_id,
                user.id,
                "breeding_dossier" if breeding_report_requested else "research_report",
            )
            yield sse_event("status", {"label": "大模型未返回可展示的说明，已保留基于受控数据生成的报告"})
            yield sse_event("complete", {"message": response_message})
        except ResearchAgentError as exc:
            yield sse_event("error", {"detail": str(exc)})
        except Exception as exc:
            error_id = uuid.uuid4().hex[:10]
            logger.exception(
                "Research chat failed: error_id=%s session_id=%s user_id=%s",
                error_id,
                research_session_id,
                user.id,
            )
            raw_detail = re.sub(r"(?i)(bearer\\s+|api[_-]?key[=:]\\s*)[^\\s,;]+", r"\\1***", str(exc)).strip()
            if raw_detail:
                detail = f"后端执行异常（{type(exc).__name__}，错误编号 {error_id}）：{raw_detail[:260]}"
            else:
                detail = f"后端执行异常（{type(exc).__name__}，错误编号 {error_id}）。请将该编号提供给管理员排查。"
            yield sse_event("error", {"detail": detail})

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/research/messages/{message_id}/report.pdf")
def download_research_message_report(
    message_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> Response:
    """Render and archive one requested assistant answer as a private PDF."""
    message = session.scalar(
        select(ResearchMessage).where(
            ResearchMessage.id == message_id,
            ResearchMessage.owner_id == user.id,
            ResearchMessage.role == "assistant",
        )
    )
    if not message:
        raise HTTPException(404, "未找到可下载的科研报告。")
    report_operation = _report_operation(message.operation_state)
    if not report_operation:
        raise HTTPException(409, "该回答未提出生成报告请求，因此未创建报告下载。")

    existing = session.scalar(select(ResearchResult).where(
        ResearchResult.owner_id == user.id,
        ResearchResult.source_message_id == message.id,
        ResearchResult.result_type == "pdf_report",
    ))
    if existing and Path(existing.storage_path).is_file():
        return Response(
            content=Path(existing.storage_path).read_bytes(),
            media_type=existing.content_type,
            headers={
                "Content-Disposition": _download_content_disposition(
                    existing.file_name,
                    f"research-report-{existing.id[:8]}.pdf",
                )
            },
        )

    question_message = session.scalar(
        select(ResearchMessage)
        .where(
            ResearchMessage.session_id == message.session_id,
            ResearchMessage.owner_id == user.id,
            ResearchMessage.role == "user",
            ResearchMessage.created_at <= message.created_at,
        )
        .order_by(ResearchMessage.created_at.desc())
        .limit(1)
    )
    question = question_message.content if question_message else "本轮育种科研问题"
    analysis_run_id = report_operation.get("analysis_run_id")
    analysis = _load_trial_analysis(session, analysis_run_id)
    report_kind = str(report_operation.get("report_kind") or "research_report")
    breeding_report_context: dict[str, Any] | None = None
    if report_kind == "breeding_dossier" or is_breeding_report_request(question):
        try:
            breeding_report_context = build_breeding_report_context(session, question, message.project_id)
        except BreedingDossierError as exc:
            raise HTTPException(422, str(exc)) from exc
        report_bytes = build_breeding_report_pdf(
            breeding_report_context,
            generated_at=datetime.now(timezone.utc),
        )
        report_title = f"{breeding_report_context['material']['material_name']} · 品种选育报告（审定辅助草稿）"
        report_file_name = f"{breeding_report_context['material']['material_name']}-品种选育报告-审定辅助草稿-{datetime.now().strftime('%Y%m%d-%H%M%S')}.pdf"
    else:
        report_bytes = build_research_report_pdf(
            question=question,
            answer=message.content,
            evidence=message.evidence or [],
            analysis=analysis,
            generated_at=datetime.now(timezone.utc),
        )
        report_title = f"{analysis.get('title')} · PDF 研究报告" if analysis else "隆耘 Agent 育种智能体 · PDF 研究报告"
        report_file_name = f"隆耘Agent育种分析报告-{datetime.now().strftime('%Y%m%d-%H%M%S')}.pdf"
    saved = _store_research_result(
        session,
        owner_id=user.id,
        project_id=message.project_id,
        source_message_id=message.id,
        session_id=message.session_id,
        analysis_run_id=str(analysis_run_id) if analysis_run_id else None,
        result_type="pdf_report",
        title=report_title,
        file_name=report_file_name,
        content_type="application/pdf",
        content=report_bytes,
        analysis=analysis,
    )
    if breeding_report_context:
        material = breeding_report_context["material"]
        saved.summary = (
            f"{material['material_name']}（{material['material_code']}）的审定辅助品种选育报告；"
            "亲本、世代与选择记录为模拟演示数据，正式申报前必须用经审核的原始档案替换。"
        )
        saved.result_metadata = {
            **(saved.result_metadata or {}),
            "report_kind": "breeding_dossier",
            "material_code": material["material_code"],
            "is_simulated_breeding_dossier": True,
            "approval_standard_reference": "江西省品种审定标准第十一条（用户提供文件）",
        }
    session.add(ResearchAudit(
        owner_id=user.id,
        project_id=message.project_id,
        session_id=message.session_id,
        action="research_result_saved",
        audit_metadata={"result_type": "pdf_report", "result_id": saved.id, "source_message_id": message.id},
    ))
    session.commit()
    return Response(
        content=report_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": _download_content_disposition(
                saved.file_name,
                f"research-report-{saved.id[:8]}.pdf",
            )
        },
    )


@app.get("/api/research/results")
def list_research_results(
    result_type: str | None = Query(default=None),
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> list[dict[str, Any]]:
    """List durable assistant outputs for the signed-in researcher only."""
    statement = select(ResearchResult).where(ResearchResult.owner_id == user.id)
    if result_type:
        statement = statement.where(ResearchResult.result_type == result_type)
    items = session.scalars(statement.order_by(ResearchResult.created_at.desc())).all()
    # A historical pre-QC ZIP must not be exposed as a completed research
    # product.  Only a version that has passed material mapping and has been
    # explicitly published may appear as a downloadable QC result package.
    genotype_result_ids = [item.analysis_run_id for item in items if item.result_type == "genotype_qc_package" and item.analysis_run_id]
    if genotype_result_ids:
        # Keep this query simple across PostgreSQL driver versions. RLS already
        # limits rows to the current researcher before this local intersection.
        ready_versions = set(session.execute(text("""
            SELECT id FROM genotype_asset_version
            WHERE status = 'analysis_ready'
        """)).scalars().all())
        items = [item for item in items if item.result_type != "genotype_qc_package" or item.analysis_run_id in ready_versions]
    return [serialize_research_result(item) for item in items]


@app.get("/api/research/results/{result_id}/download")
def download_research_result(
    result_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> Response:
    result = session.scalar(select(ResearchResult).where(
        ResearchResult.id == result_id,
        ResearchResult.owner_id == user.id,
    ))
    if not result:
        raise HTTPException(404, "未找到该研究产物，或当前账号无权访问。")
    if result.result_type == "genotype_qc_package":
        status = session.execute(text("SELECT status FROM genotype_asset_version WHERE id=:id"), {"id": result.analysis_run_id}).scalar()
        if status != "analysis_ready":
            raise HTTPException(409, "该基因型版本仍在材料映射确认阶段，尚未形成可下载的正式结果包。")
    path = Path(result.storage_path)
    if not path.is_file():
        raise HTTPException(409, "该研究产物的本地文件已不存在，请重新生成。")
    return Response(
        content=path.read_bytes(),
        media_type=result.content_type,
        headers={
            "Content-Disposition": _download_content_disposition(
                result.file_name,
                f"research-result-{result.id[:8]}",
            )
        },
    )


@app.delete("/api/research/results/{result_id}")
def delete_research_result(
    result_id: str,
    user: CurrentUser = Depends(require_researcher),
    session: Session = Depends(get_research_session),
) -> dict[str, bool]:
    result = session.scalar(select(ResearchResult).where(
        ResearchResult.id == result_id,
        ResearchResult.owner_id == user.id,
    ))
    if not result:
        raise HTTPException(404, "未找到该研究产物，或当前账号无权访问。")
    path = Path(result.storage_path)
    session.delete(result)
    session.add(ResearchAudit(
        owner_id=user.id,
        project_id=result.project_id,
        session_id=result.session_id,
        action="research_result_deleted",
        audit_metadata={"result_id": result_id, "result_type": result.result_type},
    ))
    session.commit()
    try:
        path.unlink(missing_ok=True)
        path.parent.rmdir()
    except OSError:
        pass
    return {"deleted": True}


@app.get("/api/catalog")
def catalog(user: CurrentUser = Depends(require_data_platform_user)) -> dict[str, Any]:
    return {"traits": [{"code": code, **{key: value for key, value in trait.items() if key != "patterns"}} for code, trait in TRAITS.items()]}


@app.get("/api/templates")
def list_templates(user: CurrentUser = Depends(require_data_platform_user), session: Session = Depends(get_business_project_session)) -> list[dict[str, Any]]:
    templates = session.scalars(select(DataTemplate).order_by(DataTemplate.template_code)).all()
    return [serialize_template(item, session.get(TemplateVersion, item.current_version_id)) for item in templates]


@app.get("/api/template-change-requests")
def list_template_change_requests(user: CurrentUser = Depends(require_field_admin), session: Session = Depends(get_business_project_session)) -> list[dict[str, Any]]:
    templates = {item.id: item for item in session.scalars(select(DataTemplate)).all()}
    requests = session.scalars(select(FieldChangeRequest).order_by(FieldChangeRequest.created_at.desc())).all()
    return [{"id": item.id, "template_id": item.template_id, "template_name": templates[item.template_id].template_name if item.template_id in templates else "未知模板", "source_review_id": item.source_review_id, "source_field": item.source_field, "sample_value": item.sample_value, "request_note": item.request_note, "status": item.status, "submitted_by": item.submitted_by, "created_at": item.created_at.isoformat()} for item in requests]


@app.post("/api/templates/{template_id}/versions")
def create_template_version(template_id: str, payload: TemplateVersionCreate, user: CurrentUser = Depends(require_field_admin), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    payload.actor = audit_actor(user)
    template = session.get(DataTemplate, template_id)
    current = session.get(TemplateVersion, template.current_version_id) if template else None
    if not template or not current:
        raise HTTPException(404, "标准模板不存在。")
    fields = [dict(item) for item in (current.field_definitions or [])]
    field_code = payload.field_code.strip() or f"custom_{uuid.uuid4().hex[:10]}"
    existing = next((item for item in fields if item["code"] == field_code), None)
    if payload.action == "add_field" and existing:
        raise HTTPException(409, "该标准字段已存在，请选择更新字段或补充别名。")
    if existing:
        existing.update({"name": payload.field_name, "kind": payload.field_kind, "category": payload.category, "unit": payload.unit, "required": payload.required, "min": payload.min_value, "max": payload.max_value, "severity": payload.severity})
        existing["aliases"] = list(dict.fromkeys([*(existing.get("aliases") or []), *payload.aliases]))
    else:
        fields.append({"code": field_code, "name": payload.field_name, "category": payload.category, "unit": payload.unit, "aliases": payload.aliases, "required": payload.required, "kind": payload.field_kind, "min": payload.min_value, "max": payload.max_value, "severity": payload.severity})
    match = re.match(r"v(\d+)\.(\d+)", current.version)
    next_version = f"v{match.group(1)}.{int(match.group(2)) + 1}" if match else "v1.1"
    version = TemplateVersion(template_id=template.id, version=next_version, change_summary=payload.change_summary, field_definitions=fields, created_by=payload.actor)
    session.add(version)
    session.flush()
    template.current_version_id = version.id
    if payload.request_id:
        request = session.get(FieldChangeRequest, payload.request_id)
        if request:
            request.status = "resolved"
            request.resolved_by = payload.actor
            request.resolution_note = f"已发布到 {next_version}"
            request.resolved_version_id = version.id
    session.commit()
    return serialize_template(template, version)


@app.post("/api/template-change-requests")
def create_template_change_request(payload: FieldChangeRequestCreate, user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    payload.actor = audit_actor(user)
    source = session.get(SourceReview, payload.source_review_id)
    if not source or not source.template_version_id:
        raise HTTPException(422, "该来源没有关联标准模板，无法提交字段处理请求。")
    version = session.get(TemplateVersion, source.template_version_id)
    if not version:
        raise HTTPException(422, "来源使用的模板版本不存在。")
    existing = session.scalar(select(FieldChangeRequest).where(FieldChangeRequest.source_review_id == source.id, FieldChangeRequest.source_field == payload.source_field, FieldChangeRequest.status == "pending"))
    if existing:
        return {"id": existing.id, "status": existing.status, "message": "该字段已提交管理员处理。"}
    request = FieldChangeRequest(source_review_id=source.id, template_id=version.template_id, source_field=payload.source_field, sample_value=payload.sample_value, request_note=payload.request_note, submitted_by=payload.actor)
    session.add(request)
    append_history(source, payload.actor, "提交未识别字段给管理员", {"field": payload.source_field, "sample_value": payload.sample_value})
    session.commit()
    return {"id": request.id, "status": request.status}


@app.get("/api/dashboard")
def dashboard(user: CurrentUser = Depends(require_data_platform_user), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    varieties = session.scalar(select(func.count(Variety.id))) or 0
    published = session.scalar(select(func.count(PhenotypeObservation.id)).where(PhenotypeObservation.publish_status == "published")) or 0
    pending = session.scalar(select(func.count(PhenotypeObservation.id)).where(PhenotypeObservation.publish_status == "pending")) or 0
    blocked = session.scalar(select(func.count(PhenotypeObservation.id)).where(PhenotypeObservation.quality_status == "blocked")) or 0
    return {"varieties": varieties, "published": published, "pending": pending, "blocked": blocked}


@app.get("/api/workbench")
def workbench(user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    sources = session.scalars(select(SourceReview).order_by(SourceReview.created_at.desc())).all()
    observations = session.scalars(select(PhenotypeObservation).where(PhenotypeObservation.publish_status != "published").order_by(PhenotypeObservation.created_at.desc())).all()
    variety_map = {item.id: item for item in session.scalars(select(Variety)).all()}
    versions = {item.id: item for item in session.scalars(select(TemplateVersion)).all()}
    templates = {item.id: item for item in session.scalars(select(DataTemplate)).all()}
    source_map = {item.id: item for item in sources}
    quality_rules = active_custom_quality_rules(session)
    return {"sources": [{**serialize_source(item), "template_name": templates[versions[item.template_version_id].template_id].template_name if item.template_version_id in versions and versions[item.template_version_id].template_id in templates else "历史来源（未选模板）", "template_version": versions[item.template_version_id].version if item.template_version_id in versions else "-"} for item in sources], "pending_observations": [{**serialize_observation(item), "variety_name": variety_map[item.variety_id].variety_name if item.variety_id in variety_map else "未知品种", "template_name": templates[versions[source_map[item.source_review_id].template_version_id].template_id].template_name if item.source_review_id in source_map and source_map[item.source_review_id].template_version_id in versions and versions[source_map[item.source_review_id].template_version_id].template_id in templates else "历史来源（未选模板）", "issues": validate_observation(item, custom_quality_rules=quality_rules)} for item in observations]}


@app.post("/api/imports/upload")
async def upload_import(file: UploadFile = File(...), template_version_id: str = Query(...), actor: str = Query("数据处理员-张三"), user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    actor = audit_actor(user)
    template, version = get_template_version(session, template_version_id)
    content = await file.read()
    template, version, template_auto_switched = infer_template_from_spreadsheet(session, file.filename or "upload", content, template, version)
    content_hash = hashlib.sha256(content).hexdigest()
    if ENABLE_SOURCE_DEDUPLICATION:
        existing, reason = find_duplicate_source(session, content_hash)
        if existing:
            reject_duplicate_source(existing, reason or "相同文件内容")
    source_type, raw_text, candidates = await parse_uploaded_content(file.filename or "upload", content, template, version)
    import_metadata = candidates[0] if source_type == "html" and candidates else {}
    saved_page_url = import_metadata.pop("_saved_page_url", None)
    parsing_status = import_metadata.pop("_parsing_status", "parsed")
    quality_status = import_metadata.pop("_quality_status", "pending")
    approval_context = import_metadata.pop("_approval_context", {"approval_count": 0, "ignored_approval_count": 0})
    if ENABLE_SOURCE_DEDUPLICATION:
        existing, reason = find_duplicate_source(session, content_hash, saved_page_url)
        if existing:
            reject_duplicate_source(existing, reason or "相同文件内容")
    source_id = str(uuid.uuid4())
    safe_name = re.sub(r"[^\w.\-]+", "_", file.filename or "upload")
    project_raw_dir = RAW_STORAGE_DIR / active_project_id(session)
    project_raw_dir.mkdir(parents=True, exist_ok=True)
    target = project_raw_dir / f"{source_id}_{safe_name}"
    target.write_bytes(content)
    source = SourceReview(id=source_id, source_type=source_type, source_name=file.filename or "上传文件", source_url=saved_page_url, file_path=str(target), file_hash=content_hash, raw_text=raw_text, page_or_locator="上传文件", parsing_status=parsing_status, quality_status=quality_status, template_version_id=version.id)
    append_history(source, actor, "上传并解析文件", {"file_name": file.filename, "template": template.template_name, "template_version": version.version, "template_auto_switched": template_auto_switched, "candidate_count": len(candidates), "saved_page_url": saved_page_url, "approval_context": approval_context})
    session.add(source)
    session.commit()
    if template_auto_switched:
        for candidate in candidates:
            candidate.setdefault("parser_warnings", []).insert(0, f"已根据文件表头自动切换为“{template.template_name} {version.version}”。")
    return {"source": serialize_source(source), "template": {"id": template.id, "name": template.template_name, "version": version.version, "version_id": version.id, "auto_switched": template_auto_switched}, "candidates": candidates}


@app.post("/api/imports/url")
async def import_url(payload: UrlImport, user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    payload.actor = audit_actor(user)
    template, version = get_template_version(session, payload.template_version_id)
    if template.template_code == "rice_root_phenotype":
        raise HTTPException(422, "根系表型模板请上传 Excel 或 CSV，不支持网页导入。")
    source_url = payload.url.strip()
    if ENABLE_SOURCE_DEDUPLICATION:
        existing, reason = find_duplicate_source(session, "", source_url)
        if existing:
            reject_duplicate_source(existing, reason or "相同原始网页地址")
    try:
        async with httpx.AsyncClient(timeout=15, follow_redirects=True, headers={"User-Agent": "RiceDataGovernanceDemo/1.0 (single-page preview)"}) as client:
            response = await client.get(source_url)
            response.raise_for_status()
    except Exception as exc:
        raise HTTPException(422, f"读取网页失败，请保存HTML后上传。原因：{str(exc)[:120]}") from exc
    content = response.content
    content_hash = hashlib.sha256(content).hexdigest()
    if ENABLE_SOURCE_DEDUPLICATION:
        existing, reason = find_duplicate_source(session, content_hash)
        if existing:
            reject_duplicate_source(existing, reason or "相同文件内容")
    html = decode_html_bytes(content)
    soup = BeautifulSoup(html, "html.parser")
    async with httpx.AsyncClient(timeout=15, follow_redirects=True, headers={"User-Agent": "RiceDataGovernanceDemo/1.0 (single-page preview)"}) as glyph_client:
        resolved_digits = await resolve_rice_data_digit_glyphs(soup, source_url, glyph_client)
    raw_title = extract_variety_title(soup)
    raw_text, masked_digits = extract_html_text_with_markers(soup)
    raw_text, approval_context = keep_first_approval_section(raw_text)
    name, aliases = parse_title(raw_title)
    source_id = str(uuid.uuid4())
    project_raw_dir = RAW_STORAGE_DIR / active_project_id(session)
    project_raw_dir.mkdir(parents=True, exist_ok=True)
    target = project_raw_dir / f"{source_id}.html"
    target.write_bytes(content)
    source = SourceReview(id=source_id, source_type="webpage", source_name=raw_title or source_url, source_url=source_url, file_path=str(target), file_hash=content_hash, raw_text=raw_text, page_or_locator="网页正文", parsing_status="partial" if masked_digits else "parsed", quality_status="requires_manual_check" if masked_digits else "pending", template_version_id=version.id)
    append_history(source, payload.actor, "读取单个网页", {"url": source_url, "resolved_digit_glyphs": resolved_digits, "approval_context": approval_context})
    session.add(source)
    session.commit()
    candidate = {"variety_name": name, "aliases": aliases, "raw_title": raw_title, "observations": extract_text_observations(raw_text), "source_locator": "网页正文", **extract_variety_basic_info(raw_text)}
    parser_warnings: list[str] = []
    if resolved_digits:
        parser_warnings.append(f"已识别并还原 {resolved_digits} 个网页图形数字；原始 HTML 已原样保存，可在审核时追溯。")
    if masked_digits:
        parser_warnings.append(f"网页正文仍有 {masked_digits} 个未识别图形字符，相关字段未自动标准化，请以原始网页人工核对。")
    if approval_context["ignored_approval_count"]:
        parser_warnings.append(f"检测到 {approval_context['approval_count']} 条审定记录。第一版仅处理第一条，后续 {approval_context['ignored_approval_count']} 条未解析、未写入数据库。")
    if parser_warnings:
        candidate["parser_warnings"] = parser_warnings
    return {"source": serialize_source(source), "candidates": [candidate]}


@app.post("/api/imports/{source_id}/reprocess")
async def reprocess_import(source_id: str, actor: str = Query("数据处理员-张三"), user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    actor = audit_actor(user)
    source = session.get(SourceReview, source_id)
    if not source or not source.template_version_id or not source.file_path:
        raise HTTPException(422, "该来源缺少可重新处理的原始文件或模板信息。")
    previous = session.get(TemplateVersion, source.template_version_id)
    template = session.get(DataTemplate, previous.template_id) if previous else None
    current = session.get(TemplateVersion, template.current_version_id) if template else None
    if not template or not current:
        raise HTTPException(422, "关联标准模板不可用。")
    path = Path(source.file_path)
    if not path.exists():
        raise HTTPException(422, "未找到本地原始文件，无法重新处理。")
    content = path.read_bytes()
    template, current, template_auto_switched = infer_template_from_spreadsheet(session, source.source_name, content, template, current)
    source_type, raw_text, candidates = await parse_uploaded_content(source.source_name, content, template, current)
    source.template_version_id = current.id
    source.raw_text = raw_text
    source.source_type = source_type
    append_history(source, actor, "按最新模板重新处理", {"from_version": previous.version if previous else "-", "to_version": current.version, "template_auto_switched": template_auto_switched, "candidate_count": len(candidates)})
    session.commit()
    if template_auto_switched:
        for candidate in candidates:
            candidate.setdefault("parser_warnings", []).insert(0, f"已根据文件表头自动切换为“{template.template_name} {current.version}”。")
    return {"source": serialize_source(source), "template": {"id": template.id, "name": template.template_name, "version": current.version, "version_id": current.id, "auto_switched": template_auto_switched}, "candidates": candidates}


@app.post("/api/imports/{source_id}/commit")
def commit_import(source_id: str, payload: ImportCommit, user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    payload.actor = audit_actor(user)
    source = session.get(SourceReview, source_id)
    if not source:
        raise HTTPException(404, "来源记录不存在")
    template, template_version = get_template_version(session, source.template_version_id)
    trait_catalog = ROOT_TRAITS if template.template_code == "rice_root_phenotype" else TRAITS
    name, title_aliases = parse_title(payload.variety_name or payload.raw_title)
    aliases = list(dict.fromkeys([*title_aliases, *payload.aliases]))
    if not name:
        raise HTTPException(422, "请先填写品种名称后再创建草稿")
    existing = session.scalars(select(Variety).where(Variety.normalized_name == normalize_name(name))).first()
    basic_info = {field: getattr(payload, field) for field in VARIETY_BASIC_FIELDS}
    variety = existing or Variety(
        variety_name=name,
        normalized_name=normalize_name(name),
        alias_names=aliases,
        raw_variety_title=payload.raw_title or name,
        source_review_id=source.id,
        data_status="pending",
        **basic_info,
    )
    if existing:
        variety.alias_names = list(dict.fromkeys([*(variety.alias_names or []), *aliases]))
        for field, value in basic_info.items():
            if value and not getattr(variety, field):
                setattr(variety, field, value)
    else:
        session.add(variety)
        session.flush()
    created: list[tuple[PhenotypeObservation, bool, str | None]] = []
    existing_trait_codes = {
        item.trait_code
        for item in session.scalars(select(PhenotypeObservation).where(PhenotypeObservation.variety_id == variety.id)).all()
    }
    if template.template_code == "rice_root_phenotype":
        existing_trait_codes.update(
            session.scalars(
                select(RootPhenotypeObservation.trait_code).where(RootPhenotypeObservation.variety_id == variety.id)
            ).all()
        )
    skipped_duplicates: list[dict[str, str]] = []
    for item in payload.observations:
        code = item.get("trait_code")
        trait = trait_catalog.get(code)
        if not trait:
            continue
        observation = PhenotypeObservation(variety_id=variety.id, source_review_id=source.id, trait_code=code, trait_name=item.get("trait_name") or trait["name"], trait_category=item.get("trait_category") or trait["category"], observation_type=item.get("observation_type") or "numeric", value_numeric=item.get("value_numeric"), value_text=item.get("value_text"), unit=item.get("unit") if item.get("unit") is not None else trait["unit"], original_value=item.get("original_value") or "", original_field=item.get("original_field"), source_locator=item.get("source_locator"), rule_version=template_version.version, quality_status="pending", publish_status="pending")
        if code in existing_trait_codes:
            skipped_duplicates.append({
                "trait_code": code,
                "trait_name": observation.trait_name,
                "reason": "该品种已存在同一标准字段，未重复创建",
            })
            continue
        existing_trait_codes.add(code)
        created.append((observation, bool(item.get("requires_confirmation")), item.get("conversion_suggestion")))

    all_observations = [observation for observation, _, _ in created]
    quality_rules = active_custom_quality_rules(session)
    for observation, requires_confirmation, conversion_suggestion in created:
        issues = validate_observation(observation, all_observations, quality_rules) + template_quality_issues(observation, template_version)
        if requires_confirmation:
            observation.quality_status = "blocked"
            observation.review_comment = conversion_suggestion or "单位或换算方式需要数据处理员确认。"
        elif any(issue["severity"] == "block" for issue in issues):
            observation.quality_status = "blocked"
        elif issues:
            observation.quality_status = "warning"
        else:
            observation.quality_status = "passed"
        session.add(observation)
    append_history(source, payload.actor, "创建待处理草稿", {
        "variety": variety.variety_name,
        "template": template.template_name,
        "template_version": template_version.version,
        "observation_count": len(created),
        "skipped_duplicate_count": len(skipped_duplicates),
        "skipped_duplicate_traits": [item["trait_code"] for item in skipped_duplicates],
        "existing_variety": bool(existing),
    })
    session.commit()
    return {
        "variety": serialize_variety(variety),
        "observations": [serialize_observation(item) for item, _, _ in created],
        "skipped_duplicates": skipped_duplicates,
    }


@app.post("/api/varieties")
def create_variety(payload: VarietyCreate, user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    payload.actor = audit_actor(user)
    if session.scalars(select(Variety).where(Variety.normalized_name == normalize_name(payload.variety_name))).first():
        raise HTTPException(409, "标准品种名称已存在，请在详情页中补充别名或数据。")
    variety = Variety(variety_name=payload.variety_name.strip(), normalized_name=normalize_name(payload.variety_name), alias_names=payload.aliases, raw_variety_title=payload.variety_name, variety_type=payload.variety_type, source_review_id=payload.source_review_id, data_status="draft")
    session.add(variety)
    if payload.source_review_id:
        append_history(session.get(SourceReview, payload.source_review_id), payload.actor, "手工新建品种", {"variety": variety.variety_name})
    session.commit()
    return serialize_variety(variety)


@app.post("/api/observations")
def create_observation(payload: ObservationCreate, user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    payload.actor = audit_actor(user)
    variety = session.get(Variety, payload.variety_id)
    if not variety:
        raise HTTPException(404, "品种不存在")
    trait = TRAITS.get(payload.trait_code)
    if not trait:
        raise HTTPException(422, "请选择标准性状；未识别字段应保留为原始文本。")
    if session.scalar(
        select(PhenotypeObservation.id).where(
            PhenotypeObservation.variety_id == variety.id,
            PhenotypeObservation.trait_code == payload.trait_code,
        )
    ):
        raise HTTPException(409, "该品种已存在这个标准字段。请编辑已有记录，不要重复新增。")
    item = PhenotypeObservation(variety_id=variety.id, source_review_id=payload.source_review_id, trait_code=payload.trait_code, trait_name=trait["name"], trait_category=trait["category"], observation_type="numeric" if payload.value_numeric is not None else "text", value_numeric=payload.value_numeric, value_text=payload.value_text, unit=payload.unit if payload.unit is not None else trait["unit"], original_value=payload.original_value, source_locator=payload.source_locator or "手工录入", rule_version="v1.0", quality_status="pending", publish_status="pending")
    issues = validate_observation(item, custom_quality_rules=active_custom_quality_rules(session))
    item.quality_status = "blocked" if any(issue["severity"] == "block" for issue in issues) else "warning" if issues else "passed"
    session.add(item)
    append_history(session.get(SourceReview, payload.source_review_id) if payload.source_review_id else None, payload.actor, "手工新增表型", {"trait": trait["name"], "value": payload.original_value})
    session.commit()
    return {**serialize_observation(item), "issues": issues}


@app.get("/api/manage/varieties")
def manage_varieties(user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> list[dict[str, Any]]:
    return [serialize_variety(item) for item in session.scalars(select(Variety).order_by(Variety.variety_name)).all()]


@app.post("/api/manual/record")
def create_manual_record(payload: ManualRecord, user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    payload.actor = audit_actor(user)
    trait = TRAITS.get(payload.trait_code)
    if not trait:
        raise HTTPException(422, "请选择标准性状；未知字段请在导入流程中保留为原文。")
    if not payload.source_reference.strip():
        raise HTTPException(422, "手工数据必须填写来源说明后才能保存。")
    variety = session.get(Variety, payload.variety_id) if payload.variety_id else None
    if not variety:
        if not payload.variety_name.strip():
            raise HTTPException(422, "请选择已有品种或填写新建品种名称。")
        normalized = normalize_name(payload.variety_name)
        variety = session.scalars(select(Variety).where(Variety.normalized_name == normalized)).first()
        if not variety:
            variety = Variety(variety_name=payload.variety_name.strip(), normalized_name=normalized, alias_names=payload.aliases, raw_variety_title=payload.variety_name.strip(), variety_type=payload.variety_type, data_status="draft")
            session.add(variety)
            session.flush()
    if session.scalar(
        select(PhenotypeObservation.id).where(
            PhenotypeObservation.variety_id == variety.id,
            PhenotypeObservation.trait_code == payload.trait_code,
        )
    ):
        raise HTTPException(409, "该品种已存在这个标准字段。请编辑已有记录，不要重复新增。")
    source = SourceReview(source_type="manual", source_name=payload.source_reference.strip(), raw_text=payload.source_note or payload.original_value, page_or_locator="手工补录", parsing_status="manual", quality_status="pending")
    session.add(source)
    session.flush()
    observation = PhenotypeObservation(variety_id=variety.id, source_review_id=source.id, trait_code=payload.trait_code, trait_name=trait["name"], trait_category=trait["category"], observation_type="numeric" if payload.value_numeric is not None else "text", value_numeric=payload.value_numeric, value_text=payload.value_text, unit=payload.unit if payload.unit is not None else trait["unit"], original_value=payload.original_value, source_locator="手工补录", rule_version="v1.0", quality_status="pending", publish_status="pending")
    issues = validate_observation(observation, custom_quality_rules=active_custom_quality_rules(session))
    observation.quality_status = "blocked" if any(item["severity"] == "block" for item in issues) else "warning" if issues else "passed"
    session.add(observation)
    append_history(source, payload.actor, "手工补录表型", {"variety": variety.variety_name, "trait": trait["name"], "original_value": payload.original_value, "source_reference": payload.source_reference})
    session.commit()
    return {"variety": serialize_variety(variety), "observation": serialize_observation(observation), "issues": issues}


@app.patch("/api/observations/{observation_id}")
def update_observation(observation_id: str, payload: ObservationUpdate, user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    payload.actor = audit_actor(user)
    item = session.get(PhenotypeObservation, observation_id)
    if not item:
        raise HTTPException(404, "表型记录不存在")
    before = {"value_numeric": item.value_numeric, "value_text": item.value_text, "unit": item.unit, "original_value": item.original_value}
    item.value_numeric = payload.value_numeric
    item.value_text = payload.value_text
    item.unit = payload.unit
    item.original_value = payload.original_value or item.original_value
    item.review_comment = payload.review_comment
    all_items = session.scalars(select(PhenotypeObservation).where(PhenotypeObservation.variety_id == item.variety_id)).all()
    issues = validate_observation(item, all_items, active_custom_quality_rules(session))
    item.quality_status = "blocked" if any(issue["severity"] == "block" for issue in issues) else "warning" if issues else "passed"
    item.publish_status = "pending"
    append_history(session.get(SourceReview, item.source_review_id) if item.source_review_id else None, payload.actor, "修改表型记录", {"observation_id": item.id, "before": before, "after": {"value_numeric": item.value_numeric, "value_text": item.value_text, "unit": item.unit, "original_value": item.original_value}, "reason": payload.review_comment})
    session.commit()
    return {**serialize_observation(item), "issues": issues}


@app.post("/api/observations/publish")
def publish_observations(payload: PublishRequest, user: CurrentUser = Depends(require_data_processor), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    payload.actor = audit_actor(user)
    published, blocked, skipped_duplicates = [], [], []
    quality_rules = active_custom_quality_rules(session)
    for observation_id in payload.observation_ids:
        item = session.get(PhenotypeObservation, observation_id)
        if not item:
            continue
        siblings = session.scalars(select(PhenotypeObservation).where(PhenotypeObservation.variety_id == item.variety_id)).all()
        source = session.get(SourceReview, item.source_review_id)
        source_version = session.get(TemplateVersion, source.template_version_id) if source and source.template_version_id else None
        issues = validate_observation(item, siblings, quality_rules) + template_quality_issues(item, source_version)
        if any(issue["severity"] == "block" for issue in issues) or not item.source_review_id:
            item.publish_status = "pending"
            item.quality_status = "blocked"
            blocked.append({"id": item.id, "issues": issues or [{"rule": "R003", "severity": "block", "message": "缺少来源信息。"}]})
            continue
        item.publish_status = "published"
        item.quality_status = "warning" if issues else "passed"
        variety = session.get(Variety, item.variety_id)
        if variety:
            variety.data_status = "published"
        template = session.get(DataTemplate, session.get(TemplateVersion, source.template_version_id).template_id) if source and source.template_version_id and session.get(TemplateVersion, source.template_version_id) else None
        if template and template.template_code == "rice_root_phenotype":
            existing_root = session.scalar(
                select(RootPhenotypeObservation.id).where(
                    RootPhenotypeObservation.variety_id == item.variety_id,
                    RootPhenotypeObservation.trait_code == item.trait_code,
                )
            )
            if existing_root:
                append_history(source, payload.actor, "跳过重复根系正式字段", {
                    "observation_id": item.id,
                    "trait": item.trait_name,
                    "rule": "同一品种 + 同一标准字段仅保留一条",
                })
                session.delete(item)
                skipped_duplicates.append(item.id)
                continue
            session.add(RootPhenotypeObservation(variety_id=item.variety_id, source_review_id=item.source_review_id, trait_code=item.trait_code, trait_name=item.trait_name, trait_category=item.trait_category, value_numeric=item.value_numeric, value_text=item.value_text, unit=item.unit, original_value=item.original_value, original_field=item.original_field, source_locator=item.source_locator, template_version=item.rule_version))
            session.delete(item)
        append_history(source, payload.actor, "发布表型记录", {"observation_id": item.id, "trait": item.trait_name, "issues": issues, "target_table": template.target_table if template else "phenotype_observation"})
        published.append(item.id)
    session.commit()
    return {"published": published, "blocked": blocked, "skipped_duplicates": skipped_duplicates}


@app.get("/api/rules")
def list_rules(user: CurrentUser = Depends(require_field_admin), session: Session = Depends(get_business_project_session)) -> list[dict[str, Any]]:
    rules = session.scalars(select(DataRule).order_by(DataRule.rule_code, DataRule.created_at.desc())).all()
    return [{"id": item.id, "rule_code": item.rule_code, "rule_name": item.rule_name, "rule_type": item.rule_type, "version": item.version, "severity": item.severity, "config": item.config, "status": item.status, "change_reason": item.change_reason, "created_by": item.created_by, "created_at": item.created_at.isoformat()} for item in rules]


@app.post("/api/rules")
def create_rule(payload: RuleCreate, user: CurrentUser = Depends(require_field_admin), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    payload.created_by = audit_actor(user)
    existing = session.scalars(select(DataRule).where(DataRule.rule_code == payload.rule_code).order_by(DataRule.created_at.desc())).first()
    version = "v1.0"
    if existing:
        match = re.match(r"v(\d+)\.(\d+)", existing.version)
        version = f"v{match.group(1)}.{int(match.group(2)) + 1}" if match else "v1.1"
    rule = DataRule(rule_code=payload.rule_code, rule_name=payload.rule_name, rule_type=payload.rule_type, version=version, severity=payload.severity, config=payload.config, status="published", change_reason=payload.change_reason, created_by=payload.created_by)
    session.add(rule)
    session.commit()
    return {"id": rule.id, "rule_code": rule.rule_code, "version": rule.version, "status": rule.status}


@app.post("/api/rules/{rule_id}/retire")
def retire_rule(rule_id: str, actor: str = Query("数据处理员-张三"), user: CurrentUser = Depends(require_field_admin), session: Session = Depends(get_business_project_session)) -> dict[str, Any]:
    actor = audit_actor(user)
    rule = session.get(DataRule, rule_id)
    if not rule:
        raise HTTPException(404, "规则不存在")
    rule.status = "retired"
    rule.change_reason = f"{rule.change_reason}；由{actor}停用"
    session.commit()
    return {"id": rule.id, "status": rule.status}


@app.get("/api/varieties")
def search_varieties(
    q: str = "",
    height_max: float | None = None,
    grain_weight_min: float | None = None,
    blast_max: float | None = None,
    user: CurrentUser = Depends(require_published_data_reader),
    session: Session = Depends(get_business_project_session),
) -> list[dict[str, Any]]:
    varieties = session.scalars(select(Variety).order_by(Variety.variety_name)).all()
    result = []
    for variety in varieties:
        obs = session.scalars(select(PhenotypeObservation).where(PhenotypeObservation.variety_id == variety.id, PhenotypeObservation.publish_status == "published")).all()
        if not obs:
            continue
        search_text = " ".join([variety.variety_name, *(variety.alias_names or [])]).lower()
        if q and q.lower() not in search_text:
            continue
        traits = {item.trait_code: item.value_numeric if item.value_numeric is not None else item.value_text for item in obs}
        if height_max is not None and (traits.get("plant_height") is None or float(traits["plant_height"]) > height_max):
            continue
        if grain_weight_min is not None and (traits.get("thousand_grain_weight") is None or float(traits["thousand_grain_weight"]) < grain_weight_min):
            continue
        leaf_score = traits.get("leaf_blast_score")
        if blast_max is not None and (leaf_score is None or float(leaf_score) > blast_max):
            continue
        result.append({**serialize_variety(variety), "traits": traits, "observation_count": len(obs)})
    return result


@app.get("/api/varieties/{variety_id}")
def variety_detail(
    variety_id: str,
    user: CurrentUser = Depends(require_published_data_reader),
    session: Session = Depends(get_business_project_session),
) -> dict[str, Any]:
    variety = session.get(Variety, variety_id)
    if not variety or variety.data_status != "published":
        raise HTTPException(404, "品种不存在")
    observations = session.scalars(
        select(PhenotypeObservation)
        .where(PhenotypeObservation.variety_id == variety_id, PhenotypeObservation.publish_status == "published")
        .order_by(PhenotypeObservation.trait_category, PhenotypeObservation.trait_name)
    ).all()
    source_ids = {item.source_review_id for item in observations if item.source_review_id}
    sources = [session.get(SourceReview, item) for item in source_ids]
    source_summaries = [{
        "id": item.id,
        "source_type": item.source_type,
        "source_name": item.source_name,
        "source_url": item.source_url,
        "page_or_locator": item.page_or_locator,
        "created_at": item.created_at.isoformat(),
    } for item in sources if item]
    return {**serialize_variety(variety), "observations": [{**serialize_observation(item), "issues": validate_observation(item, observations)} for item in observations], "sources": source_summaries}


@app.get("/api/sources/{source_id}/raw")
def raw_source(
    source_id: str,
    user: CurrentUser = Depends(require_data_processor),
) -> dict[str, Any]:
    raise HTTPException(404, "原始来源全文预览暂未在数据处理工作台开放。")


@app.post("/api/reports/pdf")
def pdf_report(
    payload: PdfReport,
    user: CurrentUser = Depends(require_published_data_reader),
    session: Session = Depends(get_business_project_session),
) -> Response:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    buffer = io.BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    chinese = styles["BodyText"].clone("Chinese")
    chinese.fontName = "STSong-Light"
    chinese.fontSize = 9
    title = styles["Title"].clone("ChineseTitle")
    title.fontName = "STSong-Light"
    story = [Paragraph("水稻表型查询与基础分析报告", title), Spacer(1, 6 * mm), Paragraph(f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}", chinese), Paragraph(f"查询条件：{json.dumps(payload.filters, ensure_ascii=False) or '全部已发布数据'}", chinese), Spacer(1, 4 * mm)]
    rows = [["品种", "株高(cm)", "千粒重(g)", "亩产(kg/亩)", "叶瘟等级"]]
    for item in payload.rows[:30]:
        traits = item.get("traits", {})
        rows.append([item.get("variety_name", ""), str(traits.get("plant_height", "-")), str(traits.get("thousand_grain_weight", "-")), str(traits.get("yield_per_mu", "-")), str(traits.get("leaf_blast_score", "-"))])
    table = Table(rows, colWidths=[42 * mm, 28 * mm, 28 * mm, 32 * mm, 24 * mm])
    table.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), "STSong-Light"), ("FONTSIZE", (0, 0), (-1, -1), 8), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8f2ed")), ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#b7c7bf")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fbf9")])]))
    story.append(table)
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph("说明：本报告仅基于已发布标准数据生成。原始文件不包含在下载内容中；抗病等级未在缺少正式评价体系时自动转换为抗性分类。", chinese))
    document.build(story)
    return Response(content=buffer.getvalue(), media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=rice-phenotype-report.pdf"})


async def execute_longyun_acps_partner(question: str, caller_aic: str) -> AcpsExecutionResult:
    """Run an ACPs Partner task inside the published-data-only boundary.

    External ACPs callers never inherit a browser user's private knowledge
    scope, attachments, chat history, or public-web search configuration.
    """
    with SessionLocal() as session:
        published_context, evidence_cards = await build_published_evidence_context(
            session,
            question.strip(),
            requested_by=f"ACPs:{caller_aic or 'unknown-leader'}",
        )

    completed: dict[str, Any] | None = None
    async for event in stream_research_reply(
        user_prompt=question.strip(),
        evidence_context=published_context,
        memory_state={},
        public_web_context="",
        vision_images=[],
        conversation_history=[],
        has_current_vision_images=False,
    ):
        if event["type"] == "complete":
            completed = event
    if not completed or not str(completed.get("content") or "").strip():
        raise ResearchAgentError("隆耘智能体未生成可展示的 ACPs 任务结果。")

    evidence_summary = [
        {
            key: card.get(key)
            for key in ("type", "title", "detail", "query_template", "query_parameters")
            if card.get(key) is not None
        }
        for card in evidence_cards
    ]
    return AcpsExecutionResult(
        text=str(completed["content"]).strip(),
        structured_data={
            "dataBoundary": "published-standard-data-only",
            "evidenceCount": len(evidence_summary),
            "evidence": evidence_summary,
        },
    )


@app.post("/api/acps/leader/dispatch", tags=["ACPs"])
async def acps_leader_dispatch(
    payload: AcpsLeaderDispatchRequest,
    user: CurrentUser = Depends(require_researcher),
) -> dict[str, Any]:
    """Let an authenticated researcher use Longyun as an ACPs Leader."""
    try:
        project = resolve_project_access(session, user, x_project_id)
        _set_active_project(session, project.id)
        result = await dispatch_partner_task(payload, ACPS_SETTINGS)
    except TimeoutError as exc:
        raise HTTPException(504, str(exc)) from exc
    except (httpx.HTTPError, RuntimeError, ValueError) as exc:
        raise HTTPException(502, str(exc)) from exc
    except Exception as exc:
        logger.exception("ACPs Leader dispatch failed")
        raise HTTPException(502, "ACPs Leader 调度失败，请检查 Partner、Discovery 和证书配置。") from exc
    return {
        **result,
        "requestedBy": audit_actor(user),
        "leaderAic": ACPS_SETTINGS.aic or None,
    }


mount_acps_routes(app, ACPS_SETTINGS, execute_longyun_acps_partner)
