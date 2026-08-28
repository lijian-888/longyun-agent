"""Institution-scoped object storage and structured data ingestion.

Raw files are written to one private MinIO bucket per institution. Parsed
records are written to a separate PostgreSQL database per institution while
retaining institution_id, project_id and stable entity identifiers in every
row. The module deliberately contains no FastAPI or authentication code; the
main application resolves the verified account/project boundary before calling
this service.
"""

from __future__ import annotations

import csv
import gzip
import hashlib
import io
import json
import os
import re
import tempfile
import uuid
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine, make_url


REGULAR_MAX_BYTES = 200 * 1024 * 1024
GENOTYPE_ARCHIVE_MAX_BYTES = 2 * 1024 * 1024 * 1024
UPLOAD_CHUNK_BYTES = 8 * 1024 * 1024

DATASET_LABELS = {
    "germplasm": "种质资源数据",
    "pedigree": "系谱数据",
    "phenotype": "试验/表型数据",
    "environment": "环境数据",
    "genotype": "基因型数据",
    "literature": "公开文献",
}

TABULAR_DATASETS = {"germplasm", "pedigree", "phenotype", "environment"}
SUPPORTED_SUFFIXES = {
    "germplasm": {".csv", ".xlsx", ".json"},
    "pedigree": {".csv", ".xlsx", ".json"},
    "phenotype": {".csv", ".xlsx", ".json"},
    "environment": {".csv", ".xlsx", ".json"},
    "genotype": {".vcf", ".vcf.gz", ".zip"},
    "literature": {".pdf", ".docx", ".txt"},
}

FIELD_ALIASES: dict[str, dict[str, tuple[str, ...]]] = {
    "germplasm": {
        "germplasm_id": ("germplasm_id", "material_id", "material_code", "accession_id", "种质编号", "材料编号", "材料编码", "资源编号"),
        "name": ("name", "material_name", "germplasm_name", "种质名称", "材料名称", "品种名称"),
        "species": ("species", "crop", "物种", "作物"),
        "origin": ("origin", "source", "原产地", "来源"),
    },
    "pedigree": {
        "child_id": ("child_id", "offspring_id", "material_id", "子代编号", "后代材料", "材料编号"),
        "female_parent_id": ("female_parent_id", "mother_id", "maternal_id", "母本编号", "母本"),
        "male_parent_id": ("male_parent_id", "father_id", "paternal_id", "父本编号", "父本"),
        "generation": ("generation", "世代"),
    },
    "phenotype": {
        "observation_id": ("observation_id", "record_id", "iid", "观测编号", "记录编号"),
        "germplasm_id": ("germplasm_id", "material_id", "material_code", "accession_id", "种质编号", "材料编号", "材料编码"),
        "trial_id": ("trial_id", "experiment_id", "试验编号"),
        "environment_id": ("environment_id", "site_year_id", "analysis_environment", "环境编号", "分析环境", "点位年份编号"),
        "trait_code": ("trait_code", "trait", "性状编号", "性状"),
        "value": ("value", "trait_value", "观测值", "性状值"),
        "unit": ("unit", "单位"),
    },
    "environment": {
        "environment_id": ("environment_id", "site_year_id", "环境编号", "点位年份编号"),
        "location": ("location", "site", "试验地点", "地点"),
        "year": ("year", "trial_year", "年份", "试验年份"),
        "season": ("season", "季节", "作季"),
        "soil_type": ("soil_type", "土壤类型"),
    },
}

REQUIRED_FIELDS = {
    "germplasm": ("germplasm_id", "name"),
    "pedigree": ("child_id",),
    "phenotype": ("germplasm_id", "trait_code", "value"),
    "environment": ("environment_id",),
}

IMPACTED_FEATURES = {
    "germplasm": ["种质查询", "系谱追溯", "表型关联", "基因型映射"],
    "pedigree": ["系谱图", "亲本关系分析", "育种材料档案"],
    "phenotype": ["性状查询", "区域试验分析", "GWAS"],
    "environment": ["多环境分析", "G×E 分析", "试验可追溯性"],
    "genotype": ["基因型质控", "GWAS", "群体结构分析"],
    "literature": ["文献检索", "知识库问答", "证据引用"],
}


class InstitutionDataError(ValueError):
    """An actionable data-ingestion error safe to show to the operator."""


@dataclass(frozen=True)
class InstitutionDataSettings:
    minio_endpoint: str
    minio_access_key: str
    minio_secret_key: str
    minio_secure: bool
    minio_region: str | None
    migration_database_url: str
    application_database_url: str
    application_database_role: str
    staging_dir: Path

    @classmethod
    def from_env(cls) -> "InstitutionDataSettings":
        return cls(
            minio_endpoint=os.getenv("MINIO_ENDPOINT", "minio:9000").strip(),
            minio_access_key=os.getenv("MINIO_ROOT_USER", "longyun-minio").strip(),
            minio_secret_key=os.getenv("MINIO_ROOT_PASSWORD", "").strip(),
            minio_secure=os.getenv("MINIO_SECURE", "false").strip().lower() in {"1", "true", "yes", "on"},
            minio_region=os.getenv("MINIO_REGION", "").strip() or None,
            migration_database_url=os.getenv("MIGRATION_DATABASE_URL", "").strip(),
            application_database_url=os.getenv("DATABASE_URL", "").strip(),
            application_database_role=os.getenv("APP_DATABASE_ROLE", "rice_app").strip(),
            staging_dir=Path(os.getenv("INSTITUTION_UPLOAD_STAGING_DIR", "/data/raw/institution-staging")),
        )


@dataclass(frozen=True)
class StagedUpload:
    path: Path
    file_name: str
    size_bytes: int
    sha256: str
    suffix: str


def safe_identifier(value: str, prefix: str, max_length: int = 63) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", str(value).lower()).strip("_")
    if not normalized:
        normalized = hashlib.sha256(str(value).encode("utf-8")).hexdigest()[:12]
    return f"{prefix}_{normalized}"[:max_length].rstrip("_")


def safe_object_name(value: str) -> str:
    # Browser filenames may contain Chinese characters. Sanitising the entire
    # name with an ASCII-only expression used to turn `连续性状.xlsx` into
    # `xlsx`, silently losing the dot and causing a false "no extension" 422.
    # Split and preserve the complete extension before cleaning the stem.
    name = re.split(r"[\\/]", str(value or "upload"))[-1]
    suffix = complete_suffix(name)
    stem = name[:-len(suffix)] if suffix else name
    cleaned_stem = re.sub(r"[^\w()（）\-]+", "_", stem, flags=re.UNICODE).strip("._-")
    if not cleaned_stem:
        cleaned_stem = f"upload-{hashlib.sha256(name.encode('utf-8')).hexdigest()[:12]}"
    safe_suffix = re.sub(r"[^A-Za-z0-9.]+", "", suffix.lower())
    available = max(1, 240 - len(safe_suffix))
    return f"{cleaned_stem[:available].rstrip('._-')}{safe_suffix}"


def complete_suffix(file_name: str) -> str:
    lower = str(file_name).lower()
    return ".vcf.gz" if lower.endswith(".vcf.gz") else Path(lower).suffix


def upload_limit(dataset_type: str, file_name: str) -> int:
    suffix = complete_suffix(file_name)
    if dataset_type == "genotype" and suffix in {".vcf.gz", ".zip"}:
        return GENOTYPE_ARCHIVE_MAX_BYTES
    return REGULAR_MAX_BYTES


def validate_file_contract(dataset_type: str, file_name: str, size_bytes: int) -> str:
    if dataset_type not in DATASET_LABELS:
        raise InstitutionDataError("数据类型不受支持。")
    suffix = complete_suffix(file_name)
    if suffix not in SUPPORTED_SUFFIXES[dataset_type]:
        allowed = "、".join(sorted(SUPPORTED_SUFFIXES[dataset_type]))
        raise InstitutionDataError(f"{DATASET_LABELS[dataset_type]}不支持 {suffix or '无扩展名'}；允许：{allowed}。")
    limit = upload_limit(dataset_type, file_name)
    if size_bytes > limit:
        label = "2GB" if limit == GENOTYPE_ARCHIVE_MAX_BYTES else "200MB"
        raise InstitutionDataError(f"{file_name} 超过 {label} 上限。")
    if size_bytes <= 0:
        raise InstitutionDataError("上传文件为空。")
    return suffix


async def stage_upload(upload: Any, dataset_type: str, staging_dir: Path) -> StagedUpload:
    """Stream an UploadFile to disk while enforcing the limit during transfer."""
    file_name = safe_object_name(getattr(upload, "filename", "upload") or "upload")
    suffix = complete_suffix(file_name)
    if dataset_type not in DATASET_LABELS or suffix not in SUPPORTED_SUFFIXES.get(dataset_type, set()):
        validate_file_contract(dataset_type, file_name, 1)
    limit = upload_limit(dataset_type, file_name)
    staging_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256()
    size = 0
    descriptor, temporary_name = tempfile.mkstemp(prefix="institution-import-", suffix=suffix, dir=staging_dir)
    os.close(descriptor)
    target = Path(temporary_name)
    try:
        with target.open("wb") as handle:
            while True:
                chunk = await upload.read(UPLOAD_CHUNK_BYTES)
                if not chunk:
                    break
                size += len(chunk)
                if size > limit:
                    label = "2GB" if limit == GENOTYPE_ARCHIVE_MAX_BYTES else "200MB"
                    raise InstitutionDataError(f"{file_name} 超过 {label} 上限，上传已终止。")
                digest.update(chunk)
                handle.write(chunk)
        validate_file_contract(dataset_type, file_name, size)
        return StagedUpload(target, file_name, size, digest.hexdigest(), suffix)
    except Exception:
        target.unlink(missing_ok=True)
        raise


def _decode_text(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise InstitutionDataError("文本编码无法识别，请使用 UTF-8 或 GB18030。")


def parse_tabular_file(path: Path, suffix: str, dataset_type: str | None = None) -> list[dict[str, Any]]:
    if suffix == ".csv":
        content = _decode_text(path.read_bytes())
        try:
            dialect = csv.Sniffer().sniff(content[:8192], delimiters=",\t;")
        except csv.Error:
            dialect = csv.excel
        return [dict(row) for row in csv.DictReader(io.StringIO(content), dialect=dialect)]
    if suffix == ".xlsx":
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            aliases = {
                _normalized_header(alias)
                for values in FIELD_ALIASES.get(dataset_type or "", {}).values()
                for alias in values
            }
            candidates: list[tuple[tuple[int, int, int, int], Any, int, list[str]]] = []
            for sheet_index, sheet in enumerate(workbook.worksheets):
                for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True)):
                    headers = [str(value or "").strip() for value in row]
                    non_empty = [header for header in headers if header]
                    if len(non_empty) < 2:
                        continue
                    recognized = sum(_normalized_header(header) in aliases for header in non_empty)
                    candidates.append(((recognized, len(non_empty), -sheet_index, -row_index), sheet, row_index, headers))
            if not candidates:
                raise InstitutionDataError("XLSX 中没有找到至少包含两列的表头。")
            _, sheet, header_index, headers = max(candidates, key=lambda item: item[0])
            rows = sheet.iter_rows(min_row=header_index + 2, values_only=True)
            result = [
                {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
                for row in rows
                if any(value not in (None, "") for value in row)
            ]
            workbook.close()
            if not result:
                raise InstitutionDataError("XLSX 表头下没有可导入的数据行。")
            return result
        except InstitutionDataError:
            raise
        except (StopIteration, OSError, ValueError, zipfile.BadZipFile) as exc:
            raise InstitutionDataError("无法读取 XLSX，请按标准模板重新保存。") from exc
    if suffix == ".json":
        try:
            payload = json.loads(_decode_text(path.read_bytes()))
        except json.JSONDecodeError as exc:
            raise InstitutionDataError(f"JSON 格式错误：第 {exc.lineno} 行。") from exc
        if isinstance(payload, dict):
            payload = payload.get("records", payload.get("data", [payload]))
        if not isinstance(payload, list) or not all(isinstance(item, dict) for item in payload):
            raise InstitutionDataError("JSON 必须是对象数组，或包含 records/data 数组。")
        return [dict(item) for item in payload]
    raise InstitutionDataError("当前数据不是可解析的表格格式。")


def _normalized_header(value: Any) -> str:
    return re.sub(r"[\s_\-()/（）]+", "", str(value or "").strip().lower())


def normalize_records(
    dataset_type: str,
    rows: Iterable[dict[str, Any]],
    field_mapping: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Map source headers to canonical fields; explicit mapping is source -> standard."""
    if dataset_type not in TABULAR_DATASETS:
        raise InstitutionDataError("该数据类型不使用表格字段映射。")
    explicit = {_normalized_header(source): target for source, target in (field_mapping or {}).items()}
    aliases = {
        _normalized_header(alias): canonical
        for canonical, values in FIELD_ALIASES[dataset_type].items()
        for alias in values
    }
    canonical_fields = set(FIELD_ALIASES[dataset_type])
    invalid_targets = sorted(set(explicit.values()) - canonical_fields)
    if invalid_targets:
        raise InstitutionDataError(
            f"字段映射包含当前数据类型不支持的标准字段：{'、'.join(invalid_targets)}；"
            f"允许：{'、'.join(sorted(canonical_fields))}。"
        )
    output: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        normalized: dict[str, Any] = {"_row_number": row_number, "_source": dict(row)}
        for source, value in row.items():
            source_key = _normalized_header(source)
            target = explicit.get(source_key) or aliases.get(source_key)
            if target in canonical_fields:
                normalized[target] = value.strip() if isinstance(value, str) else value
        output.append(normalized)
    if not output:
        raise InstitutionDataError("文件中没有可导入的数据行。")
    if not any(any(field in row for field in canonical_fields) for row in output):
        raise InstitutionDataError(
            f"未识别到{DATASET_LABELS[dataset_type]}标准字段；请确认数据类型，或填写“原始列名 -> 标准字段”映射。"
        )
    return output


def extract_vcf_samples(path: Path, compressed: bool) -> list[str]:
    opener = gzip.open if compressed else open
    try:
        with opener(path, "rt", encoding="utf-8", errors="replace") as handle:
            for line in handle:
                if line.startswith("#CHROM"):
                    columns = line.rstrip("\r\n").split("\t")
                    return [value.strip() for value in columns[9:] if value.strip()]
                if not line.startswith("#"):
                    break
    except (OSError, EOFError) as exc:
        raise InstitutionDataError("VCF/VCF.GZ 无法解压或读取。") from exc
    raise InstitutionDataError("VCF 缺少 #CHROM 表头。")


def extract_plink_samples(path: Path) -> list[str]:
    try:
        with zipfile.ZipFile(path) as archive:
            members = [item for item in archive.infolist() if not item.is_dir()]
            if len(members) > 2000 or sum(item.file_size for item in members) > 20 * 1024 * 1024 * 1024:
                raise InstitutionDataError("PLINK 压缩包解压后过大或文件过多。")
            if any(Path(item.filename).is_absolute() or ".." in Path(item.filename).parts for item in members):
                raise InstitutionDataError("PLINK 压缩包包含不安全路径。")
            fam_files = [item for item in members if item.filename.lower().endswith(".fam")]
            bed_stems = {Path(item.filename).stem for item in members if item.filename.lower().endswith(".bed")}
            bim_stems = {Path(item.filename).stem for item in members if item.filename.lower().endswith(".bim")}
            fam_stems = {Path(item.filename).stem for item in fam_files}
            common = bed_stems & bim_stems & fam_stems
            if len(common) != 1:
                raise InstitutionDataError("PLINK ZIP 必须包含一组同前缀 .bed/.bim/.fam 文件。")
            fam = next(item for item in fam_files if Path(item.filename).stem in common)
            content = _decode_text(archive.read(fam))
    except zipfile.BadZipFile as exc:
        raise InstitutionDataError("PLINK 文件不是有效 ZIP 压缩包。") from exc
    samples = []
    for line in content.splitlines():
        fields = line.split()
        if len(fields) >= 2:
            samples.append(fields[1])
    if not samples:
        raise InstitutionDataError("PLINK .fam 文件中没有样本。")
    return samples


def extract_literature_text(path: Path, suffix: str) -> tuple[str, list[str]]:
    warnings: list[str] = []
    if suffix == ".txt":
        return _decode_text(path.read_bytes()), warnings
    if suffix == ".docx":
        try:
            with zipfile.ZipFile(path) as archive:
                xml = _decode_text(archive.read("word/document.xml"))
            text_value = re.sub(r"<[^>]+>", " ", xml)
            return re.sub(r"\s+", " ", text_value).strip(), warnings
        except (zipfile.BadZipFile, KeyError) as exc:
            raise InstitutionDataError("DOCX 文件结构无效。") from exc
    if suffix == ".pdf":
        try:
            import fitz

            with fitz.open(path) as document:
                value = "\n".join(page.get_text("text") for page in document)
            if not value.strip():
                warnings.append("PDF 未提取到文本，原文件已保存，受影响功能：文献检索、证据引用。")
            return value, warnings
        except Exception as exc:
            raise InstitutionDataError(f"PDF 无法解析：{str(exc)[:120]}") from exc
    raise InstitutionDataError("文献格式不受支持。")


def parse_field_mapping(value: str | None) -> dict[str, str]:
    if not value:
        return {}
    try:
        payload = json.loads(value)
    except json.JSONDecodeError as exc:
        raise InstitutionDataError("字段映射必须是 JSON 对象。") from exc
    if not isinstance(payload, dict) or not all(isinstance(key, str) and isinstance(item, str) for key, item in payload.items()):
        raise InstitutionDataError("字段映射必须是“原始列名 -> 标准字段”字符串对象。")
    return payload


class MinioInstitutionStore:
    def __init__(self, settings: InstitutionDataSettings):
        self.settings = settings
        self._client: Any | None = None

    @property
    def client(self) -> Any:
        if self._client is None:
            if not self.settings.minio_secret_key:
                raise RuntimeError("MINIO_ROOT_PASSWORD 未配置。")
            try:
                from minio import Minio
            except ImportError as exc:
                raise RuntimeError("MinIO Python SDK 未安装。") from exc
            self._client = Minio(
                self.settings.minio_endpoint,
                access_key=self.settings.minio_access_key,
                secret_key=self.settings.minio_secret_key,
                secure=self.settings.minio_secure,
                region=self.settings.minio_region,
            )
        return self._client

    def ensure_private_bucket(self, bucket_name: str) -> None:
        if not self.client.bucket_exists(bucket_name):
            self.client.make_bucket(bucket_name, location=self.settings.minio_region)

    def put_file(self, bucket_name: str, object_key: str, upload: StagedUpload, content_type: str) -> None:
        self.ensure_private_bucket(bucket_name)
        self.client.fput_object(
            bucket_name,
            object_key,
            str(upload.path),
            content_type=content_type or "application/octet-stream",
            metadata={"sha256": upload.sha256, "original-file-name": quote(upload.file_name, safe="")},
        )


INSTITUTION_DATABASE_DDL = (
    """
    CREATE TABLE IF NOT EXISTS ingest_batch (
      id VARCHAR(36) PRIMARY KEY,
      institution_id VARCHAR(80) NOT NULL,
      project_id VARCHAR(36) NOT NULL,
      dataset_type VARCHAR(40) NOT NULL,
      source_file_name VARCHAR(500) NOT NULL,
      object_bucket VARCHAR(200) NOT NULL,
      object_key TEXT NOT NULL,
      file_sha256 VARCHAR(64) NOT NULL,
      file_size_bytes BIGINT NOT NULL,
      template_version_id VARCHAR(36),
      field_mapping JSONB NOT NULL DEFAULT '{}'::jsonb,
      status VARCHAR(30) NOT NULL,
      row_count INTEGER NOT NULL DEFAULT 0,
      issue_count INTEGER NOT NULL DEFAULT 0,
      created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
      completed_at TIMESTAMPTZ
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS data_entity (
      institution_id VARCHAR(80) NOT NULL,
      project_id VARCHAR(36) NOT NULL,
      entity_type VARCHAR(50) NOT NULL,
      entity_key VARCHAR(300) NOT NULL,
      source_batch_id VARCHAR(36) NOT NULL REFERENCES ingest_batch(id),
      payload JSONB NOT NULL DEFAULT '{}'::jsonb,
      created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
      updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
      PRIMARY KEY (institution_id, project_id, entity_type, entity_key)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS data_relation (
      id VARCHAR(36) PRIMARY KEY,
      institution_id VARCHAR(80) NOT NULL,
      project_id VARCHAR(36) NOT NULL,
      source_entity_type VARCHAR(50) NOT NULL,
      source_entity_key VARCHAR(300) NOT NULL,
      relation_type VARCHAR(80) NOT NULL,
      target_entity_type VARCHAR(50) NOT NULL,
      target_entity_key VARCHAR(300) NOT NULL,
      source_batch_id VARCHAR(36) NOT NULL REFERENCES ingest_batch(id),
      status VARCHAR(30) NOT NULL,
      message TEXT,
      created_at TIMESTAMPTZ NOT NULL DEFAULT now()
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS data_issue (
      id VARCHAR(36) PRIMARY KEY,
      institution_id VARCHAR(80) NOT NULL,
      project_id VARCHAR(36) NOT NULL,
      source_batch_id VARCHAR(36) NOT NULL REFERENCES ingest_batch(id),
      entity_type VARCHAR(50),
      entity_key VARCHAR(300),
      issue_type VARCHAR(50) NOT NULL,
      severity VARCHAR(30) NOT NULL,
      message TEXT NOT NULL,
      affected_features JSONB NOT NULL DEFAULT '[]'::jsonb,
      resolved BOOLEAN NOT NULL DEFAULT FALSE,
      created_at TIMESTAMPTZ NOT NULL DEFAULT now()
    )
    """,
    "CREATE INDEX IF NOT EXISTS ix_data_entity_project_key ON data_entity(project_id, entity_key)",
    "CREATE INDEX IF NOT EXISTS ix_data_relation_project_source ON data_relation(project_id, source_entity_key)",
    "CREATE INDEX IF NOT EXISTS ix_data_relation_project_target ON data_relation(project_id, target_entity_key)",
    "CREATE INDEX IF NOT EXISTS ix_data_issue_project_batch ON data_issue(project_id, source_batch_id)",
)


class InstitutionDatabaseManager:
    def __init__(self, settings: InstitutionDataSettings):
        self.settings = settings
        self._engines: dict[str, Engine] = {}

    def database_name(self, institution_code: str) -> str:
        return safe_identifier(institution_code, "longyun")

    def _target_url(self, database_name: str, *, migration: bool) -> str:
        source = self.settings.migration_database_url if migration else self.settings.application_database_url
        if not source:
            raise RuntimeError("MIGRATION_DATABASE_URL/DATABASE_URL 未配置。")
        return make_url(source).set(database=database_name).render_as_string(hide_password=False)

    def ensure_database(self, database_name: str) -> None:
        if not re.fullmatch(r"[a-z][a-z0-9_]{0,62}", database_name):
            raise RuntimeError("机构业务数据库名不安全。")
        role = self.settings.application_database_role
        if not re.fullmatch(r"[a-z_][a-z0-9_]{0,62}", role):
            raise RuntimeError("APP_DATABASE_ROLE 不安全。")
        admin_url = make_url(self.settings.migration_database_url).set(database="postgres")
        admin = create_engine(admin_url, isolation_level="AUTOCOMMIT", pool_pre_ping=True)
        try:
            with admin.connect() as connection:
                exists = connection.execute(
                    text("SELECT 1 FROM pg_database WHERE datname = :name"), {"name": database_name}
                ).scalar()
                if not exists:
                    connection.exec_driver_sql(f'CREATE DATABASE "{database_name}" OWNER "{role}"')
        finally:
            admin.dispose()
        migration_engine = create_engine(self._target_url(database_name, migration=True), pool_pre_ping=True)
        try:
            with migration_engine.begin() as connection:
                for statement in INSTITUTION_DATABASE_DDL:
                    connection.execute(text(statement))
                connection.exec_driver_sql(f'GRANT USAGE ON SCHEMA public TO "{role}"')
                connection.exec_driver_sql(f'GRANT SELECT, INSERT, UPDATE, DELETE ON ALL TABLES IN SCHEMA public TO "{role}"')
                connection.exec_driver_sql(f'GRANT USAGE, SELECT ON ALL SEQUENCES IN SCHEMA public TO "{role}"')
                connection.exec_driver_sql(
                    f'ALTER DEFAULT PRIVILEGES IN SCHEMA public '
                    f'GRANT SELECT, INSERT, UPDATE, DELETE ON TABLES TO "{role}"'
                )
                connection.exec_driver_sql(
                    f'ALTER DEFAULT PRIVILEGES IN SCHEMA public '
                    f'GRANT USAGE, SELECT ON SEQUENCES TO "{role}"'
                )
        finally:
            migration_engine.dispose()

    def engine(self, database_name: str) -> Engine:
        if database_name not in self._engines:
            self._engines[database_name] = create_engine(
                self._target_url(database_name, migration=False), pool_pre_ping=True
            )
        return self._engines[database_name]


def _entity_exists(connection: Any, institution_id: str, project_id: str, entity_type: str, entity_key: str) -> bool:
    return bool(connection.execute(text("""
        SELECT 1 FROM data_entity
        WHERE institution_id=:institution_id AND project_id=:project_id
          AND entity_type=:entity_type AND entity_key=:entity_key
    """), {
        "institution_id": institution_id,
        "project_id": project_id,
        "entity_type": entity_type,
        "entity_key": entity_key,
    }).scalar())


def _upsert_entity(
    connection: Any,
    *,
    institution_id: str,
    project_id: str,
    entity_type: str,
    entity_key: str,
    batch_id: str,
    payload: dict[str, Any],
) -> None:
    connection.execute(text("""
        INSERT INTO data_entity (
          institution_id, project_id, entity_type, entity_key, source_batch_id, payload
        ) VALUES (
          :institution_id, :project_id, :entity_type, :entity_key, :batch_id, CAST(:payload AS JSONB)
        )
        ON CONFLICT (institution_id, project_id, entity_type, entity_key)
        DO UPDATE SET source_batch_id=EXCLUDED.source_batch_id,
                      payload=EXCLUDED.payload, updated_at=now()
    """), {
        "institution_id": institution_id,
        "project_id": project_id,
        "entity_type": entity_type,
        "entity_key": entity_key,
        "batch_id": batch_id,
        "payload": json.dumps(payload, ensure_ascii=False, default=str),
    })
    # Imports do not have to arrive in dependency order. When a referenced
    # germplasm/environment is imported later, repair inbound unresolved
    # relations and resolve the source issue once no unresolved target remains.
    repaired_sources = connection.execute(text("""
        UPDATE data_relation
        SET status='linked', message=NULL
        WHERE institution_id=:institution_id AND project_id=:project_id
          AND target_entity_type=:entity_type AND target_entity_key=:entity_key
          AND status='unresolved'
        RETURNING source_entity_type, source_entity_key
    """), {
        "institution_id": institution_id,
        "project_id": project_id,
        "entity_type": entity_type,
        "entity_key": entity_key,
    }).all()
    for source_type, source_key in set(repaired_sources):
        connection.execute(text("""
            UPDATE data_issue AS issue
            SET resolved=TRUE
            WHERE issue.institution_id=:institution_id
              AND issue.project_id=:project_id
              AND issue.issue_type='association_anomaly'
              AND issue.entity_type=:source_type
              AND issue.entity_key=:source_key
              AND issue.resolved=FALSE
              AND NOT EXISTS (
                SELECT 1 FROM data_relation AS relation
                WHERE relation.institution_id=issue.institution_id
                  AND relation.project_id=issue.project_id
                  AND relation.source_entity_type=:source_type
                  AND relation.source_entity_key=:source_key
                  AND relation.status='unresolved'
              )
        """), {
            "institution_id": institution_id,
            "project_id": project_id,
            "source_type": source_type,
            "source_key": source_key,
        })


def _add_issue(
    connection: Any,
    *,
    institution_id: str,
    project_id: str,
    batch_id: str,
    dataset_type: str,
    entity_type: str | None,
    entity_key: str | None,
    issue_type: str,
    severity: str,
    message: str,
) -> dict[str, Any]:
    issue = {
        "id": str(uuid.uuid4()),
        "issue_type": issue_type,
        "severity": severity,
        "message": message,
        "entity_type": entity_type,
        "entity_key": entity_key,
        "affected_features": IMPACTED_FEATURES[dataset_type],
    }
    connection.execute(text("""
        INSERT INTO data_issue (
          id, institution_id, project_id, source_batch_id, entity_type, entity_key,
          issue_type, severity, message, affected_features
        ) VALUES (
          :id, :institution_id, :project_id, :batch_id, :entity_type, :entity_key,
          :issue_type, :severity, :message, CAST(:affected_features AS JSONB)
        )
    """), {
        **issue,
        "institution_id": institution_id,
        "project_id": project_id,
        "batch_id": batch_id,
        "affected_features": json.dumps(issue["affected_features"], ensure_ascii=False),
    })
    return issue


def _add_relation(
    connection: Any,
    *,
    institution_id: str,
    project_id: str,
    batch_id: str,
    source_type: str,
    source_key: str,
    relation_type: str,
    target_type: str,
    target_key: str,
    target_exists: bool,
) -> None:
    connection.execute(text("""
        INSERT INTO data_relation (
          id, institution_id, project_id, source_entity_type, source_entity_key,
          relation_type, target_entity_type, target_entity_key, source_batch_id, status, message
        ) VALUES (
          :id, :institution_id, :project_id, :source_type, :source_key,
          :relation_type, :target_type, :target_key, :batch_id, :status, :message
        )
    """), {
        "id": str(uuid.uuid4()),
        "institution_id": institution_id,
        "project_id": project_id,
        "source_type": source_type,
        "source_key": source_key,
        "relation_type": relation_type,
        "target_type": target_type,
        "target_key": target_key,
        "batch_id": batch_id,
        "status": "linked" if target_exists else "unresolved",
        "message": None if target_exists else f"未找到 {target_type}:{target_key}",
    })


def import_into_institution_database(
    engine: Engine,
    *,
    batch_id: str,
    institution_id: str,
    project_id: str,
    dataset_type: str,
    upload: StagedUpload,
    bucket_name: str,
    object_key: str,
    field_mapping: dict[str, str] | None = None,
    template_version_id: str | None = None,
) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    created_entities = 0
    with engine.begin() as connection:
        connection.execute(text("""
            INSERT INTO ingest_batch (
              id, institution_id, project_id, dataset_type, source_file_name,
              object_bucket, object_key, file_sha256, file_size_bytes, status
              , template_version_id, field_mapping
            ) VALUES (
              :id, :institution_id, :project_id, :dataset_type, :source_file_name,
              :bucket, :object_key, :sha256, :size_bytes, 'processing',
              :template_version_id, CAST(:field_mapping AS JSONB)
            )
        """), {
            "id": batch_id,
            "institution_id": institution_id,
            "project_id": project_id,
            "dataset_type": dataset_type,
            "source_file_name": upload.file_name,
            "bucket": bucket_name,
            "object_key": object_key,
            "sha256": upload.sha256,
            "size_bytes": upload.size_bytes,
            "template_version_id": template_version_id,
            "field_mapping": json.dumps(field_mapping or {}, ensure_ascii=False),
        })

        if dataset_type in TABULAR_DATASETS:
            rows = normalize_records(
                dataset_type,
                parse_tabular_file(upload.path, upload.suffix, dataset_type),
                field_mapping,
            )
            for row in rows:
                missing = [field for field in REQUIRED_FIELDS[dataset_type] if row.get(field) in (None, "")]
                provisional_key = str(row.get(next(iter(REQUIRED_FIELDS[dataset_type]), ""), f"row-{row['_row_number']}"))
                if missing:
                    issues.append(_add_issue(
                        connection,
                        institution_id=institution_id,
                        project_id=project_id,
                        batch_id=batch_id,
                        dataset_type=dataset_type,
                        entity_type=dataset_type,
                        entity_key=provisional_key,
                        issue_type="missing_data",
                        severity="blocked",
                        message=f"第 {row['_row_number']} 行缺少必填字段：{'、'.join(missing)}。",
                    ))
                    continue
                if dataset_type == "germplasm":
                    entity_type, entity_key = "germplasm", str(row["germplasm_id"])
                elif dataset_type == "environment":
                    entity_type, entity_key = "environment", str(row["environment_id"])
                elif dataset_type == "pedigree":
                    entity_type, entity_key = "pedigree", str(row["child_id"])
                else:
                    seed = "|".join(str(row.get(key, "")) for key in ("germplasm_id", "trial_id", "trait_code", "environment_id", "_row_number"))
                    entity_type = "phenotype_observation"
                    entity_key = str(row.get("observation_id") or hashlib.sha256(seed.encode("utf-8")).hexdigest()[:24])
                payload = {key: value for key, value in row.items() if not key.startswith("_")}
                _upsert_entity(
                    connection,
                    institution_id=institution_id,
                    project_id=project_id,
                    entity_type=entity_type,
                    entity_key=entity_key,
                    batch_id=batch_id,
                    payload=payload,
                )
                created_entities += 1

                relations: list[tuple[str, str, str]] = []
                if dataset_type == "pedigree":
                    relations.append(("child_material", "germplasm", str(row["child_id"])))
                    for field, relation in (("female_parent_id", "female_parent"), ("male_parent_id", "male_parent")):
                        if row.get(field):
                            relations.append((relation, "germplasm", str(row[field])))
                        else:
                            issues.append(_add_issue(
                                connection,
                                institution_id=institution_id,
                                project_id=project_id,
                                batch_id=batch_id,
                                dataset_type=dataset_type,
                                entity_type=entity_type,
                                entity_key=entity_key,
                                issue_type="missing_data",
                                severity="warning",
                                message=f"系谱 {entity_key} 缺少 {field}，系谱图将不完整。",
                            ))
                elif dataset_type == "phenotype":
                    relations.append(("observed_material", "germplasm", str(row["germplasm_id"])))
                    if row.get("environment_id"):
                        relations.append(("observed_environment", "environment", str(row["environment_id"])))
                for relation_type, target_type, target_key in relations:
                    exists = _entity_exists(connection, institution_id, project_id, target_type, target_key)
                    _add_relation(
                        connection,
                        institution_id=institution_id,
                        project_id=project_id,
                        batch_id=batch_id,
                        source_type=entity_type,
                        source_key=entity_key,
                        relation_type=relation_type,
                        target_type=target_type,
                        target_key=target_key,
                        target_exists=exists,
                    )
                    if not exists:
                        issues.append(_add_issue(
                            connection,
                            institution_id=institution_id,
                            project_id=project_id,
                            batch_id=batch_id,
                            dataset_type=dataset_type,
                            entity_type=entity_type,
                            entity_key=entity_key,
                            issue_type="association_anomaly",
                            severity="warning",
                            message=f"{entity_type}:{entity_key} 无法关联 {target_type}:{target_key}。",
                        ))
        elif dataset_type == "genotype":
            samples = extract_plink_samples(upload.path) if upload.suffix == ".zip" else extract_vcf_samples(upload.path, upload.suffix == ".vcf.gz")
            dataset_key = upload.sha256[:24]
            _upsert_entity(
                connection,
                institution_id=institution_id,
                project_id=project_id,
                entity_type="genotype_dataset",
                entity_key=dataset_key,
                batch_id=batch_id,
                payload={"format": upload.suffix, "sample_count": len(samples), "file_sha256": upload.sha256},
            )
            created_entities += 1
            for sample in samples:
                _upsert_entity(
                    connection,
                    institution_id=institution_id,
                    project_id=project_id,
                    entity_type="genotype_sample",
                    entity_key=sample,
                    batch_id=batch_id,
                    payload={"dataset_id": dataset_key, "sample_id": sample},
                )
                created_entities += 1
                exists = _entity_exists(connection, institution_id, project_id, "germplasm", sample)
                _add_relation(
                    connection,
                    institution_id=institution_id,
                    project_id=project_id,
                    batch_id=batch_id,
                    source_type="genotype_sample",
                    source_key=sample,
                    relation_type="sample_material",
                    target_type="germplasm",
                    target_key=sample,
                    target_exists=exists,
                )
                if not exists:
                    issues.append(_add_issue(
                        connection,
                        institution_id=institution_id,
                        project_id=project_id,
                        batch_id=batch_id,
                        dataset_type=dataset_type,
                        entity_type="genotype_sample",
                        entity_key=sample,
                        issue_type="association_anomaly",
                        severity="warning",
                        message=f"基因型样本 {sample} 未关联到同编号种质材料。",
                    ))
        else:
            extracted_text, parser_warnings = extract_literature_text(upload.path, upload.suffix)
            entity_key = upload.sha256[:24]
            _upsert_entity(
                connection,
                institution_id=institution_id,
                project_id=project_id,
                entity_type="literature_document",
                entity_key=entity_key,
                batch_id=batch_id,
                payload={
                    "file_name": upload.file_name,
                    "file_sha256": upload.sha256,
                    "text": extracted_text[:2_000_000],
                    "parser_warnings": parser_warnings,
                },
            )
            created_entities += 1
            for warning in parser_warnings:
                issues.append(_add_issue(
                    connection,
                    institution_id=institution_id,
                    project_id=project_id,
                    batch_id=batch_id,
                    dataset_type=dataset_type,
                    entity_type="literature_document",
                    entity_key=entity_key,
                    issue_type="missing_data",
                    severity="warning",
                    message=warning,
                ))

        connection.execute(text("""
            UPDATE ingest_batch
            SET status=:status, row_count=:row_count, issue_count=:issue_count, completed_at=now()
            WHERE id=:batch_id
        """), {
            "status": "completed_with_issues" if issues else "completed",
            "row_count": created_entities,
            "issue_count": len(issues),
            "batch_id": batch_id,
        })
    return {
        "status": "completed_with_issues" if issues else "completed",
        "entity_count": created_entities,
        "issue_count": len(issues),
        "issues": issues,
    }


def trace_entity(engine: Engine, institution_id: str, project_id: str, entity_key: str) -> dict[str, Any]:
    with engine.connect() as connection:
        entities = [dict(row) for row in connection.execute(text("""
            SELECT entity_type, entity_key, payload, source_batch_id, created_at, updated_at
            FROM data_entity
            WHERE institution_id=:institution_id AND project_id=:project_id AND entity_key=:entity_key
            ORDER BY entity_type
        """), {"institution_id": institution_id, "project_id": project_id, "entity_key": entity_key}).mappings()]
        relations = [dict(row) for row in connection.execute(text("""
            SELECT source_entity_type, source_entity_key, relation_type,
                   target_entity_type, target_entity_key, status, message, source_batch_id
            FROM data_relation
            WHERE institution_id=:institution_id AND project_id=:project_id
              AND (source_entity_key=:entity_key OR target_entity_key=:entity_key)
            ORDER BY created_at
        """), {"institution_id": institution_id, "project_id": project_id, "entity_key": entity_key}).mappings()]
        issues = [dict(row) for row in connection.execute(text("""
            SELECT issue_type, severity, message, affected_features, resolved, source_batch_id, created_at
            FROM data_issue
            WHERE institution_id=:institution_id AND project_id=:project_id
              AND (
                entity_key=:entity_key
                OR entity_key IN (
                  SELECT source_entity_key FROM data_relation
                  WHERE institution_id=:institution_id AND project_id=:project_id
                    AND (source_entity_key=:entity_key OR target_entity_key=:entity_key)
                )
              )
            ORDER BY created_at DESC
        """), {"institution_id": institution_id, "project_id": project_id, "entity_key": entity_key}).mappings()]
    return {"entity_key": entity_key, "entities": entities, "relations": relations, "issues": issues}


def list_batches(engine: Engine, institution_id: str, project_id: str, limit: int = 100) -> list[dict[str, Any]]:
    with engine.connect() as connection:
        return [dict(row) for row in connection.execute(text("""
            SELECT id, dataset_type, source_file_name, object_bucket, object_key,
                   file_sha256, file_size_bytes, status, row_count, issue_count,
                   template_version_id, field_mapping, created_at, completed_at
            FROM ingest_batch
            WHERE institution_id=:institution_id AND project_id=:project_id
            ORDER BY created_at DESC LIMIT :limit
        """), {"institution_id": institution_id, "project_id": project_id, "limit": limit}).mappings()]


def default_access_policy(institution_id: str, bucket_name: str) -> dict[str, Any]:
    return {
        "version": "2026-08-26",
        "effect": "private",
        "principal": f"institution:{institution_id}",
        "bucket": bucket_name,
        "actions": ["s3:GetObject", "s3:PutObject", "s3:ListBucket"],
        "enforcedBy": ["application-institution-boundary", "minio-private-bucket"],
    }


def object_key_for(institution_id: str, project_id: str, dataset_type: str, batch_id: str, file_name: str) -> str:
    date = datetime.now(timezone.utc).strftime("%Y/%m/%d")
    return f"{safe_identifier(institution_id, 'institution')}/{project_id}/{dataset_type}/{date}/{batch_id}/{safe_object_name(file_name)}"
